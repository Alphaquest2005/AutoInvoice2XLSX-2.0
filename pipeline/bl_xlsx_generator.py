#!/usr/bin/env python3
"""
BL (Bill of Lading) XLSX Generator

Generates CARICOM-format XLSX files for BL shipment processing.
All column definitions, styles, formulas, and totals structure are loaded
dynamically from config/columns.yaml and config/grouping.yaml via spec_loader.

Supports document type configuration from config/document_types.json:
  - grouping=true  (e.g. 4000-000): group header + detail row pairs
  - grouping=false (e.g. 7400-000): one row per item, no grouping

Usage:
    from bl_xlsx_generator import generate_bl_xlsx
    generate_bl_xlsx(invoice_data, matched_items, supplier_name,
                     supplier_info, output_path, document_type="7400-000")
"""

import json
import os
import logging
import re
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None
    Comment = None

logger = logging.getLogger(__name__)

import sys
import sqlite3
from pipeline.config_loader import (
    load_columns,
    load_country_codes,
    load_file_paths,
    load_financial_constants,
    load_hs_structure,
    load_library_enums,
    load_patterns,
    load_pipeline,
    load_validation_tolerances,
    load_xlsx_labels,
    load_document_types,
)

_pipeline_dir = os.path.dirname(os.path.abspath(__file__))
if _pipeline_dir not in sys.path:
    sys.path.insert(0, _pipeline_dir)

# Ensure src/ is importable so we can pull the SSOT CET category service.
_FILE_PATHS = load_file_paths()
_src_dir = os.path.join(
    os.path.dirname(_pipeline_dir), _FILE_PATHS["source_dirs"]["src"]
)
if _src_dir not in sys.path:
    sys.path.insert(0, _src_dir)

from spec_loader import get_spec
from autoinvoice.domain.services.cet_category import category_for as _category_for

# ─── Load spec + config (singletons) ──────────────────────
_spec = get_spec()
_FIN = load_financial_constants()
_COLUMNS_CFG = load_columns()
_LIBENUMS = load_library_enums()
_PATTERNS = load_patterns()
_PIPELINE_CFG = load_pipeline()
_TOLERANCES = load_validation_tolerances()
_LABELS = load_xlsx_labels()
_DOCTYPES = load_document_types()
_COUNTRY = load_country_codes()
_DEFAULT_COUNTRY = _COUNTRY["default_origin"]
_BASE_CURRENCY = _FIN["base_currency"]
_HS = load_hs_structure()
_HS_SLICE_HEADING = _HS["slice"]["heading"]
_HS_SLICE_CHAPTER = _HS["slice"]["chapter"]
_HS_TARIFF_LEN = _HS["tariff_length"]
# Primary zero-tariff sentinel: "00000000". Taken from hs_structure.yaml
# so the padding strings below stay derived, not hardcoded.
_HS_ZERO_CODE = next(
    (s for s in _HS["zero_sentinels"] if isinstance(s, str) and s),
    "0" * _HS_TARIFF_LEN,  # magic-ok: zero-digit padding char
)
# Padding strings: zero-fill for chapter/heading level lookups.
_HS_ZERO_PAD_HEADING = _HS_ZERO_CODE[_HS_SLICE_HEADING:]     # "0000"
_HS_ZERO_PAD_CHAPTER = _HS_ZERO_CODE[_HS_SLICE_CHAPTER:]     # "000000"
_STYLES = _COLUMNS_CFG.get("styles", {})
_FILL_TYPE_SOLID = _LIBENUMS["openpyxl"]["fill_type"]["SOLID"]
_SHEET_CFG = _LIBENUMS["openpyxl"]["sheet"]
_NUMFMT = _LIBENUMS["openpyxl"]["number_format"]
_ALIGN_CENTER = _LIBENUMS["openpyxl"]["alignment"]["CENTER"]
_COMMENT_AUTHOR = _COLUMNS_CFG["comment_author"]

# ─── Populate-on trigger tokens (match columns.yaml populate_on values) ───
# These tokens are compared against the populate_on string on each column
# config entry in columns.yaml. Keep in sync with columns.yaml.
_POPULATE_FIRST_GROUP = "first_group_per_invoice"  # magic-ok: columns.yaml populate_on token

# ─── Formula placeholder tokens ───────────────────────────
# Brace-wrapped placeholder strings that appear in columns.yaml
# formula_templates and grouping.yaml totals_section patterns. Used by
# the hand-rolled replacer in _resolve_totals_formula and the ungrouped
# totals writer. Keep names in sync with those YAML values.
_PH_ALL_P_REFS          = "{all_P_refs}"          # magic-ok: formula_templates placeholder
_PH_FIRST_ROW           = "{first_row}"           # magic-ok: formula_templates placeholder
_PH_LAST_DATA_ROW       = "{last_data_row}"       # magic-ok: formula_templates placeholder
_PH_SUBTOTAL_ROW        = "{subtotal_row}"        # magic-ok: formula_templates placeholder
_PH_SUBTOTAL_GROUPED    = "{subtotal_grouped_row}"  # magic-ok: formula_templates placeholder
_PH_SUBTOTAL_DETAILS    = "{subtotal_details_row}"  # magic-ok: formula_templates placeholder
_PH_ADJUSTMENTS_ROW     = "{adjustments_row}"     # magic-ok: formula_templates placeholder
_PH_NET_TOTAL_ROW       = "{net_total_row}"       # magic-ok: formula_templates placeholder
_PH_GROUP_P_REFS        = "{group_P_refs}"        # magic-ok: formula_templates placeholder
_PH_GROUP_Q_REFS        = "{group_Q_refs}"        # magic-ok: formula_templates placeholder

# Value-interpolation tokens used in grouping.yaml totals_section patterns.
# When a totals row's column_P is exactly one of these, the resolver returns
# the corresponding numeric value from invoice_data rather than a formula.
_VAL_FREIGHT         = "${freight_value}"     # magic-ok: grouping.yaml totals value token
_VAL_INSURANCE       = "${insurance_value}"   # magic-ok: grouping.yaml totals value token
_VAL_OTHER_COST      = "${other_cost_value}"  # magic-ok: grouping.yaml totals value token
_VAL_DEDUCTION       = "${deduction_value}"   # magic-ok: grouping.yaml totals value token
_VAL_INVOICE_TOTAL   = "${invoice_total}"     # magic-ok: grouping.yaml totals value token

# Row-type tokens from grouping.yaml totals_section entries.
_ROW_TYPE_BLANK      = "blank_row"            # magic-ok: grouping.yaml row type token

# ─── Build openpyxl style objects from spec ────────────────
def _build_styles():
    """Build openpyxl Font/Fill/Border objects from spec config."""
    hs = _spec.header_style
    gs = _spec.group_style
    ds = _spec.detail_style

    header_fill = PatternFill(
        start_color=hs['fill_color'],
        end_color=hs['fill_color'],
        fill_type=_FILL_TYPE_SOLID,
    )
    header_font = Font(
        bold=hs['font_bold'],
        color=hs['font_color'],
        size=hs['font_size'],
    )
    group_fill = PatternFill(
        start_color=gs['fill_color'],
        end_color=gs['fill_color'],
        fill_type=gs.get('fill_type', _FILL_TYPE_SOLID),
    )
    group_font = Font(
        bold=gs['font_bold'],
        size=gs['font_size'],
    )
    detail_font = Font(
        bold=ds['font_bold'],
        size=ds['font_size'],
    )
    bold_font = Font(
        bold=True,
        size=_spec.totals_default_formatting['font_size'],
    )
    border = Border(
        left=Side(style=_spec.border_style),
        right=Side(style=_spec.border_style),
        top=Side(style=_spec.border_style),
        bottom=Side(style=_spec.border_style),
    )
    return header_fill, header_font, group_fill, group_font, detail_font, bold_font, border

HEADER_FILL, HEADER_FONT, GROUP_FILL, GROUP_FONT, DETAIL_FONT, BOLD_FONT, THIN_BORDER = _build_styles()

# Uncertainty markers — applied to cells whose value was recovered from OCR
# via the orphan-price scan or absorbed into ADJUSTMENTS.  Visible so a
# reviewer immediately sees which numbers are reconstructed vs directly
# extracted from the invoice text.
_UNCERTAIN_STYLE = _STYLES.get("uncertain", {})
_RECOVERED_STYLE = _STYLES.get("recovered", {})
UNCERTAIN_FILL = PatternFill(
    start_color=_UNCERTAIN_STYLE.get("fill_color"),
    end_color=_UNCERTAIN_STYLE.get("fill_color"),
    fill_type=_UNCERTAIN_STYLE.get("fill_type", _FILL_TYPE_SOLID),
)
RECOVERED_FILL = PatternFill(
    start_color=_RECOVERED_STYLE.get("fill_color"),
    end_color=_RECOVERED_STYLE.get("fill_color"),
    fill_type=_RECOVERED_STYLE.get("fill_type", _FILL_TYPE_SOLID),
)


# ─── CET description & rate cache ──────────────────────────
_cet_desc_cache: Dict[str, str] = {}
_cet_rate_cache: Dict[str, Optional[float]] = {}
_cet_db_loaded = False

# XCD exchange rate + statutory CARICOM duty rates.  All loaded from
# config/financial_constants.yaml (SSOT). Change the YAML, not this file.
XCD_RATE = _FIN["xcd_rate"]
CSC_RATE = _FIN["csc_rate"]
VAT_RATE = _FIN["vat_rate"]
_DEFAULT_CET_RATE = _FIN["default_cet_rate"]
_PCT = _FIN["pct_conversion"]
_COMPOSITE_CET_COEF = _FIN["composite_duty"]["cet_coefficient"]
_COMPOSITE_BASE_COEF = _FIN["composite_duty"]["base_coefficient"]

def _parse_cet_rate(rate_str: Optional[str]) -> Optional[float]:
    """Parse a duty_rate string from the CET database into a float (0.0–1.0).
    Returns None if unparseable."""
    if not rate_str:
        return None
    rate_str = rate_str.strip()
    if rate_str.lower() in _PIPELINE_CFG["cet_db"]["free_rate_literals"]:
        return 0.0
    # Letter categories (A, C, D) — cannot resolve to a numeric rate
    if rate_str.isalpha():
        return None
    # Range like "0-5%" — take the upper bound as conservative estimate
    if '-' in rate_str and '%' in rate_str:
        try:
            upper = rate_str.replace('%', '').split('-')[1].strip()
            return float(upper) / float(_PCT)
        except (ValueError, IndexError):
            return None
    # Decimal rate (0.2 = 20%, 0.05 = 5%)
    try:
        val = float(rate_str)
        # Values > 1 are percentages (e.g. "20"), values <= 1 are already ratios
        return val if val <= 1.0 else val / float(_PCT)
    except ValueError:
        return None

def _load_cet_descriptions() -> None:
    """Load CET code descriptions and duty rates from the database."""
    global _cet_db_loaded
    if _cet_db_loaded:
        return
    db_path = os.path.join(
        os.path.dirname(_pipeline_dir),
        _FILE_PATHS["references"]["cet_database"],
    )
    _cet_query = _PIPELINE_CFG["cet_db"]["query"]
    _ro_mode = _PIPELINE_CFG["cet_db"]["ro_uri_mode"]
    if not os.path.exists(db_path):
        logger.warning(f"[CET] Database not found: {db_path}")
        _cet_db_loaded = True
        return
    try:
        conn = sqlite3.connect(f'file:{db_path}?mode={_ro_mode}', uri=True)
        rows = conn.execute(_cet_query).fetchall()
        for code, desc, rate in rows:
            if desc:
                _cet_desc_cache[code] = desc
            _cet_rate_cache[code] = _parse_cet_rate(rate)
        conn.close()
        logger.info(f"[CET] Loaded {len(_cet_desc_cache)} descriptions, {sum(1 for v in _cet_rate_cache.values() if v is not None)} rates from database")
        _cet_db_loaded = True
    except Exception as e:
        logger.warning(f"[CET] Read-only connect failed: {e}, trying temp copy")
        try:
            import shutil, tempfile
            tmp = os.path.join(
                tempfile.gettempdir(),
                _FILE_PATHS["references"]["cet_database_temp_name"],
            )
            shutil.copy2(db_path, tmp)
            conn = sqlite3.connect(tmp)
            rows = conn.execute(_cet_query).fetchall()
            for code, desc, rate in rows:
                if desc:
                    _cet_desc_cache[code] = desc
                _cet_rate_cache[code] = _parse_cet_rate(rate)
            conn.close()
            logger.info(f"[CET] Loaded {len(_cet_desc_cache)} descriptions from temp copy")
            _cet_db_loaded = True
        except Exception as e2:
            logger.error(f"[CET] Temp copy also failed: {e2}")

def get_cet_category(tariff_code: str) -> str:
    """Look up the CET description for a tariff code.

    Delegates to :func:`autoinvoice.domain.services.cet_category.category_for`
    — the SSOT implementation of the leaf-→subheading-→heading-→chapter walk.
    This module only supplies the in-memory description cache.
    """
    _load_cet_descriptions()
    return _category_for(tariff_code, _cet_desc_cache)


def get_cet_rate(tariff_code: str) -> Optional[float]:
    """Look up the CET duty rate for a tariff code. Returns rate as 0.0–1.0.
    Falls back through heading → chapter → prefix. Returns 0.20 default for
    consumer goods if not found (most personal imports are 20% CET)."""
    _load_cet_descriptions()
    if not tariff_code or tariff_code == _HS_ZERO_CODE:
        return _DEFAULT_CET_RATE  # default consumer goods rate

    # Exact match
    rate = _cet_rate_cache.get(tariff_code)
    if rate is not None:
        return rate

    # Heading level (first 4 digits + 0000)
    heading = tariff_code[:_HS_SLICE_HEADING] + _HS_ZERO_PAD_HEADING
    rate = _cet_rate_cache.get(heading)
    if rate is not None:
        return rate

    # Chapter level (first 2 digits + 000000)
    chapter = tariff_code[:_HS_SLICE_CHAPTER] + _HS_ZERO_PAD_CHAPTER
    rate = _cet_rate_cache.get(chapter)
    if rate is not None:
        return rate

    # Prefix search — find any code with same heading-length prefix that has a rate
    prefix = tariff_code[:_HS_SLICE_HEADING]
    for code, r in _cet_rate_cache.items():
        if code.startswith(prefix) and r is not None:
            return r

    # Default: 20% for consumer goods (most common for personal imports)
    return _DEFAULT_CET_RATE


def calculate_duties(cif_usd: float, cet_rate: float,
                     customs_freight: float = 0, insurance: float = 0) -> Dict:
    """Calculate ASYCUDA duties for a given CIF value.

    Returns dict with: cif_usd, cif_xcd, cet, csc, vat, total_duties,
    effective_rate, cet_rate_used.
    """
    cif_total_usd = round(cif_usd + customs_freight + insurance, 2)
    cif_xcd = round(cif_total_usd * XCD_RATE, 2)

    cet = round(cif_xcd * cet_rate, 2)
    csc = round(cif_xcd * CSC_RATE, 2)
    vat = round((cif_xcd + cet + csc) * VAT_RATE, 2)
    total_duties = round(cet + csc + vat, 2)
    effective_rate = total_duties / cif_xcd if cif_xcd else 0

    return {
        'cif_usd': cif_total_usd,
        'cif_xcd': cif_xcd,
        'cet_rate': cet_rate,
        'cet': cet,
        'csc_rate': CSC_RATE,
        'csc': csc,
        'vat_rate': VAT_RATE,
        'vat': vat,
        'total_duties': total_duties,
        'effective_rate': effective_rate,
    }


def _normalize_date(date_str: str) -> str:
    """Normalize date string to the format specified in columns.yaml (YYYY-MM-DD)."""
    if not date_str or not isinstance(date_str, str):
        return date_str or ''
    date_str = date_str.strip()
    iso_regex = _PATTERNS["date_iso_anchored"]
    us_regex = _PATTERNS["date_us_anchored"]
    out_fmt = _PATTERNS["date_output_format"]
    # Already in YYYY-MM-DD
    if re.match(iso_regex, date_str):
        return date_str
    from datetime import datetime
    # Try common date formats, output as YYYY-MM-DD per spec
    for fmt in _PATTERNS["date_parse_formats"]:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime(out_fmt)
        except ValueError:
            continue
    # Handle M/D/YYYY, M/DD/YYYY, MM/D/YYYY
    m = re.match(us_regex, date_str)
    if m:
        from datetime import datetime
        try:
            dt = datetime(int(m.group(3)), int(m.group(1)), int(m.group(2)))  # magic-ok: regex group index (year)
            return dt.strftime(out_fmt)
        except ValueError:
            pass
    return date_str


def _resolve_value(template: str, context: dict) -> Any:
    """Resolve a ${...} template string against a context dict.
    Returns the resolved value, or None if the template is null/empty.
    Handles prefix+template patterns like "    ${supplier_item_desc}".
    """
    if template is None:
        return None
    if not isinstance(template, str):
        return template
    if template.startswith('${'):
        # Simple template: ${key}
        key = template[2:-1] if template.endswith('}') else template[2:]
        return context.get(key)
    # Check for prefix + ${key} pattern (e.g. "    ${supplier_item_desc}")
    idx = template.find('${')
    if idx > 0:
        prefix = template[:idx]
        rest = template[idx:]
        key = rest[2:-1] if rest.endswith('}') else rest[2:]
        val = context.get(key, '')
        return f"{prefix}{val}"
    return template  # Literal value like "4000-000" or "USD"


def load_document_type_config(document_type: str) -> Dict:
    """Load document type settings via the SSOT config loader.

    Falls back to a best-effort default matching the JSON 'default' key if
    the doc_type is absent from the config.
    """
    try:
        doc_types = _DOCTYPES.get("document_types", {})
        if document_type in doc_types:
            return doc_types[document_type]
        logger.warning(
            f"Document type '{document_type}' not found in config, "
            f"using default settings"
        )
    except Exception as e:  # pragma: no cover — config loader is solid
        logger.warning(f"Failed to read document_types config: {e}")

    default_dt = _DOCTYPES.get("default")
    unknown_desc = _LABELS["defaults"]["document_type_unknown_description"]
    return {
        "grouping": document_type == default_dt,
        "description": unknown_desc,
    }


_DEFAULT_BL_DOC_TYPE = "7400-000"  # magic-ok: canonical doc_type literal, validated against load_document_types() on import
assert _DEFAULT_BL_DOC_TYPE in _DOCTYPES["document_types"], (
    "generate_bl_xlsx default document_type must exist in document_types.json"  # magic-ok: assertion message
)
_DEFAULT_REFERENCE_LABEL = _LABELS["reference"]["default_label"]


def generate_bl_xlsx(
    invoice_data: Dict,
    matched_items: List[Dict],
    supplier_name: str,
    supplier_info: Dict,
    output_path: str,
    document_type: str = _DEFAULT_BL_DOC_TYPE,
    reference_items: Optional[List[Dict]] = None,
    reference_label: str = _DEFAULT_REFERENCE_LABEL,
    reference_adjustments: Optional[Dict] = None,
) -> str:
    """Generate a CARICOM-format XLSX file for one BL invoice.

    Args:
        reference_items: Optional items from other declarations in the same
            invoice.  Rendered as a read-only reference section below the
            totals so the reviewer can see the full invoice and verify
            that the split is correct.  These items are NOT included in
            the subtotals or variance check.
        reference_label: Header label for the reference section.
    """
    if not openpyxl:
        raise ImportError("openpyxl is required. Run: pip install openpyxl")

    doc_config = load_document_type_config(document_type)
    grouping = doc_config.get('grouping', False)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _SHEET_CFG["default_title"]

    _write_headers(ws)

    if not matched_items:
        # Even with no main items, write reference section if provided
        # (e.g. declaration with no items split to it but needs full invoice context)
        if reference_items:
            _write_reference_section(ws, reference_items, reference_label,
                                     invoice_data, matched_items,
                                     reference_adjustments)
            for col_idx, width in _spec.column_widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = width
        wb.save(output_path)
        return output_path

    # Single-item invoices: a "group of 1" adds no information.
    # The grouped layout would emit a group-header row + 1 detail row +
    # SUBTOTAL (GROUPED)/SUBTOTAL (DETAILS)/GROUP VERIFICATION rows, all
    # showing the same number.  Render ungrouped for cleaner output.
    # The doc_type tag (e.g. 4000-000) is preserved on the detail row.
    if grouping and len(matched_items) == 1:
        grouping = False

    if grouping:
        row_num = _write_items_grouped(ws, matched_items, invoice_data,
                                       supplier_name, supplier_info,
                                       document_type)
        _write_subtotals_grouped(ws, row_num, len(matched_items), invoice_data)
    else:
        row_num = _write_items_ungrouped(ws, matched_items, invoice_data,
                                         supplier_name, supplier_info,
                                         document_type)
        _write_subtotals_ungrouped(ws, row_num, len(matched_items), invoice_data)

    # Duty estimation section — classification cross-check
    _write_duty_estimation_section(ws, matched_items, invoice_data)

    # Reference items section — appended after totals for context
    if reference_items:
        _write_reference_section(ws, reference_items, reference_label,
                                 invoice_data, matched_items,
                                 reference_adjustments)

    # Column widths from spec
    for col_idx, width in _spec.column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = _SHEET_CFG["freeze_panes_cell"]

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    logger.info(f"Generated BL XLSX: {output_path}")

    return output_path


# ─── Ungrouped Mode (e.g. 7400-000) ──────────────────────

def _write_items_ungrouped(
    ws, matched_items: List[Dict], invoice_data: Dict,
    supplier_name: str, supplier_info: Dict, document_type: str,
) -> int:
    """Write one row per item — no group/detail pairs."""
    row_num = 2
    col_count = _spec.col_count()

    # Pre-resolve column indices
    COL_F = _spec.col_index('F')    # magic-ok: columns.yaml schema key
    COL_Q = _spec.col_index('Q')    # magic-ok: columns.yaml schema key
    COL_R = _spec.col_index('R')    # magic-ok: columns.yaml schema key
    COL_AK = _spec.col_index('AK')  # magic-ok: columns.yaml schema key

    q_formula = _spec.formula_spec('Q')  # magic-ok: columns.yaml schema key (formula = "=O{row}*K{row}")
    r_formula = _spec.formula_spec('R')  # magic-ok: columns.yaml schema key (formula = "=P{row}-Q{row}")

    for idx, item in enumerate(matched_items):
        tariff_code = str(item.get('tariff_code', _HS_ZERO_CODE))
        cet_desc = get_cet_category(tariff_code)
        spec_cat = _spec.category_name(tariff_code)
        if spec_cat != _spec.category_default:
            category = spec_cat
        else:
            category = cet_desc or item.get('category', '') or spec_cat

        # Build context for value resolution
        ctx = _build_item_context(item, invoice_data, supplier_name, supplier_info,
                                  document_type, tariff_code, category,
                                  group_label=None, group_qty=None, group_total_cost=None)

        # Write columns using spec
        for letter, cfg in _spec._col_list:
            col_idx = cfg['index']
            populate_on = cfg.get('populate_on')

            # First-row-only columns
            if populate_on == _POPULATE_FIRST_GROUP and row_num != 2:
                continue

            # Formula columns
            formula = cfg.get('formula')
            if formula:
                ws.cell(row=row_num, column=col_idx, value=formula.format(row=row_num))
                continue

            # Static value (e.g., "USD")
            static = cfg.get('value')
            if static and not static.startswith('${'):
                ws.cell(row=row_num, column=col_idx, value=static)
                continue

            # For ungrouped, use detail_value if available, otherwise group_value
            val_template = cfg.get('detail_value') or cfg.get('group_value') or static
            if val_template is None:
                # Check default
                default = cfg.get('default')
                if default is not None and (populate_on != _POPULATE_FIRST_GROUP or row_num == 2):
                    ws.cell(row=row_num, column=col_idx, value=default)
                continue

            val = _resolve_value(val_template, ctx)
            if val is not None:
                ws.cell(row=row_num, column=col_idx, value=val)

        ws.cell(row=row_num, column=COL_AK, value=tariff_code)

        # Apply styling
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = DETAIL_FONT
            cell.border = THIN_BORDER
        for col in _spec.currency_columns_all:
            ws.cell(row=row_num, column=col).number_format = _spec.currency_format

        # Uncertainty marker for orphan-price recovered items (ungrouped).
        dq = (item.get('data_quality') or '').strip()
        if dq and Comment is not None:
            col_o = _spec.col_index('O')  # magic-ok: columns.yaml schema key
            col_p = _spec.col_index('P')  # magic-ok: columns.yaml schema key
            for col_idx in (col_o, col_p):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.fill = RECOVERED_FILL
                cell.comment = Comment(
                    f"Data quality: {dq}\n"
                    f"This value was recovered from OCR text via the "
                    f"orphan-price scan. Review and correct if wrong.",
                    _COMMENT_AUTHOR,
                )

        row_num += 1

    return row_num


# ─── Grouped Mode (e.g. 4000-000) ────────────────────────

def _write_items_grouped(
    ws, matched_items: List[Dict], invoice_data: Dict,
    supplier_name: str, supplier_info: Dict, document_type: str,
) -> int:
    """Write group row + detail rows, grouped by tariff code."""
    from collections import OrderedDict

    groups = OrderedDict()
    for item in matched_items:
        tariff_code = str(item.get('tariff_code', _HS_ZERO_CODE))
        groups.setdefault(tariff_code, []).append(item)

    row_num = 2
    is_first_row = True
    col_count = _spec.col_count()

    COL_AK = _spec.col_index('AK')  # magic-ok: columns.yaml schema key

    q_formula = _spec.formula_spec('Q')  # magic-ok: columns.yaml schema key
    r_formula = _spec.formula_spec('R')  # magic-ok: columns.yaml schema key

    for tariff_code, group_items in groups.items():
        group_qty = sum(it['quantity'] for it in group_items)
        group_total_cost = sum(it['total_cost'] for it in group_items)
        first_item = group_items[0]
        n_items = len(group_items)

        # Category: spec mappings -> CET -> item category
        spec_cat = _spec.category_name(tariff_code)
        cet_desc = get_cet_category(tariff_code)
        category = spec_cat if spec_cat != _spec.category_default else (cet_desc or first_item.get('category', '') or spec_cat)

        # Build group label
        group_desc = first_item['supplier_item_desc']
        if n_items > 1:
            group_label = f"{group_desc} (+{n_items - 1} more) ({n_items} items)"
        else:
            group_label = f"{group_desc} (1 items)"

        # Average cost — full precision per spec (NO rounding)
        average_unit_cost = group_total_cost / group_qty if group_qty else ''

        # Build context
        ctx = _build_item_context(first_item, invoice_data, supplier_name, supplier_info,
                                  document_type, tariff_code, category,
                                  group_label=group_label, group_qty=group_qty,
                                  group_total_cost=group_total_cost)
        ctx['average_unit_cost'] = average_unit_cost
        ctx['sum_quantity'] = group_qty
        ctx['sum_total_cost'] = group_total_cost

        # ── GROUP ROW ──
        for letter, cfg in _spec._col_list:
            col_idx = cfg['index']
            populate_on = cfg.get('populate_on')

            # First-row-only columns
            if populate_on == _POPULATE_FIRST_GROUP and not is_first_row:
                continue

            # Formula columns — use spec formula for group rows too
            formula = cfg.get('formula')
            if formula:
                ws.cell(row=row_num, column=col_idx, value=formula.format(row=row_num))
                continue

            # Static value
            static = cfg.get('value')
            if static and not static.startswith('${'):
                ws.cell(row=row_num, column=col_idx, value=static)
                continue

            # Group value
            val_template = cfg.get('group_value') or static
            if val_template is None:
                default = cfg.get('default')
                if default is not None and (populate_on != _POPULATE_FIRST_GROUP or is_first_row):
                    ws.cell(row=row_num, column=col_idx, value=default)
                continue

            val = _resolve_value(val_template, ctx)
            if val is not None:
                ws.cell(row=row_num, column=col_idx, value=val)

        ws.cell(row=row_num, column=COL_AK, value=tariff_code)

        if is_first_row:
            is_first_row = False

        # Apply group row styling
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = GROUP_FILL
            cell.font = GROUP_FONT
            cell.border = THIN_BORDER
        for col in _spec.currency_columns_all:
            ws.cell(row=row_num, column=col).number_format = _spec.currency_format
        row_num += 1

        # ── DETAIL ROWS ──
        for item in group_items:
            detail_ctx = {
                'tariff_code': tariff_code,
                'supplier_item': item['supplier_item'],
                'supplier_item_desc': item['supplier_item_desc'],
                'item_quantity': item['quantity'],
                'unit_price': item['unit_price'],
                'item_total_cost': item['total_cost'],
                'uom': item.get('uom', _LABELS["defaults"]["uom"]),
            }

            detail_blank = set(_spec.detail_blank_columns())

            for letter, cfg in _spec._col_list:
                col_idx = cfg['index']

                # Skip columns that must be blank on detail rows
                if letter in detail_blank:
                    continue

                # Formula columns
                formula = cfg.get('formula')
                if formula:
                    ws.cell(row=row_num, column=col_idx, value=formula.format(row=row_num))
                    continue

                # Static value
                static = cfg.get('value')
                if static and not static.startswith('${'):
                    ws.cell(row=row_num, column=col_idx, value=static)
                    continue

                # Detail value
                val_template = cfg.get('detail_value')
                if val_template is None:
                    continue

                val = _resolve_value(val_template, detail_ctx)
                if val is not None:
                    ws.cell(row=row_num, column=col_idx, value=val)

            # Apply detail styling
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = DETAIL_FONT
                cell.border = THIN_BORDER
            for col in _spec.currency_columns_detail:
                ws.cell(row=row_num, column=col).number_format = _spec.currency_format

            # Uncertainty marker: if this item was recovered from OCR by
            # the orphan-price scanner, fill its P (total_cost) and O
            # (unit_cost) cells with the RECOVERED_FILL and attach a
            # comment so a reviewer can see it at a glance.
            dq = (item.get('data_quality') or '').strip()
            if dq and Comment is not None:
                col_o = _spec.col_index('O')  # magic-ok: columns.yaml schema key
                col_p = _spec.col_index('P')  # magic-ok: columns.yaml schema key
                for col_idx in (col_o, col_p):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.fill = RECOVERED_FILL
                    cell.comment = Comment(
                        f"{_LABELS['recovered_comment']['prefix']} {dq}\n"
                        f"{_LABELS['recovered_comment']['body']}",
                        _COMMENT_AUTHOR,
                    )
            row_num += 1

    return row_num


def _build_item_context(item, invoice_data, supplier_name, supplier_info,
                        document_type, tariff_code, category,
                        group_label=None, group_qty=None, group_total_cost=None) -> dict:
    """Build a context dict for resolving ${...} templates in column specs."""
    # Invoice-level fields
    insurance = invoice_data.get('insurance', 0)
    credits = invoice_data.get('credits', 0)
    insurance_or_credits = insurance if insurance else (-credits if credits else 0)

    tax = invoice_data.get('tax', 0) or 0
    other_cost = invoice_data.get('other_cost', 0) or 0
    tax_plus_other_cost = tax + other_cost

    discount = invoice_data.get('discount', 0) or 0
    free_shipping = invoice_data.get('free_shipping', 0) or 0
    discount_plus_free_shipping = discount + free_shipping

    return {
        # Document
        'document_type': document_type,
        # Invoice
        'invoice_number': str(invoice_data.get('invoice_num', '')),
        'invoice_date': _normalize_date(invoice_data.get('invoice_date', '')),
        'invoice_total': invoice_data.get('invoice_total', 0),
        # Category
        'category_name': category,
        'tariff_code': tariff_code,
        # PO fields
        'po_number': item.get('po_number', ''),
        'po_item_ref': item.get('po_item_ref', ''),
        'po_item_desc': item.get('po_item_desc', ''),
        # Supplier item
        'supplier_item': item.get('supplier_item', ''),
        'supplier_item_desc': item.get('supplier_item_desc', ''),
        # Quantities and costs
        'item_quantity': item.get('quantity', 0),
        'unit_price': item.get('unit_price', 0),
        'item_total_cost': item.get('total_cost', 0),
        'uom': item.get('uom', _LABELS["defaults"]["uom"]),
        # Group aggregates
        'group_label': group_label or '',
        'sum_quantity': group_qty,
        'sum_total_cost': group_total_cost,
        'average_unit_cost': (group_total_cost / group_qty) if group_qty else '',
        # Invoice-level costs
        'freight': invoice_data.get('freight', 0) or 0,
        'insurance_or_credits': insurance_or_credits if insurance_or_credits else 0,
        'tax_plus_other_cost': tax_plus_other_cost if tax_plus_other_cost else 0,
        'discount_plus_free_shipping': discount_plus_free_shipping if discount_plus_free_shipping else 0,
        'packages': 1,  # magic-ok: canonical single-package default; invoice-level override comes via context
        # Supplier info
        'supplier_code': supplier_info.get('code', ''),
        'supplier_name': supplier_name,
        'supplier_address': supplier_info.get('address', ''),
        'country_code': supplier_info.get('country', _DEFAULT_COUNTRY),
    }


# ─── Headers ─────────────────────────────────────────────

def _write_headers(ws) -> None:
    """Write column headers from spec."""
    hs = _spec.header_style
    for col_idx, header in enumerate(_spec.headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            horizontal=hs['alignment'],
            wrap_text=hs['wrap_text'],
        )
        cell.border = THIN_BORDER


# ─── Subtotals (Ungrouped) ───────────────────────────────

def _write_subtotals_ungrouped(ws, row_num: int, item_count: int,
                                invoice_data: Dict = None) -> None:
    """Write subtotals for ungrouped mode — no group verification needed."""
    if item_count == 0:
        return

    first_row = 2
    last_item_row = first_row + item_count - 1
    col_count = _spec.col_count()
    label_col = _spec.ungrouped_totals_label_column
    COL_P = _spec.col_index('P')  # magic-ok: columns.yaml schema key

    p_refs = "+".join([f"P{first_row + i}" for i in range(item_count)])

    totals_rows = _spec.ungrouped_totals_rows
    row_refs = {}  # label -> row_num for formula cross-references

    for row_cfg in totals_rows:
        label = row_cfg.get('label', '')
        formula_p = row_cfg.get('column_P', '')

        ws.cell(row=row_num, column=label_col, value=label)

        # Resolve formula with references
        resolved = formula_p
        if _PH_ALL_P_REFS in resolved:
            resolved = resolved.replace(_PH_ALL_P_REFS, p_refs)
        if _PH_FIRST_ROW in resolved:
            resolved = resolved.replace(_PH_FIRST_ROW, str(first_row))
        if _PH_SUBTOTAL_ROW in resolved:
            resolved = resolved.replace(_PH_SUBTOTAL_ROW, str(row_refs.get('SUBTOTAL', row_num)))
        if _PH_ADJUSTMENTS_ROW in resolved:
            resolved = resolved.replace(_PH_ADJUSTMENTS_ROW, str(row_refs.get('ADJUSTMENTS', row_num)))
        if _PH_NET_TOTAL_ROW in resolved:
            resolved = resolved.replace(_PH_NET_TOTAL_ROW, str(row_refs.get('NET TOTAL', row_num)))

        if resolved:
            ws.cell(row=row_num, column=COL_P, value=resolved)

        row_refs[label] = row_num

        # Style
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
        ws.cell(row=row_num, column=COL_P).number_format = _spec.currency_format

        row_num += 1

    # INVOICE NOTES row for ungrouped mode (see grouped writer for rationale)
    if invoice_data:
        notes = invoice_data.get('data_quality_notes') or []
        uncertain = bool(invoice_data.get('invoice_total_uncertain'))
        if notes or uncertain:
            row_num += 1
            note_text = (
                ' | '.join(notes)
                if notes
                else _LABELS["totals"]["INVOICE_TOTAL_UNCERTAIN"]
            )
            notes_prefix = _LABELS["totals"]["INVOICE_NOTES_PREFIX"]
            label_cell = ws.cell(
                row=row_num, column=label_col,
                value=f"{notes_prefix}: {note_text}"
            )
            label_cell.font = Font(
                bold=True, size=BOLD_FONT.size,
                color=_UNCERTAIN_STYLE.get("font_color"),
            )
            label_cell.fill = UNCERTAIN_FILL
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row_num, column=col)
                if col != label_col:
                    cell.fill = UNCERTAIN_FILL
                cell.border = THIN_BORDER
            row_num += 1


# ─── Subtotals (Grouped) ─────────────────────────────────

def _write_subtotals_grouped(ws, row_num: int, item_count: int,
                              invoice_data: Dict = None) -> None:
    """Write subtotals for grouped mode from spec totals_section."""
    if item_count == 0:
        return

    first_row = 2
    last_data_row = row_num - 1
    col_count = _spec.col_count()
    label_col = _spec.totals_label_column
    COL_P = _spec.col_index('P')  # magic-ok: columns.yaml schema key
    COL_Q = _spec.col_index('Q')  # magic-ok: columns.yaml schema key

    # Identify group vs detail rows by fill color
    group_fill_color = _spec.group_style['fill_color']
    group_rows = []
    detail_rows = []
    for r in range(2, row_num):
        fill = ws.cell(row=r, column=1).fill
        is_group = (fill and fill.start_color and fill.start_color.rgb
                    and group_fill_color in str(fill.start_color.rgb))
        if is_group:
            group_rows.append(r)
        else:
            detail_rows.append(r)

    # Fallback "0" — an Excel literal zero used when a worksheet has no group rows.
    group_p_refs = "+".join([f"P{r}" for r in group_rows]) if group_rows else "0"  # magic-ok: empty-group Excel literal
    group_q_refs = "+".join([f"Q{r}" for r in group_rows]) if group_rows else "0"  # magic-ok: empty-group Excel literal

    # Get invoice-level values for the totals section
    freight_val = invoice_data.get('freight', 0) or 0 if invoice_data else 0
    insurance = invoice_data.get('insurance', 0) or 0 if invoice_data else 0
    credits_val = invoice_data.get('credits', 0) or 0 if invoice_data else 0
    insurance_val = insurance if insurance else (-credits_val if credits_val else 0)
    tax = invoice_data.get('tax', 0) or 0 if invoice_data else 0
    other_cost = invoice_data.get('other_cost', 0) or 0 if invoice_data else 0
    other_cost_val = tax + other_cost
    discount = invoice_data.get('discount', 0) or 0 if invoice_data else 0
    free_shipping = invoice_data.get('free_shipping', 0) or 0 if invoice_data else 0
    deduction_val = discount + free_shipping
    invoice_total = invoice_data.get('invoice_total', 0) if invoice_data else 0

    row_refs = {}  # label -> row_num

    for row_cfg in _spec.totals_rows:
        # Handle blank rows
        if row_cfg.get('type') == _ROW_TYPE_BLANK:
            row_num += 1
            continue

        label = row_cfg.get('label', '')
        formula_p = row_cfg.get('column_P', '')
        formula_q = row_cfg.get('column_Q', '')
        formatting = row_cfg.get('formatting', {})

        ws.cell(row=row_num, column=label_col, value=label)

        # Resolve P formula
        resolved_p = _resolve_totals_formula(
            formula_p, first_row, last_data_row,
            group_p_refs, group_q_refs, row_refs,
            freight_val, insurance_val, other_cost_val, deduction_val,
            invoice_total, 'P',  # magic-ok: columns.yaml schema key
        )
        if resolved_p is not None:
            ws.cell(row=row_num, column=COL_P, value=resolved_p)

        # Resolve Q formula
        if formula_q:
            resolved_q = _resolve_totals_formula(
                formula_q, first_row, last_data_row,
                group_p_refs, group_q_refs, row_refs,
                freight_val, insurance_val, other_cost_val, deduction_val,
                invoice_total, 'Q',  # magic-ok: columns.yaml schema key
            )
            if resolved_q is not None:
                ws.cell(row=row_num, column=COL_Q, value=resolved_q)

        row_refs[label] = row_num

        # Apply formatting
        font_color = formatting.get('font_color')
        font_bold = formatting.get('font_bold', True)

        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            if font_color:
                cell.font = Font(bold=font_bold, size=BOLD_FONT.size, color=font_color)
            else:
                cell.font = BOLD_FONT
            cell.border = THIN_BORDER

        ws.cell(row=row_num, column=COL_P).number_format = _spec.currency_format
        ws.cell(row=row_num, column=COL_Q).number_format = _spec.currency_format

        row_num += 1

    # INVOICE NOTES row — appended when the invoice carries data_quality
    # notes (orphan recovery, uncertain totals, variance absorbed into
    # ADJUSTMENTS).  Visible marker so reviewers can see AT A GLANCE that
    # this block has reconstructed numbers.
    if invoice_data:
        notes = invoice_data.get('data_quality_notes') or []
        uncertain = bool(invoice_data.get('invoice_total_uncertain'))
        if notes or uncertain:
            row_num += 1  # blank row for readability
            note_text = (
                ' | '.join(notes)
                if notes
                else _LABELS["totals"]["INVOICE_TOTAL_UNCERTAIN"]
            )
            notes_prefix = _LABELS["totals"]["INVOICE_NOTES_PREFIX"]
            label_cell = ws.cell(
                row=row_num, column=label_col,
                value=f"{notes_prefix}: {note_text}"
            )
            label_cell.font = Font(
                bold=True, size=BOLD_FONT.size,
                color=_UNCERTAIN_STYLE.get("font_color"),
            )
            label_cell.fill = UNCERTAIN_FILL
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row_num, column=col)
                if col != label_col:
                    cell.fill = UNCERTAIN_FILL
                cell.border = THIN_BORDER
            row_num += 1


def _resolve_totals_formula(formula: str, first_row: int, last_data_row: int,
                             group_p_refs: str, group_q_refs: str,
                             row_refs: dict,
                             freight_val, insurance_val, other_cost_val,
                             deduction_val, invoice_total,
                             col_letter: str = 'P') -> Any:  # magic-ok: columns.yaml schema key default
    """Resolve a totals formula template into a concrete Excel formula or value."""
    if not formula:
        return None

    # Literal value references
    if formula == _VAL_FREIGHT:
        return freight_val
    if formula == _VAL_INSURANCE:
        return insurance_val
    if formula == _VAL_OTHER_COST:
        return other_cost_val
    if formula == _VAL_DEDUCTION:
        return deduction_val
    if formula == _VAL_INVOICE_TOTAL:
        return invoice_total

    resolved = formula
    # Replace group refs
    if _PH_GROUP_P_REFS in resolved:
        resolved = resolved.replace(_PH_GROUP_P_REFS, group_p_refs)
    if _PH_GROUP_Q_REFS in resolved:
        resolved = resolved.replace(_PH_GROUP_Q_REFS, group_q_refs)
    # Replace row references
    if _PH_FIRST_ROW in resolved:
        resolved = resolved.replace(_PH_FIRST_ROW, str(first_row))
    if _PH_LAST_DATA_ROW in resolved:
        resolved = resolved.replace(_PH_LAST_DATA_ROW, str(last_data_row))
    if _PH_SUBTOTAL_GROUPED in resolved:
        resolved = resolved.replace(_PH_SUBTOTAL_GROUPED,
                                     str(row_refs.get('SUBTOTAL (GROUPED)', first_row)))
    if _PH_SUBTOTAL_DETAILS in resolved:
        resolved = resolved.replace(_PH_SUBTOTAL_DETAILS,
                                     str(row_refs.get('SUBTOTAL (DETAILS)', first_row)))
    if _PH_ADJUSTMENTS_ROW in resolved:
        resolved = resolved.replace(_PH_ADJUSTMENTS_ROW,
                                     str(row_refs.get('ADJUSTMENTS', first_row)))
    if _PH_NET_TOTAL_ROW in resolved:
        resolved = resolved.replace(_PH_NET_TOTAL_ROW,
                                     str(row_refs.get('NET TOTAL', first_row)))
    return resolved


# ─── Duty Estimation Section ───────────────────────────────

_DUTY_STYLE = _STYLES.get("duty_section", {})
DUTY_HEADER_FILL = PatternFill(
    start_color=_DUTY_STYLE.get("header_fill_color"),
    end_color=_DUTY_STYLE.get("header_fill_color"),
    fill_type=_FILL_TYPE_SOLID,
)
DUTY_HEADER_FONT = Font(
    bold=True, size=_DUTY_STYLE.get("font_size"),
    color=_DUTY_STYLE.get("header_font_color"),
)
DUTY_LABEL_FONT = Font(
    bold=False, size=_DUTY_STYLE.get("font_size"),
    color=_DUTY_STYLE.get("label_font_color"),
)
DUTY_VALUE_FONT = Font(
    bold=True, size=_DUTY_STYLE.get("font_size"),
    color=_DUTY_STYLE.get("value_font_color"),
)
DUTY_WARN_FONT = Font(
    bold=True, size=_DUTY_STYLE.get("font_size"),
    color=_DUTY_STYLE.get("warn_font_color"),
)


def _write_duty_estimation_section(
    ws, matched_items: List[Dict], invoice_data: Dict,
) -> None:
    """Write a duty estimation section below totals for classification cross-check.

    Calculates expected ASYCUDA duties (CET + CSC + VAT) based on CIF and
    tariff codes.  When the client's declared duty value is provided via
    invoice_data['_client_declared_duties'], shows the variance — a large
    difference signals potential tariff misclassification.
    """
    col_count = _spec.col_count()
    COL_J = _spec.col_index('J')  # magic-ok: columns.yaml schema key
    COL_P = _spec.col_index('P')  # magic-ok: columns.yaml schema key

    duty_labels = _LABELS["duty"]
    currency_fmt = _NUMFMT["CURRENCY_USD"]
    percent_fmt = _NUMFMT["PERCENT_1DP"]

    # Find the last used row
    row_num = ws.max_row + 2  # magic-ok: 2-row gap is fixed layout spec

    # Gather values
    invoice_total = invoice_data.get('invoice_total', 0) or 0
    customs_freight = invoice_data.get('_customs_freight', 0) or 0
    customs_insurance = invoice_data.get('_customs_insurance', 0) or 0
    client_declared = invoice_data.get('_client_declared_duties')

    # Calculate weighted CET rate from items' tariff codes
    total_cost = 0
    weighted_rate = 0
    tariff_rates = {}  # tariff -> (rate, cost)
    for item in matched_items:
        tc = str(item.get('tariff_code', _HS_ZERO_CODE))
        cost = float(item.get('total_cost', 0) or 0)
        rate = get_cet_rate(tc)
        if rate is None:
            rate = _DEFAULT_CET_RATE
        total_cost += cost
        weighted_rate += cost * rate
        if tc not in tariff_rates:
            tariff_rates[tc] = [rate, 0]
        tariff_rates[tc][1] += cost

    avg_cet_rate = (
        (weighted_rate / total_cost) if total_cost > 0 else _DEFAULT_CET_RATE
    )

    # Calculate duties
    duties = calculate_duties(
        cif_usd=invoice_total,
        cet_rate=avg_cet_rate,
        customs_freight=customs_freight,
        insurance=customs_insurance,
    )

    # ── Header row ──
    ws.cell(row=row_num, column=COL_J, value=duty_labels["section_header"])
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = DUTY_HEADER_FILL
        cell.font = DUTY_HEADER_FONT
        cell.border = THIN_BORDER
    row_num += 1

    def _write_duty_row(label, value, fmt=currency_fmt, font=None):
        nonlocal row_num
        ws.cell(row=row_num, column=COL_J, value=label).font = font or DUTY_LABEL_FONT
        val_cell = ws.cell(row=row_num, column=COL_P, value=value)
        val_cell.font = font or DUTY_VALUE_FONT
        val_cell.number_format = fmt
        for col in range(1, col_count + 1):
            ws.cell(row=row_num, column=col).border = THIN_BORDER
        row_num += 1

    # CIF breakdown
    _write_duty_row(duty_labels["cif_usd_label"], duties['cif_usd'])
    _write_duty_row(
        duty_labels["cif_xcd_label_tpl"].format(xcd_rate=XCD_RATE),
        duties['cif_xcd'],
    )

    # Per-tariff CET rates (show what rates were used)
    if len(tariff_rates) == 1:
        tc, (rate, _) = list(tariff_rates.items())[0]
        _write_duty_row(
            duty_labels["cet_single_tpl"].format(pct=rate * _PCT, tc=tc),
            duties['cet'],
        )
    else:
        _write_duty_row(
            duty_labels["cet_weighted_tpl"].format(pct=avg_cet_rate * _PCT),
            duties['cet'],
        )
        for tc, (rate, cost) in sorted(tariff_rates.items()):
            share = (cost / total_cost * _PCT) if total_cost else 0
            _write_duty_row(
                duty_labels["cet_row_tpl"].format(
                    tc=tc, pct=rate * _PCT, share=share,
                ),
                round(duties['cif_xcd'] * (cost / total_cost) * rate, 2) if total_cost else 0,
            )

    _write_duty_row(
        duty_labels["csc_tpl"].format(pct=CSC_RATE * _PCT), duties['csc'],
    )
    _write_duty_row(
        duty_labels["vat_tpl"].format(pct=VAT_RATE * _PCT), duties['vat'],
    )

    # Total
    ws.cell(
        row=row_num, column=COL_J, value=duty_labels["estimated_total"],
    ).font = DUTY_HEADER_FONT
    val_cell = ws.cell(row=row_num, column=COL_P, value=duties['total_duties'])
    val_cell.font = DUTY_HEADER_FONT
    val_cell.number_format = currency_fmt
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = DUTY_HEADER_FILL
        cell.border = THIN_BORDER
    row_num += 1

    # Effective rate
    _write_duty_row(
        duty_labels["effective_rate"],
        duties['effective_rate'],
        fmt=percent_fmt,
    )

    duty_variance_warn = _TOLERANCES["duty_variance_warn_usd"]
    cet_mismatch_threshold = _TOLERANCES["cet_rate_mismatch_threshold"]

    # Client declared value comparison
    if client_declared is not None and client_declared > 0:
        _write_duty_row(duty_labels["client_declared"], client_declared)
        variance = round(client_declared - duties['total_duties'], 2)
        variance_font = (
            DUTY_WARN_FONT if abs(variance) > duty_variance_warn else DUTY_VALUE_FONT
        )
        _write_duty_row(
            duty_labels["duty_variance"], variance, font=variance_font,
        )

        # Reverse-engineer client's implied CET rate
        # total = cif_xcd * (composite_cet_coef * cet_rate + composite_base_coef)
        # cet_rate = (total/cif_xcd - composite_base_coef) / composite_cet_coef
        implied_cet_valid = False
        if duties['cif_xcd'] > 0 and client_declared > 0:
            implied_r = (
                (client_declared / duties['cif_xcd'] - _COMPOSITE_BASE_COEF)
                / _COMPOSITE_CET_COEF
            )
            if 0 <= implied_r <= 1.0:
                implied_cet_valid = True
                _write_duty_row(
                    duty_labels["implied_cet_rate"],
                    implied_r,
                    fmt=percent_fmt,
                )
                if abs(implied_r - avg_cet_rate) > cet_mismatch_threshold:
                    ws.cell(
                        row=row_num, column=COL_J,
                        value=duty_labels["cet_mismatch_warn_tpl"].format(
                            sys_pct=avg_cet_rate * _PCT,
                            cli_pct=implied_r * _PCT,
                        ),
                    ).font = DUTY_WARN_FONT
                    for col in range(1, col_count + 1):
                        ws.cell(row=row_num, column=col).border = THIN_BORDER
                    row_num += 1

        # ── Reverse-calculated CIF (Fix C-2) ──────────────────────────────
        # When the client's declared duty is much lower (or higher) than the
        # estimate for the items CURRENTLY in this XLSX, the gap is usually
        # driven by item allocation — items that should belong to a
        # different waybill are being charged here (or vice versa).
        #
        # Reverse-calc:  cif_implied_xcd = declared_duty / composite_rate
        # where composite_rate = 1.15*cet + 0.219 at the system's avg CET.
        # Comparing cif_implied vs actual CIF lets the user see whether the
        # item split across per-declaration XLSX files matches reality.
        composite = _COMPOSITE_CET_COEF * avg_cet_rate + _COMPOSITE_BASE_COEF
        if composite > 0:
            cif_implied_xcd = client_declared / composite
            cif_implied_usd = cif_implied_xcd / XCD_RATE
            items_implied_usd = max(0.0, cif_implied_usd - (customs_freight or 0) - (customs_insurance or 0))
            _write_duty_row(
                duty_labels["implied_cif_tpl"].format(pct=avg_cet_rate * _PCT),
                round(cif_implied_xcd, 2),
            )
            _write_duty_row(
                duty_labels["implied_items_usd"],
                round(items_implied_usd, 2),
            )
            # Allocation check — compare to the items actually on this sheet.
            actual_items_usd = float(sum(
                float(item.get('total_cost', 0) or 0) for item in matched_items
            ))
            if actual_items_usd > 0 and items_implied_usd > 0:
                # If the two values diverge by more than the configured
                # relative-gap tolerance, the per-declaration item split
                # likely mis-assigned items between waybills.
                alloc_tolerance = _TOLERANCES["item_allocation_relative_gap"]
                rel_gap = abs(actual_items_usd - items_implied_usd) / max(
                    actual_items_usd, items_implied_usd
                )
                if rel_gap > alloc_tolerance:
                    direction = (
                        duty_labels["alloc_over"]
                        if actual_items_usd > items_implied_usd
                        else duty_labels["alloc_under"]
                    )
                    ws.cell(
                        row=row_num, column=COL_J,
                        value=duty_labels["alloc_warn_tpl"].format(
                            actual=actual_items_usd,
                            implied=items_implied_usd,
                            direction=direction,
                        ),
                    ).font = DUTY_WARN_FONT
                    for col in range(1, col_count + 1):
                        ws.cell(row=row_num, column=col).border = THIN_BORDER
                    row_num += 1


# ─── Reference Section (other declaration items) ──────────

# Muted styling for reference items — visible but clearly not part of totals.
_REF_STYLE = _STYLES.get("reference_section", {})
REF_HEADER_FILL = PatternFill(
    start_color=_REF_STYLE.get("header_fill_color"),
    end_color=_REF_STYLE.get("header_fill_color"),
    fill_type=_FILL_TYPE_SOLID,
)
REF_HEADER_FONT = Font(
    bold=True, size=_REF_STYLE.get("font_size"),
    color=_REF_STYLE.get("header_font_color"),
)
REF_DETAIL_FONT = Font(
    bold=False, size=_REF_STYLE.get("font_size"),
    color=_REF_STYLE.get("detail_font_color"),
)


def _write_reference_section(
    ws, reference_items: List[Dict], label: str,
    invoice_data: Optional[Dict] = None,
    main_items: Optional[List[Dict]] = None,
    reference_adjustments: Optional[Dict] = None,
) -> None:
    """Append a read-only reference section showing items from other declarations.

    These items are rendered below the totals block so they do NOT affect
    any formulas.  The section uses muted green styling to distinguish it
    from the active data above.

    When ``invoice_data`` and ``main_items`` are provided, a combined
    variance check is appended so the reviewer can verify the full invoice
    reconciliation on each sheet.

    ``reference_adjustments`` optionally provides the other declaration's
    prorated adjustments: ``{'freight': ..., 'insurance': ...,
    'other_cost': ..., 'deduction': ...}``.  When present, the combined
    section sums main + reference values directly (no reverse proration).
    """
    from collections import OrderedDict

    if not reference_items:
        return

    # Find the current last row
    row_num = ws.max_row + 2  # blank row separator
    col_count = _spec.col_count()
    COL_J = _spec.col_index('J')  # magic-ok: columns.yaml schema key
    COL_K = _spec.col_index('K')  # magic-ok: columns.yaml schema key
    COL_N = _spec.col_index('N')  # magic-ok: columns.yaml schema key
    COL_O = _spec.col_index('O')  # magic-ok: columns.yaml schema key
    COL_P = _spec.col_index('P')  # magic-ok: columns.yaml schema key

    def _ref_style_row(r, font=REF_HEADER_FONT, fill=None):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = font
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
        ws.cell(row=r, column=COL_P).number_format = _spec.currency_format

    # Section header
    ws.cell(row=row_num, column=COL_J, value=label)
    _ref_style_row(row_num, fill=REF_HEADER_FILL)
    row_num += 1

    ref_labels = _LABELS["reference"]

    # Group reference items by tariff code
    groups: OrderedDict = OrderedDict()
    for item in reference_items:
        tc = str(item.get('tariff_code', _HS_ZERO_CODE))
        groups.setdefault(tc, []).append(item)

    for tariff_code, items in groups.items():
        n = len(items)
        group_total = sum(it.get('total_cost', 0) for it in items)
        first = items[0]
        cet_desc = get_cet_category(tariff_code)
        category = cet_desc or first.get('category', '')

        desc = first.get('supplier_item_desc', '')
        if n > 1:
            group_label = ref_labels["group_label_multi_tpl"].format(
                desc=desc, more=n - 1, n=n,
            )
        else:
            group_label = ref_labels["group_label_single_tpl"].format(
                desc=desc, n=n,
            )

        # Group row — no date (col D), no invoice# (col C)
        ws.cell(row=row_num, column=_spec.col_index('E'), value=category)  # magic-ok: columns.yaml schema key
        ws.cell(row=row_num, column=_spec.col_index('F'), value=tariff_code)  # magic-ok: columns.yaml schema key
        ws.cell(row=row_num, column=COL_J, value=group_label)
        ws.cell(row=row_num, column=COL_K, value=sum(it.get('quantity', 1) for it in items))
        ws.cell(row=row_num, column=COL_N, value=_BASE_CURRENCY)
        avg_cost = group_total / n if n else 0
        ws.cell(row=row_num, column=COL_O, value=avg_cost)
        ws.cell(row=row_num, column=COL_P, value=group_total)
        _ref_style_row(row_num, fill=REF_HEADER_FILL)
        for col in _spec.currency_columns_all:
            ws.cell(row=row_num, column=col).number_format = _spec.currency_format
        row_num += 1

        # Detail rows
        for item in items:
            ws.cell(row=row_num, column=_spec.col_index('I'),  # magic-ok: columns.yaml schema key
                    value=item.get('supplier_item', ''))
            ws.cell(row=row_num, column=COL_J,
                    value=item.get('supplier_item_desc', ''))
            ws.cell(row=row_num, column=COL_K, value=item.get('quantity', 1))
            ws.cell(row=row_num, column=COL_N, value=_BASE_CURRENCY)
            ws.cell(row=row_num, column=COL_O, value=item.get('unit_price', 0))
            ws.cell(row=row_num, column=COL_P, value=item.get('total_cost', 0))
            _ref_style_row(row_num, font=REF_DETAIL_FONT)
            for col in _spec.currency_columns_detail:
                ws.cell(row=row_num, column=col).number_format = _spec.currency_format
            row_num += 1

    # Reference subtotal
    ref_total = sum(it.get('total_cost', 0) for it in reference_items)
    ws.cell(row=row_num, column=COL_J, value=ref_labels["subtotal"])
    ws.cell(row=row_num, column=COL_P, value=ref_total)
    _ref_style_row(row_num)
    ref_subtotal_row = row_num
    row_num += 1

    # Reference adjustments — the other declaration's prorated freight, etc.
    ref_adj = reference_adjustments or {}
    ref_freight_row = ref_insurance_row = ref_other_row = ref_deduction_row = None
    if ref_adj:
        ws.cell(row=row_num, column=COL_J, value=ref_labels["freight"])
        ws.cell(row=row_num, column=COL_P, value=ref_adj.get('freight', 0))
        _ref_style_row(row_num)
        ref_freight_row = row_num
        row_num += 1

        ws.cell(row=row_num, column=COL_J, value=ref_labels["insurance"])
        ws.cell(row=row_num, column=COL_P, value=ref_adj.get('insurance', 0))
        _ref_style_row(row_num)
        ref_insurance_row = row_num
        row_num += 1

        ws.cell(row=row_num, column=COL_J, value=ref_labels["other_cost"])
        ws.cell(row=row_num, column=COL_P, value=ref_adj.get('other_cost', 0))
        _ref_style_row(row_num)
        ref_other_row = row_num
        row_num += 1

        ws.cell(row=row_num, column=COL_J, value=ref_labels["deduction"])
        ws.cell(row=row_num, column=COL_P, value=ref_adj.get('deduction', 0))
        _ref_style_row(row_num)
        ref_deduction_row = row_num
        row_num += 1

        ws.cell(row=row_num, column=COL_J, value=ref_labels["adjustments"])
        ws.cell(row=row_num, column=COL_P,
                value=f'=P{ref_freight_row}+P{ref_insurance_row}+P{ref_other_row}-P{ref_deduction_row}')
        _ref_style_row(row_num)
        ref_adj_row = row_num
        row_num += 1

        ws.cell(row=row_num, column=COL_J, value=ref_labels["net_total"])
        ws.cell(row=row_num, column=COL_P,
                value=f'=P{ref_subtotal_row}+P{ref_adj_row}')
        _ref_style_row(row_num)
        row_num += 1

    # ── Combined variance check (all formulas) ───────────────
    # Shows the full invoice reconciliation across both declarations.
    # When reference_adjustments is present, sums main + reference directly.
    # Otherwise falls back to reverse proration.
    if invoice_data and main_items is not None:
        tlabels = _LABELS["totals"]
        _main_keys = (
            tlabels["SUBTOTAL_GROUPED"],
            tlabels["SUBTOTAL"],
            tlabels["TOTAL_INTERNAL_FREIGHT"],
            tlabels["TOTAL_INSURANCE"],
            tlabels["TOTAL_OTHER_COST"],
            tlabels["TOTAL_DEDUCTION"],
            tlabels["ADJUSTMENTS"],
            tlabels["NET_TOTAL"],
        )
        # Find key rows from the main totals section by label
        main_rows = {}
        for r in range(2, ref_subtotal_row):
            lbl = str(ws.cell(row=r, column=COL_J).value or '')
            if lbl in _main_keys:
                main_rows[lbl] = r

        # Support both grouped (separate freight/insurance rows) and
        # ungrouped (single ADJUSTMENTS row) modes.
        subtotal_row = (
            main_rows.get(tlabels["SUBTOTAL_GROUPED"])
            or main_rows.get(tlabels["SUBTOTAL"])
        )
        freight_row = main_rows.get(tlabels["TOTAL_INTERNAL_FREIGHT"])
        insurance_row = main_rows.get(tlabels["TOTAL_INSURANCE"])
        other_cost_row = main_rows.get(tlabels["TOTAL_OTHER_COST"])
        deduction_row = main_rows.get(tlabels["TOTAL_DEDUCTION"])
        adjustments_row = main_rows.get(tlabels["ADJUSTMENTS"])  # ungrouped mode

        if not subtotal_row:
            return  # can't build formulas without the main section

        full_invoice_total = invoice_data.get('_full_invoice_total', 0)

        row_num += 1  # blank separator

        # COMBINED ITEMS TOTAL = main subtotal + reference subtotal
        ws.cell(row=row_num, column=COL_J, value=ref_labels["combined_items_total"])
        ws.cell(row=row_num, column=COL_P,
                value=f'=P{subtotal_row}+P{ref_subtotal_row}')
        _ref_style_row(row_num)
        combined_row = row_num
        row_num += 1

        # Grouped mode has separate freight/insurance/other/deduction rows.
        # Ungrouped mode has a single ADJUSTMENTS row instead.
        is_grouped = bool(freight_row)

        if is_grouped and ref_adj and ref_freight_row:
            # Direct sum: main adjustment rows + reference adjustment rows
            ws.cell(row=row_num, column=COL_J, value=ref_labels["full_invoice_freight"])
            ws.cell(row=row_num, column=COL_P,
                    value=f'=P{freight_row}+P{ref_freight_row}')
            _ref_style_row(row_num)
            full_freight_row = row_num
            row_num += 1

            ws.cell(row=row_num, column=COL_J, value=ref_labels["full_invoice_insurance"])
            ws.cell(row=row_num, column=COL_P,
                    value=f'=P{insurance_row}+P{ref_insurance_row}')
            _ref_style_row(row_num)
            full_insurance_row = row_num
            row_num += 1

            ws.cell(row=row_num, column=COL_J, value=ref_labels["full_invoice_other"])
            ws.cell(row=row_num, column=COL_P,
                    value=f'=P{other_cost_row}+P{ref_other_row}')
            _ref_style_row(row_num)
            full_other_row = row_num
            row_num += 1

            ws.cell(row=row_num, column=COL_J, value=ref_labels["full_invoice_deduction"])
            ws.cell(row=row_num, column=COL_P,
                    value=f'=P{deduction_row}+P{ref_deduction_row}')
            _ref_style_row(row_num)
            full_deduction_row = row_num
            row_num += 1

            # FULL ADJUSTMENTS = freight + insurance + other_cost - deduction
            ws.cell(row=row_num, column=COL_J, value=ref_labels["full_adjustments"])
            ws.cell(row=row_num, column=COL_P,
                    value=f'=P{full_freight_row}+P{full_insurance_row}+P{full_other_row}-P{full_deduction_row}')
            _ref_style_row(row_num)
            full_adj_row = row_num
            row_num += 1

        elif is_grouped:
            # Grouped mode, no reference adjustments — reverse proration
            ws.cell(row=row_num, column=COL_J, value=ref_labels["full_invoice_freight"])
            ws.cell(row=row_num, column=COL_P,
                    value=f'=P{freight_row}/P{subtotal_row}*P{combined_row}')
            _ref_style_row(row_num)
            full_freight_row = row_num
            row_num += 1

            ws.cell(row=row_num, column=COL_J, value=ref_labels["full_invoice_insurance"])
            ws.cell(row=row_num, column=COL_P,
                    value=f'=P{insurance_row}/P{subtotal_row}*P{combined_row}')
            _ref_style_row(row_num)
            full_insurance_row = row_num
            row_num += 1

            ws.cell(row=row_num, column=COL_J, value=ref_labels["full_invoice_other"])
            ws.cell(row=row_num, column=COL_P,
                    value=f'=P{other_cost_row}/P{subtotal_row}*P{combined_row}')
            _ref_style_row(row_num)
            full_other_row = row_num
            row_num += 1

            ws.cell(row=row_num, column=COL_J, value=ref_labels["full_invoice_deduction"])
            ws.cell(row=row_num, column=COL_P,
                    value=f'=P{deduction_row}/P{subtotal_row}*P{combined_row}')
            _ref_style_row(row_num)
            full_deduction_row = row_num
            row_num += 1

            # FULL ADJUSTMENTS = freight + insurance + other_cost - deduction
            ws.cell(row=row_num, column=COL_J, value=ref_labels["full_adjustments"])
            ws.cell(row=row_num, column=COL_P,
                    value=f'=P{full_freight_row}+P{full_insurance_row}+P{full_other_row}-P{full_deduction_row}')
            _ref_style_row(row_num)
            full_adj_row = row_num
            row_num += 1

        else:
            # Ungrouped mode — single ADJUSTMENTS row in main section.
            # Sum main adjustments + reference adjustments directly.
            if ref_adj and ref_freight_row:
                # Show individual full-invoice lines using reference values
                # + main's portion from invoice_data
                ws.cell(row=row_num, column=COL_J, value=ref_labels["full_invoice_freight"])
                ws.cell(row=row_num, column=COL_P,
                        value=f'=T2+P{ref_freight_row}')
                _ref_style_row(row_num)
                full_freight_row = row_num
                row_num += 1

                ws.cell(row=row_num, column=COL_J, value=ref_labels["full_invoice_insurance"])
                ws.cell(row=row_num, column=COL_P,
                        value=f'=U2+P{ref_insurance_row}')
                _ref_style_row(row_num)
                full_insurance_row = row_num
                row_num += 1

                ws.cell(row=row_num, column=COL_J, value=ref_labels["full_invoice_other"])
                ws.cell(row=row_num, column=COL_P,
                        value=f'=V2+P{ref_other_row}')
                _ref_style_row(row_num)
                full_other_row = row_num
                row_num += 1

                ws.cell(row=row_num, column=COL_J, value=ref_labels["full_invoice_deduction"])
                ws.cell(row=row_num, column=COL_P,
                        value=f'=W2+P{ref_deduction_row}')
                _ref_style_row(row_num)
                full_deduction_row = row_num
                row_num += 1

                ws.cell(row=row_num, column=COL_J, value=ref_labels["full_adjustments"])
                ws.cell(row=row_num, column=COL_P,
                        value=f'=P{full_freight_row}+P{full_insurance_row}+P{full_other_row}-P{full_deduction_row}')
                _ref_style_row(row_num)
                full_adj_row = row_num
                row_num += 1
            elif adjustments_row:
                # No reference adjustments — reverse proration via ratio
                ws.cell(row=row_num, column=COL_J, value=ref_labels["full_adjustments"])
                ws.cell(row=row_num, column=COL_P,
                        value=f'=P{adjustments_row}/P{subtotal_row}*P{combined_row}')
                _ref_style_row(row_num)
                full_adj_row = row_num
                row_num += 1
            else:
                return  # can't build combined section

        # COMBINED NET TOTAL = combined items + full adjustments
        ws.cell(row=row_num, column=COL_J, value=ref_labels["combined_net_total"])
        ws.cell(row=row_num, column=COL_P,
                value=f'=P{combined_row}+P{full_adj_row}')
        _ref_style_row(row_num)
        combined_net_row = row_num
        row_num += 1

        # FULL INVOICE TOTAL — the original un-split invoice total (input value)
        ws.cell(row=row_num, column=COL_J, value=ref_labels["full_invoice_total"])
        ws.cell(row=row_num, column=COL_P, value=full_invoice_total)
        _ref_style_row(row_num)
        full_total_row = row_num
        row_num += 1

        # COMBINED VARIANCE CHECK = invoice total - net total (formula)
        ws.cell(row=row_num, column=COL_J, value=ref_labels["combined_variance_check"])
        ws.cell(row=row_num, column=COL_P,
                value=f'=P{full_total_row}-P{combined_net_row}')
        _ref_style_row(row_num)


# ─── BL Package Update ─────────────────────────────────────

def update_xlsx_packages(xlsx_path: str, packages: int,
                         freight: float = 0, insurance: float = 0) -> None:
    """Update BL-level fields on an existing XLSX file (row 2)."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    COL_X = _spec.col_index('X')  # magic-ok: columns.yaml schema key (Packages)
    ws.cell(row=2, column=COL_X, value=packages)
    bl_extra = _LABELS["bl_extra_columns"]
    if freight:
        # BL freight stored in extra column (38) to avoid corrupting variance
        ws.cell(row=1, column=_spec.col_count() + 1, value=bl_extra["freight_header"])
        ws.cell(row=2, column=_spec.col_count() + 1, value=round(freight, 2))
    if insurance:
        ws.cell(row=1, column=_spec.col_count() + 2, value=bl_extra["insurance_header"])
        ws.cell(row=2, column=_spec.col_count() + 2, value=round(insurance, 2))
    wb.save(xlsx_path)
