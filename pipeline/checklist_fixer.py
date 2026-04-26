"""Mechanical auto-fixers for ``shipment_checklist`` findings.

Per Joseph's instruction (2026-04-25): a checklist failure must trigger an
in-app fix attempt before the email is sent. Only when every option fails
does the email go out, with the residual issue flagged inline so it can be
corrected in app.

This module owns the **mechanical** half of that contract — deterministic
fixers that operate on the saved ``email_params`` dict (and the per-invoice
XLSX files it points at) without invoking the LLM. The LLM/text re-extraction
half lives in ``checklist_llm_fixers`` and is registered separately.

Public surface
--------------
``attempt_fixes(email_params, failures, *, output_dir, results, args) -> dict``
    Mutates ``email_params`` in-place. For item-level fixers, mutates the
    referenced XLSX file in place. Returns
    ``{'fixed': [check_keys], 'skipped': [check_keys]}``. Caller is
    responsible for re-saving ``email_params`` to disk and re-running
    ``shipment_checklist`` after this returns.

Notes
-----
* ``consignee_code_missing`` is intentionally NOT registered. Per
  ``feedback_consignee_code_warn_ok.md`` Joseph wants that warn to remain
  visible per-shipment; suppressing it would hide the broker fill-in step.
* Each fixer is idempotent — calling ``attempt_fixes`` twice with the same
  inputs produces the same output. The orchestration deduplicates findings
  by ``check`` so a fixer never runs twice in one pass.
"""
from __future__ import annotations

import logging
import os
import re
from typing import Callable, Dict, List, Optional

logger = logging.getLogger(__name__)


# ── Shared regex / constants (mirrors xlsx_validator.py) ─────────────
_BL_DOC_SUFFIX_RE = re.compile(   # magic-ok: BL/waybill document-type suffix
    r'[-_ ]+(Declaration|Manifest|WorkSheet|Invoice|Packing(?:\s*List)?)\s*$',
    re.IGNORECASE,
)
_MAN_REG_TWO_GROUPS_RE = re.compile(r'(\d{4})\D+(\d+)')   # magic-ok: extract YYYY + NN from any separator
_STALE_PATH_MARKERS = ('_rerun_tmp', '/tmp/', '\\Temp\\', '/var/folders/')   # magic-ok: stale-path heuristics


# ── Helpers ──────────────────────────────────────────────────────────


def _resolve_path(path: str, base_dir: str) -> str:
    """Resolve a stale absolute path by basename in ``base_dir``.

    Mirrors ``scripts/audit_shipment_checklists.py:_resolve_path``.
    """
    if os.path.exists(path):
        return path
    if not base_dir:
        return path
    cand = os.path.join(base_dir, os.path.basename(path))
    if os.path.exists(cand):
        return cand
    return path


def _xlsx_paths_in_ep(ep: dict) -> List[str]:
    """Per-invoice XLSX paths from ``email_params`` — skips combined files."""
    out = []
    for p in ep.get('attachment_paths') or []:
        if not p.lower().endswith('.xlsx'):
            continue
        b = os.path.basename(p).lower()
        if b.endswith('_combined.xlsx') or '-combined' in b:
            continue
        if os.path.exists(p):
            out.append(p)
    return out


# ── email_params (dict) fixers ───────────────────────────────────────


def _fix_bl_has_doc_suffix(ep: dict, _finding: dict, **_kw) -> bool:
    waybill = str(ep.get('waybill', ''))
    if not waybill:
        return False
    new = _BL_DOC_SUFFIX_RE.sub('', waybill).strip()
    if new and new != waybill:
        ep['waybill'] = new
        return True
    return False


def _fix_stale_tmp_path(ep: dict, _finding: dict, *, output_dir: str = '', **_kw) -> bool:
    paths = list(ep.get('attachment_paths') or [])
    if not paths:
        return False
    new_paths = [_resolve_path(p, output_dir) for p in paths]
    if new_paths == paths:
        return False   # nothing was re-resolvable → leave for residual flag
    ep['attachment_paths'] = new_paths
    # "Fixed" means every resolved path now points at a real file on disk.
    # We deliberately do NOT re-test the legacy markers here: pytest's
    # tmp_path on Linux often lives under /tmp/, which would falsely re-fire
    # the stale-marker heuristic against the legitimately-resolved file.
    return all(os.path.exists(p) for p in new_paths)


def _fix_worksheet_pdf_in_attachments(ep: dict, _finding: dict, **_kw) -> bool:
    paths = list(ep.get('attachment_paths') or [])
    new = []
    dropped = False
    for p in paths:
        b = os.path.basename(p).lower()
        if b.endswith('.pdf') and 'worksheet' in b:
            dropped = True
            continue
        new.append(p)
    if dropped:
        ep['attachment_paths'] = new
        return True
    return False


def _fix_duplicate_attachment_basenames(ep: dict, _finding: dict, **_kw) -> bool:
    paths = list(ep.get('attachment_paths') or [])
    seen = set()
    new = []
    for p in paths:
        b = os.path.basename(p)
        if b in seen:
            continue
        seen.add(b)
        new.append(p)
    if len(new) != len(paths):
        ep['attachment_paths'] = new
        return True
    return False


def _fix_expected_entries_mismatch(ep: dict, _finding: dict, **_kw) -> bool:
    paths = list(ep.get('attachment_paths') or [])
    xlsx = []
    for p in paths:
        if not p.lower().endswith('.xlsx'):
            continue
        b = os.path.basename(p).lower()
        if b.endswith('_combined.xlsx') or '-combined' in b:
            continue
        xlsx.append(p)
    if not xlsx:
        return False
    new_count = len(xlsx)
    if ep.get('expected_entries') == new_count:
        return False
    ep['expected_entries'] = new_count
    return True


def _fix_country_origin_invalid(ep: dict, _finding: dict, *, results=None, **_kw) -> bool:
    if not results:
        return False
    countries = []
    for r in results:
        si = getattr(r, 'supplier_info', None)
        if not isinstance(si, dict):
            continue
        c = si.get('country')
        if c and isinstance(c, str) and len(c) == 2 and c.isalpha():
            countries.append(c.upper())
    if not countries:
        return False
    from collections import Counter
    mode = Counter(countries).most_common(1)[0][0]
    # Try CARICOM mapping if available; fall back to raw ISO.
    try:
        from run import _iso_to_caricom_country  # type: ignore
        mode = _iso_to_caricom_country(mode) or mode
    except Exception:
        pass
    if not (isinstance(mode, str) and len(mode) == 2 and mode.isalpha()):
        return False
    if ep.get('country_origin') == mode:
        return False
    ep['country_origin'] = mode
    return True


def _fix_man_reg_invalid(ep: dict, _finding: dict, **_kw) -> bool:
    raw = str(ep.get('man_reg', '')).strip()
    if not raw:
        return False
    m = _MAN_REG_TWO_GROUPS_RE.search(raw)
    if not m:
        return False
    normalised = f"{m.group(1)} {m.group(2)}"
    if normalised == raw:
        return False
    ep['man_reg'] = normalised
    return True


# ── XLSX item fixers ─────────────────────────────────────────────────


def _fix_item_payment_line_as_item(ep: dict, _finding: dict, **_kw) -> bool:
    """Drop payment-method rows from per-invoice XLSX and recompute S2.

    Mirrors the row-classification logic in
    ``xlsx_validator._inspect_xlsx_items`` so the same rows that triggered the
    finding are the ones removed.
    """
    try:
        import openpyxl
    except ImportError:
        return False
    from xlsx_validator import (   # noqa: WPS433 — runtime import, validator owns SSOT
        COL_INV_TOTAL,
        COL_SUPP_DESC,
        COL_TOTAL_COST,
        _DUTY_ESTIMATION_PREFIXES,
        _FORMULA_LABELS,
        _PAYMENT_KEYWORDS,
    )
    fixed_any = False
    for xp in _xlsx_paths_in_ep(ep):
        try:
            wb = openpyxl.load_workbook(xp)
            ws = wb.active
            payment_rows = []
            for row in range(2, ws.max_row + 1):
                desc_v = ws.cell(row=row, column=COL_SUPP_DESC).value
                sku_v = ws.cell(row=row, column=9).value   # column I = SKU
                desc = str(desc_v).strip() if desc_v else ''
                sku = str(sku_v).strip() if sku_v else ''
                if desc in _FORMULA_LABELS:
                    continue
                if any(desc.startswith(p) for p in _DUTY_ESTIMATION_PREFIXES):
                    continue
                desc_l = desc.lower()
                if (sku.upper() == 'PAYMENT'
                        or any(kw in desc_l for kw in _PAYMENT_KEYWORDS)):
                    payment_rows.append(row)
            if not payment_rows:
                wb.close()
                continue
            for row in sorted(payment_rows, reverse=True):
                ws.delete_rows(row)
            # Recompute S2 (InvoiceTotal) from the surviving P-column values.
            new_total = 0.0
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=COL_TOTAL_COST).value
                try:
                    new_total += float(v) if v is not None else 0.0
                except (TypeError, ValueError):
                    pass
            ws.cell(row=2, column=COL_INV_TOTAL).value = round(new_total, 2)
            wb.save(xp)
            wb.close()
            fixed_any = True
        except Exception as e:
            logger.warning(f'payment_line_as_item fix failed for {xp}: {e}')
    return fixed_any


def _fix_item_description_starts_with_tariff(ep: dict, _finding: dict, **_kw) -> bool:
    try:
        import openpyxl
    except ImportError:
        return False
    from xlsx_validator import (   # noqa: WPS433
        COL_SUPP_DESC,
        _DUTY_ESTIMATION_PREFIXES,
        _FORMULA_LABELS,
        _TARIFF_PREFIX_RE,
    )
    fixed_any = False
    for xp in _xlsx_paths_in_ep(ep):
        try:
            wb = openpyxl.load_workbook(xp)
            ws = wb.active
            mutated = False
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=COL_SUPP_DESC)
                desc_v = cell.value
                desc = str(desc_v).strip() if desc_v else ''
                if not desc or desc in _FORMULA_LABELS:
                    continue
                if any(desc.startswith(p) for p in _DUTY_ESTIMATION_PREFIXES):
                    continue
                if _TARIFF_PREFIX_RE.match(desc):
                    cell.value = desc[8:].lstrip(' -')   # 8-digit tariff + sep
                    mutated = True
            if mutated:
                wb.save(xp)
                fixed_any = True
            wb.close()
        except Exception as e:
            logger.warning(f'description_starts_with_tariff fix failed for {xp}: {e}')
    return fixed_any


# ── Registry ─────────────────────────────────────────────────────────

_FIXERS: Dict[str, Callable[..., bool]] = {
    # email_params dict fixers
    'bl_has_doc_suffix': _fix_bl_has_doc_suffix,
    'stale_tmp_path': _fix_stale_tmp_path,
    'worksheet_pdf_in_attachments': _fix_worksheet_pdf_in_attachments,
    'duplicate_attachment_basenames': _fix_duplicate_attachment_basenames,
    'expected_entries_mismatch': _fix_expected_entries_mismatch,
    'country_origin_invalid': _fix_country_origin_invalid,
    'man_reg_invalid': _fix_man_reg_invalid,
    # XLSX item fixers
    'item_payment_line_as_item': _fix_item_payment_line_as_item,
    'item_description_starts_with_tariff': _fix_item_description_starts_with_tariff,
}

# Plug LLM/text re-extraction fixers into the same registry. The LLM module
# uses ``setdefault`` so mechanical entries above always take precedence
# over an LLM fix for the same kind. Failure to import is non-fatal: the
# mechanical pass still works on its own.
try:
    from checklist_llm_fixers import register_into as _register_llm_fixers
    _register_llm_fixers(_FIXERS)
except Exception as _e:   # pragma: no cover - defensive only
    logger.warning(f'checklist_llm_fixers not registered: {_e}')


# Findings the user has explicitly opted out of auto-fixing. Such findings
# are silently passed through ``attempt_fixes`` (no entry in either
# ``fixed`` or ``skipped``) so audits can still surface them by kind.
_NO_FIX_KINDS = frozenset({
    'consignee_code_missing',   # feedback_consignee_code_warn_ok.md
})


def attempt_fixes(
    email_params: dict,
    failures: list,
    *,
    output_dir: str = '',
    results: Optional[list] = None,
    args=None,
) -> dict:
    """Run registered fixers against ``failures``; mutate ``email_params``.

    Returns ``{'fixed': [check_keys_in_apply_order],
              'skipped': [check_keys_with_no_or_failed_fixer]}``.

    Each finding kind is processed at most once per call. Caller re-runs
    ``shipment_checklist`` afterwards to see what residual remains.
    """
    fixed_kinds: List[str] = []
    skipped_kinds: List[str] = []
    seen_kinds = set()

    for f in failures or []:
        kind = (f.get('check') or '').strip()
        if not kind or kind in _NO_FIX_KINDS:
            continue
        if kind in seen_kinds:
            continue
        seen_kinds.add(kind)
        fixer = _FIXERS.get(kind)
        if fixer is None:
            skipped_kinds.append(kind)
            continue
        try:
            applied = fixer(
                email_params, f,
                output_dir=output_dir,
                results=results,
                args=args,
            )
        except Exception as e:
            logger.warning(f'checklist_fixer: {kind} raised: {e}')
            applied = False
        if applied:
            fixed_kinds.append(kind)
        else:
            skipped_kinds.append(kind)

    return {'fixed': fixed_kinds, 'skipped': skipped_kinds}
