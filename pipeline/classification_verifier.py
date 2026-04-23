"""
LLM-based post-classification cross-check: verifies that item descriptions
semantically align with their assigned CARICOM CET tariff codes.

Sends batches of item-description + tariff-code + CET-description to the LLM
and asks it to identify misclassifications. Much more accurate than keyword
matching because the LLM understands:
  - What a product actually IS vs what it sounds like
  - Material vs function classification rules
  - Word confusion traps (e.g. "descaler" ≠ "weighing scale")
  - Chapter-level product domain knowledge

Usage:
    from classification_verifier import verify_classifications
    flags = verify_classifications(xlsx_path, base_dir)
"""

import json
import logging
import os
import sqlite3
import time
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)

# ── LLM verification prompt ───────────────────────────────────────────────
VERIFY_SYSTEM_PROMPT = """You are a CARICOM CET tariff classification auditor.

Your job: review item descriptions paired with their assigned HS tariff codes
and identify MISCLASSIFICATIONS — cases where the item clearly does NOT belong
under the assigned code.

HS code structure:
- Chapter (first 2 digits): broad product category (e.g. 34=soap/polish, 73=iron/steel, 84=machinery, 85=electrical)
- Heading (4 digits): product group
- Subheading (6 digits): specific product
- Tariff line (8 digits): CET end-node with duty rate

Common misclassification patterns to watch for:
1. WORD CONFUSION: "polisher" (power tool → ch.84/85) classified as "polish" (chemical → ch.34)
2. MATERIAL MISMATCH: "stainless steel" item under copper (ch.74) or plastic (ch.39) chapter
3. FUNCTION vs MATERIAL: finished articles should be classified by function, not material
4. CATCH-ALL ABUSE: items dumped into "Other" codes (like 85369000 "Marine Electrical") that clearly belong in specific headings
5. CHAPTER ABSURDITY: items under completely wrong chapters (e.g. water strainer under "printed books" ch.49)

CARICOM CET specific heading clarifications:
- 85.08 = VACUUM CLEANERS only (not power tools, not polishers, not sanders)
- 84.67 = Tools for working in the hand, with self-contained electric motor (sanders, polishers, drills, saws)
  - 84672900 = Other hand tools with motor (includes hand-held sanders, polishers, dual-action tools)
- 84.65 = Machine tools for working wood/cork/bone (bench/industrial machines, NOT hand-held tools)
- 34.05 = Polishes, creams and similar preparations (CHEMICAL products, not power tools)

Classification rules:
- Items are classified by PRIMARY FUNCTION, not material (exception: raw materials chapters 73-79)
- Chemicals/cleaners go under ch.34 (cleaning preparations) or ch.38 (chemical products), NOT under the equipment they clean
- Hand-held power tools (sanders, polishers, grinders) go under 84.67, NOT ch.34 (polish preparations) and NOT 85.08 (vacuum cleaners)
- Fishing tackle (hooks, lures, sinkers, reels) goes under ch.95 (sport/fishing), not ch.73 (metal articles)
- Nautical navigation lights go under ch.94 (lighting) not ch.85 (electrical)

For each item, respond ONLY with items that are DEFINITELY misclassified.
Do NOT flag items that are in a reasonable/debatable classification.

Respond with JSON:
{"misclassified": [
  {"row": <row_num>, "current_code": "<current>", "reason": "<brief why it's wrong>", "suggested_code": "<8_digit_code>", "confidence": "<high|medium>"}
],
"ok_count": <number_of_items_that_pass>}"""


def _load_cet_descriptions(db_path: str) -> Dict[str, str]:
    """Load HS code → description mapping from CET database."""
    if not os.path.exists(db_path):
        return {}
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute('SELECT hs_code, description FROM cet_codes')
        result = {row[0]: (row[1] or '') for row in cur.fetchall()}
        conn.close()
        return result
    except Exception as e:
        logger.warning(f"Failed to load CET descriptions: {e}")
        return {}


def _load_cet_codes(db_path: str) -> set:
    """Load set of valid CET leaf codes."""
    if not os.path.exists(db_path):
        return set()
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute('SELECT hs_code FROM cet_codes WHERE is_leaf = 1')
        result = {row[0] for row in cur.fetchall()}
        conn.close()
        return result
    except Exception:
        return set()


def _get_chapter_description(chapter: int) -> str:
    """Human-readable HS chapter name."""
    chapters = {
        27: "Mineral fuels/oils/lubricants", 28: "Inorganic chemicals",
        30: "Pharmaceuticals", 32: "Paints/varnishes/inks",
        33: "Cosmetics/perfumery", 34: "Soap/polish/wax/cleaning preparations",
        35: "Adhesives/enzymes", 36: "Explosives/pyrotechnics",
        38: "Chemical products n.e.s.", 39: "Plastics & articles thereof",
        40: "Rubber & articles thereof", 42: "Leather/bags/luggage",
        44: "Wood & articles thereof", 48: "Paper & paperboard",
        49: "Printed books/newspapers/maps", 56: "Wadding/cordage/ropes/nets",
        61: "Knitted/crocheted apparel", 63: "Other textile articles",
        68: "Stone/plaster/cement/abrasives", 69: "Ceramic products",
        70: "Glass & glassware", 73: "Articles of iron or steel",
        74: "Copper & articles thereof", 76: "Aluminum & articles thereof",
        78: "Lead & articles thereof", 79: "Zinc & articles thereof",
        82: "Tools/cutlery of base metal", 83: "Misc articles of base metal",
        84: "Machinery & mechanical appliances", 85: "Electrical machinery & equipment",
        87: "Vehicles", 89: "Ships/boats/floating structures",
        90: "Optical/measuring/medical instruments", 92: "Musical instruments",
        94: "Furniture/bedding/lamps/lighting", 95: "Toys/games/sports/fishing",
        96: "Miscellaneous manufactured articles",
    }
    return chapters.get(chapter, f"Chapter {chapter}")


def _load_supplier_categories(base_dir: str) -> Dict[str, str]:
    """Load SKU → supplier website category mapping (e.g. from Budget Marine sitemap)."""
    path = os.path.join(base_dir, 'data', 'bm_product_matches.json')
    if not os.path.exists(path):
        return {}
    try:
        with open(path) as f:
            matches = json.load(f)
        return {sku: m.get('category', '') for sku, m in matches.items() if m.get('category')}
    except Exception:
        return {}


def _build_verification_batch(items: list, cet_descriptions: Dict[str, str],
                               batch_size: int = 40,
                               supplier_categories: Dict[str, str] = None) -> list:
    """
    Build LLM prompt batches from items.

    Each batch contains up to batch_size items formatted for the LLM.
    Includes supplier website category when available for richer context.
    """
    batches = []
    for i in range(0, len(items), batch_size):
        batch_items = items[i:i + batch_size]
        lines = []
        for item in batch_items:
            tariff = item['tariff']
            desc = item['description']
            chapter = int(tariff[:2])
            cet_desc = cet_descriptions.get(tariff, '(unknown)')
            chapter_name = _get_chapter_description(chapter)
            line = (f"Row {item['row']}: \"{desc}\" → {tariff} "
                    f"(CET: \"{cet_desc}\", Chapter {chapter}: {chapter_name})")
            # Add supplier website category for extra context
            sku = item.get('sku', '')
            if supplier_categories and sku and sku in supplier_categories:
                bm_cat = supplier_categories[sku].replace('/', ' > ')
                line += f" [Supplier store: {bm_cat}]"
            lines.append(line)
        batches.append({
            'items': batch_items,
            'prompt': '\n'.join(lines),
        })
    return batches


def verify_classifications(xlsx_path: str, base_dir: str = '.',
                           max_batches: int = 10) -> List[dict]:
    """
    LLM-based verification of all classifications in an XLSX file.

    Sends items in batches to the LLM for semantic cross-checking.

    Returns list of flagged items:
        [{'row': int, 'tariff': str, 'description': str,
          'reason': str, 'suggested_code': str, 'confidence': str}, ...]
    """
    try:
        import openpyxl
    except ImportError:
        return []

    db_path = os.path.join(base_dir, 'data', 'cet.db')
    cet_descriptions = _load_cet_descriptions(db_path)
    cet_codes = _load_cet_codes(db_path)

    if not os.path.exists(xlsx_path):
        return []

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Load supplier website categories for richer context
    supplier_categories = _load_supplier_categories(base_dir)

    # Collect all classified item rows
    items = []
    for row in range(2, ws.max_row + 1):
        tariff = str(ws.cell(row, 6).value or '').strip()
        description = str(ws.cell(row, 10).value or '').strip()
        sku = str(ws.cell(row, 9).value or '').strip()
        qty = ws.cell(row, 11).value

        if not isinstance(qty, (int, float)):
            continue
        if not tariff or len(tariff) != 8 or not tariff.isdigit():
            continue
        if not description:
            continue

        items.append({
            'row': row,
            'tariff': tariff,
            'description': description[:80],
            'sku': sku,
        })

    wb.close()

    if not items:
        return []

    # Get LLM client
    try:
        from core.llm_client import get_llm_client
        llm = get_llm_client()
    except Exception as e:
        logger.warning(f"LLM client unavailable for classification verification: {e}")
        return []

    # Build and process batches
    batches = _build_verification_batch(items, cet_descriptions,
                                         supplier_categories=supplier_categories)
    all_flags = []
    fname = os.path.basename(xlsx_path)
    total_batches = min(len(batches), max_batches)
    file_start = time.monotonic()

    logger.info(
        f"[VERIFY] Starting {fname}: {len(items)} items in "
        f"{total_batches} batch(es)"
    )

    for batch_idx, batch in enumerate(batches[:max_batches]):
        batch_start = time.monotonic()
        cache_extra = f"verify_class:{fname}:batch{batch_idx}:{len(batch['items'])}"

        logger.info(
            f"[VERIFY] {fname} batch {batch_idx + 1}/{total_batches} "
            f"({len(batch['items'])} items) — starting LLM call"
        )

        user_message = (
            f"Verify these {len(batch['items'])} item classifications from "
            f"a Budget Marine (marine chandlery) invoice.\n"
            f"Flag ONLY items that are CLEARLY misclassified:\n\n"
            f"{batch['prompt']}"
        )

        try:
            response = llm.call_json(
                user_message=user_message,
                system_prompt=VERIFY_SYSTEM_PROMPT,
                max_tokens=4096,
                cache_key_extra=cache_extra,
            )
        except Exception as e:
            batch_elapsed = time.monotonic() - batch_start
            logger.warning(
                f"[VERIFY] {fname} batch {batch_idx + 1}/{total_batches} "
                f"LLM call failed after {batch_elapsed:.1f}s: {e}"
            )
            continue

        batch_elapsed = time.monotonic() - batch_start

        if not response:
            logger.warning(
                f"[VERIFY] {fname} batch {batch_idx + 1}/{total_batches} "
                f"got empty response after {batch_elapsed:.1f}s"
            )
            continue

        ok_count = response.get('ok_count', 0)
        misclassified = response.get('misclassified', [])

        logger.info(
            f"[VERIFY] {fname} batch {batch_idx + 1}/{total_batches} done in "
            f"{batch_elapsed:.1f}s: {ok_count} OK, {len(misclassified)} flagged"
        )

        for mc in misclassified:
            row = mc.get('row')
            if row is None:
                continue

            suggested = str(mc.get('suggested_code', '')).strip()

            # Validate suggested code against CET if available
            valid_suggestion = True
            if suggested and cet_codes:
                if suggested not in cet_codes:
                    fallback = suggested[:6] + '00'
                    if fallback in cet_codes:
                        suggested = fallback
                    else:
                        valid_suggestion = False

            flag = {
                'row': row,
                'tariff': mc.get('current_code', ''),
                'description': next(
                    (i['description'] for i in batch['items'] if i['row'] == row), ''
                ),
                'reason': mc.get('reason', ''),
                'confidence': mc.get('confidence', 'medium'),
            }
            if valid_suggestion and suggested:
                cet_desc = cet_descriptions.get(suggested, '')
                flag['suggested_code'] = suggested
                flag['suggested_description'] = cet_desc
            all_flags.append(flag)

    file_elapsed = time.monotonic() - file_start
    logger.info(
        f"[VERIFY] Finished {fname}: {len(all_flags)} flags from "
        f"{len(items)} items in {file_elapsed:.1f}s "
        f"({file_elapsed / max(total_batches, 1):.1f}s/batch avg)"
    )

    return all_flags


def verify_and_fix(xlsx_path: str, base_dir: str = '.',
                   auto_fix: bool = False) -> dict:
    """
    Verify classifications and optionally auto-fix high-confidence misclassifications.

    Args:
        xlsx_path: Path to XLSX file
        base_dir: Project base directory
        auto_fix: If True, automatically apply high-confidence corrections

    Returns:
        {'flags': [...], 'fixed': [...], 'needs_review': [...]}
    """
    flags = verify_classifications(xlsx_path, base_dir)

    if not flags:
        return {'flags': [], 'fixed': [], 'needs_review': []}

    fixed = []
    needs_review = []

    if auto_fix:
        try:
            import openpyxl
            db_path = os.path.join(base_dir, 'data', 'cet.db')
            cet_codes = _load_cet_codes(db_path)

            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            changed = False

            for flag in flags:
                suggested = flag.get('suggested_code', '')
                if (flag['confidence'] == 'high' and suggested
                        and len(suggested) == 8 and suggested.isdigit()
                        and (not cet_codes or suggested in cet_codes)):
                    row = flag['row']
                    old = ws.cell(row, 6).value
                    ws.cell(row, 6).value = suggested
                    ws.cell(row, 37).value = suggested
                    flag['old_code'] = str(old or '')
                    fixed.append(flag)
                    changed = True
                    logger.info(
                        f"Auto-fixed row {row}: {old} → {suggested} "
                        f"(\"{flag['description'][:40]}\")"
                    )
                else:
                    needs_review.append(flag)

            if changed:
                wb.save(xlsx_path)
            wb.close()

        except Exception as e:
            logger.error(f"Auto-fix failed: {e}")
            needs_review = flags
    else:
        needs_review = flags

    return {'flags': flags, 'fixed': fixed, 'needs_review': needs_review}


def verify_all(output_dir: str, base_dir: str = '.',
               auto_fix: bool = False) -> Dict[str, dict]:
    """
    Verify classifications across all XLSX files in an output directory.

    Returns: {filename: {'flags': [...], 'fixed': [...], 'needs_review': [...]}}
    """
    import glob
    results = {}
    total_flags = 0
    total_fixed = 0

    xlsx_files = sorted(glob.glob(os.path.join(output_dir, '*.xlsx')))
    num_files = len(xlsx_files)
    run_start = time.monotonic()
    logger.info(f"[VERIFY-ALL] Starting classification verification across {num_files} files")

    for file_idx, xlsx_path in enumerate(xlsx_files):
        fname = os.path.basename(xlsx_path)
        logger.info(
            f"[VERIFY-ALL] File {file_idx + 1}/{num_files}: {fname}"
        )
        result = verify_and_fix(xlsx_path, base_dir, auto_fix=auto_fix)
        if result['flags']:
            results[fname] = result
            total_flags += len(result['flags'])
            total_fixed += len(result['fixed'])
        elapsed = time.monotonic() - run_start
        remaining = num_files - (file_idx + 1)
        if file_idx > 0:
            avg_per_file = elapsed / (file_idx + 1)
            eta = avg_per_file * remaining
            logger.info(
                f"[VERIFY-ALL] Progress: {file_idx + 1}/{num_files} files done, "
                f"{elapsed:.0f}s elapsed, ~{eta:.0f}s remaining"
            )

    run_elapsed = time.monotonic() - run_start
    if total_flags:
        logger.warning(
            f"[VERIFY-ALL] Complete in {run_elapsed:.0f}s: {total_flags} flags, "
            f"{total_fixed} auto-fixed across {len(results)} files"
        )
    else:
        logger.info(f"[VERIFY-ALL] Complete in {run_elapsed:.0f}s: all items pass")

    return results


def print_verification_report(results: Dict[str, dict]) -> int:
    """Print a human-readable verification report. Returns total flag count."""
    if not results:
        print("  CLASSIFICATION CROSS-CHECK: All items pass.")
        return 0

    total_flags = sum(len(r['flags']) for r in results.values())
    total_fixed = sum(len(r['fixed']) for r in results.values())
    total_review = sum(len(r['needs_review']) for r in results.values())
    high = sum(1 for r in results.values() for f in r['flags']
               if f.get('confidence') == 'high')
    medium = total_flags - high

    print(f"\n  CLASSIFICATION CROSS-CHECK: {total_flags} flags "
          f"({high} high, {medium} medium)")
    if total_fixed:
        print(f"  Auto-fixed: {total_fixed}")
    if total_review:
        print(f"  Needs review: {total_review}")
    print(f"  {'─' * 100}")

    for fname, result in sorted(results.items()):
        for f in result.get('fixed', []):
            print(f"  ✓ FIXED {fname} row {f['row']:>3}: "
                  f"{f.get('old_code', '?')} → {f['suggested_code']} "
                  f"\"{f['description'][:45]}\"")
            print(f"     Reason: {f['reason']}")

        for f in result.get('needs_review', []):
            conf = f.get('confidence', '?').upper()
            marker = '!!' if conf == 'HIGH' else '? '
            print(f"  {marker} {fname} row {f['row']:>3}: "
                  f"{f['tariff']} \"{f['description'][:45]}\"")
            print(f"     {conf}: {f['reason']}")
            if f.get('suggested_code'):
                sug_desc = f.get('suggested_description', '')
                print(f"     Suggest: {f['suggested_code']} ({sug_desc[:50]})")
    print()
    return total_flags
