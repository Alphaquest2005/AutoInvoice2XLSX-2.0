#!/usr/bin/env python3
"""
Costing Sheet Generator - Generate costing sheets from ASYCUDA XML declarations.

This module parses ASYCUDA XML files and generates Excel costing sheets with:
- Detailed product data from linked invoice XLSX files
- Tax breakdowns (CET, EVL, CSC, VAT, EXT) apportioned to each product
- Cost apportionment formulas for brokerage, trucking, etc.
- Unit cost calculations with markup
"""

import xml.etree.ElementTree as ET
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
from pathlib import Path
import json
import sys
import os
import glob

# Add openpyxl to path if needed
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)


@dataclass
class ProductDetail:
    """Represents a single product from an invoice."""
    invoice_number: str
    tariff_code: str
    description: str
    quantity: float
    unit_cost: float
    total_cost: float
    supplier_item: str = ""
    category: str = ""


@dataclass
class XmlItem:
    """Represents a grouped item from ASYCUDA XML."""
    line_number: int
    hs_code: str
    description: str
    quantity: float
    cif_value: float
    invoice_value: float
    # Source invoice info from Free_text_1 (format: invoice#|line#)
    source_invoice: str = ""
    source_line: str = ""
    # Tax rates (percentages)
    cet_rate: float = 0.0
    evl_rate: float = 0.0
    csc_rate: float = 0.0
    vat_rate: float = 0.0
    ext_rate: float = 0.0
    # Tax amounts
    cet: float = 0.0
    evl: float = 0.0
    csc: float = 0.0
    vat: float = 0.0
    ext: float = 0.0
    # Linked invoice references from Attached_documents
    invoice_refs: List[str] = field(default_factory=list)
    # Product details from invoices
    products: List[ProductDetail] = field(default_factory=list)


def get_text(element: Optional[ET.Element], default: str = None) -> Optional[str]:
    """Safely get text content from an element."""
    if element is None:
        return default
    text = element.text
    if text is None or text.strip() == '' or text.strip().lower() == 'null':
        return default
    return text.strip()


def get_float(element: Optional[ET.Element], default: float = 0.0) -> float:
    """Safely get float value from an element."""
    text = get_text(element)
    if text is None:
        return default
    try:
        return float(text)
    except ValueError:
        return default


# Cache for invoice file searches to avoid repeated file scans
_invoice_file_cache: Dict[str, Dict[str, List[str]]] = {}


def find_invoice_files(workspace_dir: str, invoice_ref: str) -> List[str]:
    """
    Find invoice XLSX files that contain the given invoice reference.
    Uses 'contains' matching to handle prefixes.
    Results are cached to avoid repeated file system scans.
    """
    global _invoice_file_cache

    # Check cache first
    if workspace_dir in _invoice_file_cache:
        if invoice_ref in _invoice_file_cache[workspace_dir]:
            return _invoice_file_cache[workspace_dir][invoice_ref]
    else:
        _invoice_file_cache[workspace_dir] = {}

    matches = []
    output_dir = Path(workspace_dir) / "output"

    if not output_dir.exists():
        _invoice_file_cache[workspace_dir][invoice_ref] = matches
        return matches

    # First try filename matching only (fast)
    for xlsx_path in output_dir.glob("*.xlsx"):
        if invoice_ref.lower() in xlsx_path.stem.lower():
            matches.append(str(xlsx_path))

    # If no matches found by filename, try content matching (slow, do sparingly)
    if not matches:
        for xlsx_path in output_dir.glob("*.xlsx"):
            if str(xlsx_path) in matches:
                continue
            try:
                wb = load_workbook(xlsx_path, read_only=True, data_only=True)
                ws = wb.active
                # Check first few rows for invoice number
                found = False
                for row in range(1, min(10, ws.max_row + 1)):
                    if found:
                        break
                    for col in range(1, min(20, ws.max_column + 1)):
                        cell_val = ws.cell(row=row, column=col).value
                        if cell_val and invoice_ref in str(cell_val):
                            matches.append(str(xlsx_path))
                            found = True
                            break
                wb.close()
            except Exception:
                pass

    _invoice_file_cache[workspace_dir][invoice_ref] = matches
    return matches


# Cache for extracted products to avoid re-reading files
_product_cache: Dict[str, List[ProductDetail]] = {}


def extract_products_from_invoice(xlsx_path: str, invoice_ref: str, source_line: str = "") -> List[ProductDetail]:
    """
    Extract product details from an invoice XLSX file.

    Args:
        xlsx_path: Path to the invoice XLSX file
        invoice_ref: Invoice reference/number to filter by
        source_line: Optional line number(s) from Free_text_1 to filter exact rows
                    Can be a single number like "5" or a range like "3-7"
    """
    global _product_cache

    # Create cache key
    cache_key = f"{xlsx_path}|{invoice_ref}|{source_line}"
    if cache_key in _product_cache:
        return _product_cache[cache_key]

    products = []

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        # Find header row and column indices
        headers = {}
        header_row = 1
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if val:
                headers[str(val).lower().strip()] = col

        # Map expected columns
        invoice_col = headers.get('invoice #', headers.get('invoice', headers.get('invoice number', 3)))
        tariff_col = headers.get('tariff code', headers.get('hs code', headers.get('tariff', 6)))
        desc_col = headers.get('description', headers.get('supplier item desc', 12))
        qty_col = headers.get('quantity', headers.get('qty', 11))
        unit_cost_col = headers.get('unit cost', headers.get('unit price', 15))
        total_cost_col = headers.get('total cost', headers.get('total', 16))
        supplier_item_col = headers.get('supplier item #', headers.get('supplier item', 9))
        category_col = headers.get('category', 5)
        # Look for line number column
        line_col = headers.get('line #', headers.get('line', headers.get('row', None)))

        # Parse source_line to determine which rows to include
        # Format can be: "5" (single line), "3-7" (range), "3,5,7" (list)
        target_lines = set()
        if source_line:
            if '-' in source_line:
                # Range like "3-7"
                parts = source_line.split('-')
                try:
                    start = int(parts[0].strip())
                    end = int(parts[1].strip())
                    target_lines = set(range(start, end + 1))
                except (ValueError, IndexError):
                    pass
            elif ',' in source_line:
                # List like "3,5,7"
                for part in source_line.split(','):
                    try:
                        target_lines.add(int(part.strip()))
                    except ValueError:
                        pass
            else:
                # Single line number
                try:
                    target_lines.add(int(source_line.strip()))
                except ValueError:
                    pass

        # Extract products
        data_row_num = 0  # Track row number within data (excluding header)
        for row in range(2, ws.max_row + 1):
            data_row_num += 1  # 1-based index for data rows

            # If we have target_lines, filter to only those rows
            if target_lines:
                # Check by line column if available
                if line_col:
                    row_line = ws.cell(row=row, column=line_col).value
                    try:
                        if int(row_line) not in target_lines:
                            continue
                    except (ValueError, TypeError):
                        # If line column not numeric, fall back to row position
                        if data_row_num not in target_lines:
                            continue
                else:
                    # No line column, use row position
                    if data_row_num not in target_lines:
                        continue

            # Check if this row belongs to the invoice we're looking for
            row_invoice = ws.cell(row=row, column=invoice_col).value
            # Filter by invoice reference if row has one
            if row_invoice and invoice_ref and invoice_ref not in str(row_invoice):
                continue

            # Get values
            tariff = ws.cell(row=row, column=tariff_col).value or ""
            desc = ws.cell(row=row, column=desc_col).value or ""
            qty = ws.cell(row=row, column=qty_col).value or 0
            unit_cost = ws.cell(row=row, column=unit_cost_col).value or 0
            total_cost = ws.cell(row=row, column=total_cost_col).value or 0
            supplier_item = ws.cell(row=row, column=supplier_item_col).value or ""
            category = ws.cell(row=row, column=category_col).value or ""

            # Skip empty rows
            if not tariff and not desc:
                continue

            # Convert to proper types
            try:
                qty = float(qty) if qty else 0
            except (ValueError, TypeError):
                qty = 0
            try:
                unit_cost = float(unit_cost) if unit_cost else 0
            except (ValueError, TypeError):
                unit_cost = 0
            try:
                total_cost = float(total_cost) if total_cost else 0
            except (ValueError, TypeError):
                total_cost = 0

            # Calculate total if not provided
            if total_cost == 0 and qty > 0 and unit_cost > 0:
                total_cost = qty * unit_cost

            products.append(ProductDetail(
                invoice_number=str(row_invoice) if row_invoice else invoice_ref,
                tariff_code=str(tariff),
                description=str(desc)[:100],
                quantity=qty,
                unit_cost=unit_cost,
                total_cost=total_cost,
                supplier_item=str(supplier_item),
                category=str(category),
            ))

        wb.close()
    except Exception as e:
        print(f"Warning: Error reading {xlsx_path}: {e}", file=sys.stderr)

    _product_cache[cache_key] = products
    return products


def parse_xml_items(xml_path: str, workspace_dir: str) -> Tuple[str, float, List[XmlItem]]:
    """
    Parse an ASYCUDA XML file and extract items with linked invoice details.

    Returns:
        Tuple of (declaration_number, total_cif, list of XmlItem)
    """
    tree = ET.parse(xml_path)
    root = tree.getroot()

    # Get declaration number
    identification = root.find('Identification')
    registration = identification.find('Registration') if identification else None
    decl_number = get_text(registration.find('Number')) if registration else Path(xml_path).stem

    # Get total CIF
    valuation = root.find('Valuation')
    total_cif = get_float(valuation.find('Total_CIF')) if valuation else 0.0

    items = []
    line_number = 1

    for item_elem in root.findall('Item'):
        # Get tarification/HS code
        tarification = item_elem.find('Tarification')
        hs_code_elem = tarification.find('HScode') if tarification else None
        hs_code = get_text(hs_code_elem.find('Precision_4')) if hs_code_elem else ""

        # Get description
        goods_desc = item_elem.find('Goods_description')
        description = get_text(goods_desc.find('Commercial_Description')) if goods_desc else ""
        if not description:
            description = get_text(goods_desc.find('Description_of_goods')) if goods_desc else ""

        # Parse Free_text_1 for source invoice# and line# (format: "invoice#|line#")
        source_invoice = ""
        source_line = ""
        free_text_1 = get_text(item_elem.find('.//Free_text_1'))
        if free_text_1 and '|' in free_text_1:
            parts = free_text_1.split('|')
            source_invoice = parts[0].strip()
            source_line = parts[1].strip() if len(parts) > 1 else ""

        # Get quantity
        supp_unit = tarification.find('Supplementary_unit') if tarification else None
        quantity = get_float(supp_unit.find('Suppplementary_unit_quantity')) if supp_unit else 1.0
        if quantity == 0:
            quantity = 1.0

        # Get valuation data
        val_item = item_elem.find('Valuation_item')
        if val_item is None:
            continue

        # CIF value
        cif_value = get_float(val_item.find('Total_CIF_itm'))

        # Invoice value
        item_invoice = val_item.find('Item_Invoice')
        invoice_value = get_float(item_invoice.find('Amount_national_currency')) if item_invoice else 0.0

        # Parse taxes
        cet = evl = csc = vat = ext = 0.0
        cet_rate = evl_rate = csc_rate = vat_rate = ext_rate = 0.0
        taxation = item_elem.find('Taxation')
        if taxation is not None:
            for tax_line in taxation.findall('Taxation_line'):
                tax_code = get_text(tax_line.find('Duty_tax_code'))
                tax_amount = get_float(tax_line.find('Duty_tax_amount'))
                tax_rate = get_float(tax_line.find('Duty_tax_rate'))

                if tax_code == 'CET':
                    cet = tax_amount
                    cet_rate = tax_rate
                elif tax_code == 'EVL':
                    evl = tax_amount
                    evl_rate = tax_rate
                elif tax_code == 'CSC':
                    csc = tax_amount
                    csc_rate = tax_rate
                elif tax_code == 'VAT':
                    vat = tax_amount
                    vat_rate = tax_rate
                elif tax_code == 'EXT':
                    ext = tax_amount
                    ext_rate = tax_rate

        # Get attached invoice references
        invoice_refs = []
        for attached in item_elem.findall('.//Attached_documents'):
            doc_code = get_text(attached.find('Attached_document_code'))
            doc_name = get_text(attached.find('Attached_document_name'))
            doc_ref = get_text(attached.find('Attached_document_reference'))

            # IV05 = Commercial Invoice
            if doc_code == 'IV05' and doc_ref:
                invoice_refs.append(doc_ref)

        # Find and load product details from linked invoices
        products = []
        # Use source_invoice for precise matching if available, otherwise use attached doc refs
        search_refs = [source_invoice] if source_invoice else invoice_refs
        for inv_ref in search_refs:
            if not inv_ref:
                continue
            xlsx_files = find_invoice_files(workspace_dir, inv_ref)
            for xlsx_path in xlsx_files:
                file_products = extract_products_from_invoice(xlsx_path, inv_ref, source_line)
                products.extend(file_products)

        items.append(XmlItem(
            line_number=line_number,
            hs_code=hs_code or f"ITEM-{line_number}",
            description=description[:100] if description else "",
            quantity=quantity,
            cif_value=cif_value,
            invoice_value=invoice_value,
            source_invoice=source_invoice,
            source_line=source_line,
            cet_rate=cet_rate,
            evl_rate=evl_rate,
            csc_rate=csc_rate,
            vat_rate=vat_rate,
            ext_rate=ext_rate,
            cet=cet,
            evl=evl,
            csc=csc,
            vat=vat,
            ext=ext,
            invoice_refs=invoice_refs,
            products=products,
        ))
        line_number += 1

    return decl_number, total_cif, items


def generate_costing_sheet(xml_path: str, output_path: str = None, workspace_dir: str = None) -> str:
    """
    Generate a costing sheet Excel file from an ASYCUDA XML.

    Args:
        xml_path: Path to the ASYCUDA XML file
        output_path: Optional output path
        workspace_dir: Workspace directory to search for invoice files

    Returns:
        Path to the generated Excel file
    """
    # Determine workspace directory
    if workspace_dir is None:
        # Try to find workspace relative to XML path
        xml_parent = Path(xml_path).parent
        for parent in [xml_parent] + list(xml_parent.parents):
            if (parent / 'output').exists():
                workspace_dir = str(parent)
                break
            if parent.name == 'workspace':
                workspace_dir = str(parent)
                break
        if workspace_dir is None:
            workspace_dir = str(xml_parent.parent.parent)  # Fallback

    # Parse the XML
    decl_number, total_cif, items = parse_xml_items(xml_path, workspace_dir)

    if not items:
        raise ValueError(f"No items found in XML file: {xml_path}")

    # Generate output path if not provided
    if output_path is None:
        xml_name = Path(xml_path).stem
        output_dir = Path(xml_path).parent
        output_path = str(output_dir / f"{xml_name}_Costing.xlsx")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Costing"

    # Styles
    header_font = Font(bold=True)
    currency_format = '#,##0.00'
    percent_format = '0.0%'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    group_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # Green for group rows

    # Row 1: Title
    ws['A1'] = f"Costing Sheet - {decl_number}"
    ws['A1'].font = Font(bold=True, size=14)

    # Row 3: Total CIF reference
    ws['B3'] = 'Total CIF'
    ws['B3'].font = header_font
    ws['C3'] = total_cif
    ws['C3'].number_format = currency_format

    # Row 5: User-editable totals for apportionment
    ws['B5'] = 'Totals (Enter values to apportion)'
    ws['B5'].font = header_font

    # Editable cells for additional costs (shifted by 1 column for Type column)
    editable_cols = {
        'T': ('Brokerage', 0),
        'U': ('Truckage', 0),
        'V': ('Overtime', 0),
        'W': ('Port', 0),
        'X': ('Storage', 0),
        'Y': ('Handling', 0),
        'Z': ('Inspection', 0),
    }

    for col, (name, default) in editable_cols.items():
        ws[f'{col}5'] = default
        ws[f'{col}5'].fill = input_fill
        ws[f'{col}5'].number_format = currency_format
        ws[f'{col}5'].border = thin_border

    # Markup percentage
    ws['AC5'] = 0.5  # 50% default markup
    ws['AC5'].fill = input_fill
    ws['AC5'].number_format = percent_format
    ws['AC5'].border = thin_border

    # Row 6: Headers
    headers = [
        ('A', 'Type'),
        ('B', 'Entry#'),
        ('C', 'Line#'),
        ('D', 'Invoice#'),
        ('E', 'Tariff Code'),
        ('F', 'Description'),
        ('G', 'Qty'),
        ('H', 'Unit Cost'),
        ('I', 'Total Cost'),
        ('J', 'Cost Factor'),
        ('K', 'CET Rate'),
        ('L', 'CET'),
        ('M', 'EVL Rate'),
        ('N', 'EVL'),
        ('O', 'CSC Rate'),
        ('P', 'CSC'),
        ('Q', 'VAT Rate'),
        ('R', 'VAT'),
        ('S', 'EXT'),
        ('T', 'Brokerage'),
        ('U', 'Truckage'),
        ('V', 'Overtime'),
        ('W', 'Port'),
        ('X', 'Storage'),
        ('Y', 'Handling'),
        ('Z', 'Inspection'),
        ('AA', 'Total Cost'),
        ('AB', 'Unit Cost'),
        ('AC', 'Markup'),
        ('AD', 'Unit Price'),
        ('AE', 'COGS'),
    ]

    for col, header in headers:
        cell = ws[f'{col}6']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Set column widths
    col_widths = {
        'A': 8, 'B': 15, 'C': 6, 'D': 15, 'E': 12, 'F': 40, 'G': 8, 'H': 10, 'I': 12,
        'J': 10, 'K': 8, 'L': 10, 'M': 8, 'N': 10, 'O': 8, 'P': 10, 'Q': 8, 'R': 10, 'S': 10,
        'T': 10, 'U': 10, 'V': 10, 'W': 10, 'X': 10, 'Y': 10, 'Z': 10,
        'AA': 12, 'AB': 12, 'AC': 10, 'AD': 12, 'AE': 12,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Data rows - show BOTH grouped rows AND detail rows
    data_start_row = 7
    current_row = data_start_row
    detail_rows = []  # Track detail rows for sum formula

    for item in items:
        # First, always show the GROUPED row (highlighted in green)
        ws[f'A{current_row}'] = 'GROUP'
        ws[f'B{current_row}'] = decl_number
        ws[f'C{current_row}'] = item.line_number
        ws[f'D{current_row}'] = item.source_invoice or (', '.join(item.invoice_refs) if item.invoice_refs else 'N/A')
        ws[f'E{current_row}'] = item.hs_code
        ws[f'F{current_row}'] = item.description
        ws[f'G{current_row}'] = item.quantity
        ws[f'H{current_row}'] = item.invoice_value / item.quantity if item.quantity > 0 else item.invoice_value
        ws[f'I{current_row}'] = item.invoice_value

        # Tax rates and amounts for grouped row
        ws[f'K{current_row}'] = item.cet_rate / 100 if item.cet_rate else 0
        ws[f'L{current_row}'] = item.cet
        ws[f'M{current_row}'] = item.evl_rate / 100 if item.evl_rate else 0
        ws[f'N{current_row}'] = item.evl
        ws[f'O{current_row}'] = item.csc_rate / 100 if item.csc_rate else 0
        ws[f'P{current_row}'] = item.csc
        ws[f'Q{current_row}'] = item.vat_rate / 100 if item.vat_rate else 0
        ws[f'R{current_row}'] = item.vat
        ws[f'S{current_row}'] = item.ext

        # Format grouped row
        for col in ['H', 'I', 'L', 'N', 'P', 'R', 'S']:
            ws[f'{col}{current_row}'].number_format = currency_format
        for col in ['K', 'M', 'O', 'Q']:
            ws[f'{col}{current_row}'].number_format = '0%'

        # Highlight grouped row in green
        for col in [c[0] for c in headers]:
            ws[f'{col}{current_row}'].fill = group_fill
            ws[f'{col}{current_row}'].border = thin_border

        current_row += 1

        # Then, show detail rows if products exist
        if item.products:
            group_total_cost = sum(p.total_cost for p in item.products)
            if group_total_cost == 0:
                group_total_cost = item.invoice_value

            for product in item.products:
                # Calculate this product's share of the group's taxes
                if group_total_cost > 0:
                    product_factor = product.total_cost / group_total_cost
                else:
                    product_factor = 1.0 / len(item.products)

                detail_rows.append(current_row)

                ws[f'A{current_row}'] = 'DETAIL'
                ws[f'B{current_row}'] = decl_number
                ws[f'C{current_row}'] = item.line_number
                ws[f'D{current_row}'] = product.invoice_number
                ws[f'E{current_row}'] = product.tariff_code or item.hs_code
                ws[f'F{current_row}'] = product.description or item.description
                ws[f'G{current_row}'] = product.quantity
                ws[f'H{current_row}'] = product.unit_cost
                ws[f'I{current_row}'] = product.total_cost

                # Cost factor for overall apportionment (based on detail rows only)
                ws[f'J{current_row}'] = f'=I{current_row}/SUM({{DETAIL_SUM}})'
                ws[f'J{current_row}'].number_format = '0.0000%'

                # Apportioned taxes based on product's share within the group
                ws[f'K{current_row}'] = item.cet_rate / 100 if item.cet_rate else 0
                ws[f'L{current_row}'] = item.cet * product_factor
                ws[f'M{current_row}'] = item.evl_rate / 100 if item.evl_rate else 0
                ws[f'N{current_row}'] = item.evl * product_factor
                ws[f'O{current_row}'] = item.csc_rate / 100 if item.csc_rate else 0
                ws[f'P{current_row}'] = item.csc * product_factor
                ws[f'Q{current_row}'] = item.vat_rate / 100 if item.vat_rate else 0
                ws[f'R{current_row}'] = item.vat * product_factor
                ws[f'S{current_row}'] = item.ext * product_factor

                # Apportioned additional costs (using overall cost factor)
                for col in ['T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
                    ws[f'{col}{current_row}'] = f'=J{current_row}*${col}$5'

                # Total cost
                ws[f'AA{current_row}'] = f'=I{current_row}+L{current_row}+N{current_row}+P{current_row}+R{current_row}+S{current_row}+T{current_row}+U{current_row}+V{current_row}+W{current_row}+X{current_row}+Y{current_row}+Z{current_row}'

                # Unit cost
                ws[f'AB{current_row}'] = f'=IF(G{current_row}>0,AA{current_row}/G{current_row},0)'

                # Markup
                ws[f'AC{current_row}'] = f'=AB{current_row}*$AC$5'

                # Unit Price
                ws[f'AD{current_row}'] = f'=AB{current_row}+AC{current_row}'

                # COGS
                ws[f'AE{current_row}'] = f'=AD{current_row}*G{current_row}'

                # Format cells
                for col in ['H', 'I', 'L', 'N', 'P', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE']:
                    ws[f'{col}{current_row}'].number_format = currency_format
                for col in ['K', 'M', 'O', 'Q']:
                    ws[f'{col}{current_row}'].number_format = '0%'

                for col in [c[0] for c in headers]:
                    ws[f'{col}{current_row}'].border = thin_border

                current_row += 1

    last_data_row = current_row - 1

    # Build detail sum formula for cost factor calculation
    if detail_rows:
        # Create a formula that sums only detail rows
        detail_sum_parts = [f'I{r}' for r in detail_rows]
        # Group contiguous rows for efficiency
        detail_sum = '+'.join(detail_sum_parts) if len(detail_rows) <= 20 else f'I{detail_rows[0]}:I{detail_rows[-1]}'
    else:
        detail_sum = f'I{data_start_row}:I{last_data_row}'

    # Update formulas with actual detail sum references
    for row in detail_rows:
        cell = ws[f'J{row}']
        if cell.value and '{DETAIL_SUM}' in str(cell.value):
            cell.value = cell.value.replace('{DETAIL_SUM}', detail_sum)

    # Add totals row (only sum detail rows)
    total_row = current_row
    ws[f'A{total_row}'] = 'TOTALS'
    ws[f'A{total_row}'].font = header_font

    # Sum columns for detail rows only
    sum_cols = ['G', 'I', 'L', 'N', 'P', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AE']
    for col in sum_cols:
        # Build SUMIF formula to only sum DETAIL rows
        ws[f'{col}{total_row}'] = f'=SUMIF($A${data_start_row}:$A${last_data_row},"DETAIL",{col}{data_start_row}:{col}{last_data_row})'
        ws[f'{col}{total_row}'].number_format = currency_format
        ws[f'{col}{total_row}'].font = header_font
        ws[f'{col}{total_row}'].border = thin_border

    # Freeze panes
    ws.freeze_panes = 'A7'

    # Save
    wb.save(output_path)

    return output_path


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='Generate costing sheet from ASYCUDA XML')
    parser.add_argument('xml_file', help='Path to ASYCUDA XML file')
    parser.add_argument('--output', '-o', help='Output Excel file path')
    parser.add_argument('--workspace', '-w', help='Workspace directory to search for invoice files')

    args = parser.parse_args()

    try:
        output = generate_costing_sheet(args.xml_file, args.output, args.workspace)
        result = {
            'success': True,
            'output_path': output,
            'message': f'Costing sheet generated: {output}'
        }
        print(json.dumps(result))
    except Exception as e:
        result = {
            'success': False,
            'error': str(e)
        }
        print(json.dumps(result))
        sys.exit(1)
