#!/usr/bin/env python3
"""
Regenerate existing XLSX files to include the DUTY ESTIMATION section.

Reads each XLSX, extracts invoice_data/matched_items/reference_items from
existing rows, then calls generate_bl_xlsx() to regenerate in-place.
Skips files that already have the duty estimation section.

Usage:
    python pipeline/patch_duty_estimation.py [--folder FOLDER] [--dry-run]
"""

import argparse
import glob
import json
import os
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

import openpyxl

from bl_xlsx_generator import generate_bl_xlsx


def _has_duty_section(ws) -> bool:
    """Check if worksheet already has a DUTY ESTIMATION section."""
    for row in ws.iter_rows(min_col=10, max_col=10, values_only=True):
        if row[0] and 'DUTY ESTIMATION' in str(row[0]):
            return True
    return False


def _extract_data_from_xlsx(ws):
    """Extract invoice_data, matched_items, reference_items from existing XLSX.

    Returns (invoice_data, matched_items, reference_items, supplier_name,
             supplier_info, document_type).
    """
    # Map header names to column indices
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip()] = col

    col_a = headers.get('Document Type', 1)
    col_c = headers.get('Supplier Invoice#', 3)
    col_d = headers.get('Date', 4)
    col_e = headers.get('Category', 5)
    col_f = headers.get('TariffCode', 6)
    col_g = headers.get('PO Item Number', 7)
    col_h = headers.get('PO Item Description', 8)
    col_i = headers.get('Supplier Item Number', 9)
    col_j = headers.get('Supplier Item Description', 10)
    col_k = headers.get('Quantity', 11)
    col_l = headers.get('Per Unit', 12)
    col_m = headers.get('UNITS', 13)
    col_n = headers.get('Currency', 14)
    col_o = headers.get('Cost', 15)
    col_p = headers.get('Total Cost', 16)
    col_q = headers.get('Total', 17)
    col_s = headers.get('InvoiceTotal', 19)
    col_t = headers.get('Total Internal Freight', 20)
    col_u = headers.get('Total Insurance', 21)
    col_v = headers.get('Total Other Cost', 22)
    col_w = headers.get('Total Deduction', 23)
    col_x = headers.get('Packages', 24)
    col_z = headers.get('Supplier Code', 26)
    col_aa = headers.get('Supplier Name', 27)
    col_ab = headers.get('Supplier Address', 28)
    col_ac = headers.get('Country Code', 29)
    col_ak = headers.get('GroupBy', 37)

    # Read first data row for invoice-level info
    doc_type = ws.cell(row=2, column=col_a).value or '4000-000'
    invoice_num = ws.cell(row=2, column=col_c).value or ''
    invoice_date = ws.cell(row=2, column=col_d).value or ''
    invoice_total_val = ws.cell(row=2, column=col_s).value
    invoice_total = float(invoice_total_val) if invoice_total_val else 0
    supplier_code = ws.cell(row=2, column=col_z).value or ''
    supplier_name = ws.cell(row=2, column=col_aa).value or ''
    supplier_address = ws.cell(row=2, column=col_ab).value or ''
    country_code = ws.cell(row=2, column=col_ac).value or ''
    packages = ws.cell(row=2, column=col_x).value

    matched_items = []
    reference_items = []
    total_freight = 0
    total_insurance = 0
    total_other = 0
    total_deduction = 0
    full_invoice_total = 0
    ref_freight = 0
    ref_insurance = 0
    ref_other = 0
    ref_deduction = 0
    in_reference = False
    is_grouped = ws.cell(row=2, column=col_ak).value is not None

    current_group_tariff = None

    for row_idx in range(2, ws.max_row + 1):
        j_val = ws.cell(row=row_idx, column=col_j).value
        j_str = str(j_val) if j_val else ''

        # Summary/total rows
        if j_str.startswith('SUBTOTAL') or j_str == 'GROUP VERIFICATION':
            continue
        if j_str.startswith('TOTAL INTERNAL FREIGHT'):
            v = ws.cell(row=row_idx, column=col_p).value
            total_freight = float(v) if v else 0
            continue
        if j_str.startswith('TOTAL INSURANCE'):
            v = ws.cell(row=row_idx, column=col_p).value
            total_insurance = float(v) if v else 0
            continue
        if j_str.startswith('TOTAL OTHER COST'):
            v = ws.cell(row=row_idx, column=col_p).value
            total_other = float(v) if v else 0
            continue
        if j_str.startswith('TOTAL DEDUCTION'):
            v = ws.cell(row=row_idx, column=col_p).value
            total_deduction = float(v) if v else 0
            continue
        if j_str in ('ADJUSTMENTS', 'NET TOTAL', 'VARIANCE CHECK', ''):
            continue
        if 'INVOICE TOTAL' in j_str and 'FULL' not in j_str:
            continue
        if 'Items on other declaration' in j_str:
            in_reference = True
            continue
        # Reference adjustment rows
        if j_str == 'REFERENCE FREIGHT':
            v = ws.cell(row=row_idx, column=col_p).value
            ref_freight = float(v) if v and not isinstance(v, str) else 0
            continue
        if j_str == 'REFERENCE INSURANCE':
            v = ws.cell(row=row_idx, column=col_p).value
            ref_insurance = float(v) if v and not isinstance(v, str) else 0
            continue
        if j_str == 'REFERENCE OTHER COST':
            v = ws.cell(row=row_idx, column=col_p).value
            ref_other = float(v) if v and not isinstance(v, str) else 0
            continue
        if j_str == 'REFERENCE DEDUCTION':
            v = ws.cell(row=row_idx, column=col_p).value
            ref_deduction = float(v) if v and not isinstance(v, str) else 0
            continue
        if j_str in ('REFERENCE SUBTOTAL', 'REFERENCE ADJUSTMENTS', 'REFERENCE NET TOTAL'):
            continue
        if j_str == 'FULL INVOICE TOTAL':
            v = ws.cell(row=row_idx, column=col_p).value
            full_invoice_total = float(v) if v and not isinstance(v, str) else 0
            continue
        if j_str.startswith('COMBINED') or j_str.startswith('FULL '):
            continue
        if 'DUTY ESTIMATION' in j_str:
            break

        # Data rows
        tariff = ws.cell(row=row_idx, column=col_f).value
        cost = ws.cell(row=row_idx, column=col_p).value
        qty = ws.cell(row=row_idx, column=col_k).value
        category = ws.cell(row=row_idx, column=col_e).value
        item_num = ws.cell(row=row_idx, column=col_i).value
        unit_cost = ws.cell(row=row_idx, column=col_o).value
        units = ws.cell(row=row_idx, column=col_m).value
        currency = ws.cell(row=row_idx, column=col_n).value

        if in_reference:
            # Reference items: group header rows have category + tariff
            if tariff and cost:
                reference_items.append({
                    'tariff_code': str(tariff),
                    'total_cost': float(cost),
                    'quantity': int(qty) if qty else 1,
                    'supplier_item_desc': j_str,
                    'supplier_item': str(item_num) if item_num else '',
                    'category': str(category) if category else '',
                    'unit_price': float(unit_cost) if unit_cost else 0,
                    'uom': str(units) if units else 'NMB',
                    'currency': str(currency) if currency else 'USD',
                })
            continue

        # Regular data rows
        if tariff:
            current_group_tariff = str(tariff)

        if cost is not None:
            item_tariff = str(tariff) if tariff else current_group_tariff or '00000000'
            item = {
                'tariff_code': item_tariff,
                'total_cost': float(cost),
                'quantity': int(qty) if qty else 1,
                'supplier_item_desc': j_str,
                'supplier_item': str(item_num) if item_num else '',
                'category': str(category) if category else '',
                'unit_price': float(unit_cost) if unit_cost else float(cost),
                'uom': str(units) if units else 'NMB',
                'currency': str(currency) if currency else 'USD',
            }
            matched_items.append(item)

    # Build reference adjustments if any were found
    ref_adjustments = None
    if ref_freight or ref_insurance or ref_other or ref_deduction:
        ref_adjustments = {
            'freight': ref_freight,
            'insurance': ref_insurance,
            'other_cost': ref_other,
            'deduction': ref_deduction,
        }

    invoice_data = {
        'invoice_number': invoice_num,
        'invoice_date': str(invoice_date) if invoice_date else '',
        'invoice_total': invoice_total,
        'total_freight': total_freight,
        'total_insurance': total_insurance,
        'total_other_cost': total_other,
        'total_deduction': total_deduction,
        'packages': packages,
        'country_origin': country_code,
        '_customs_freight': 0,
        '_customs_insurance': 0,
        '_full_invoice_total': full_invoice_total,
    }

    supplier_info = {
        'code': supplier_code,
        'name': supplier_name,
        'address': supplier_address,
        'country_code': country_code,
    }

    return invoice_data, matched_items, reference_items, supplier_name, supplier_info, doc_type, ref_adjustments


def patch_xlsx(xlsx_path: str, dry_run: bool = False) -> str:
    """Regenerate XLSX with duty estimation section.

    Returns status string.
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        return f'error: {e}'

    ws = wb.active

    if _has_duty_section(ws):
        wb.close()
        return 'skip: already has duty section'

    invoice_data, matched_items, reference_items, supplier_name, supplier_info, doc_type, ref_adj = \
        _extract_data_from_xlsx(ws)
    wb.close()

    if not matched_items:
        return 'skip: no items found'

    if dry_run:
        total_cost = sum(i['total_cost'] for i in matched_items)
        return f'would patch: {len(matched_items)} items, total=${total_cost:.2f}'

    # Regenerate the XLSX using generate_bl_xlsx
    try:
        generate_bl_xlsx(
            invoice_data=invoice_data,
            matched_items=matched_items,
            supplier_name=supplier_name,
            supplier_info=supplier_info,
            output_path=xlsx_path,
            document_type=doc_type,
            reference_items=reference_items if reference_items else None,
            reference_label='Items on other declarations',
            reference_adjustments=ref_adj,
        )
        return f'patched: {len(matched_items)} items'
    except Exception as e:
        return f'error regenerating: {e}'


def main():
    parser = argparse.ArgumentParser(description='Regenerate XLSX files with duty estimation')
    parser.add_argument('--folder', help='Process only this folder name')
    parser.add_argument('--file', help='Process a single XLSX file')
    parser.add_argument('--dry-run', action='store_true', help='Show what would be done')
    parser.add_argument('--base', default='workspace/output/downloads-regression-emails',
                        help='Base directory')
    args = parser.parse_args()

    if args.file:
        xlsx_files = [args.file]
    else:
        base = os.path.join(BASE_DIR, args.base)
        if args.folder:
            patterns = [os.path.join(base, args.folder, '*.xlsx')]
        else:
            patterns = [os.path.join(base, '**', '*.xlsx')]

        xlsx_files = []
        for pat in patterns:
            xlsx_files.extend(glob.glob(pat, recursive=True))

    patched = 0
    skipped = 0
    errors = 0

    for i, xlsx_path in enumerate(sorted(xlsx_files)):
        name = os.path.basename(xlsx_path)
        result = patch_xlsx(xlsx_path, dry_run=args.dry_run)
        status = result.split(':')[0]
        print(f'[{i+1}/{len(xlsx_files)}] {name}: {result}')

        if status == 'patched' or status == 'would patch':
            patched += 1
        elif status == 'skip':
            skipped += 1
        else:
            errors += 1

    print(f'\nDone: {patched} patched, {skipped} skipped, {errors} errors')


if __name__ == '__main__':
    main()
