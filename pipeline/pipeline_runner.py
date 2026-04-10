#!/usr/bin/env python3
"""
CARICOM Invoice Processing Pipeline - Enhanced Runner

This script orchestrates the invoice processing pipeline with JSON output
support for the Electron app integration.

Usage:
    python pipeline_runner.py --input invoice.pdf --output result.xlsx
    python pipeline_runner.py --input invoice.pdf --output result.xlsx --json-output
    python pipeline_runner.py --input invoice.pdf --output result.xlsx --stage classify
    python pipeline_runner.py --validate result.xlsx --json-output
    python pipeline_runner.py --reclassify-items 0,3,5 --input parsed.json --json-output
"""

import argparse
import json
import logging
import os
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import yaml
except ImportError:
    yaml = None

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def progress(stage: str, item: int = None, total: int = None, message: str = None):
    """Emit progress in JSON format for the Electron bridge."""
    data = {"stage": stage}
    if item is not None:
        data["item"] = item
    if total is not None:
        data["total"] = total
    if message:
        data["message"] = message
    print(f"PROGRESS:{json.dumps(data)}", flush=True)


class PipelineRunner:
    """
    Data-driven pipeline orchestrator.
    Reads configuration from YAML, executes stages in order.
    """

    def __init__(self, config_path: str):
        self.config_path = Path(config_path)
        self.base_dir = self.config_path.parent.parent

        with open(config_path) as f:
            if yaml:
                self.config = yaml.safe_load(f)
            else:
                # Fallback: try JSON
                self.config = json.load(f)

        self.stages = self.config.get('stages', [])
        self.settings = self.config.get('settings', {})
        self.work_dir = None
        self.context = {}

    def run(self, input_file: str, output_file: str, stage_filter: str = None) -> Dict:
        """Execute the full pipeline or a specific stage."""
        logger.info(f"Starting pipeline: {self.config.get('pipeline', {}).get('name', 'Pipeline')}")
        progress("init", message="Starting pipeline")

        self.work_dir = Path(tempfile.mkdtemp(prefix='caricom_'))

        self.context = {
            'input_file': input_file,
            'output_file': output_file,
            'work_dir': str(self.work_dir),
            'base_dir': str(self.base_dir),
            'settings': self.settings
        }

        report = {
            'status': 'success',
            'started': datetime.now().isoformat(),
            'input': input_file,
            'output': output_file,
            'stages': [],
            'errors': [],
            'warnings': []
        }

        # Detect if input is already an XLSX file (already processed)
        # In this case, parse the XLSX, re-group, and regenerate
        input_is_xlsx = input_file.lower().endswith(('.xlsx', '.xls'))
        if input_is_xlsx:
            logger.info("Input is XLSX - reprocessing (parse → classify → group → generate)")
            progress("xlsx_parse", message="XLSX detected - extracting data for reprocessing")

            # Import xlsx_parser
            try:
                script_dir = self.base_dir / 'pipeline'
                script_dir_str = str(script_dir)
                if script_dir_str not in sys.path:
                    sys.path.insert(0, script_dir_str)

                import xlsx_parser

                # Step 1: Parse XLSX to extract items
                parsed_path = str(self.work_dir / 'xlsx_parsed.json')
                parse_result = xlsx_parser.run(input_file, parsed_path)

                report['stages'].append({
                    'name': 'xlsx_parse',
                    'type': 'script',
                    'status': parse_result.get('status', 'error'),
                    'started': datetime.now().isoformat(),
                    'completed': datetime.now().isoformat(),
                    'total_items': parse_result.get('total_items', 0),
                    'input': input_file,
                    'output': parsed_path,
                })

                if parse_result.get('status') != 'success':
                    report['status'] = 'failed'
                    report['errors'].append(f"XLSX parse failed: {parse_result.get('error', 'unknown')}")
                    report['completed'] = datetime.now().isoformat()
                    progress("done", message="Pipeline failed")
                    return report

                # Step 2: Classify items (using rule engine)
                progress("classify", message="Classifying items")
                classified_path = str(self.work_dir / 'classified.json')

                # Find classify stage config
                classify_stage = next((s for s in self.stages if s['name'] == 'classify'), None)
                rules_path = None
                if classify_stage:
                    rules_path = self._resolve_path(classify_stage.get('rules'))

                if classify_stage and rules_path and os.path.exists(rules_path):
                    classify_result = self._run_rule_engine_stage({
                        **classify_stage,
                        'input': parsed_path,
                        'output': classified_path,
                    }, {
                        'name': 'classify',
                        'type': 'rule_engine',
                        'status': 'success',
                        'started': datetime.now().isoformat()
                    })
                    report['stages'].append(classify_result)
                else:
                    # No classify stage or rules not found - use existing classifications from parsed data
                    import shutil
                    shutil.copy2(parsed_path, classified_path)
                    report['stages'].append({
                        'name': 'classify',
                        'type': 'passthrough',
                        'status': 'success',
                        'message': 'Using existing classifications from XLSX',
                        'started': datetime.now().isoformat(),
                        'completed': datetime.now().isoformat(),
                    })

                # Step 3: Group items
                progress("group", message="Grouping items by tariff code")
                grouped_path = str(self.work_dir / 'grouped.json')

                import grouping_engine
                group_result = grouping_engine.run(classified_path, grouped_path)
                report['stages'].append({
                    'name': 'group',
                    'type': 'script',
                    'status': group_result.get('status', 'error'),
                    'started': datetime.now().isoformat(),
                    'completed': datetime.now().isoformat(),
                    'total_groups': group_result.get('total_groups', 0),
                    'total_items': group_result.get('total_items', 0),
                    'input': classified_path,
                    'output': grouped_path,
                })

                if group_result.get('status') != 'success':
                    report['status'] = 'failed'
                    report['errors'].append(f"Grouping failed: {group_result.get('error', 'unknown')}")
                    report['completed'] = datetime.now().isoformat()
                    progress("done", message="Pipeline failed")
                    return report

                # Step 4: Generate XLSX
                progress("generate_xlsx", message="Generating XLSX output")
                import xlsx_generator
                # Pass context with original input_file for proper versioned naming
                gen_result = xlsx_generator.run(grouped_path, output_file, context=self.context)
                report['stages'].append({
                    'name': 'generate_xlsx',
                    'type': 'script',
                    'status': gen_result.get('status', 'error'),
                    'started': datetime.now().isoformat(),
                    'completed': datetime.now().isoformat(),
                    'output': gen_result.get('output', output_file),
                    'total_rows': gen_result.get('total_rows', 0),
                    'variance_check': gen_result.get('variance_check', 0),
                    'group_verification': gen_result.get('group_verification', 0),
                })

                # Update output path if versioned
                if gen_result.get('output') and gen_result['output'] != output_file:
                    report['output'] = gen_result['output']

                if gen_result.get('status') != 'success':
                    report['status'] = 'failed'
                    report['errors'].append(f"XLSX generation failed: {gen_result.get('error', 'unknown')}")

                # Step 5: Validate output
                progress("verify", message="Validating output")
                validation_result = self.validate_only(report['output'])
                report['stages'].append({
                    'name': 'verify',
                    'type': 'validator',
                    'status': 'success' if validation_result.get('valid', False) else 'warning',
                    'started': datetime.now().isoformat(),
                    'completed': datetime.now().isoformat(),
                    **validation_result
                })

                if not validation_result.get('valid', True):
                    report['warnings'].extend(validation_result.get('errors', []))

                report['completed'] = datetime.now().isoformat()
                report['xlsx_reprocess'] = True
                progress("done", message="XLSX reprocessing complete")
                return report

            except Exception as e:
                logger.error(f"XLSX reprocessing failed: {e}")
                report['status'] = 'failed'
                report['errors'].append(f"XLSX reprocessing error: {str(e)}")
                report['completed'] = datetime.now().isoformat()
                progress("done", message="Pipeline failed")
                return report

        try:
            for stage in self.stages:
                if not stage.get('enabled', True):
                    continue

                stage_name = stage['name']

                # If filtering to a specific stage, skip others
                if stage_filter and stage_name != stage_filter:
                    continue

                progress(stage_name, message=f"Running stage: {stage_name}")
                stage_result = self._execute_stage(stage)
                report['stages'].append(stage_result)

                # Update report output if stage produced a different path (e.g. version increment)
                if stage_result.get('output') and stage_result['status'] == 'success':
                    if stage_name == 'generate_xlsx' and stage_result['output'] != report['output']:
                        report['output'] = stage_result['output']
                        # Update context so subsequent stages (verify, learn) use the actual file
                        self.context['output_file'] = stage_result['output']

                if stage_result['status'] == 'error':
                    on_failure = stage.get('on_failure', 'abort')
                    if on_failure == 'abort':
                        report['status'] = 'failed'
                        report['errors'].append(f"Stage {stage_name} failed: {stage_result.get('error', 'unknown')}")
                        break
                    elif on_failure == 'flag_for_review':
                        report['warnings'].append(f"Stage {stage_name} flagged for review")

        except Exception as e:
            logger.error(f"Pipeline error: {e}")
            report['status'] = 'error'
            report['errors'].append(str(e))

        finally:
            report['completed'] = datetime.now().isoformat()

        progress("done", message="Pipeline complete")
        return report

    def _execute_stage(self, stage: Dict) -> Dict:
        """Execute a single pipeline stage."""
        stage_name = stage['name']
        stage_type = stage['type']

        result = {
            'name': stage_name,
            'type': stage_type,
            'status': 'success',
            'started': datetime.now().isoformat()
        }

        try:
            if stage_type == 'script':
                result = self._run_script_stage(stage, result)
            elif stage_type == 'rule_engine':
                result = self._run_rule_engine_stage(stage, result)
            elif stage_type == 'validator':
                result = self._run_validator_stage(stage, result)
            else:
                raise ValueError(f"Unknown stage type: {stage_type}")

        except Exception as e:
            logger.error(f"Stage {stage_name} failed: {e}")
            result['status'] = 'error'
            result['error'] = str(e)

        finally:
            result['completed'] = datetime.now().isoformat()

        return result

    def _run_script_stage(self, stage: Dict, result: Dict) -> Dict:
        """Execute a Python script stage."""
        script_name = stage.get('script')
        input_path = self._resolve_path(stage.get('input'))
        output_path = self._resolve_path(stage.get('output'))

        # Try to import and run the stage module
        script_dir = self.base_dir / 'pipeline'
        module_name = script_name.replace('.py', '')
        module_path = script_dir / script_name

        # Ensure pipeline directory is in sys.path for inter-module imports
        script_dir_str = str(script_dir)
        if script_dir_str not in sys.path:
            sys.path.insert(0, script_dir_str)

        if module_path.exists():
            import importlib.util
            spec = importlib.util.spec_from_file_location(module_name, str(module_path))
            if spec and spec.loader:
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)

                if hasattr(module, 'run'):
                    stage_output = module.run(
                        input_path=input_path,
                        output_path=output_path,
                        config=stage,
                        context=self.context
                    )
                    if stage_output:
                        result.update(stage_output)
                else:
                    logger.warning(f"Module {module_name} has no 'run' function")
        else:
            logger.info(f"Script not found: {module_path} - skipping")
            result['message'] = f"Script {script_name} not yet implemented"

        result['input'] = input_path
        # Preserve output path from stage module if it set one (e.g. version increment)
        if 'output' not in result:
            result['output'] = output_path
        return result

    def _run_rule_engine_stage(self, stage: Dict, result: Dict) -> Dict:
        """Execute rule engine classification with web lookup fallback."""
        rules_path = self._resolve_path(stage.get('rules'))
        input_path = self._resolve_path(stage.get('input'))
        output_path = self._resolve_path(stage.get('output'))

        if rules_path and os.path.exists(rules_path):
            with open(rules_path) as f:
                rules = json.load(f)
            result['rules_loaded'] = len(rules.get('rules', []))

            # Run classification using classifier.py which includes web lookup fallback
            if input_path and os.path.exists(input_path):
                try:
                    # Import classifier module
                    script_dir = self.base_dir / 'pipeline'
                    classifier_path = script_dir / 'classifier.py'
                    if classifier_path.exists():
                        import importlib.util
                        spec = importlib.util.spec_from_file_location('classifier', str(classifier_path))
                        if spec and spec.loader:
                            classifier_module = importlib.util.module_from_spec(spec)
                            spec.loader.exec_module(classifier_module)

                            # Build config from stage settings
                            stage_config = {
                                'web_verify': stage.get('web_verify', {'enabled': True}),
                                'base_dir': str(self.base_dir),
                            }

                            # Run classifier with web fallback
                            classify_result = classifier_module.run(
                                input_path=input_path,
                                output_path=output_path,
                                config=stage_config,
                                context=self.context
                            )

                            # Emit progress for each item
                            with open(input_path) as f:
                                data = json.load(f)
                            item_list = data if isinstance(data, list) else data.get('items', [])
                            total = len(item_list)
                            for i in range(total):
                                progress("classify", item=i + 1, total=total)

                            result['items_classified'] = classify_result.get('items_classified', 0)
                            result['unmatched'] = classify_result.get('items_unmatched', 0)
                            result['input'] = input_path
                            result['output'] = output_path
                            return result

                except Exception as e:
                    logger.warning(f"Classifier module failed, falling back to RuleEngine: {e}")

                # Fallback to basic RuleEngine if classifier fails
                engine = RuleEngine(rules_path)
                with open(input_path) as f:
                    data = json.load(f)

                classified = []
                item_list = data if isinstance(data, list) else data.get('items', [])
                total = len(item_list)

                for i, item in enumerate(item_list):
                    progress("classify", item=i + 1, total=total)
                    desc = item.get('description', '')
                    match = engine.classify(desc)
                    item['classification'] = match or {'code': 'UNKNOWN', 'confidence': 0}
                    classified.append(item)

                if output_path:
                    os.makedirs(os.path.dirname(output_path), exist_ok=True)
                    # Preserve invoice_metadata and other top-level fields
                    output_data = {'items': classified}
                    if isinstance(data, dict):
                        for key in data:
                            if key != 'items':
                                output_data[key] = data[key]
                    with open(output_path, 'w') as f:
                        json.dump(output_data, f, indent=2)

                result['items_classified'] = total
                result['unmatched'] = sum(1 for i in classified if i['classification'].get('code') == 'UNKNOWN')

        result['input'] = input_path
        result['output'] = output_path
        return result

    def _run_validator_stage(self, stage: Dict, result: Dict) -> Dict:
        """Execute validation stage."""
        input_path = self._resolve_path(stage.get('input'))
        output_path = self._resolve_path(stage.get('output'))
        checks = stage.get('checks', [])

        if input_path and input_path.endswith('.xlsx') and os.path.exists(input_path):
            validator = Validator()
            validation = validator.validate_xlsx(input_path)
            result.update(validation)
        elif input_path and os.path.exists(input_path):
            # For JSON validation stages, pass data through to output
            # and run code_validator if available
            try:
                script_dir = self.base_dir / 'pipeline'
                cv_path = script_dir / 'code_validator.py'
                if cv_path.exists():
                    import importlib.util
                    spec = importlib.util.spec_from_file_location('code_validator', str(cv_path))
                    if spec and spec.loader:
                        module = importlib.util.module_from_spec(spec)
                        spec.loader.exec_module(module)
                        if hasattr(module, 'run'):
                            stage_output = module.run(
                                input_path=input_path,
                                output_path=output_path,
                                config=stage,
                                context=self.context
                            )
                            if stage_output:
                                result.update(stage_output)
                            result['input'] = input_path
                            return result
            except Exception as e:
                logger.warning(f"code_validator failed: {e}, passing data through")

            # Fallback: copy input to output to preserve pipeline data flow
            if output_path:
                import shutil
                os.makedirs(os.path.dirname(output_path), exist_ok=True)
                shutil.copy2(input_path, output_path)

            result['checks_performed'] = len(checks)
            result['checks'] = {
                (check.get('name', str(check)) if isinstance(check, dict) else str(check)): 'skipped'
                for check in checks
            }
        else:
            result['checks_performed'] = len(checks)
            result['checks'] = {}
            for check in checks:
                if isinstance(check, dict):
                    check_name = check.get('name', list(check.keys())[0])
                    result['checks'][check_name] = 'pending'
                else:
                    result['checks'][check] = 'pending'

        result['input'] = input_path
        return result

    def _resolve_path(self, path: Optional[str]) -> Optional[str]:
        """Resolve path with variable substitution."""
        if not path:
            return None

        for key, value in self.context.items():
            if isinstance(value, str):
                path = path.replace(f'${{{key}}}', value)

        if not os.path.isabs(path):
            path = str(self.base_dir / path)

        return path

    def validate_only(self, xlsx_path: str) -> Dict:
        """Run validation on existing Excel file."""
        validator = Validator()
        return validator.validate_xlsx(xlsx_path)

    def reclassify_items(self, input_file: str, indices: List[int]) -> Dict:
        """Re-run classification for specific items."""
        rules_path = str(self.base_dir / 'rules' / 'classification_rules.json')

        if not os.path.exists(rules_path):
            return {'error': 'classification_rules.json not found'}

        if not os.path.exists(input_file):
            return {'error': f'Input file not found: {input_file}'}

        engine = RuleEngine(rules_path)

        with open(input_file) as f:
            data = json.load(f)

        items = data if isinstance(data, list) else data.get('items', [])
        reclassified = []

        for idx in indices:
            if 0 <= idx < len(items):
                item = items[idx]
                desc = item.get('description', '')
                old_class = item.get('classification', {}).get('code', 'NONE')
                match = engine.classify(desc)
                item['classification'] = match or {'code': 'UNKNOWN', 'confidence': 0}
                reclassified.append({
                    'index': idx,
                    'description': desc,
                    'old_code': old_class,
                    'new_code': item['classification']['code'],
                })

        # Save back
        if isinstance(data, dict):
            data['items'] = items
        with open(input_file, 'w') as f:
            json.dump(data if isinstance(data, dict) else items, f, indent=2)

        return {
            'status': 'success',
            'reclassified': reclassified,
            'total': len(reclassified),
        }


class RuleEngine:
    """Data-driven rule engine for classification."""

    def __init__(self, rules_path: str):
        with open(rules_path) as f:
            self.config = json.load(f)

        self.rules = sorted(
            self.config.get('rules', []),
            key=lambda r: r.get('priority', 0),
            reverse=True
        )
        self.noise_words = set(
            self.config.get('word_analysis', {}).get('noise_words', [])
        )

    def classify(self, description: str) -> Optional[Dict]:
        """Apply rules to classify an item."""
        desc_upper = description.upper()

        for rule in self.rules:
            if self._matches_rule(desc_upper, rule):
                return {
                    'code': rule['code'],
                    'category': rule.get('category', 'PRODUCTS'),
                    'confidence': rule.get('confidence', 0.8),
                    'rule_id': rule.get('id'),
                    'notes': rule.get('notes')
                }

        return None

    def _matches_rule(self, desc: str, rule: Dict) -> bool:
        """Check if description matches rule patterns."""
        patterns = rule.get('patterns', [])
        exclude = rule.get('exclude', [])

        for excl in exclude:
            if excl.upper() in desc:
                return False

        for pattern in patterns:
            if pattern.upper() in desc:
                return True

        return False


class Validator:
    """Excel output validator."""

    def validate_xlsx(self, filepath: str) -> Dict:
        try:
            import openpyxl
        except ImportError:
            return {'error': 'openpyxl not installed', 'valid': False}

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        report = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'checks': {}
        }

        # Check VARIANCE CHECK
        variance = self._find_cell_value(ws, 'VARIANCE CHECK', 16)
        report['checks']['variance_check'] = variance
        if variance is not None and abs(float(variance)) > 0.001:
            report['valid'] = False
            report['errors'].append(f"VARIANCE CHECK = ${variance}, expected $0.00")

        # Check GROUP VERIFICATION
        group_var = self._find_cell_value(ws, 'GROUP VERIFICATION', 16)
        report['checks']['group_verification'] = group_var
        if group_var is not None and abs(float(group_var)) > 0.001:
            report['valid'] = False
            report['errors'].append(f"GROUP VERIFICATION = ${group_var}, expected $0.00")

        # Scan for formula errors
        error_types = ['#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#N/A']
        errors_found = []

        for row in range(1, ws.max_row + 1):
            for col in range(1, min(ws.max_column + 1, 40)):
                cell = ws.cell(row=row, column=col)
                if cell.value and str(cell.value) in error_types:
                    col_letter = openpyxl.utils.get_column_letter(col)
                    errors_found.append({
                        'cell': f"{col_letter}{row}",
                        'error': str(cell.value)
                    })

        report['checks']['formula_errors'] = len(errors_found)
        if errors_found:
            report['valid'] = False
            report['errors'].extend([f"{e['error']} at {e['cell']}" for e in errors_found[:10]])

        return report

    def _find_cell_value(self, ws, label: str, value_col: int) -> Optional[float]:
        for row in range(1, ws.max_row + 1):
            for col in range(1, 15):
                cell = ws.cell(row=row, column=col)
                if cell.value and label.upper() in str(cell.value).upper():
                    val = ws.cell(row=row, column=value_col).value
                    if val is not None:
                        try:
                            return float(val)
                        except (ValueError, TypeError):
                            return None
        return None


def main():
    parser = argparse.ArgumentParser(
        description='CARICOM Invoice Processing Pipeline'
    )
    parser.add_argument('--input', '-i', help='Input PDF file')
    parser.add_argument('--output', '-o', help='Output Excel file')
    parser.add_argument('--config', '-c', help='Pipeline configuration file')
    parser.add_argument('--validate', '-v', help='Validate existing Excel file only')
    parser.add_argument('--stage', '-s', help='Run only this specific stage')
    parser.add_argument('--reclassify-items', help='Comma-separated item indices to reclassify')
    parser.add_argument('--json-output', action='store_true', help='Output results as JSON')
    parser.add_argument('--dry-run', action='store_true', help='Show what would be done')

    args = parser.parse_args()

    # Determine config path
    script_dir = Path(__file__).parent
    base_dir = script_dir.parent

    if args.config:
        config_path = Path(args.config)
    else:
        config_path = base_dir / 'config' / 'pipeline.yaml'

    if not config_path.exists():
        msg = f"Config not found at {config_path}"
        if args.json_output:
            print(json.dumps({'error': msg}))
        else:
            print(f"Error: {msg}")
        sys.exit(2)

    runner = PipelineRunner(str(config_path))

    # Validate only
    if args.validate:
        result = runner.validate_only(args.validate)
        if args.json_output:
            print(f"REPORT:{json.dumps(result)}")
        else:
            print(json.dumps(result, indent=2))
        sys.exit(0 if result.get('valid', False) else 1)

    # Reclassify items
    if args.reclassify_items:
        if not args.input:
            msg = "Must provide --input for reclassify"
            if args.json_output:
                print(json.dumps({'error': msg}))
            else:
                print(f"Error: {msg}")
            sys.exit(1)

        indices = [int(x.strip()) for x in args.reclassify_items.split(',')]
        result = runner.reclassify_items(args.input, indices)
        if args.json_output:
            print(f"REPORT:{json.dumps(result)}")
        else:
            print(json.dumps(result, indent=2))
        sys.exit(0 if result.get('status') == 'success' else 1)

    # Full or filtered pipeline run
    if not args.input or not args.output:
        parser.print_help()
        sys.exit(1)

    if args.dry_run:
        stages = [s for s in runner.stages if s.get('enabled', True)]
        if args.stage:
            stages = [s for s in stages if s['name'] == args.stage]
        if args.json_output:
            print(json.dumps({'stages': [s['name'] for s in stages]}))
        else:
            print("DRY RUN - Would execute:")
            for s in stages:
                print(f"  - {s['name']} ({s['type']})")
        sys.exit(0)

    report = runner.run(args.input, args.output, args.stage)

    if args.json_output:
        print(f"REPORT:{json.dumps(report)}")
    else:
        print(json.dumps(report, indent=2))

    if report['status'] == 'success':
        sys.exit(0)
    elif report['status'] == 'failed':
        sys.exit(1)
    else:
        sys.exit(2)


if __name__ == '__main__':
    main()
