#!/usr/bin/env python3
"""
PO Matcher Module

Reads Purchase Order XLSX files and matches parsed invoice items
to PO line items using a scoring-based algorithm.

This module is format-agnostic — it works with any parsed invoice data
regardless of the invoice format.

Usage:
    po_items = POReader.read_po_xlsx('/path/to/po.xlsx')
    matcher = POMatcher(po_items)
    matched = matcher.match_invoice(invoice_data, 'invoice_filename.pdf')
"""

import json
import os
import re
import logging
from typing import Any, Dict, List, Optional, Set

try:
    import openpyxl
except ImportError:
    openpyxl = None

logger = logging.getLogger(__name__)


def _load_item_aliases(base_dir: str = '.') -> Dict[str, List[Dict]]:
    """Load item alias mappings from data/item_aliases.json.

    Returns a dict keyed by normalized supplier item code,
    where each value is a list of {bm_code, supplier_prefix, supplier_item}.
    """
    alias_path = os.path.join(base_dir, 'data', 'item_aliases.json')
    if not os.path.exists(alias_path):
        logger.debug("No item_aliases.json found at %s", alias_path)
        return {}
    try:
        with open(alias_path, 'r') as f:
            data = json.load(f)
        raw = data.get('aliases', {})
        # Build normalized lookup: strip dashes/spaces, uppercase
        normalized = {}
        for key, entries in raw.items():
            norm_key = key.upper().replace('-', '').replace(' ', '')
            if norm_key not in normalized:
                normalized[norm_key] = []
            normalized[norm_key].extend(entries)
        logger.info("Loaded %d alias keys from %s", len(normalized), alias_path)
        return normalized
    except Exception as e:
        logger.warning("Failed to load item aliases: %s", e)
        return {}


# Module-level cache so aliases are loaded once
_alias_cache: Optional[Dict[str, List[Dict]]] = None
_alias_base_dir: Optional[str] = None


def _get_aliases(base_dir: str = '.') -> Dict[str, List[Dict]]:
    """Get cached alias lookup dict."""
    global _alias_cache, _alias_base_dir
    if _alias_cache is None or _alias_base_dir != base_dir:
        _alias_cache = _load_item_aliases(base_dir)
        _alias_base_dir = base_dir
    return _alias_cache

# Default scoring weights
DEFAULT_SCORING = {
    'exact_sku_match': 100,
    'alias_sku_match': 90,
    'partial_sku_match': 50,
    'catalog_match': 80,
    'short_code_match': 60,
    'exact_quantity_match': 20,
    'close_quantity_match': 10,
    'exact_price_match': 30,
    'close_price_match': 15,
    'exact_total_match': 25,
    'match_threshold': 30,
}


class POReader:
    """Reads and parses PO XLSX files into structured data."""

    @staticmethod
    def _detect_xlsx_format(ws) -> str:
        """Detect XLSX format from header row.

        Returns 'po' for classic PO format or 'commercial_invoice' for
        Odoo commercial invoice export format.
        """
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c).lower() if c else '' for c in row]
            # Commercial invoice format has "Number" in col 0 and
            # "Invoice lines/Product/Internal Reference" in col 4
            if headers and headers[0] == 'number' and any('invoice lines' in h for h in headers):
                return 'commercial_invoice'
        return 'po'

    @staticmethod
    def read_po_xlsx(po_path: str) -> List[Dict]:
        """
        Read PO XLSX and return list of PO item dicts.

        Supports two formats:
          1. Classic PO format (10 cols): Partner, Product/Internal Reference,
             Description, Demand, Unit Price, Total, Combined Ratio,
             Order Reference, UOM Combined, UoM
          2. Commercial invoice export (12 cols): Number, Reference,
             Invoice Partner Display Name, Company Name,
             Invoice lines/Product/Internal Reference,
             Invoice lines/Product/Name, Country of Origin, Vendor,
             Unit Price, Quantity, Subtotal, Total

        Args:
            po_path: Path to PO XLSX file

        Returns:
            List of PO item dicts with keys: supplier, product_ref,
            short_code, supplier_sku, description, clean_desc, demand,
            unit_price, total, ratio, order_ref, po_number, invoice_refs, uom
        """
        if not openpyxl:
            raise ImportError("openpyxl is required. Run: pip install openpyxl")

        wb = openpyxl.load_workbook(po_path)
        ws = wb.active

        fmt = POReader._detect_xlsx_format(ws)
        if fmt == 'commercial_invoice':
            items = POReader._read_commercial_invoice_xlsx(ws)
        else:
            items = POReader._read_classic_po_xlsx(ws)

        wb.close()
        return items

    @staticmethod
    def _read_commercial_invoice_xlsx(ws) -> List[Dict]:
        """Parse Odoo commercial invoice export format.

        Columns: Number(0), Reference(1), Partner(2), Company(3),
        Product/Internal Reference(4), Product/Name(5),
        Country of Origin(6), Vendor(7), Unit Price(8),
        Quantity(9), Subtotal(10), Total(11)
        """
        po_items = []
        current_invoice = None
        current_po_ref = None
        current_supplier = None

        for row in ws.iter_rows(values_only=True, min_row=2):
            if len(row) < 10:
                continue

            inv_num = row[0]
            reference = row[1]
            partner = row[2]

            # Invoice number only appears on first row of each invoice group
            if inv_num:
                current_invoice = str(inv_num)
                current_po_ref = str(reference) if reference else ''
                if partner:
                    current_supplier = str(partner)

            if not current_invoice:
                continue

            product_ref = str(row[4]) if row[4] else ''
            description = str(row[5]) if row[5] else ''
            unit_price = row[8] if isinstance(row[8], (int, float)) else 0
            quantity = row[9] if isinstance(row[9], (int, float)) else 0
            subtotal = row[10] if isinstance(row[10], (int, float)) else 0

            if not description or not quantity:
                continue

            # Build invoice_refs from the invoice number
            # PDF filename: INV_SX_2026_004235.pdf  ←  XLSX: INV/SX/2026/004235
            inv_ref_underscore = current_invoice.replace('/', '_')
            invoice_refs = [inv_ref_underscore]

            # Extract PO number from Reference field (e.g. "PO/GD/07089- TBM TROPICAL")
            po_number = ''
            if current_po_ref:
                po_match = re.match(r'(PO/\w+/\d+)', current_po_ref)
                if po_match:
                    po_number = po_match.group(1)

            # Extract short code from product_ref
            if '/' in product_ref:
                short_code = product_ref.split('/')[-1]
            else:
                short_code = product_ref

            # For commercial invoices, supplier_sku is the product_ref itself
            supplier_sku = short_code

            po_items.append({
                'supplier': current_supplier,
                'product_ref': product_ref,
                'short_code': short_code,
                'supplier_sku': supplier_sku,
                'description': description,
                'clean_desc': description,
                'demand': quantity,
                'unit_price': unit_price,
                'total': subtotal,
                'ratio': 1,
                'order_ref': current_po_ref,
                'po_number': po_number,
                'invoice_refs': invoice_refs,
                'uom': 'Unit',
                'matched': False,
            })

        return po_items

    @staticmethod
    def _read_classic_po_xlsx(ws) -> List[Dict]:
        """Parse classic PO format (10 columns)."""
        po_items = []
        current_supplier = None

        for row in ws.iter_rows(values_only=True, min_row=2):
            if len(row) < 10:
                continue

            partner, product_ref, description, demand, unit_price, total, \
                ratio, order_ref, uom_combined, uom = row[:10]

            if partner:
                current_supplier = partner
            if not description or not isinstance(demand, (int, float)):
                continue

            # Extract PO number and invoice refs from order_ref
            # Format: "PO/GD/07178 (0098813384)" → po_number="PO/GD/07178", invoice_refs=["0098813384"]
            invoice_refs = []
            po_number = ""
            if order_ref:
                order_ref_str = str(order_ref)
                po_match = re.match(r'(PO/GD/\d+)', order_ref_str)
                if po_match:
                    po_number = po_match.group(1)
                ref_match = re.search(r'\(([^)]+)\)', order_ref_str)
                if ref_match:
                    refs_str = ref_match.group(1)
                    for ref in refs_str.split('/'):
                        invoice_refs.append(ref.strip())

            # Extract the short item code from product_ref (part after last /)
            product_ref_str = str(product_ref) if product_ref else ""
            if '/' in product_ref_str:
                short_code = product_ref_str.split('/')[-1]
            else:
                short_code = product_ref_str

            # Extract supplier's item number from description brackets [XXX]
            desc_str = str(description) if description else ""
            bracket_match = re.search(r'\[([^\]]+)\]', desc_str)
            supplier_sku = bracket_match.group(1) if bracket_match else short_code

            # Clean description
            clean_desc = re.sub(r'\[[^\]]*\]\s*', '', desc_str).strip()
            clean_desc = re.sub(r'\s*\[\d+\s*\|\s*\w+\]', '', clean_desc).strip()

            po_items.append({
                'supplier': current_supplier,
                'product_ref': product_ref_str,
                'short_code': short_code,
                'supplier_sku': supplier_sku,
                'description': desc_str,
                'clean_desc': clean_desc,
                'demand': demand,
                'unit_price': unit_price if isinstance(unit_price, (int, float)) else 0,
                'total': total if isinstance(total, (int, float)) else 0,
                'ratio': ratio if isinstance(ratio, (int, float)) else 1,
                'order_ref': str(order_ref) if order_ref else "",
                'po_number': po_number,
                'invoice_refs': invoice_refs,
                'uom': uom or "Unit",
                'matched': False,
            })

        return po_items


class POMatcher:
    """
    Matches parsed invoice items to Purchase Order items using scoring.

    The scoring algorithm compares each PDF-extracted item against candidate
    PO items (filtered by invoice reference) and picks the best match above
    a configurable threshold.
    """

    def __init__(self, po_items: List[Dict], scoring_config: Dict = None,
                 base_dir: str = '.'):
        """
        Args:
            po_items: List of PO item dicts from POReader.read_po_xlsx()
            scoring_config: Optional override for scoring weights/thresholds
            base_dir: Project root for loading alias data
        """
        self.po_items = po_items
        self.scoring = {**DEFAULT_SCORING, **(scoring_config or {})}
        self.base_dir = base_dir
        self._aliases = _get_aliases(base_dir)

        # Build supplier → invoice ref mapping
        self._supplier_map = {}
        for item in po_items:
            s = item['supplier']
            if s not in self._supplier_map:
                self._supplier_map[s] = {'refs': set()}
            for ref in item['invoice_refs']:
                self._supplier_map[s]['refs'].add(re.sub(r'-[A-Z]$', '', ref))

    def determine_supplier(self, pdf_filename: str) -> Optional[str]:
        """
        Determine supplier name from PDF filename using PO invoice_refs.

        Args:
            pdf_filename: PDF filename (e.g., '8444174.pdf')

        Returns:
            Supplier name or None
        """
        invoice_ref = self._clean_invoice_ref(pdf_filename)
        norm_ref = invoice_ref.replace('-', '')

        # Exact match first
        for supplier, data in self._supplier_map.items():
            if invoice_ref in data['refs']:
                return supplier

        # Normalized match (ignore hyphens)
        for supplier, data in self._supplier_map.items():
            for ref in data['refs']:
                norm_r = ref.replace('-', '')
                if norm_r == norm_ref:
                    return supplier

        # Partial match
        for supplier, data in self._supplier_map.items():
            for ref in data['refs']:
                norm_r = ref.replace('-', '')
                if (invoice_ref in ref or ref in invoice_ref or
                        norm_ref in norm_r or norm_r in norm_ref):
                    return supplier

        return None

    def match_invoice(self, invoice_data: Dict, pdf_filename: str) -> List[Dict]:
        """
        Match a parsed invoice's items against PO items.

        Args:
            invoice_data: Parsed invoice data with 'items' list.
                          Each item should have: supplier_item (or sku),
                          description, quantity, unit_price (or unit_cost), total (or total_cost)
            pdf_filename: PDF filename for invoice ref lookup

        Returns:
            List of matched items with both PO and supplier data
        """
        invoice_ref = self._clean_invoice_ref(pdf_filename)
        candidate_po_items = self._find_candidate_po_items(invoice_ref)
        pdf_items = invoice_data.get('items', [])
        matched_items = []

        if not pdf_items:
            # No items extracted from PDF — use PO items directly
            for po_item in candidate_po_items:
                refs_clean = [re.sub(r'-[A-Z]$', '', r) for r in po_item['invoice_refs']]
                if len(refs_clean) == 1 or invoice_ref == refs_clean[0]:
                    matched_items.append({
                        'po_item_ref': po_item['product_ref'],
                        'po_item_desc': po_item['clean_desc'],
                        'po_number': po_item['po_number'],
                        'supplier_item': po_item['supplier_sku'],
                        'supplier_item_desc': po_item['clean_desc'],
                        'quantity': po_item['demand'],
                        'unit_price': po_item['unit_price'],
                        'total_cost': po_item['total'],
                        'uom': po_item['uom'],
                        'match_score': 0,
                    })
            return matched_items

        # Two-pass matching: SKU matches first, then price/qty matches.
        # This prevents price-only matches from stealing PO entries that
        # should go to items with exact SKU matches.
        used_po = set()
        used_pdf = set()
        sku_threshold = self.scoring['exact_sku_match']  # 100
        threshold = self.scoring['match_threshold']
        match_results = {}  # pdf_idx -> (po_idx, score)

        # Pass 1: Match items that have SKU-based matches (score >= sku_threshold)
        for pdf_idx, pdf_item in enumerate(pdf_items):
            best_match = None
            best_score = 0
            for po_idx, po_item in enumerate(candidate_po_items):
                if po_idx in used_po:
                    continue
                score = self._score_match(pdf_item, po_item)
                if score > best_score:
                    best_score = score
                    best_match = po_idx
            if best_match is not None and best_score >= sku_threshold:
                used_po.add(best_match)
                used_pdf.add(pdf_idx)
                match_results[pdf_idx] = (best_match, best_score)

        # Pass 2: Match remaining items by overall scoring
        for pdf_idx, pdf_item in enumerate(pdf_items):
            if pdf_idx in used_pdf:
                continue
            best_match = None
            best_score = 0
            for po_idx, po_item in enumerate(candidate_po_items):
                if po_idx in used_po:
                    continue
                score = self._score_match(pdf_item, po_item)
                if score > best_score:
                    best_score = score
                    best_match = po_idx
            if best_match is not None and best_score >= threshold:
                used_po.add(best_match)
                used_pdf.add(pdf_idx)
                match_results[pdf_idx] = (best_match, best_score)

        # Build result list in original PDF item order
        for pdf_idx, pdf_item in enumerate(pdf_items):
            if pdf_idx in match_results:
                po_idx, best_score = match_results[pdf_idx]
                po_item = candidate_po_items[po_idx]

                supplier_item = (pdf_item.get('supplier_item') or
                                 pdf_item.get('sku') or
                                 f'ITEM-{pdf_idx + 1:03d}')
                supplier_desc = (pdf_item.get('description') or
                                 po_item['clean_desc'])
                unit_price = (pdf_item.get('unit_price') or
                              pdf_item.get('unit_cost') or
                              po_item['unit_price'])
                total_cost = (pdf_item.get('total') or
                              pdf_item.get('total_cost') or
                              po_item['total'])

                matched_items.append({
                    'po_item_ref': po_item['product_ref'],
                    'po_item_desc': po_item['clean_desc'],
                    'po_number': po_item['po_number'],
                    'supplier_item': supplier_item,
                    'supplier_item_desc': supplier_desc,
                    'quantity': pdf_item.get('quantity', po_item['demand']),
                    'unit_price': unit_price,
                    'total_cost': total_cost,
                    'uom': po_item['uom'],
                    'match_score': best_score,
                })
            else:
                # Unmatched PDF item — include with empty PO fields
                supplier_item = (pdf_item.get('supplier_item') or
                                 pdf_item.get('sku') or
                                 f'ITEM-{pdf_idx + 1:03d}')
                unit_price = (pdf_item.get('unit_price') or
                              pdf_item.get('unit_cost') or 0)
                total_cost = (pdf_item.get('total') or
                              pdf_item.get('total_cost') or 0)

                matched_items.append({
                    'po_item_ref': '',
                    'po_item_desc': '',
                    'po_number': '',
                    'supplier_item': supplier_item,
                    'supplier_item_desc': pdf_item.get('description', ''),
                    'quantity': pdf_item.get('quantity', 0),
                    'unit_price': unit_price,
                    'total_cost': total_cost,
                    'uom': 'Unit',
                    'match_score': 0,
                })

        return matched_items

    def _find_candidate_po_items(self, invoice_ref: str) -> List[Dict]:
        """Find PO items that reference this invoice."""
        candidates = []
        # Normalize: strip hyphens/dashes for comparison
        # e.g. PDF filename "2026001763" should match PO ref "2026-001763"
        norm_ref = invoice_ref.replace('-', '')
        for po_item in self.po_items:
            for ref in po_item['invoice_refs']:
                clean_ref = re.sub(r'-[A-Z]$', '', ref)
                norm_clean = clean_ref.replace('-', '')
                if (clean_ref == invoice_ref or
                        invoice_ref in clean_ref or
                        clean_ref in invoice_ref or
                        norm_clean == norm_ref or
                        norm_ref in norm_clean or
                        norm_clean in norm_ref):
                    candidates.append(po_item)
                    break
        return candidates

    def _score_match(self, pdf_item: Dict, po_item: Dict) -> int:
        """Calculate match score between a PDF item and a PO item."""
        score = 0
        scoring = self.scoring

        # Get normalized SKUs
        po_sku = self._normalize_sku(po_item['supplier_sku'])
        pdf_sku = self._normalize_sku(
            pdf_item.get('supplier_item') or pdf_item.get('sku') or ''
        )

        # SKU matching (require min 3 chars for substring matches to
        # prevent false positives from very short parsed SKUs like "8")
        sku_matched = False
        if po_sku and pdf_sku:
            if po_sku == pdf_sku:
                score += scoring['exact_sku_match']
                sku_matched = True
            elif (min(len(po_sku), len(pdf_sku)) >= 3 and
                  (po_sku in pdf_sku or pdf_sku in po_sku)):
                score += scoring['partial_sku_match']
                sku_matched = True
            elif self._aliases and self._check_alias_match(pdf_sku, po_sku):
                score += scoring['alias_sku_match']
                sku_matched = True

        # Legacy SKU matching (e.g. Hilco "Legacy Item #" = old supplier code)
        # Compare against both full po_sku and short_code (strips BM prefix like CRO/)
        if not sku_matched:
            legacy_sku = self._normalize_sku(pdf_item.get('legacy_sku', ''))
            if legacy_sku:
                po_short = self._normalize_sku(po_item['short_code'])
                if (po_sku and legacy_sku == po_sku) or \
                   (po_short and legacy_sku == po_short):
                    score += scoring['exact_sku_match']
                    sku_matched = True

        # Catalog number match (for formats like Donovan with catalog|code)
        catalog_num = pdf_item.get('catalog_num', '')
        if catalog_num and len(catalog_num) >= 3:
            if catalog_num in po_item['supplier_sku'] or \
               catalog_num in po_item['description']:
                score += scoring['catalog_match']

        # Short code match (same min-length guard)
        po_short = self._normalize_sku(po_item['short_code'])
        if po_short and pdf_sku:
            if (po_short == pdf_sku or
                    (min(len(po_short), len(pdf_sku)) >= 3 and
                     (po_short in pdf_sku or pdf_sku in po_short))):
                score += scoring['short_code_match']

        # Quantity match
        pdf_qty = pdf_item.get('quantity', 0)
        po_qty = po_item.get('demand', 0)
        if pdf_qty and po_qty:
            if pdf_qty == po_qty:
                score += scoring['exact_quantity_match']
            elif abs(pdf_qty - po_qty) <= 1:
                score += scoring['close_quantity_match']

        # Unit price match
        pdf_price = pdf_item.get('unit_price') or pdf_item.get('unit_cost') or 0
        po_price = po_item.get('unit_price', 0)
        if pdf_price and po_price:
            if abs(pdf_price - po_price) < 0.02:
                score += scoring['exact_price_match']
            elif abs(pdf_price - po_price) < 0.50:
                score += scoring['close_price_match']

        # Total match
        pdf_total = pdf_item.get('total') or pdf_item.get('total_cost') or 0
        po_total = po_item.get('total', 0)
        if pdf_total and po_total:
            if abs(pdf_total - po_total) < 0.02:
                score += scoring['exact_total_match']

        return score

    def _check_alias_match(self, pdf_sku: str, po_sku: str) -> bool:
        """Check if pdf_sku and po_sku map to each other via the alias table.

        The alias table maps BM internal codes to supplier codes (prefix/item).
        A match means: one SKU is the BM code and the other is the supplier code
        (or vice versa) for the same product.
        """
        # Check if pdf_sku resolves to po_sku's BM code (or vice versa)
        for sku_a, sku_b in [(pdf_sku, po_sku), (po_sku, pdf_sku)]:
            entries = self._aliases.get(sku_a, [])
            for entry in entries:
                bm_norm = entry['bm_code'].upper().replace('-', '').replace(' ', '')
                supplier_norm = entry['supplier_item'].upper().replace('-', '').replace(' ', '')
                if sku_b == bm_norm or sku_b == supplier_norm:
                    return True
        return False

    @staticmethod
    def _normalize_sku(sku: str) -> str:
        """Normalize SKU for comparison."""
        if not sku:
            return ''
        return sku.upper().replace('-', '').replace(' ', '')

    @staticmethod
    def _clean_invoice_ref(pdf_filename: str) -> str:
        """Extract clean invoice reference from PDF filename."""
        invoice_ref = os.path.splitext(os.path.basename(pdf_filename))[0]
        invoice_ref = re.sub(r'\s*\(\d+\)$', '', invoice_ref)
        return invoice_ref
