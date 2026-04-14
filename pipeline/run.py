#!/usr/bin/env python3
"""
Unified pipeline entry point.

Auto-detects workflow type from input:
  - Folder with BL PDF + PO XLSX → BL batch processing (multi-invoice)
  - Folder without BL            → Batch mode (each PDF individually)
  - Single PDF file              → Single invoice processing

Usage:
    # BL mode (auto-detected: folder has BL PDF + PO XLSX)
    python pipeline/run.py --input-dir workspace/documents/Fw_BL_TSCW18489131

    # Single invoice mode
    python pipeline/run.py --input invoice.pdf --output output.xlsx

    # With email sending
    python pipeline/run.py --input-dir workspace/documents/Fw_BL_TSCW18489131 --send-email

    # JSON output for TypeScript consumption
    python pipeline/run.py --input-dir folder --json-output
"""

import argparse
import json
import logging
import os
import re
import sys
from typing import Optional

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Add pipeline directory to path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

from stages.supplier_resolver import (
    init as init_resolver,
    load_supplier_db,
    save_supplier_db,
    load_classification_rules,
    find_po_file,
    find_bl_pdf,
    get_pdf_files,
    classify_input_pdfs,
    extract_pdf_text,
    extract_manifest_metadata,
    extract_freight_invoice_data,
)
from stages.invoice_processor import process_single_invoice, InvoiceResult
from stages.bl_allocator import allocate_bl_packages, BLAllocation
import send_history


def _lookup_consignee_code(consignee_name: str) -> tuple:
    """
    Look up consignee code and address from config/shipment_rules.yaml.

    Normalises names for matching: lowercased, '&' ↔ 'and', strips punctuation.
    Returns (code, address) or ('', '') if no match.
    """
    if not consignee_name:
        return ('', '')

    rules_path = os.path.join(BASE_DIR, 'config', 'shipment_rules.yaml')
    try:
        import yaml
        with open(rules_path, 'r') as f:
            rules = yaml.safe_load(f) or {}
    except Exception:
        return ('', '')

    consignees = rules.get('consignees', {})
    if not consignees:
        return ('', '')

    def _normalise(s):
        s = s.lower().strip()
        s = s.replace('&', ' and ').replace('.', ' ').replace("'", '')
        # collapse whitespace
        return ' '.join(s.split())

    query = _normalise(consignee_name)
    # Split into significant words for partial matching
    query_words = set(query.split()) - {'and', 'the', 'of', 'inc', 'ltd', 'co'}

    best_code = ''
    best_address = ''
    best_score = 0

    for code, info in consignees.items():
        name = info.get('name', '')
        norm = _normalise(name)
        norm_words = set(norm.split()) - {'and', 'the', 'of', 'inc', 'ltd', 'co'}

        # Exact normalised match
        if query == norm:
            return (str(code), info.get('address', ''))

        # Word overlap score — require significant overlap
        if query_words and norm_words:
            overlap = query_words & norm_words
            score = len(overlap) / max(len(query_words), len(norm_words))
            if score > best_score and score >= 0.5:
                best_score = score
                best_code = str(code)
                best_address = info.get('address', '')

    if best_code:
        logger.info(f"Consignee '{consignee_name}' matched shipment_rules → code={best_code}")
    return (best_code, best_address)


def _inject_handwritten_duties(decl_meta: dict, results: list) -> bool:
    """Inject handwritten customs duty values into invoice_data for XLSX duty estimation.

    Declaration metadata from pdf_splitter.extract_declaration_handwriting() stores
    handwritten customs values in _customs_value_ec (EC$) and _customs_value_usd.
    The XLSX duty estimation section reads invoice_data['_client_declared_duties']
    (expected in XCD/EC$) and invoice_data['_customs_freight'].

    This function bridges the two: it converts the extracted handwritten value to a
    float and injects it into each result's invoice_data so the duty estimation
    section can display the client-vs-estimated variance.

    Returns True if any values were injected (caller should regenerate XLSX).
    """
    if not decl_meta or not results:
        return False

    # Extract handwritten duty value (prefer EC$ since ASYCUDA duties are in XCD)
    client_duties = None
    raw_ec = decl_meta.get('_customs_value_ec')
    raw_usd = decl_meta.get('_customs_value_usd')

    if raw_ec:
        try:
            client_duties = float(str(raw_ec).replace(',', '').replace('$', '').strip())
        except (ValueError, TypeError):
            pass

    # If only USD value available, convert to XCD
    if client_duties is None and raw_usd:
        try:
            usd_val = float(str(raw_usd).replace(',', '').replace('$', '').strip())
            # XCD_RATE = 2.7169 (Eastern Caribbean Dollar per USD)
            client_duties = round(usd_val * 2.7169, 2)
            logger.info(f"Converted handwritten USD ${usd_val} to EC$ {client_duties}")
        except (ValueError, TypeError):
            pass

    if client_duties is None or client_duties <= 0:
        return False

    # Extract freight from declaration metadata for CIF calculation
    decl_freight = None
    raw_freight = decl_meta.get('freight')
    if raw_freight:
        try:
            decl_freight = float(str(raw_freight).replace(',', '').strip())
        except (ValueError, TypeError):
            pass

    logger.info(f"Injecting handwritten client duties EC${client_duties} into "
                f"{len(results)} invoice result(s)")

    injected = False
    for r in results:
        inv_data = getattr(r, 'invoice_data', None)
        if inv_data is None:
            continue
        # Only set if not already populated (don't overwrite manual/other sources)
        if inv_data.get('_client_declared_duties') is None:
            inv_data['_client_declared_duties'] = client_duties
            injected = True
        if decl_freight is not None and inv_data.get('_customs_freight') is None:
            inv_data['_customs_freight'] = decl_freight
    return injected


def _extract_consignee(args) -> str:
    """
    Extract consignee name from all available document sources.

    Priority order:
      1. Manifest metadata (consignee_name) — from ASYCUDA waybill manifest
      2. Declaration metadata (consignee) — from simplified declaration / pdf_splitter
      3. Bill of Lading PDF — parsed on demand via bl_parser
    """
    decl_meta = getattr(args, '_declaration_metadata', {})

    # Placeholder values that mean "no consignee" — skip these
    _PLACEHOLDER_CONSIGNEES = {'SAME AS CONSIGNEE', 'SAME AS SHIPPER', 'SAME AS ABOVE',
                               'AS PER SHIPPER', 'TO ORDER', 'TO THE ORDER OF',
                               'NOTIFY', 'NOTIFY:', 'NOTIFY PARTY', 'NOTIFY PARTY:'}

    def _is_valid_consignee(name: str) -> bool:
        """Reject names that are clearly not consignee names."""
        if not name or len(name) < 3:
            return False
        upper = name.upper().strip()
        if upper in _PLACEHOLDER_CONSIGNEES:
            return False
        # Reject if it starts with "Notify" — OCR artifact from manifest
        if re.match(r'^NOTIFY\b', upper):
            return False
        return True

    # 1. Manifest metadata (set by extract_manifest_metadata in _prepare_invoice_pdfs)
    consignee = (decl_meta.get('consignee_name') or '').strip()
    if _is_valid_consignee(consignee):
        logger.info(f"Consignee from manifest: {consignee}")
        return consignee

    # 2. Declaration metadata (set by extract_declaration_metadata in _prepare_invoice_pdfs)
    consignee = (decl_meta.get('consignee') or '').strip()
    if _is_valid_consignee(consignee):
        logger.info(f"Consignee from declaration: {consignee}")
        return consignee

    # 3. Bill of Lading PDF — parse it for consignee
    classification = getattr(args, '_classification', {})
    bl_files = classification.get('bill_of_lading', [])
    if bl_files and hasattr(args, 'input_dir'):
        bl_path = os.path.join(args.input_dir, bl_files[0])
        if os.path.exists(bl_path):
            try:
                from bl_parser import parse_bl_pdf
                bl_data = parse_bl_pdf(bl_path)
                consignee = (bl_data.get('consignee') or '').strip()
                if consignee:
                    logger.info(f"Consignee from Bill of Lading: {consignee}")
                    return consignee
            except Exception as e:
                logger.debug(f"BL consignee extraction failed: {e}")

            # Fallback: extract consignee from BL text via regex
            try:
                from stages.supplier_resolver import extract_pdf_text
                bl_text = extract_pdf_text(bl_path)
                if bl_text:
                    # Match "CONSIGNEE(NOT NEGOTIABLE...)\nNAME" — the header section only
                    consignee_match = re.search(
                        r'CONSIGNEE\s*\(NOT NEGOTIABLE[^)]*\)[^\n]*\n\s*(.+)',
                        bl_text, re.IGNORECASE
                    )
                    if not consignee_match:
                        # Simpler: "CONSIGNEE\nNAME" (header on its own line)
                        consignee_match = re.search(
                            r'^CONSIGNEE\s*$[^\S\n]*\n\s*(.+)',
                            bl_text, re.IGNORECASE | re.MULTILINE
                        )
                    if consignee_match:
                        name = consignee_match.group(1).strip().split('\n')[0].strip()
                        if name and len(name) > 2 and not re.match(r'^\d', name):
                            logger.info(f"Consignee from BL text: {name}")
                            return name
            except Exception as e:
                logger.debug(f"BL text consignee extraction failed: {e}")

    # 4. Declaration PDFs not yet parsed via pdf_splitter — try them
    decl_files = classification.get('declaration', [])
    if decl_files and hasattr(args, 'input_dir'):
        try:
            from pdf_splitter import extract_declaration_metadata
            for df in decl_files:
                decl_path = os.path.join(args.input_dir, df)
                if os.path.exists(decl_path):
                    meta = extract_declaration_metadata(decl_path)
                    consignee = (meta.get('consignee') or '').strip()
                    if consignee:
                        logger.info(f"Consignee from declaration PDF ({df}): {consignee}")
                        return consignee
        except Exception as e:
            logger.debug(f"Declaration consignee extraction failed: {e}")

    # 5. Invoice PDFs — extract consignee name from "Ship To" / "Bill To" in text
    #    Handles two common layouts:
    #    (a) "Ship To:\n  Company Name\n  Address" — name is AFTER the label
    #    (b) "Company Name\n  BILL TO:  SHIP TO:\n  City" — name is BEFORE the label (multi-column)
    invoice_files = classification.get('invoice', [])
    if invoice_files and hasattr(args, 'input_dir'):
        try:
            from stages.supplier_resolver import extract_pdf_text
            for inv_file in invoice_files[:3]:  # Check first 3 invoices max
                inv_path = os.path.join(args.input_dir, inv_file)
                if not os.path.exists(inv_path):
                    continue
                text = extract_pdf_text(inv_path)
                if not text:
                    continue

                # Pattern (b): multi-column layout where company name is on the
                # line BEFORE "BILL TO: SHIP TO:" — e.g.:
                #   BEAUTY EXPO BEAUTY EXPO
                #   BILL TO: SHIP TO:
                #   ST. GEORGE'S ST. GEORGE'S
                multi_col = re.search(
                    r'^(.+)\n\s*BILL\s*TO\s*:\s+SHIP\s*TO\s*:',
                    text, re.IGNORECASE | re.MULTILINE
                )
                if multi_col:
                    # The line before contains the company name (possibly duplicated for both columns)
                    raw = multi_col.group(1).strip()
                    # De-duplicate: "BEAUTY EXPO BEAUTY EXPO" → "BEAUTY EXPO"
                    words = raw.split()
                    half = len(words) // 2
                    if half > 0 and words[:half] == words[half:]:
                        raw = ' '.join(words[:half])
                    if raw and not re.match(r'^[\d#]', raw) and len(raw) > 2:
                        logger.info(f"Consignee from invoice multi-column Ship-To ({inv_file}): {raw}")
                        return raw

                # Pattern (a-sold): "SOLD TO:" or "BILL TO:" followed by name on next line
                # Handles: "SOLD TO:\nBEAUTY EXPO INVOICE NUMBER 91005"
                sold_to_match = re.search(
                    r'(?:SOLD\s+TO|BILL\s+TO)\s*[:\-]?\s*\n\s*(.+)',
                    text, re.IGNORECASE
                )
                if sold_to_match:
                    name = sold_to_match.group(1).strip()
                    # Strip trailing metadata on same line (e.g. "BEAUTY EXPO INVOICE NUMBER 91005")
                    name = re.split(r'\s+(?:INVOICE|ORDER|DATE|ACCOUNT|PO\s*#|REF)', name, flags=re.IGNORECASE)[0].strip()
                    if name and not re.match(r'^[\d#]', name) and len(name) > 2:
                        logger.info(f"Consignee from invoice Sold-To ({inv_file}): {name}")
                        return name

                # Pattern (a): "Ship To:", "Deliver To:", etc. followed by name on the next line
                ship_to_match = re.search(
                    r'(?:Ship\s*(?:ped\s+)?To|Deliver(?:y)?\s*(?:To|Address)|'
                    r'Shipping\s+Address|Consignee)\s*[:\-]?\s*\n\s*(.+)',
                    text, re.IGNORECASE
                )
                if ship_to_match:
                    name = ship_to_match.group(1).strip()
                    # Clean up: take first line only, strip c/o suffix for cleaner matching
                    name = name.split('\n')[0].strip()
                    # Skip if it's just a number/PO Box/street (not a name)
                    if name and not re.match(r'^[\d#]', name) and len(name) > 2:
                        logger.info(f"Consignee from invoice Ship-To ({inv_file}): {name}")
                        # Extract address from subsequent lines of Ship To block
                        full_block = re.search(
                            r'(?:Ship\s*(?:ped\s+)?To|Deliver(?:y)?\s*(?:To|Address)|'
                            r'Shipping\s+Address|Consignee)\s*[:\-]?\s*\n'
                            r'\s*(.+(?:\n.+)*?)'
                            r'\n\s*(?:PRO|Article|BOL|Item|Weight|Carrier|$)',
                            text, re.IGNORECASE
                        )
                        if full_block:
                            lines = [ln.strip() for ln in full_block.group(1).split('\n') if ln.strip()]
                            if len(lines) > 1:
                                addr = ', '.join(lines[1:])
                                args._invoice_consignee_address = addr
                                logger.info(f"Consignee address from invoice: {addr}")
                        return name
        except Exception as e:
            logger.debug(f"Invoice ship-to extraction failed: {e}")

    return ''


def _extract_supplier_for_doc_type(args) -> str:
    """Extract supplier name from invoice PDFs for doc-type rule matching."""
    classification = getattr(args, '_classification', {})
    invoice_files = classification.get('invoice', [])
    if not invoice_files or not hasattr(args, 'input_dir'):
        return ''

    try:
        from stages.supplier_resolver import extract_pdf_text
        from format_registry import FormatRegistry

        registry = FormatRegistry(BASE_DIR)

        # Prioritize actual invoice PDFs over packing lists, pallet lists, etc.
        skip_patterns = ('packing', 'pallet', 'caricom', 'manifest', 'declaration')
        actual_invoices = [f for f in invoice_files
                          if not any(p in f.lower() for p in skip_patterns)]
        # Fall back to all files if no actual invoices found
        candidates = (actual_invoices or invoice_files)[:5]

        for inv_file in candidates:
            inv_path = os.path.join(args.input_dir, inv_file)
            if not os.path.exists(inv_path):
                continue
            text = extract_pdf_text(inv_path)
            if not text:
                continue

            # Try format detection — spec may define supplier_name
            spec = registry.detect_format(text)
            if spec:
                meta = spec.get('metadata', {})
                supplier_spec = meta.get('supplier_name', {})
                if isinstance(supplier_spec, dict) and supplier_spec.get('value'):
                    return supplier_spec['value']

            # If spec has supplier_name with patterns/fallback, use the fallback
            if isinstance(supplier_spec, dict) and supplier_spec.get('fallback'):
                return supplier_spec['fallback']

            # Fallback: check for known supplier patterns in text
            for rule_text in ['Budget Marine', 'Amazon', 'Walmart', 'West Marine',
                              'Reef Lifestyle', 'Flip Flop']:
                if rule_text.lower() in text.lower():
                    return rule_text
    except Exception as e:
        logger.debug(f"Supplier extraction for doc-type failed: {e}")

    return ''


def _match_rule(rule, name_lower: str) -> bool:
    """Check if a name matches a rule's primary match string or any alias."""
    # Check primary match
    match_str = rule.get('match', '').lower()
    if match_str and match_str in name_lower:
        return True
    # Check aliases
    for alias in rule.get('aliases', []):
        if alias.lower() in name_lower:
            return True
    return False


def resolve_doc_type(args) -> str:
    """
    Auto-resolve document type from consignee/supplier name using config/document_types.json rules.

    Checks consignee name (from manifest, declaration, BL) against each rule's
    primary match string and aliases. Falls back to supplier name from invoice PDFs.
    If --doc-type was explicitly set (not 'auto'), keep the user's choice.
    Falls back to the config default (4000-000).
    """
    # User explicitly specified a doc type — honour it
    if args.doc_type != 'auto':
        return args.doc_type

    config_path = os.path.join(BASE_DIR, 'config', 'document_types.json')
    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return '4000-000'

    default_type = config.get('default', '4000-000')
    rules = config.get('consignee_rules', [])

    # Check consignee first
    consignee = _extract_consignee(args)
    consignee_lower = consignee.lower().strip()

    if consignee_lower:
        for rule in rules:
            if _match_rule(rule, consignee_lower):
                logger.info(f"Consignee '{consignee}' matched rule '{rule.get('match')}' → doc_type={rule['doc_type']}")
                args._matched_consignee_rule = rule
                args._consignee_match_type = 'consignee'
                return rule['doc_type']

    # Check supplier name from invoice PDFs as fallback.
    # IMPORTANT: A supplier match only determines doc_type — it does NOT
    # mean the *consignee* is the same entity.  E.g. Budget Marine may be
    # the supplier/seller while the actual consignee (Ship-To) is "FLIP FLOP".
    supplier = _extract_supplier_for_doc_type(args)
    supplier_lower = supplier.lower().strip()

    if supplier_lower:
        for rule in rules:
            if _match_rule(rule, supplier_lower):
                logger.info(f"Supplier '{supplier}' matched rule '{rule.get('match')}' → doc_type={rule['doc_type']} (supplier-only match, consignee metadata NOT applied)")
                args._matched_consignee_rule = rule
                args._consignee_match_type = 'supplier'
                return rule['doc_type']

    if not consignee_lower and not supplier_lower:
        logger.info(f"No consignee/supplier found in any document → doc_type={default_type}")
    else:
        logger.info(f"Consignee '{consignee}', Supplier '{supplier}' — no rule matched → doc_type={default_type}")
    return default_type


def _maybe_combine_entries(args, results, output_dir):
    """
    If combine_entries is enabled for this consignee and there are multiple
    invoice results, combine all XLSX files into a single combined XLSX.
    Returns the (possibly modified) results list.
    """
    if len(results) < 2:
        return results

    # Check combine_entries from matched consignee rule or config default
    config_path = os.path.join(BASE_DIR, 'config', 'document_types.json')
    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        config = {}

    combine_default = config.get('combine_entries_default', True)
    matched_rule = getattr(args, '_matched_consignee_rule', None)
    combine = matched_rule.get('combine_entries', combine_default) if matched_rule else combine_default

    if not combine:
        logger.info("combine_entries=false — keeping separate XLSX files")
        return results

    # Collect XLSX paths from results, deduplicate by absolute path
    xlsx_paths_raw = [r.xlsx_path for r in results if r.xlsx_path and os.path.exists(r.xlsx_path)]
    seen = set()
    xlsx_paths = []
    for p in xlsx_paths_raw:
        ap = os.path.abspath(p)
        if ap not in seen:
            seen.add(ap)
            xlsx_paths.append(p)
    if len(xlsx_paths) < 2:
        return results

    # Build combined output path
    bl_number = getattr(args, 'bl', '')
    if not bl_number and results:
        # Use first invoice number when no BL number available
        bl_number = re.sub(r'[<>:"/\\|?*]', '_', results[0].invoice_num)
    if not bl_number:
        bl_number = 'combined'
    combined_path = os.path.join(output_dir, f'{bl_number}_combined.xlsx')

    print(f"\n    Combining {len(xlsx_paths)} invoices into single XLSX...")

    # Collect OCR quality notes from each result for the OCR Notes sheet
    ocr_notes = []
    for r in results:
        inv_data = getattr(r, 'invoice_data', {}) or {}
        ocr_q = inv_data.get('ocr_quality', {})
        if ocr_q:
            ocr_notes.append({
                'pdf_file': getattr(r, 'pdf_file', ''),
                'score': ocr_q.get('score', 0),
                'rating': ocr_q.get('rating', 'unknown'),
                'details': ocr_q.get('details', ''),
                'raw_text': inv_data.get('raw_text', '')[:2000],
            })

    from xlsx_combiner import combine_xlsx_files
    result = combine_xlsx_files(xlsx_paths, combined_path,
                                ocr_notes=ocr_notes if ocr_notes else None)

    if result.get('status') != 'success':
        logger.warning(f"XLSX combine failed: {result.get('error', 'unknown')}")
        return results

    print(f"    Combined XLSX: {os.path.basename(combined_path)}")
    print(f"    Grand Total: ${result['grand_total']:.2f}")
    if abs(result.get('grand_variance', 0)) > 0.01:
        print(f"    ⚠ Grand Variance: ${result['grand_variance']:.2f}")
    else:
        print(f"    Grand Variance: $0.00 ✓")

    # Replace the first result's xlsx_path with the combined file,
    # stash all PDF paths so they're all attached to the email
    all_pdf_paths = []
    total_packages = 0
    total_items = []
    for r in results:
        pdf_path = getattr(r, 'pdf_output_path', None)
        if pdf_path and os.path.exists(pdf_path):
            all_pdf_paths.append(pdf_path)
        total_packages += getattr(r, 'packages', 1)
        total_items.extend(getattr(r, 'matched_items', []))

    combined_result = results[0]
    combined_result.xlsx_path = combined_path
    combined_result._combined_pdf_paths = all_pdf_paths
    combined_result.packages = total_packages

    # Update packages cell in combined XLSX to reflect total
    try:
        from openpyxl import load_workbook as _lw
        wb = _lw(combined_path)
        ws = wb.active
        ws.cell(row=2, column=24, value=total_packages)  # Col X = Packages
        wb.save(combined_path)
        wb.close()
    except Exception as e:
        logger.warning(f"Could not update packages in combined XLSX: {e}")

    # Remove individual XLSX files (now merged into combined)
    for xlsx_path in xlsx_paths:
        if xlsx_path != combined_path and os.path.exists(xlsx_path):
            os.remove(xlsx_path)

    return [combined_result]


def detect_mode(args) -> str:
    """
    Detect processing mode from CLI arguments.

    Content-based detection: reads each PDF to classify as BL, invoice,
    declaration, etc. Stores the classification on args._classification
    so downstream functions don't re-read every PDF.

    Returns: 'bl', 'batch', or 'single'
    """
    # Convert single-file --input to --input-dir mode FIRST,
    # so the classification block below processes it
    if args.input and not args.input_dir:
        input_file = os.path.abspath(args.input)
        input_base = os.path.splitext(os.path.basename(input_file))[0]
        temp_dir = os.path.join(BASE_DIR, 'workspace', 'shipments', f'Shipment_ {input_base}')
        os.makedirs(temp_dir, exist_ok=True)
        dest = os.path.join(temp_dir, os.path.basename(input_file))
        if not os.path.exists(dest):
            import shutil
            shutil.copy2(input_file, dest)
        args.input_dir = temp_dir
        args.input = None
        if not args.output_dir or args.output_dir == os.path.join(BASE_DIR, 'workspace', 'shipments'):
            args.output_dir = temp_dir

    if args.input_dir:
        # Classify all PDFs by content (single pass)
        print("\n[0] Classifying documents by content...")
        classification = classify_input_pdfs(args.input_dir)

        # Check for _bl_hint.json (written by createCombinedFolder in TypeScript)
        # This tells us exactly which PDFs are the real BL, preventing misclassification
        # of shipping docs (e.g. "Tropical Doc") from the invoice folder as BLs.
        bl_hint = _load_bl_hint(args.input_dir)
        if bl_hint:
            hint_bl_pdfs = set(bl_hint.get('bl_pdfs', []))
            if hint_bl_pdfs:
                # Move any misclassified BL files back to invoice if they're NOT in the hint
                real_bls = []
                reclassified = []
                for f in classification.get('bill_of_lading', []):
                    if f in hint_bl_pdfs:
                        real_bls.append(f)
                    else:
                        reclassified.append(f)
                        classification.setdefault('invoice', []).append(f)
                classification['bill_of_lading'] = real_bls

                # Also ensure hint BL files ARE in the BL list (even if classifier missed them)
                existing_bls = set(real_bls)
                for f in hint_bl_pdfs:
                    if f not in existing_bls and os.path.exists(os.path.join(args.input_dir, f)):
                        classification['bill_of_lading'].append(f)
                        # Remove from other categories
                        for cat in ('invoice', 'unknown', 'packing_list', 'declaration', 'manifest'):
                            if f in classification.get(cat, []):
                                classification[cat].remove(f)

                if reclassified:
                    print(f"    BL hint: reclassified {len(reclassified)} file(s) from BL to invoice")
                    for f in reclassified:
                        print(f"      {f}")

                # Apply BL number from hint
                if bl_hint.get('bl_number') and not args.bl:
                    args.bl = bl_hint['bl_number']
                    print(f"    BL number from hint: {args.bl}")

        # Filename-based rescue: scanned declarations/manifests with no extractable
        # text get classified as "unknown". Reclassify by filename before logging.
        _DECL_HINTS = ('declaration', 'simplified')
        _MANIFEST_HINTS = ('manifest',)
        rescued = []
        for uf in list(classification.get('unknown', [])):
            uf_lower = uf.lower()
            if any(h in uf_lower for h in _DECL_HINTS):
                classification['unknown'].remove(uf)
                classification.setdefault('declaration', []).append(uf)
                rescued.append((uf, 'declaration'))
            elif any(h in uf_lower for h in _MANIFEST_HINTS):
                classification['unknown'].remove(uf)
                classification.setdefault('manifest', []).append(uf)
                rescued.append((uf, 'manifest'))
        if rescued:
            for fname, cat in rescued:
                logger.info(f"Reclassified {fname} → {cat} (filename-based)")

        args._classification = classification

        # Log what we found
        for doc_type, files in classification.items():
            if files:
                print(f"    {doc_type}: {len(files)} file(s)")

        bl_files = classification.get('bill_of_lading', [])
        decl_files = classification.get('declaration', [])
        bl_pdf = os.path.join(args.input_dir, bl_files[0]) if bl_files else None

        # Simplified Declaration serves same role as BL for shipment identification
        if not bl_pdf and decl_files:
            bl_pdf = os.path.join(args.input_dir, decl_files[0])
            logger.info(f"Using declaration as shipment document: {decl_files[0]}")

        # Auto-detect BL number from BL content or filename
        if bl_pdf and not args.bl:
            _auto_detect_bl_number(args, bl_pdf)

        po_file = find_po_file(args.input_dir, args.bl or '')
        if bl_pdf and po_file:
            return 'bl'
        if bl_pdf and not po_file:
            logger.warning("BL/Declaration found but no PO XLSX — running BL mode without PO matching")
            return 'bl'
        # No BL or declaration — still use BL mode (batch), pipeline handles gracefully
        return 'bl'
    return 'bl'


def _load_bl_hint(input_dir: str) -> dict:
    """Load _bl_hint.json if present in the input directory."""
    hint_path = os.path.join(input_dir, '_bl_hint.json')
    if os.path.exists(hint_path):
        try:
            with open(hint_path, 'r') as f:
                hint = json.load(f)
            logger.info(f"BL hint loaded: {hint}")
            return hint
        except (json.JSONDecodeError, IOError) as e:
            logger.warning(f"Failed to load BL hint: {e}")
    return {}


def _looks_like_bl_number(value: str) -> bool:
    """Return True if value looks like a real BL/waybill number, not a description."""
    import re
    if not value or len(value) < 4:
        return False
    # Descriptions have spaces, ampersands, or are too long — not BL numbers
    if ' ' in value.strip() or '&' in value or len(value) > 25:
        return False
    # Must be mostly alphanumeric (allow hyphens, underscores)
    if not re.match(r'^[A-Za-z0-9_-]+$', value):
        return False
    return True


def _auto_detect_bl_number(args, bl_pdf: str) -> None:
    """Extract BL number from BL PDF content, falling back to filename patterns."""
    import re

    # Try to extract from PDF content first (with OCR fallback for scanned BLs)
    text = extract_pdf_text(bl_pdf)
    if not text or not text.strip():
        # Scanned BL — try OCR via bl_parser's _extract_text
        try:
            from bl_parser import _extract_text as _bl_extract_text
            text = _bl_extract_text(bl_pdf)
        except ImportError:
            pass
    if text:
        text_upper = text.upper()
        # Common BL number patterns in content
        # Carrier-specific patterns first (most reliable), then generic
        bl_patterns = [
            r'(TSCW\d+)',               # Tropical Shipping container
            r'(TS\s*\d{7,})',           # Tropical Shipping (e.g. TS 0750206)
            r'(HBL\d+)',                # Caribconex house BL
            r'(MEDU\d+[A-Z]*\d*)',      # Mediterranean Shipping
            r'(HDMU[A-Z0-9]+)',         # Hyundai
            r'(MAEU\d+)',               # Maersk
            r'(COSU\d+)',               # COSCO
            r'B/L\s*(?:NO\.?|NUMBER|#)[:\s]*([A-Z0-9]+)',
            r'BL\s*(?:NO\.?|NUMBER|#)[:\s]*([A-Z0-9]+)',
        ]
        for pattern in bl_patterns:
            match = re.search(pattern, text_upper)
            if match:
                args.bl = match.group(1).strip()
                logger.info(f"Auto-detected BL number from content: {args.bl}")
                return

    # Fallback: extract from filename
    bl_basename = os.path.splitext(os.path.basename(bl_pdf))[0]
    bl_match = re.search(r'TSCW\d+', bl_basename, re.IGNORECASE)
    if bl_match:
        args.bl = bl_match.group(0)
        logger.info(f"Auto-detected BL number from filename: {args.bl}")
        return

    bl_stripped = re.sub(r'[-_ ]*BL$', '', bl_basename, flags=re.IGNORECASE).strip()
    if bl_stripped:
        bl_stripped = re.sub(r'^.*BILL\s*OF\s*LADING\s*[-_ ]*', '', bl_stripped, flags=re.IGNORECASE).strip()
    if bl_stripped and _looks_like_bl_number(bl_stripped):
        args.bl = bl_stripped
        logger.info(f"Auto-detected BL number from filename: {args.bl}")
    elif bl_stripped:
        logger.info(f"Filename '{bl_stripped}' does not look like a BL number — skipping")


def _prepare_invoice_pdfs(args, classification: dict) -> list:
    """
    Pre-process PDFs before invoice processing:
    1. Split combined declaration+invoice PDFs
    2. Return list of PDF paths ready for process_single_invoice

    Also extracts declaration metadata (waybill, freight, packages, etc.)
    and stores it on args._declaration_metadata.
    """
    import tempfile

    invoice_files = classification.get('invoice', [])
    declaration_files = classification.get('declaration', [])
    unknown_files = classification.get('unknown', [])
    freight_invoice_files = classification.get('freight_invoice', [])

    args._declaration_metadata = {}
    args._all_declarations = []  # List of (metadata_dict, pdf_path) for multi-declaration support

    # Extract freight invoice data (standalone freight/landing charges documents)
    # This data takes precedence over BL/manifest freight in email params
    args._freight_invoice_data = {}
    for fi_file in freight_invoice_files:
        fi_path = os.path.join(args.input_dir, fi_file)
        fi_data = extract_freight_invoice_data(fi_path)
        if fi_data:
            args._freight_invoice_data = fi_data
            logger.info(f"Freight invoice data from {fi_file}: {fi_data}")
            break

    # Extract metadata from standalone declaration or manifest PDFs
    # Only used as fallback — split declarations (from combined PDFs) take priority
    # because standalone extraction from multi-page PDFs only finds the first waybill
    metadata_candidates = declaration_files + classification.get('manifest', [])
    # Filename-based fallback: scanned declaration PDFs often have no extractable text,
    # so the classifier marks them "unknown".  If the filename clearly contains
    # "Declaration" or "Manifest", treat it as a declaration metadata candidate too.
    _DECL_FILENAME_HINTS = ('declaration', 'manifest', 'simplified')
    for uf in unknown_files:
        uf_lower = uf.lower()
        if any(hint in uf_lower for hint in _DECL_FILENAME_HINTS):
            if uf not in metadata_candidates:
                metadata_candidates.append(uf)
                logger.info(f"Unknown file {uf} has declaration-like filename — "
                            "adding to metadata candidates")
    _standalone_decl_sources = set()  # Track which source PDFs had standalone extraction
    if metadata_candidates:
        try:
            import hashlib as _hl
            _seen_hashes = set()
            from pdf_splitter import extract_declaration_metadata
            for decl_file in metadata_candidates:
                decl_path = os.path.join(args.input_dir, decl_file)
                # Deduplicate by content hash (same PDF with different names)
                try:
                    h = _hl.md5(open(decl_path, 'rb').read()).hexdigest()
                    if h in _seen_hashes:
                        continue
                    _seen_hashes.add(h)
                except OSError:
                    pass
                meta = extract_declaration_metadata(decl_path)
                if meta and any(v for v in meta.values()):
                    args._all_declarations.append((meta, decl_path))
                    _standalone_decl_sources.add(decl_file)
                    logger.info(f"Declaration metadata from {decl_file}: {meta}")
        except ImportError:
            logger.debug("pdf_splitter not available for declaration metadata")
    # Set primary declaration metadata (backward compatibility)
    if args._all_declarations:
        args._declaration_metadata = args._all_declarations[0][0]

    # Also check declaration/manifest/unknown PDFs — they might be combined documents
    # containing invoice pages that need to be split out and processed
    manifest_files = classification.get('manifest', [])
    pdfs_to_process = invoice_files + declaration_files + unknown_files + manifest_files

    # Deduplicate by content hash — identical PDFs with different names should only be processed once
    import hashlib
    _seen_hashes = set()
    _unique_pdfs = []
    for f in pdfs_to_process:
        fpath = os.path.join(args.input_dir, f)
        try:
            h = hashlib.md5(open(fpath, 'rb').read()).hexdigest()
            if h in _seen_hashes:
                logger.info(f"Skipping duplicate PDF (same content): {f}")
                continue
            _seen_hashes.add(h)
        except OSError:
            pass
        _unique_pdfs.append(f)
    pdfs_to_process = _unique_pdfs

    final_invoice_paths = []
    split_dir = os.path.join(args.output_dir, '_split_temp')

    is_invoice_file = set(invoice_files)
    is_unknown_file = set(unknown_files)
    is_manifest_file = set(manifest_files)

    for f in pdfs_to_process:
        pdf_path = os.path.join(args.input_dir, f)
        originally_invoice = f in is_invoice_file
        originally_manifest = f in is_manifest_file

        # Check if this PDF has multiple pages with mixed content
        # Skip page-level analysis for single-page PDFs (no splitting possible)
        try:
            from pdf_splitter import analyze_pdf, split_pdf_multi_invoice
            import pdfplumber as _pb
            with _pb.open(pdf_path) as _check_pdf:
                num_pages = len(_check_pdf.pages)
            if num_pages <= 1:
                # Single page — no splitting possible
                # Include if originally classified as invoice, unknown, or manifest
                if originally_invoice or f in is_unknown_file or originally_manifest:
                    final_invoice_paths.append(pdf_path)
                continue

            pages, used_ocr, used_heuristic, page_texts = analyze_pdf(pdf_path)

            has_declaration = any(p.doc_type == 'declaration' for p in pages)
            has_invoice = any(p.doc_type == 'invoice' for p in pages)

            if has_declaration and has_invoice and len(pages) > 1:
                # This is a combined PDF — split it
                print(f"    Splitting combined PDF: {f} "
                      f"({sum(1 for p in pages if p.doc_type == 'invoice')} invoice + "
                      f"{sum(1 for p in pages if p.doc_type == 'declaration')} declaration pages)")

                os.makedirs(split_dir, exist_ok=True)
                split_result = split_pdf_multi_invoice(pdf_path, pages, split_dir, page_texts=page_texts)

                # Extract declaration metadata from EACH split declaration PDF
                # so we can send separate emails for each simplified declaration.
                # Split declarations are more accurate than standalone extraction
                # (standalone only finds the first waybill in multi-page PDFs),
                # so replace any standalone declarations from this source file.
                if split_result.get('declarations'):
                    try:
                        from pdf_splitter import extract_declaration_metadata
                        split_decls = []
                        for decl_path in split_result['declarations']:
                            meta = extract_declaration_metadata(decl_path)
                            if meta and any(v for v in meta.values()):
                                split_decls.append((meta, decl_path))
                                logger.info(f"Declaration metadata from split: {os.path.basename(decl_path)}: {meta}")
                        if split_decls:
                            # Remove standalone declarations from this source file
                            # (split declarations supersede the standalone extraction)
                            if f in _standalone_decl_sources:
                                args._all_declarations = [
                                    d for d in args._all_declarations
                                    if os.path.basename(d[1]) != f
                                ]
                            args._all_declarations.extend(split_decls)
                            if not args._declaration_metadata:
                                args._declaration_metadata = args._all_declarations[0][0]
                    except Exception:
                        pass

                # Include split invoice pages if the original file was classified
                # as an invoice or manifest. Manifests (e.g. combined Amazon order +
                # Simplified Declaration) contain commercial invoice pages.
                # Declaration-only files (e.g. CARICOM docs) may contain invoice-like
                # pages (CARICOM Invoice forms, Shipper's Letters) but these are
                # customs forms, not commercial invoices.
                # Include split invoice pages from any combined PDF that has both
                # declaration and invoice content — regardless of initial classification.
                # Combined declaration+invoice PDFs from WebSource are often classified
                # as "declaration" even though they contain commercial invoice pages.
                for inv_path in split_result.get('invoices', []):
                    # Check if this split page is actually a freight invoice
                    try:
                        inv_text = extract_pdf_text(inv_path)
                        from stages.supplier_resolver import _is_freight_invoice
                        if _is_freight_invoice(inv_text):
                            # Extract freight data from split page (if not already from standalone)
                            if not args._freight_invoice_data:
                                fi_data = extract_freight_invoice_data(inv_path)
                                if fi_data:
                                    args._freight_invoice_data = fi_data
                                    logger.info(f"Freight invoice data from split page: {fi_data}")
                            logger.info(f"Skipping freight invoice page: {os.path.basename(inv_path)}")
                            continue
                    except Exception:
                        pass
                    final_invoice_paths.append(inv_path)
                continue

            if has_declaration and not has_invoice:
                # Pure declaration — skip (already extracted metadata above)
                logger.info(f"Skipping pure declaration: {f}")
                continue

        except ImportError:
            # pdf_splitter not available — include invoices and manifests
            logger.warning(
                "pdf_splitter unavailable (missing PyPDF2/pypdf?) — "
                f"cannot split combined PDF {f}, treating as-is"
            )
            if not originally_invoice and not originally_manifest:
                continue
        except Exception as e:
            logger.warning(f"PDF analysis/split failed for {f}, using as-is: {e}")
            if not originally_invoice and not originally_manifest:
                continue

        # Not a combined PDF — add to invoice list
        # For declaration/unknown files that page analysis says have invoice content, include them
        final_invoice_paths.append(pdf_path)

    # Deduplicate declarations by waybill number (keep the split version which
    # has per-page metadata, discard standalone extractions of the same waybill)
    if len(args._all_declarations) > 1:
        seen_waybills = set()
        unique_decls = []
        for meta, path in args._all_declarations:
            wb = meta.get('waybill', '')
            if wb and wb in seen_waybills:
                logger.info(f"Skipping duplicate declaration (same waybill {wb}): {os.path.basename(path)}")
                continue
            if wb:
                seen_waybills.add(wb)
            unique_decls.append((meta, path))
        args._all_declarations = unique_decls
        if args._all_declarations:
            args._declaration_metadata = args._all_declarations[0][0]

    return final_invoice_paths


def run_bl_mode(args) -> dict:
    """
    BL batch processing: multi-invoice + PO + BL PDF.

    This is the main workflow for documents.websource emails containing
    a Bill of Lading with multiple supplier invoices and a PO XLSX.
    """
    print("=" * 80)
    print(f"Processing BL"
          f"{' #' + args.bl if args.bl else ''}"
          f"  |  Document Type: {args.doc_type if args.doc_type != 'auto' else '(auto)'}"
          f"  |  Mode: BL Batch")
    print("=" * 80)

    # Imports needed for BL mode
    try:
        from format_registry import FormatRegistry
        from po_matcher import POReader, POMatcher
    except ImportError as e:
        print(f"ERROR: Missing required module: {e}")
        sys.exit(1)

    # Use pre-computed classification from detect_mode
    classification = getattr(args, '_classification', None)

    # 1. Find and load PO XLSX
    po_file = args.po_file or find_po_file(args.input_dir, args.bl or '')
    matcher = None
    if po_file and os.path.exists(po_file):
        print(f"\n[1] Loading PO data from {os.path.basename(po_file)}...")
        po_items = POReader.read_po_xlsx(po_file)
        print(f"    {len(po_items)} PO line items loaded")
        matcher = POMatcher(po_items, base_dir=BASE_DIR)
    else:
        print(f"\n[1] No PO XLSX found — running without PO matching")

    # 2. Initialize format registry
    print("\n[2] Initializing format registry...")
    registry = FormatRegistry(BASE_DIR)
    print(f"    {len(registry.list_formats())} format specs loaded")

    # Reset auto-format batch cache for this run
    try:
        from stages.auto_format import reset_batch_cache
        reset_batch_cache()
    except ImportError:
        pass

    # 3. Load supplier database and classification rules
    supplier_db = load_supplier_db()
    rules, noise_words = load_classification_rules()
    print(f"    {len(rules)} classification rules loaded")

    # Clean output directory before generating fresh files
    # (skip if output_dir == input_dir to avoid deleting source PDFs)
    output_dir = args.output_dir
    args._output_dir = output_dir  # Store for _save_email_params
    input_dir_real = os.path.realpath(args.input_dir) if args.input_dir else ''
    output_dir_real = os.path.realpath(output_dir) if output_dir else ''
    if os.path.exists(output_dir):
        if input_dir_real != output_dir_real:
            # Different dirs: clean all files from output
            for f in os.listdir(output_dir):
                fp = os.path.join(output_dir, f)
                if os.path.isfile(fp):
                    os.remove(fp)
        else:
            # Same dir: only clean generated output files, preserve source PDFs
            # Source PDFs are listed in the classification result
            source_pdfs = set()
            if classification:
                for cat_files in classification.values():
                    source_pdfs.update(cat_files)
            for f in os.listdir(output_dir):
                fp = os.path.join(output_dir, f)
                if os.path.isdir(fp) and f == '_split_temp':
                    # Preserve OCR sidecar .txt files (expensive to regenerate)
                    # but remove split PDFs so they get re-split fresh
                    for sf in os.listdir(fp):
                        sfp = os.path.join(fp, sf)
                        if os.path.isfile(sfp) and not sf.endswith('.txt'):
                            os.remove(sfp)
                elif os.path.isfile(fp) and f not in source_pdfs:
                    # Remove generated files (.xlsx, .pdf copies, _email_params, etc.)
                    os.remove(fp)
    os.makedirs(output_dir, exist_ok=True)

    # 4. Pre-process: split combined PDFs, get invoice paths
    invoice_paths = _prepare_invoice_pdfs(args, classification or {})

    # Auto-resolve document type from consignee
    args.doc_type = resolve_doc_type(args)
    print(f"    Document Type: {args.doc_type}")

    print(f"\n[3] Processing {len(invoice_paths)} PDF invoices...\n")

    # ── Phase 1: Process each invoice ──
    results = []
    failures = []
    all_attachments = []
    for pdf_path in invoice_paths:
        result = process_single_invoice(
            pdf_path=pdf_path,
            registry=registry,
            matcher=matcher,
            rules=rules,
            noise_words=noise_words,
            supplier_db=supplier_db,
            output_dir=output_dir,
            document_type=args.doc_type,
            verbose=args.verbose,
        )
        if result:
            if len(result.matched_items) == 0:
                if result.format_name == '_default':
                    # Unrecognised doc (BOL, manifest, etc.) — skip silently
                    print(f"    Skipping non-invoice: {os.path.basename(pdf_path)} (no format match)")
                else:
                    failures.append({
                        'pdf_path': pdf_path,
                        'pdf_file': os.path.basename(pdf_path),
                        'reason': f'No items extracted (format: {result.format_name})',
                        'invoice_num': result.invoice_num,
                    })
                    print(f"    WARNING: {os.path.basename(pdf_path)} — 0 items extracted, flagged for review")
            else:
                results.append(result)
                all_attachments.extend([result.pdf_output_path, result.xlsx_path])
        else:
            failures.append({
                'pdf_path': pdf_path,
                'pdf_file': os.path.basename(pdf_path),
                'reason': 'Text extraction failed',
            })

    # Promote successful auto-generated format specs
    _promote_auto_specs(results)

    # Apply any reviewer-edited Proposed Fixes YAML dropped into input_dir
    _maybe_apply_fixes(args, results)

    # Copy failed PDFs to Unprocessed/ folder
    _handle_failures(failures, output_dir, args.input_dir)

    # Save updated supplier database
    save_supplier_db(supplier_db)

    # ── Phase 2: BL allocation ──
    original_invoice_count = len(results)  # total invoices before combining
    bl_files = (classification or {}).get('bill_of_lading', [])
    bl_pdf_path = os.path.join(args.input_dir, bl_files[0]) if bl_files else None
    bl_alloc = allocate_bl_packages(
        bl_pdf_path=bl_pdf_path,
        invoice_results=results,
        output_dir=output_dir,
        bl_number=args.bl or '',
    )
    args._bl_alloc = bl_alloc  # store for consolidation report cross-reference

    # ── Phase 2.5: Combine entries if configured ──
    args._pre_combine_results = list(results)  # preserve for consolidation cross-reference
    results = _maybe_combine_entries(args, results, output_dir)

    # Rebuild attachments from (possibly combined) results
    all_attachments = []
    for r in results:
        # Combined entries stash all PDFs in _combined_pdf_paths
        pdf_paths = getattr(r, '_combined_pdf_paths', None) or [r.pdf_output_path]
        all_attachments.extend(pdf_paths)
        all_attachments.append(r.xlsx_path)

    # Add BL PDF to attachments
    if bl_alloc and bl_alloc.bl_output_path:
        all_attachments.append(bl_alloc.bl_output_path)

    # ── Print summary ──
    # Use pre-combine results so each source invoice shows as its own row
    # instead of collapsing into the single combined entry.
    _summary_results = getattr(args, '_pre_combine_results', None) or results
    _print_summary(_summary_results, bl_alloc)

    # ── Phase 2.5: Manifest metadata ──
    manifest_meta = _apply_manifest_metadata(args, classification, output_dir, all_attachments)

    # Fallback: if manifest metadata extraction failed (e.g. Simplified Declaration
    # without ASYCUDA format), use declaration metadata from pdf_splitter
    decl_meta = getattr(args, '_declaration_metadata', {})
    if not manifest_meta:
        if decl_meta and any(v for v in decl_meta.values()):
            manifest_meta = decl_meta
            logger.info(f"Using declaration metadata as manifest fallback: {decl_meta}")
    elif decl_meta:
        # Merge missing fields from declaration metadata into manifest metadata
        # (e.g. weight from simplified declaration when manifest parser doesn't extract it)
        for key in ('weight', 'packages', 'freight', 'consignee', 'office', 'fob_value'):
            if not manifest_meta.get(key) and decl_meta.get(key):
                manifest_meta[key] = decl_meta[key]
                logger.info(f"Supplemented manifest {key} from declaration: {decl_meta[key]}")

    # ── Inject handwritten customs values into invoice_data for duty estimation ──
    # The declaration metadata may contain _customs_value_ec / _customs_value_usd
    # extracted by LLM vision from pencil annotations on the Simplified Declaration.
    # Wire these into invoice_data['_client_declared_duties'] so the XLSX duty
    # estimation section can show the variance comparison.
    if _inject_handwritten_duties(decl_meta, results):
        # Regenerate XLSX files so the duty estimation section includes the
        # client declared duties comparison.  The initial XLSX was generated
        # before declaration metadata was available.
        try:
            from bl_xlsx_generator import generate_bl_xlsx
            for r in results:
                if r.xlsx_path and os.path.exists(r.xlsx_path):
                    generate_bl_xlsx(
                        r.invoice_data,
                        r.matched_items,
                        os.path.basename(r.xlsx_path).rsplit('.', 1)[0],
                        r.supplier_info,
                        r.xlsx_path,
                        document_type=getattr(args, 'doc_type', 'auto'),
                    )
                    print(f"    {r.invoice_num}: XLSX regenerated with client duty comparison")
        except Exception as e:
            logger.warning(f"XLSX regeneration for duty comparison failed: {e}")

    # Apply manifest packages to XLSX when BL allocator didn't set them
    if manifest_meta and manifest_meta.get('packages') and results:
        bl_has_packages = bl_alloc and getattr(bl_alloc, 'packages', None)
        if not bl_has_packages:
            try:
                manifest_pkgs = int(manifest_meta['packages'])
                for r in results:
                    if r.xlsx_path and os.path.exists(r.xlsx_path):
                        from openpyxl import load_workbook as _lw
                        wb = _lw(r.xlsx_path)
                        ws = wb.active
                        # Set manifest packages on row 2, clear all other rows
                        # (combined XLSX may have packages on multiple invoice header rows)
                        ws.cell(row=2, column=24, value=manifest_pkgs)
                        for row in range(3, ws.max_row + 1):
                            if ws.cell(row, 24).value is not None:
                                ws.cell(row, 24).value = None
                        wb.save(r.xlsx_path)
                        wb.close()
                        r.packages = manifest_pkgs
                        print(f"    Updated packages in {os.path.basename(r.xlsx_path)}: {manifest_pkgs} (from manifest)")
            except (ValueError, TypeError, Exception) as e:
                logger.warning(f"Could not apply manifest packages to XLSX: {e}")

    # ── Phase 2.7: Consolidation report (package-to-invoice mapping) ──
    consolidation_meta = _apply_consolidation_report(
        args, classification, results, output_dir, all_attachments)
    if consolidation_meta:
        # Consolidation packages override manifest packages (more granular)
        if not manifest_meta:
            manifest_meta = {}
        if consolidation_meta.get('total_packages') and not manifest_meta.get('packages'):
            manifest_meta['packages'] = str(consolidation_meta['total_packages'])
        if consolidation_meta.get('total_weight') and not manifest_meta.get('weight'):
            manifest_meta['weight'] = str(consolidation_meta['total_weight'])
        if consolidation_meta.get('total_freight') and not manifest_meta.get('freight'):
            manifest_meta['freight'] = str(consolidation_meta['total_freight'])
        if consolidation_meta.get('waybill') and not manifest_meta.get('waybill'):
            manifest_meta['waybill'] = consolidation_meta['waybill']

    # ── Phase 2.8: Final package cap enforcement ──
    # After ALL package-writing stages (BL, combine, manifest, consolidation),
    # verify no XLSX has more packages than the authoritative source allows.
    _enforce_package_cap(results, bl_alloc, manifest_meta)

    # ── XLSX validation + auto-fix ──
    validation = None
    try:
        from xlsx_validator import validate_and_fix
        validation = validate_and_fix(results, BASE_DIR, bl_alloc=bl_alloc,
                                      manifest_meta=manifest_meta)
    except Exception as e:
        logger.warning(f"XLSX validation failed: {e}")
        _validate_xlsx_variance(results)  # fallback to print-only

    # ── Phase 3: Email params + optional send ──
    email_sent = False
    email_params_path = ''
    all_email_params_paths = []
    checklist = None
    if results:
        all_declarations = getattr(args, '_all_declarations', [])

        if len(all_declarations) > 1:
            # Multiple simplified declarations — save separate email params for each
            print(f"\n    Multiple declarations ({len(all_declarations)}) — generating separate emails for each")

            # Separate invoice/xlsx attachments from declaration/manifest attachments
            invoice_attachments = [p for p in all_attachments
                                   if not any(tag in os.path.basename(p).lower()
                                              for tag in ('declaration', 'manifest'))]

            for idx, (decl_meta, decl_pdf_path) in enumerate(all_declarations):
                import shutil
                # Name declaration PDF as {waybill}-Declaration.pdf (matches manifest convention)
                decl_waybill = decl_meta.get('waybill', '')
                if decl_waybill:
                    decl_out_name = f"{decl_waybill}-Declaration.pdf"
                else:
                    decl_out_name = os.path.basename(decl_pdf_path)
                decl_out_path = os.path.join(output_dir, decl_out_name)
                if os.path.abspath(decl_pdf_path) != os.path.abspath(decl_out_path):
                    shutil.copy2(decl_pdf_path, decl_out_path)

                decl_attachments = list(invoice_attachments) + [decl_out_path]

                saved_bl = args.bl
                saved_man_reg = getattr(args, 'man_reg', '')
                args.bl = decl_waybill or saved_bl
                args.man_reg = decl_meta.get('man_reg', '') or saved_man_reg

                decl_manifest_meta = dict(manifest_meta or {})
                for key_src, key_dst in [('waybill', 'waybill'), ('consignee', 'consignee'),
                                          ('freight', 'freight'), ('packages', 'packages'),
                                          ('weight', 'weight'), ('man_reg', 'man_reg')]:
                    if decl_meta.get(key_src):
                        decl_manifest_meta[key_dst] = decl_meta[key_src]

                params_path = _save_email_params(args, results, bl_alloc, decl_attachments,
                                                  decl_manifest_meta, output_dir,
                                                  total_invoices=original_invoice_count)
                if idx == 0:
                    import shutil as _sh
                    _sh.copy2(params_path, params_path + '.bak')
                else:
                    indexed_path = os.path.join(output_dir, f'_email_params_{idx + 1}.json')
                    os.replace(params_path, indexed_path)
                    params_path = indexed_path
                    bak = os.path.join(output_dir, '_email_params.json.bak')
                    if os.path.exists(bak):
                        _sh.copy2(bak, os.path.join(output_dir, '_email_params.json'))

                all_email_params_paths.append(params_path)
                waybill_label = decl_meta.get('waybill', f'Declaration {idx+1}')
                print(f"    Declaration {idx+1}: {waybill_label} → {os.path.basename(params_path)}")

                args.bl = saved_bl
                args.man_reg = saved_man_reg

            email_params_path = all_email_params_paths[0]

            # Clean up backup file
            bak = os.path.join(output_dir, '_email_params.json.bak')
            if os.path.exists(bak):
                os.remove(bak)

            # Save Proposed Fixes sidecar once per shipment (all declarations
            # share the same underlying results; one fixes email covers the
            # whole waybill).
            _maybe_save_proposed_fixes(results, email_params_path, output_dir)

            if args.send_email:
                for pp in all_email_params_paths:
                    try:
                        from xlsx_validator import shipment_checklist
                        with open(pp) as f:
                            ep = json.load(f)
                        cl = shipment_checklist(ep, validation)
                        if cl and not cl['passed']:
                            print(f"    Email BLOCKED for {os.path.basename(pp)}: {cl['blocker_count']} blocker(s)")
                            continue
                    except Exception:
                        pass
                    _send_email_from_params(pp, args)
                    email_sent = True
                # Send the Proposed Fixes sidecar once per shipment (not per declaration)
                _send_proposed_fixes_sidecar(output_dir)
        else:
            # Single declaration (or none) — original flow
            decl_meta = getattr(args, '_declaration_metadata', {})
            if decl_meta:
                if decl_meta.get('waybill') and not args.bl:
                    args.bl = decl_meta['waybill']
                if decl_meta.get('man_reg') and not args.man_reg:
                    args.man_reg = decl_meta['man_reg']
                # Merge OCR-extracted declaration fields into manifest_meta so
                # _save_email_params can apply consignee, freight, weight,
                # packages, office from scanned Simplified Declaration forms.
                # Without this, scanned declarations lose these fields because
                # extract_manifest_metadata (no OCR) returns empty for them.
                if manifest_meta is None:
                    manifest_meta = {}
                for key_src, key_dst in [('consignee', 'consignee'),
                                          ('freight', 'freight'),
                                          ('packages', 'packages'),
                                          ('weight', 'weight'),
                                          ('office', 'office')]:
                    if decl_meta.get(key_src) and not manifest_meta.get(key_dst):
                        manifest_meta[key_dst] = decl_meta[key_src]
            # Attach the single declaration PDF (renamed {waybill}-Declaration.pdf)
            # so the broker receives the customs form alongside the invoice + xlsx.
            if all_declarations:
                import shutil as _sh
                decl_meta_single, decl_pdf_path_single = all_declarations[0]
                decl_wb_single = decl_meta_single.get('waybill', '')
                if decl_wb_single:
                    decl_out_name_single = f"{decl_wb_single}-Declaration.pdf"
                else:
                    decl_out_name_single = os.path.basename(decl_pdf_path_single)
                decl_out_path_single = os.path.join(output_dir, decl_out_name_single)
                if os.path.abspath(decl_pdf_path_single) != os.path.abspath(decl_out_path_single):
                    _sh.copy2(decl_pdf_path_single, decl_out_path_single)
                if decl_out_path_single not in all_attachments:
                    all_attachments.append(decl_out_path_single)
            # Always save email params for TypeScript to pick up
            email_params_path = _save_email_params(args, results, bl_alloc, all_attachments,
                                                    manifest_meta, output_dir,
                                                    total_invoices=original_invoice_count)
            all_email_params_paths = [email_params_path]
            # Save Proposed Fixes sidecar if any invoice carries uncertainty.
            _maybe_save_proposed_fixes(results, email_params_path, output_dir)
            # ── Shipment pre-send checklist ──
            checklist = None
            try:
                from xlsx_validator import shipment_checklist
                with open(email_params_path) as f:
                    email_params = json.load(f)
                checklist = shipment_checklist(email_params, validation)
            except Exception as e:
                logger.warning(f"Shipment checklist failed: {e}")

            # Only send email directly when --send-email (legacy CLI mode)
            if args.send_email:
                if checklist and not checklist['passed']:
                    print(f"    Email BLOCKED by checklist: {checklist['blocker_count']} blocker(s). Fix issues first.")
                else:
                    email_sent = _send_bl_email(args, results, bl_alloc, all_attachments, manifest_meta,
                                                total_invoices=original_invoice_count)

    # Archive any reviewer-edited Proposed Fixes YAML that was applied this run
    _maybe_archive_applied_fixes(args, args.bl or args.waybill or '')

    # ── JSON output for TypeScript ──
    report = _build_report(results, bl_alloc, email_sent, failures)
    if all_email_params_paths:
        report['email_params_paths'] = all_email_params_paths
    if email_params_path:
        report['email_params_path'] = email_params_path
    if validation:
        report['validation'] = validation
    if checklist:
        report['checklist'] = checklist
    if getattr(args, '_fixes_reports', None):
        report['fixes_applied'] = args._fixes_reports
    if args.json_output:
        print(f"\nREPORT:JSON:{json.dumps(report)}")

    return report


def run_single_mode(args) -> dict:
    """
    Single invoice processing via pipeline_runner.py.
    """
    try:
        from pipeline_runner import PipelineRunner
    except ImportError:
        print("ERROR: pipeline_runner module not found")
        sys.exit(1)

    config_path = args.config or os.path.join(BASE_DIR, 'config', 'pipeline.yaml')
    runner = PipelineRunner(config_path)

    output_path = args.output
    if not output_path:
        input_base = os.path.splitext(os.path.basename(args.input))[0]
        output_dir = args.output_dir or os.path.join(BASE_DIR, 'workspace', 'shipments')
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{input_base}.xlsx")

    result = runner.run(args.input, output_path)

    if args.json_output:
        print(f"\nREPORT:JSON:{json.dumps(result)}")

    return result


def run_batch_mode(args) -> dict:
    """
    Batch mode: process all PDFs in a folder through the same pipeline as BL mode.

    Same as BL mode but without BL allocation step. Handles any batch of
    shipping documents — invoices are parsed, matched, classified, converted
    to XLSX, and emailed as a single shipment.
    """
    print("=" * 80)
    print(f"Processing Batch Shipment  |  Document Type: {args.doc_type if args.doc_type != 'auto' else '(auto)'}  |  Mode: Batch")
    print("=" * 80)

    try:
        from format_registry import FormatRegistry
        from po_matcher import POReader, POMatcher
    except ImportError as e:
        print(f"ERROR: Missing required module: {e}")
        sys.exit(1)

    # Use pre-computed classification from detect_mode
    classification = getattr(args, '_classification', None)

    # 1. Find and load PO XLSX (optional for batch mode)
    po_file = args.po_file or find_po_file(args.input_dir, args.bl or '')
    matcher = None
    if po_file and os.path.exists(po_file):
        print(f"\n[1] Loading PO data from {os.path.basename(po_file)}...")
        po_items = POReader.read_po_xlsx(po_file)
        print(f"    {len(po_items)} PO line items loaded")
        matcher = POMatcher(po_items, base_dir=BASE_DIR)
    else:
        print(f"\n[1] No PO XLSX found — running without PO matching")

    # 2. Initialize format registry
    print("\n[2] Initializing format registry...")
    registry = FormatRegistry(BASE_DIR)
    print(f"    {len(registry.list_formats())} format specs loaded")

    # Reset auto-format batch cache for this run
    try:
        from stages.auto_format import reset_batch_cache
        reset_batch_cache()
    except ImportError:
        pass

    # 3. Load supplier database and classification rules
    supplier_db = load_supplier_db()
    rules, noise_words = load_classification_rules()
    print(f"    {len(rules)} classification rules loaded")

    # Clean output directory
    output_dir = args.output_dir
    args._output_dir = output_dir  # Store for _save_batch_email_params
    if os.path.exists(output_dir):
        for f in os.listdir(output_dir):
            fp = os.path.join(output_dir, f)
            if os.path.isfile(fp):
                os.remove(fp)
    os.makedirs(output_dir, exist_ok=True)

    # 4. Pre-process: split combined PDFs, get invoice paths
    invoice_paths = _prepare_invoice_pdfs(args, classification or {})
    if not invoice_paths:
        print(f"No invoice PDFs found in {args.input_dir}")
        return {'status': 'error', 'error': 'No invoice PDFs found'}

    # Auto-resolve document type from consignee
    args.doc_type = resolve_doc_type(args)
    print(f"    Document Type: {args.doc_type}")

    print(f"\n[3] Processing {len(invoice_paths)} PDF invoices...\n")

    # Process each invoice through the full pipeline
    results = []
    failures = []
    all_attachments = []
    for pdf_path in invoice_paths:
        result = process_single_invoice(
            pdf_path=pdf_path,
            registry=registry,
            matcher=matcher,
            rules=rules,
            noise_words=noise_words,
            supplier_db=supplier_db,
            output_dir=output_dir,
            document_type=args.doc_type,
            verbose=args.verbose,
        )
        if result:
            if len(result.matched_items) == 0:
                if result.format_name == '_default':
                    # Unrecognised doc (BOL, manifest, etc.) — skip silently
                    print(f"    Skipping non-invoice: {os.path.basename(pdf_path)} (no format match)")
                else:
                    failures.append({
                        'pdf_path': pdf_path,
                        'pdf_file': os.path.basename(pdf_path),
                        'reason': f'No items extracted (format: {result.format_name})',
                        'invoice_num': result.invoice_num,
                    })
                    print(f"    WARNING: {os.path.basename(pdf_path)} — 0 items extracted, flagged for review")
            else:
                results.append(result)
                all_attachments.extend([result.pdf_output_path, result.xlsx_path])
        else:
            failures.append({
                'pdf_path': pdf_path,
                'pdf_file': os.path.basename(pdf_path),
                'reason': 'Text extraction failed',
            })

    # Promote successful auto-generated format specs
    _promote_auto_specs(results)

    # Apply any reviewer-edited Proposed Fixes YAML dropped into input_dir
    _maybe_apply_fixes(args, results)

    # Copy failed PDFs to Unprocessed/ folder
    _handle_failures(failures, output_dir, args.input_dir)

    # Save updated supplier database
    save_supplier_db(supplier_db)

    # Print summary (no BL allocation for batch mode)
    _print_summary(results, bl_alloc=None)

    # Manifest metadata
    manifest_meta = _apply_manifest_metadata(args, classification, output_dir, all_attachments)

    # Fallback: if no manifest (e.g. Simplified Declaration shipments) use
    # declaration metadata extracted from pdf_splitter so downstream email
    # params inherit waybill, office, consignee, freight etc.
    decl_meta = getattr(args, '_declaration_metadata', {})
    if not manifest_meta:
        if decl_meta and any(v for v in decl_meta.values()):
            manifest_meta = decl_meta
            logger.info(f"Using declaration metadata as manifest fallback: {decl_meta}")
    elif decl_meta:
        for key in ('office', 'weight', 'packages', 'freight', 'consignee', 'fob_value'):
            if not manifest_meta.get(key) and decl_meta.get(key):
                manifest_meta[key] = decl_meta[key]
                logger.info(f"Supplemented manifest {key} from declaration: {decl_meta[key]}")

    # Inject handwritten customs values into invoice_data for duty estimation (batch mode)
    if _inject_handwritten_duties(decl_meta, results):
        try:
            from bl_xlsx_generator import generate_bl_xlsx
            for r in results:
                if r.xlsx_path and os.path.exists(r.xlsx_path):
                    generate_bl_xlsx(
                        r.invoice_data,
                        r.matched_items,
                        os.path.basename(r.xlsx_path).rsplit('.', 1)[0],
                        r.supplier_info,
                        r.xlsx_path,
                        document_type=getattr(args, 'doc_type', 'auto'),
                    )
                    print(f"    {r.invoice_num}: XLSX regenerated with client duty comparison")
        except Exception as e:
            logger.warning(f"XLSX regeneration for duty comparison failed: {e}")

    # Apply manifest packages to XLSX (batch mode has no BL allocator)
    if manifest_meta and manifest_meta.get('packages') and results:
        try:
            manifest_pkgs = int(manifest_meta['packages'])
            for r in results:
                if r.xlsx_path and os.path.exists(r.xlsx_path):
                    from openpyxl import load_workbook as _lw
                    wb = _lw(r.xlsx_path)
                    ws = wb.active
                    # Set manifest packages on row 2, clear all other rows
                    ws.cell(row=2, column=24, value=manifest_pkgs)
                    for row in range(3, ws.max_row + 1):
                        if ws.cell(row, 24).value is not None:
                            ws.cell(row, 24).value = None
                    wb.save(r.xlsx_path)
                    wb.close()
                    r.packages = manifest_pkgs
                    print(f"    Updated packages in {os.path.basename(r.xlsx_path)}: {manifest_pkgs} (from manifest)")
        except (ValueError, TypeError, Exception) as e:
            logger.warning(f"Could not apply manifest packages to XLSX: {e}")

    # Consolidation report (package-to-invoice mapping)
    consolidation_meta = _apply_consolidation_report(
        args, classification, results, output_dir, all_attachments)
    if consolidation_meta:
        if not manifest_meta:
            manifest_meta = {}
        if consolidation_meta.get('total_packages') and not manifest_meta.get('packages'):
            manifest_meta['packages'] = str(consolidation_meta['total_packages'])
        if consolidation_meta.get('total_weight') and not manifest_meta.get('weight'):
            manifest_meta['weight'] = str(consolidation_meta['total_weight'])
        if consolidation_meta.get('total_freight') and not manifest_meta.get('freight'):
            manifest_meta['freight'] = str(consolidation_meta['total_freight'])
        if consolidation_meta.get('waybill') and not manifest_meta.get('waybill'):
            manifest_meta['waybill'] = consolidation_meta['waybill']

    # Final package cap enforcement
    _enforce_package_cap(results, None, manifest_meta)

    # XLSX validation + auto-fix
    validation = None
    try:
        from xlsx_validator import validate_and_fix
        validation = validate_and_fix(results, BASE_DIR, bl_alloc=None,
                                      manifest_meta=manifest_meta)
    except Exception as e:
        logger.warning(f"XLSX validation failed: {e}")
        _validate_xlsx_variance(results)  # fallback to print-only

    # Email params + optional send
    email_sent = False
    email_params_path = ''
    all_email_params_paths = []
    if results:
        all_declarations = getattr(args, '_all_declarations', [])

        if len(all_declarations) > 1:
            # Multiple simplified declarations — save separate email params for each
            # Each email reuses the same invoice PDF(s) + XLSX(es) but with different
            # declaration metadata (waybill, freight, packages, etc.)
            print(f"\n    Multiple declarations ({len(all_declarations)}) — generating separate emails for each")

            # Separate invoice/xlsx attachments from declaration/manifest attachments
            invoice_attachments = [p for p in all_attachments
                                   if not any(tag in os.path.basename(p).lower()
                                              for tag in ('declaration', 'manifest'))]

            for idx, (decl_meta, decl_pdf_path) in enumerate(all_declarations):
                # Copy declaration PDF to output dir with {waybill}-Declaration.pdf naming
                import shutil
                decl_waybill = decl_meta.get('waybill', '')
                if decl_waybill:
                    decl_out_name = f"{decl_waybill}-Declaration.pdf"
                else:
                    decl_out_name = os.path.basename(decl_pdf_path)
                decl_out_path = os.path.join(output_dir, decl_out_name)
                if os.path.abspath(decl_pdf_path) != os.path.abspath(decl_out_path):
                    shutil.copy2(decl_pdf_path, decl_out_path)

                # Build per-declaration attachments: declaration PDF + all invoice/xlsx files
                decl_attachments = list(invoice_attachments) + [decl_out_path]

                # Override args with this declaration's metadata
                saved_waybill = getattr(args, 'waybill', '')
                saved_man_reg = getattr(args, 'man_reg', '')
                saved_bl = getattr(args, 'bl', '')
                args.waybill = decl_waybill or saved_waybill
                args.man_reg = decl_meta.get('man_reg', '') or saved_man_reg
                args.bl = decl_waybill or saved_bl

                # Build manifest_meta override from this declaration
                decl_manifest_meta = dict(manifest_meta or {})
                if decl_meta.get('waybill'):
                    decl_manifest_meta['waybill'] = decl_meta['waybill']
                if decl_meta.get('consignee'):
                    decl_manifest_meta['consignee_name'] = decl_meta['consignee']
                if decl_meta.get('freight'):
                    decl_manifest_meta['freight'] = decl_meta['freight']
                if decl_meta.get('packages'):
                    decl_manifest_meta['packages'] = decl_meta['packages']
                if decl_meta.get('weight'):
                    decl_manifest_meta['weight'] = decl_meta['weight']
                if decl_meta.get('man_reg'):
                    decl_manifest_meta['man_reg'] = decl_meta['man_reg']

                # Save email params: _email_params.json for first, _email_params_2.json etc
                # _save_batch_email_params always writes to _email_params.json,
                # so rename immediately to avoid overwriting on next iteration
                params_path = _save_batch_email_params(
                    args, results, decl_attachments, decl_manifest_meta, output_dir)
                if idx == 0:
                    # First declaration keeps _email_params.json (backward compat)
                    # but save a copy since next iteration will overwrite it
                    import shutil as _sh
                    _sh.copy2(params_path, params_path + '.bak')
                else:
                    indexed_path = os.path.join(output_dir, f'_email_params_{idx + 1}.json')
                    os.replace(params_path, indexed_path)
                    params_path = indexed_path
                    # Restore first declaration's params file if it was overwritten
                    bak = os.path.join(output_dir, '_email_params.json.bak')
                    if os.path.exists(bak):
                        _sh.copy2(bak, os.path.join(output_dir, '_email_params.json'))


                all_email_params_paths.append(params_path)
                waybill_label = decl_meta.get('waybill', f'Declaration {idx+1}')
                print(f"    Declaration {idx+1}: {waybill_label} → {os.path.basename(params_path)}")

                # Restore args
                args.waybill = saved_waybill
                args.man_reg = saved_man_reg
                args.bl = saved_bl

            email_params_path = all_email_params_paths[0]

            # Clean up backup file
            bak = os.path.join(output_dir, '_email_params.json.bak')
            if os.path.exists(bak):
                os.remove(bak)

            # Send emails for each declaration if --send-email
            if args.send_email:
                for params_path in all_email_params_paths:
                    try:
                        from xlsx_validator import shipment_checklist
                        with open(params_path) as f:
                            ep = json.load(f)
                        cl = shipment_checklist(ep, validation)
                        if cl and not cl['passed']:
                            print(f"    Email BLOCKED for {os.path.basename(params_path)}: {cl['blocker_count']} blocker(s)")
                            continue
                    except Exception:
                        pass
                    _send_email_from_params(params_path, args)
                    email_sent = True
        else:
            # Single declaration (or none) — original flow
            decl_meta = getattr(args, '_declaration_metadata', {})
            if decl_meta:
                if decl_meta.get('waybill') and not args.waybill:
                    args.waybill = decl_meta['waybill']
                if decl_meta.get('man_reg') and not args.man_reg:
                    args.man_reg = decl_meta['man_reg']
                # Merge OCR-extracted declaration fields into manifest_meta so
                # _save_batch_email_params can apply consignee, freight, weight,
                # packages, office from scanned Simplified Declaration forms.
                if manifest_meta is None:
                    manifest_meta = {}
                for key_src, key_dst in [('consignee', 'consignee'),
                                          ('freight', 'freight'),
                                          ('packages', 'packages'),
                                          ('weight', 'weight'),
                                          ('office', 'office')]:
                    if decl_meta.get(key_src) and not manifest_meta.get(key_dst):
                        manifest_meta[key_dst] = decl_meta[key_src]
            # Attach the single declaration PDF (renamed {waybill}-Declaration.pdf)
            # so the broker receives the customs form alongside the invoice + xlsx.
            if all_declarations:
                import shutil as _sh
                decl_meta_single, decl_pdf_path_single = all_declarations[0]
                decl_wb_single = decl_meta_single.get('waybill', '')
                if decl_wb_single:
                    decl_out_name_single = f"{decl_wb_single}-Declaration.pdf"
                else:
                    decl_out_name_single = os.path.basename(decl_pdf_path_single)
                decl_out_path_single = os.path.join(output_dir, decl_out_name_single)
                if os.path.abspath(decl_pdf_path_single) != os.path.abspath(decl_out_path_single):
                    _sh.copy2(decl_pdf_path_single, decl_out_path_single)
                if decl_out_path_single not in all_attachments:
                    all_attachments.append(decl_out_path_single)
            # Always save email params for TypeScript to pick up
            email_params_path = _save_batch_email_params(args, results, all_attachments,
                                                          manifest_meta, output_dir)
            all_email_params_paths = [email_params_path]
            _maybe_save_proposed_fixes(results, email_params_path, output_dir)
            # ── Shipment pre-send checklist ──
            checklist = None
            try:
                from xlsx_validator import shipment_checklist
                with open(email_params_path) as f:
                    email_params = json.load(f)
                checklist = shipment_checklist(email_params, validation)
            except Exception as e:
                logger.warning(f"Shipment checklist failed: {e}")

            # Only send email directly when --send-email (legacy CLI mode)
            if args.send_email:
                if checklist and not checklist['passed']:
                    print(f"    Email BLOCKED by checklist: {checklist['blocker_count']} blocker(s). Fix issues first.")
                else:
                    email_sent = _send_batch_email(args, results, all_attachments, manifest_meta)

    # Archive any reviewer-edited Proposed Fixes YAML that was applied this run
    _maybe_archive_applied_fixes(args, args.waybill or args.bl or '')

    # JSON output for TypeScript
    report = _build_report(results, bl_alloc=None, email_sent=email_sent, failures=failures)
    if all_email_params_paths:
        report['email_params_paths'] = all_email_params_paths
    if email_params_path:
        report['email_params_path'] = email_params_path
    if validation:
        report['validation'] = validation
    if checklist:
        report['checklist'] = checklist
    if getattr(args, '_fixes_reports', None):
        report['fixes_applied'] = args._fixes_reports
    if args.json_output:
        print(f"\nREPORT:JSON:{json.dumps(report)}")

    return report


def _print_summary(results: list, bl_alloc) -> None:
    """Print processing summary table."""
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"{'PDF File':<30} {'Supplier':<30} {'Format':<12} "
          f"{'Items':>6} {'Match':>6} {'Class':>6} {'Total':>12}")
    print("-" * 108)
    for r in results:
        print(f"{r.pdf_file:<30} {r.supplier_info.get('name', ''):<30} "
              f"{r.format_name:<12} "
              f"{len(r.matched_items):>6} {r.matched_count:>6} "
              f"{r.classified_count:>6} ${r.invoice_data.get('invoice_total', 0):>10.2f}")
    print("-" * 108)
    total_items = sum(len(r.matched_items) for r in results)
    total_matched = sum(r.matched_count for r in results)
    total_classified = sum(r.classified_count for r in results)
    grand_total = sum(r.invoice_data.get('invoice_total', 0) for r in results)
    print(f"{'TOTAL':<30} {'':<30} {'':<12} "
          f"{total_items:>6} {total_matched:>6} {total_classified:>6}")

    if bl_alloc:
        print(f"\nBL Freight: ${bl_alloc.freight:.2f}  "
              f"Packages: {bl_alloc.packages}  Weight: {bl_alloc.weight} KG  "
              f"Insurance: ${bl_alloc.insurance:.2f}")


def _print_xlsx_variance_detail(ws, inv_total: float, sum_items: float,
                                adjustments: float, net_total: float,
                                variance: float) -> None:
    """Print detailed line-by-line item dump from XLSX when variance is detected."""
    try:
        # Detect description column from header row
        desc_col = 10  # Default to J
        for col in (10, 12):
            val = ws.cell(row=1, column=col).value
            if val and 'desc' in str(val).lower():
                desc_col = col
                break

        w = 100
        print("  " + "-" * w)
        print(f"  DETAILED ITEM BREAKDOWN (from XLSX):")
        print(f"  {'#':>3} | {'Description':<48} | {'Qty':>5} | {'Unit':>8} | {'Total':>10}")
        print("  " + "-" * w)

        item_idx = 0
        running_sum = 0.0
        for row in range(2, ws.max_row + 1):
            tc = ws.cell(row, 16).value  # Column P = TotalCost
            desc = ws.cell(row, desc_col).value
            qty = ws.cell(row, 11).value   # Column K = Quantity
            unit = ws.cell(row, 15).value  # Column O = UnitCost

            # Stop at formula rows (totals section)
            if isinstance(tc, str) and tc.startswith('='):
                break

            if not desc or not isinstance(tc, (int, float)):
                continue

            desc_str = str(desc).upper()
            # Skip group rows (blue background)
            fill = ws.cell(row, 1).fill
            is_group = (fill and fill.start_color and
                        'D9E1F2' in str(getattr(fill.start_color, 'rgb', '') or ''))
            if is_group:
                continue

            item_idx += 1
            running_sum += float(tc)
            qty_str = str(qty) if qty else ''
            unit_val = float(unit) if isinstance(unit, (int, float)) else 0
            print(f"  {item_idx:>3} | {str(desc)[:48]:<48} | {qty_str:>5} | "
                  f"${unit_val:>7.2f} | ${float(tc):>9.2f}")

        print("  " + "-" * w)
        print(f"  {'':>3} | {'ITEMS SUM':<48} | {'':>5} | {'':>8} | ${running_sum:>9.2f}")
        if abs(adjustments) > 0.001:
            print(f"  {'':>3} | {'ADJUSTMENTS':<48} | {'':>5} | {'':>8} | ${adjustments:>9.2f}")
        print(f"  {'':>3} | {'NET TOTAL':<48} | {'':>5} | {'':>8} | ${net_total:>9.2f}")
        print(f"  {'':>3} | {'INVOICE TOTAL':<48} | {'':>5} | {'':>8} | ${inv_total:>9.2f}")
        print(f"  {'':>3} | {'VARIANCE':<48} | {'':>5} | {'':>8} | ${variance:>9.2f}")
        print("  " + "-" * w)
    except Exception as e:
        print(f"  (variance detail failed: {e})")


def _validate_xlsx_variance(results: list) -> None:
    """
    Post-generation XLSX variance check.

    Opens each generated XLSX and verifies the reconciliation formula:
      InvoiceTotal = Sum(ItemCosts) + Freight + OtherCost(tax) - Deductions + Insurance
    Reports PASS/FAIL per file.  Mirrors the XLSX subtotal formulas exactly:
      Adjustments  = T(freight) + U(insurance) + V(tax) - W(deductions)
      Net Total    = Subtotal(items) + Adjustments
      Variance     = InvoiceTotal(S) - Net Total     → must be $0.00
    """
    try:
        import openpyxl
    except ImportError:
        return  # openpyxl not available — skip silently

    if not results:
        return

    fail_count = 0
    print()
    print("VARIANCE CHECK")
    print("-" * 108)
    print(f"  {'File':<30} {'InvTotal':>10} {'Items':>10} {'Freight':>8} "
          f"{'Tax':>8} {'Insur':>8} {'Deduct':>8} {'Net':>10} {'Result':>8}")
    print("  " + "-" * 104)

    for r in results:
        xlsx_path = r.xlsx_path
        if not xlsx_path or not os.path.exists(xlsx_path):
            continue

        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=False)
            ws = wb.active
        except Exception:
            continue

        # Read invoice-level fields from row 2 (first data row)
        def _num(v):
            return v if isinstance(v, (int, float)) else 0

        inv_total = _num(ws.cell(2, 19).value)   # S = InvoiceTotal
        freight = _num(ws.cell(2, 20).value)      # T = Total Internal Freight
        insurance = _num(ws.cell(2, 21).value)    # U = Total Insurance (or -credits)
        tax = _num(ws.cell(2, 22).value)          # V = Total Other Cost (tax)
        deduction = _num(ws.cell(2, 23).value)    # W = Total Deduction

        # Sum item TotalCost (column P = 16), stop at formula rows
        sum_items = 0.0
        for row in range(2, ws.max_row + 1):
            tc = ws.cell(row, 16).value
            if isinstance(tc, (int, float)):
                sum_items += tc
            elif isinstance(tc, str) and tc.startswith('='):
                break

        adjustments = freight + insurance + tax - deduction
        net_total = round(sum_items + adjustments, 2)
        variance = round(inv_total - net_total, 2)

        fname = os.path.basename(xlsx_path)
        if abs(variance) > 0.01:
            fail_count += 1
            status = f"FAIL ${variance:+.2f}"
        else:
            status = "PASS"

        print(f"  {fname:<30} {inv_total:>10.2f} {sum_items:>10.2f} "
              f"{freight:>8.2f} {tax:>8.2f} {insurance:>8.2f} "
              f"{deduction:>8.2f} {net_total:>10.2f} {status:>8}")

        # Detailed line-by-line XLSX item dump when variance detected
        if abs(variance) > 0.01:
            _print_xlsx_variance_detail(ws, inv_total, sum_items, adjustments,
                                       net_total, variance)

        wb.close()

    print("  " + "-" * 104)
    if fail_count == 0:
        print(f"  All {len(results)} invoices PASS — variance = $0.00")
    else:
        print(f"  {fail_count} of {len(results)} invoices FAIL variance check")
        print(f"  Fix: ensure format specs extract freight/tax/credits/deductions")
    print()


def _promote_auto_specs(results: list) -> None:
    """Promote auto-generated format specs that produced successful results."""
    try:
        from workflow.format_spec_generator import promote_spec
    except ImportError:
        return

    promoted = set()
    for r in results:
        spec_path = r.invoice_data.get('_auto_spec_path', '')
        if spec_path and spec_path not in promoted and os.path.exists(spec_path):
            new_path = promote_spec(spec_path)
            if new_path:
                promoted.add(spec_path)
                print(f"    Promoted auto-generated format spec: {os.path.basename(new_path)}")


def _handle_failures(failures: list, output_dir: str, input_dir: str) -> None:
    """Copy failed PDFs to Unprocessed/ subfolder with a failure manifest."""
    if not failures:
        return

    import shutil
    unprocessed_dir = os.path.join(output_dir, 'Unprocessed')
    os.makedirs(unprocessed_dir, exist_ok=True)

    manifest_lines = [f"Source: {input_dir}", ""]
    for f in failures:
        src = f['pdf_path']
        if os.path.exists(src):
            shutil.copy2(src, os.path.join(unprocessed_dir, f['pdf_file']))
        manifest_lines.append(f"{f['pdf_file']}: {f['reason']}")

    manifest_path = os.path.join(unprocessed_dir, '_failures.txt')
    with open(manifest_path, 'w') as mf:
        mf.write('\n'.join(manifest_lines))

    print(f"\nWARNING: {len(failures)} invoice(s) failed — copied to Unprocessed/")
    for f in failures:
        print(f"    {f['pdf_file']}: {f['reason']}")


def _apply_manifest_metadata(args, classification: dict, output_dir: str,
                              all_attachments: list) -> dict:
    """
    Extract manifest metadata and copy manifest PDF to output.

    Returns manifest metadata dict (empty if no manifest found).
    """
    import shutil

    manifest_files = (classification or {}).get('manifest', [])
    if not manifest_files:
        return {}

    manifest_path = os.path.join(args.input_dir, manifest_files[0])
    manifest_meta = extract_manifest_metadata(manifest_path)
    if not manifest_meta:
        return {}

    print(f"\n[Manifest] ASYCUDA Waybill metadata from {manifest_files[0]}:")
    for k, v in manifest_meta.items():
        print(f"    {k}: {v}")

    # Copy manifest PDF to output: {waybill}-Manifest.pdf
    waybill = manifest_meta.get('waybill', '')
    if waybill:
        manifest_out_name = f"{waybill}-Manifest.pdf"
        manifest_out_path = os.path.join(output_dir, manifest_out_name)
        if os.path.abspath(manifest_path) != os.path.abspath(manifest_out_path):
            shutil.copy2(manifest_path, manifest_out_path)
            print(f"    Manifest PDF copied: {manifest_out_name}")
        all_attachments.append(manifest_out_path)

    return manifest_meta


def _apply_consolidation_report(args, classification: dict, results: list,
                                 output_dir: str, all_attachments: list) -> dict:
    """
    Extract consolidation report data and cross-reference with processed invoices.

    The consolidation report (RECAP) maps packages, weight, and freight to invoice
    numbers. This function:
    1. Parses the RECAP to extract per-invoice line items
    2. Matches RECAP invoice refs to processed invoice results
    3. Assigns packages/weight/freight from RECAP to matching invoices
    4. Flags missing invoices (in RECAP but not processed, or vice versa)

    Returns consolidation metadata dict (empty if no report found).
    """
    import shutil

    recap_files = (classification or {}).get('consolidation_report', [])
    if not recap_files:
        return {}

    recap_path = os.path.join(args.input_dir, recap_files[0])
    recap_text = extract_pdf_text(recap_path)
    if not recap_text:
        logger.warning(f"Could not extract text from consolidation report: {recap_files[0]}")
        return {}

    print(f"\n[Consolidation] Processing {recap_files[0]}...")

    # Parse consolidation line items
    # Format: tracking origin invoice_ref packages unit weight volume [value] P freight freight
    recap_lines = []
    total_packages = 0
    total_weight = 0.0
    total_freight = 0.0

    for line in recap_text.split('\n'):
        line = line.strip()
        if not line:
            continue

        # Match detail lines: 6-digit tracking + invoice ref pattern
        m = re.match(
            r'(\d{6})\s+\w+.*?(?:CENTER|CENTRE)\s+'
            r'(\d{3}-\d{7}-\d{3,7})\s+'
            r'(\d+)\s+\w+\s+'
            r'([\d.]+)\s+'          # weight
            r'([\d.]+)'             # volume
            r'(?:\s+([\d,.]+))?'    # optional declared value
            r'\s+P\s+'
            r'([\d,.]+)',           # freight
            line
        )
        if m:
            invoice_ref = m.group(2)
            packages = int(m.group(3))
            weight = float(m.group(4))
            freight_str = m.group(7).replace(',', '')
            freight = float(freight_str)
            value_str = m.group(6)
            declared_value = float(value_str.replace(',', '')) if value_str else 0.0

            recap_lines.append({
                'tracking_ref': m.group(1),
                'invoice_ref': invoice_ref,
                'packages': packages,
                'weight': weight,
                'volume': float(m.group(5)),
                'declared_value': declared_value,
                'freight': freight,
            })
            total_packages += packages
            total_weight += weight
            total_freight += freight
            continue

        # Match totals line: NW/R total_packages total_weight ...
        m_total = re.match(r'(\d+)W/R\s+(\d+)\s+([\d.]+)\s+([\d.]+)', line)
        if m_total:
            total_packages = int(m_total.group(2))
            total_weight = float(m_total.group(3))

    if not recap_lines:
        logger.warning("No line items found in consolidation report")
        return {}

    # Extract waybill from recap
    waybill_match = re.search(r'HBL(\d+)', recap_text)
    waybill = waybill_match.group(1) if waybill_match else ''

    print(f"    Waybill: HBL{waybill}")
    print(f"    {len(recap_lines)} package line(s), {total_packages} total packages, "
          f"{total_weight:.2f} kg, ${total_freight:.2f} freight")

    # Group by invoice reference (multiple package lines per invoice)
    from collections import defaultdict
    invoice_recap = defaultdict(lambda: {'packages': 0, 'weight': 0.0, 'freight': 0.0,
                                          'declared_value': 0.0, 'tracking_refs': []})
    for item in recap_lines:
        ref = item['invoice_ref']
        invoice_recap[ref]['packages'] += item['packages']
        invoice_recap[ref]['weight'] += item['weight']
        invoice_recap[ref]['freight'] += item['freight']
        invoice_recap[ref]['declared_value'] += item['declared_value']
        invoice_recap[ref]['tracking_refs'].append(item['tracking_ref'])

    print(f"    Invoice references in RECAP: {len(invoice_recap)}")
    for ref, data in invoice_recap.items():
        print(f"      {ref}: {data['packages']} pkg, {data['weight']:.2f} kg, "
              f"${data['freight']:.2f} freight"
              + (f", ${data['declared_value']:.2f} value" if data['declared_value'] else ''))

    # Cross-reference with processed invoices
    # Build lookup: normalised invoice number → result
    # Use pre-combination results if available (combined results lose individual invoice numbers)
    all_results = getattr(args, '_pre_combine_results', results) or results
    result_by_invoice = {}
    all_invoice_nums = set()
    for r in all_results:
        inv_num = getattr(r, 'invoice_num', '') or getattr(r, 'invoice_number', '') or ''
        if inv_num:
            result_by_invoice[inv_num] = r
            all_invoice_nums.add(inv_num)
            # Also index by truncated form (RECAP often truncates order numbers)
            # e.g. "113-7797988-8941842" → "113-7797988-894"
            for trunc_len in (15, 14, 13):
                if len(inv_num) > trunc_len:
                    result_by_invoice[inv_num[:trunc_len]] = r

    # Determine the authoritative package cap from BL or manifest
    # Consolidation must NEVER assign more packages than BL/manifest says
    bl_alloc = getattr(args, '_bl_alloc', None)
    bl_pkg_cap = None
    if bl_alloc and getattr(bl_alloc, 'packages', None):
        try:
            bl_pkg_cap = int(bl_alloc.packages)
        except (ValueError, TypeError):
            pass
    if bl_pkg_cap is None:
        # Try manifest
        manifest_meta_check = getattr(args, '_manifest_meta', None) or {}
        if manifest_meta_check.get('packages'):
            try:
                bl_pkg_cap = int(manifest_meta_check['packages'])
            except (ValueError, TypeError):
                pass

    # Match RECAP refs to invoices
    matched = []
    unmatched_recap = []
    recap_assigned_total = 0

    for ref, data in invoice_recap.items():
        # Try exact match first, then prefix match
        matched_result = result_by_invoice.get(ref)
        if not matched_result:
            # Try matching RECAP ref as prefix of full invoice number
            for inv_num, r in result_by_invoice.items():
                if inv_num.startswith(ref) or ref.startswith(inv_num[:len(ref)]):
                    matched_result = r
                    break

        if matched_result:
            matched.append((ref, data, matched_result))
            recap_assigned_total += data['packages']
        else:
            unmatched_recap.append(ref)

    # Validate: consolidation total must not exceed BL/manifest cap
    if bl_pkg_cap and recap_assigned_total > bl_pkg_cap:
        print(f"    ⚠ RECAP total packages ({recap_assigned_total}) exceeds "
              f"BL/manifest ({bl_pkg_cap}) — using BL/manifest allocation instead")
        # Don't override BL allocations
    else:
        # Apply packages from consolidation to matched invoices
        for ref, data, matched_result in matched:
            pkgs = data['packages']

            # Cap individual assignment: can't exceed BL cap minus others already assigned
            if bl_pkg_cap:
                other_assigned = sum(d['packages'] for r2, d, mr in matched
                                     if mr is not matched_result)
                max_for_this = bl_pkg_cap - other_assigned
                if pkgs > max_for_this:
                    print(f"    ⚠ Capping {ref} packages: {pkgs} → {max_for_this} "
                          f"(BL/manifest cap: {bl_pkg_cap})")
                    pkgs = max(max_for_this, 1)

            matched_result.packages = pkgs
            inv_display = getattr(matched_result, 'invoice_num', '') or \
                         getattr(matched_result, 'invoice_number', '?')
            print(f"    ✓ {ref} → {inv_display} ({pkgs} packages)")

            # Update XLSX with packages
            if hasattr(matched_result, 'xlsx_path') and matched_result.xlsx_path and \
               os.path.exists(matched_result.xlsx_path):
                try:
                    from openpyxl import load_workbook as _lw
                    wb = _lw(matched_result.xlsx_path)
                    ws = wb.active
                    ws.cell(row=2, column=24, value=pkgs)
                    # Clear packages from other rows
                    for row in range(3, ws.max_row + 1):
                        if ws.cell(row, 24).value is not None:
                            ws.cell(row, 24).value = None
                    wb.save(matched_result.xlsx_path)
                    wb.close()
                except Exception as e:
                    logger.warning(f"Could not update packages in XLSX for {ref}: {e}")

    # Check for invoices NOT in RECAP
    matched_inv_nums = set()
    for ref, data, r in matched:
        inv = getattr(r, 'invoice_num', '') or getattr(r, 'invoice_number', '') or ''
        matched_inv_nums.add(inv)

    unmatched_invoices = []
    for inv_num in all_invoice_nums:
        if inv_num and inv_num not in matched_inv_nums:
            unmatched_invoices.append(inv_num)

    # Report discrepancies
    if unmatched_recap:
        print(f"\n    ⚠ RECAP refs with NO matching invoice:")
        for ref in unmatched_recap:
            data = invoice_recap[ref]
            print(f"      {ref} ({data['packages']} pkg, ${data['declared_value']:.2f})")
        print(f"      → These invoices may be missing from the shipment!")

    if unmatched_invoices:
        print(f"\n    ⚠ Processed invoices NOT in RECAP:")
        for inv in unmatched_invoices:
            print(f"      {inv}")
        print(f"      → These invoices may not be part of this consolidation")

    if not unmatched_recap and not unmatched_invoices:
        print(f"    ✓ All invoices match between RECAP and processed documents")

    # Copy RECAP PDF to output
    if waybill:
        recap_out_name = f"HBL{waybill}-Consolidation.pdf"
    else:
        recap_out_name = os.path.basename(recap_path)
    recap_out_path = os.path.join(output_dir, recap_out_name)
    if os.path.abspath(recap_path) != os.path.abspath(recap_out_path):
        shutil.copy2(recap_path, recap_out_path)
        print(f"    Consolidation PDF copied: {recap_out_name}")
    all_attachments.append(recap_out_path)

    return {
        'waybill': waybill,
        'total_packages': total_packages,
        'total_weight': total_weight,
        'total_freight': total_freight,
        'invoice_recap': dict(invoice_recap),
        'unmatched_recap_refs': unmatched_recap,
        'unmatched_invoices': unmatched_invoices,
    }


def _enforce_package_cap(results: list, bl_alloc, manifest_meta: dict) -> None:
    """
    Final safety check: ensure total packages across all XLSX files does not
    exceed the authoritative cap from BL or manifest.

    This runs AFTER all package-writing stages (BL allocator, combiner,
    manifest override, consolidation report) as a last line of defense.
    If the total exceeds the cap, trim from the largest allocation and
    rewrite the XLSX files.
    """
    if not results:
        return

    # Determine authoritative package cap (BL > manifest)
    pkg_cap = None
    cap_source = None

    if bl_alloc and getattr(bl_alloc, 'packages', None):
        try:
            pkg_cap = int(bl_alloc.packages)
            cap_source = 'BL'
        except (ValueError, TypeError):
            pass

    if pkg_cap is None and manifest_meta and manifest_meta.get('packages'):
        try:
            pkg_cap = int(manifest_meta['packages'])
            cap_source = 'manifest'
        except (ValueError, TypeError):
            pass

    if pkg_cap is None:
        return  # No authoritative source — nothing to enforce

    # Read actual packages from each result
    actual_packages = []
    for r in results:
        pkgs = getattr(r, 'packages', 1) or 1
        actual_packages.append(pkgs)

    actual_total = sum(actual_packages)

    if actual_total <= pkg_cap:
        return  # Within cap — all good

    # ── Over cap: trim and fix ──
    print(f"\n[PACKAGE CAP] Total packages ({actual_total}) exceeds "
          f"{cap_source} cap ({pkg_cap}) — correcting...")

    # Redistribute: proportionally reduce to fit cap, minimum 1 per entry
    n = len(results)
    if pkg_cap < n:
        # More entries than packages — each gets 1
        corrected = [1] * n
    else:
        # Proportional reduction
        corrected = list(actual_packages)
        surplus = actual_total - pkg_cap

        while surplus > 0:
            # Trim from the largest allocation (keep minimum 1)
            max_idx = max(range(n), key=lambda i: corrected[i])
            if corrected[max_idx] <= 1:
                break  # Can't reduce further
            trim = min(corrected[max_idx] - 1, surplus)
            corrected[max_idx] -= trim
            surplus -= trim

    # Apply corrections
    for i, r in enumerate(results):
        old_pkg = actual_packages[i]
        new_pkg = corrected[i]
        if old_pkg != new_pkg:
            r.packages = new_pkg
            inv_id = getattr(r, 'invoice_num', '') or getattr(r, 'invoice_number', '?')
            print(f"    {inv_id}: {old_pkg} → {new_pkg} packages")

            # Update XLSX file
            xlsx_path = getattr(r, 'xlsx_path', '')
            if xlsx_path and os.path.exists(xlsx_path):
                try:
                    from openpyxl import load_workbook as _lw
                    wb = _lw(xlsx_path)
                    ws = wb.active
                    ws.cell(row=2, column=24, value=new_pkg)
                    # Clear stale package values from other rows
                    for row in range(3, ws.max_row + 1):
                        if ws.cell(row, 24).value is not None:
                            ws.cell(row, 24).value = None
                    wb.save(xlsx_path)
                    wb.close()
                except Exception as e:
                    logger.warning(f"Could not fix packages in XLSX: {e}")

    final_total = sum(corrected)
    print(f"    Corrected: {actual_total} → {final_total} "
          f"(cap: {pkg_cap} from {cap_source})")


def _dedupe_attachments(all_attachments: list) -> list:
    """Deduplicate attachment list by path and by file content (MD5).

    Handles: same path listed twice, and different-named files with identical content.
    Preserves order, keeps the first occurrence.
    """
    import hashlib
    seen_paths = set()
    seen_hashes = set()
    unique = []
    for p in all_attachments:
        if not p or not os.path.exists(p):
            continue
        abspath = os.path.abspath(p)
        if abspath in seen_paths:
            continue
        seen_paths.add(abspath)
        # Hash file content to catch same-content-different-name duplicates
        try:
            h = hashlib.md5(open(abspath, 'rb').read()).hexdigest()
            if h in seen_hashes:
                logger.info(f"Skipping duplicate attachment (same content): {os.path.basename(p)}")
                continue
            seen_hashes.add(h)
        except Exception:
            pass
        unique.append(p)
    return unique


_OFFICE_LOCATION_MAP = None
_OFFICE_ALIAS_MAP = None

def _load_office_location_map() -> tuple:
    """Load config/office_locations.yaml → ({office_code: location}, {alias: code})."""
    global _OFFICE_LOCATION_MAP, _OFFICE_ALIAS_MAP
    if _OFFICE_LOCATION_MAP is not None:
        return _OFFICE_LOCATION_MAP, _OFFICE_ALIAS_MAP
    try:
        import yaml
        path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                            'config', 'office_locations.yaml')
        with open(path) as f:
            data = yaml.safe_load(f) or {}
        _OFFICE_LOCATION_MAP = {str(k).upper(): str(v)
                                for k, v in (data.get('offices') or {}).items()}
        _OFFICE_ALIAS_MAP = {str(k).upper(): str(v).upper()
                             for k, v in (data.get('office_aliases') or {}).items()}
    except Exception as e:
        logger.warning(f"Could not load office_locations.yaml: {e}")
        _OFFICE_LOCATION_MAP = {}
        _OFFICE_ALIAS_MAP = {}
    return _OFFICE_LOCATION_MAP, _OFFICE_ALIAS_MAP


def _extract_office_from_bl(args) -> str:
    """Extract office/port from BL port of discharge when no manifest/declaration."""
    classification = getattr(args, '_classification', {})
    bl_files = classification.get('bill_of_lading', [])
    if not bl_files or not hasattr(args, 'input_dir'):
        return ''
    try:
        from stages.supplier_resolver import extract_pdf_text
        bl_path = os.path.join(args.input_dir, bl_files[0])
        if not os.path.exists(bl_path):
            return ''
        bl_text = extract_pdf_text(bl_path)
        if not bl_text:
            return ''
        # "PORT OF DISCHARGE\nST GEORGES SEAPORT (GRENADA)"
        m = re.search(r'PORT\s+OF\s+DISCHARGE[^\n]*\n\s*([A-Z][A-Z ]+?)(?:\s+SEAPORT|\s+PORT|\s*\()', bl_text, re.IGNORECASE)
        if m:
            port = m.group(1).strip()
            logger.info(f"Office from BL port of discharge: {port}")
            return port
    except Exception as e:
        logger.debug(f"BL office extraction failed: {e}")
    return ''


def _resolve_location_office(args, manifest_meta: dict) -> tuple:
    """Return (location, office) from manifest + config mapping + args.

    Office: manifest/declaration customs office takes precedence over --office.
    If the manifest provides a full name (e.g. 'ST GEORGES'), it is resolved
    to the canonical code (e.g. 'GDSGO') via office_aliases.
    Falls back to BL port of discharge when no manifest/declaration.
    Location: looked up from config/office_locations.yaml by office code;
              falls back to --location (CLI default: 'WebSource').
    """
    offices_map, alias_map = _load_office_location_map()
    raw_office = (manifest_meta or {}).get('office') or getattr(args, 'office', '') or ''

    # Fallback: extract office from BL port of discharge when no manifest
    if not raw_office:
        raw_office = _extract_office_from_bl(args)

    office_key = raw_office.strip().upper()

    # Resolve alias (full name → code): "ST GEORGES" → "GDSGO"
    if office_key and office_key in alias_map:
        office_key = alias_map[office_key]

    # Look up location by (possibly resolved) office code
    mapped = offices_map.get(office_key) if office_key else None
    location = mapped or getattr(args, 'location', '') or 'WebSource'

    # Return the canonical code (not the raw full name)
    office = office_key if office_key in offices_map else raw_office
    return location, office


def _maybe_apply_fixes(args, results: list) -> None:
    """Auto-discover + apply reviewer-edited Proposed Fixes YAMLs.

    Called after Phase 1 processing so each invoice's matched_items and
    invoice_data are in memory but the combine/email steps have not yet
    started.  Stores the discovered YAML paths on ``args._fixes_yaml_paths``
    (so they can be archived after the email is sent) and the per-invoice
    change reports on ``args._fixes_reports`` (Phase 4 candidate journal).

    Any failure here is logged and swallowed — a bad fixes file must not
    block a legitimate shipment from being processed.
    """
    args._fixes_reports = []
    args._fixes_yaml_paths = []
    try:
        import apply_fixes
        fixes_paths = apply_fixes.discover_fixes(getattr(args, 'input_dir', '') or '')
        if not fixes_paths:
            return
        docs = []
        for fp in fixes_paths:
            try:
                docs.append(apply_fixes.load_fixes_yaml(fp))
                args._fixes_yaml_paths.append(fp)
            except Exception as e:
                logger.warning(f"fixes: could not load {fp}: {e}")
        fixes_map = apply_fixes.build_fixes_map(docs)
        if not fixes_map:
            return
        print(f"\n[3.5] Applying reviewer fixes ({len(fixes_map)} invoice(s))...")
        reports = apply_fixes.apply_fixes_to_results(results, fixes_map)
        args._fixes_reports = reports
        # Phase 4: append to the learned-fixes candidate journal so
        # recurring fix patterns can be promoted into format specs later.
        try:
            waybill = getattr(args, 'bl', '') or getattr(args, 'waybill', '') or ''
            apply_fixes.log_fix_candidates(reports, results, waybill, BASE_DIR)
        except Exception as e:
            logger.warning(f"fixes: candidate journal write failed: {e}")
        # Regenerate XLSX for any invoice that was touched
        from bl_xlsx_generator import generate_bl_xlsx
        touched = {r['invoice_num'] for r in reports}
        for r in results:
            if r.invoice_num in touched:
                try:
                    generate_bl_xlsx(
                        r.invoice_data,
                        r.matched_items,
                        os.path.basename(r.xlsx_path).rsplit('.', 1)[0],
                        r.supplier_info,
                        r.xlsx_path,
                        document_type=getattr(args, 'doc_type', 'auto'),
                    )
                    print(f"    {r.invoice_num}: XLSX regenerated with fixes")
                except Exception as e:
                    logger.warning(
                        f"fixes: failed to regenerate XLSX for {r.invoice_num}: {e}"
                    )
    except Exception as e:
        logger.warning(f"fixes replay failed: {e}")


def _maybe_archive_applied_fixes(args, waybill: str) -> None:
    """Move applied Proposed Fixes YAMLs to ``data/learned_fixes/YYYY-MM/``.

    Called after the shipment email has been saved/sent so we only archive
    fixes that were actually committed.  Silent if no fixes were applied.
    """
    paths = getattr(args, '_fixes_yaml_paths', None) or []
    if not paths:
        return
    try:
        import apply_fixes
        for p in paths:
            try:
                dest = apply_fixes.archive_fixes_yaml(p, waybill, BASE_DIR)
                if dest:
                    print(f"    Archived fixes: {os.path.basename(dest)}")
            except Exception as e:
                logger.warning(f"fixes: could not archive {p}: {e}")
    except Exception as e:
        logger.warning(f"fixes archive stage failed: {e}")


def _maybe_save_proposed_fixes(results: list, email_params_path: str, output_dir: str) -> None:
    """Detect uncertain invoices and write a _proposed_fixes_params.json sidecar.

    Scans ``results`` for invoices that carry uncertainty markers
    (``data_quality_notes``, ``invoice_total_uncertain``, or per-item
    ``data_quality``) and, if any are present, writes the YAML patch plus
    the email-params sidecar so ``send_shipment_email.py`` can send a
    second email to the reviewer mailbox.  Failures are logged and
    swallowed — they must never break the shipment email flow.
    """
    try:
        import proposed_fixes
        waybill = 'UNKNOWN'
        if email_params_path and os.path.exists(email_params_path):
            try:
                with open(email_params_path) as f:
                    waybill = json.load(f).get('waybill', 'UNKNOWN')
            except Exception:
                pass
        uncertain = proposed_fixes.detect_uncertain_invoices(results)
        if not uncertain:
            return
        artefacts = proposed_fixes.save_fixes_artifacts(waybill, uncertain, output_dir)
        if artefacts:
            print(
                f"    Proposed Fixes: {len(uncertain)} uncertain invoice(s) → "
                f"{os.path.basename(artefacts['params_path'])}"
            )
    except Exception as e:
        logger.warning(f"proposed_fixes save failed: {e}")


def _save_email_params(args, results: list, bl_alloc, all_attachments: list,
                       manifest_meta: dict = None, output_dir: str = '',
                       total_invoices: int = 0) -> str:
    """
    Compute email params from pipeline results and save to _email_params.json.
    Does NOT send — email sending is a separate step (send_shipment_email.py).
    Returns the path to the saved params file.
    """
    # Determine predominant country of origin
    countries = [r.supplier_info.get('country', 'US') for r in results]
    country_origin = max(set(countries), key=countries.count)

    # Use BL freight if available, else sum of invoice freight
    if bl_alloc:
        email_freight = bl_alloc.freight
        email_packages = bl_alloc.packages
        email_weight = bl_alloc.weight
        # Fallback when BL doesn't provide packages/weight
        if not email_packages or email_packages == '0':
            email_packages = str(len(results))
        if not email_weight or email_weight == '0':
            total_qty = sum(
                sum(int(m.get('quantity', 0) or 0) for m in r.matched_items)
                for r in results
            )
            email_weight = str(round(total_qty * 0.01, 2)) if total_qty > 0 else '1'
    else:
        email_freight = sum(r.freight for r in results)
        email_packages = str(len(results))
        email_weight = '0'

    bl_number = args.bl or (bl_alloc.bl_data.get('bl_number', '') if bl_alloc else '') or ''
    # Fallback: use invoice number from first result as waybill (for PDFs without declarations)
    if not bl_number and results:
        first_inv_num = getattr(results[0], 'invoice_num', '') or ''
        if first_inv_num and first_inv_num not in ('unknown', 'combined'):
            bl_number = re.sub(r'[<>:"/\\|?*]', '_', first_inv_num)
    if not bl_number:
        bl_number = 'Next Shipment'
    consignee_address = ''
    consignee_name = _extract_consignee(args) or args.consignee

    # Use consignee_code/address from matched rule ONLY when the match was on
    # the actual consignee name (not a supplier-only match).  A supplier match
    # tells us the doc_type but the consignee is a different entity.
    matched_rule = getattr(args, '_matched_consignee_rule', None)
    match_type = getattr(args, '_consignee_match_type', None)

    if match_type == 'consignee' and matched_rule:
        # Direct consignee match — use rule's code + address
        consignee_code = matched_rule.get('consignee_code', '') or args.consignee_code
        if matched_rule.get('consignee_address'):
            consignee_address = matched_rule['consignee_address']
    elif match_type == 'supplier' and matched_rule and not consignee_name:
        # Supplier match with no consignee extracted — use rule's consignee data
        # as fallback (e.g. Budget Marine St. Maarten ships to Budget Marine Grenada,
        # and the BL OCR was too poor to extract the consignee)
        consignee_code = matched_rule.get('consignee_code', '') or args.consignee_code
        consignee_name = matched_rule.get('consignee_name', '') or matched_rule.get('match', '').title() or ''
        if matched_rule.get('consignee_address'):
            consignee_address = matched_rule['consignee_address']
        if consignee_name:
            logger.info(f"Using consignee from supplier rule fallback: {consignee_name}")
    else:
        # No match — use CLI arg / leave blank
        consignee_code = args.consignee_code

    # Fallback: look up consignee code from shipment_rules.yaml
    if not consignee_code and consignee_name:
        sr_code, sr_address = _lookup_consignee_code(consignee_name)
        if sr_code:
            consignee_code = sr_code
            if sr_address and not consignee_address:
                consignee_address = sr_address

    # If no address from rule, use address extracted from invoice (Bill To block)
    if not consignee_address:
        consignee_address = getattr(args, '_invoice_consignee_address', '') or ''

    # Manifest/declaration overrides
    if manifest_meta:
        if manifest_meta.get('waybill') and _looks_like_bl_number(manifest_meta['waybill']):
            bl_number = manifest_meta['waybill']
        if manifest_meta.get('man_reg'):
            args.man_reg = manifest_meta['man_reg']
        if manifest_meta.get('consignee_address'):
            consignee_address = manifest_meta['consignee_address']
        _placeholders = {'SAME AS CONSIGNEE', 'SAME AS SHIPPER', 'SAME AS ABOVE',
                         'AS PER SHIPPER', 'TO ORDER', 'TO THE ORDER OF'}
        manifest_consignee = (manifest_meta.get('consignee') or '').strip()
        if manifest_consignee and manifest_consignee.upper() not in _placeholders and not consignee_name:
            consignee_name = manifest_consignee
        if manifest_meta.get('packages'):
            email_packages = manifest_meta['packages']
        if manifest_meta.get('weight'):
            email_weight = manifest_meta['weight']
        # Use manifest freight when no BL freight is available
        if manifest_meta.get('freight') and (not email_freight or float(email_freight) == 0):
            email_freight = manifest_meta['freight']

    # Freight invoice takes precedence over all other freight sources
    # Use the freight line item (excludes landing charges), fall back to total
    freight_inv = getattr(args, '_freight_invoice_data', {})
    if freight_inv:
        if freight_inv.get('freight'):
            email_freight = freight_inv['freight']
        elif freight_inv.get('total'):
            email_freight = freight_inv['total']
        if freight_inv.get('packages'):
            email_packages = freight_inv['packages']
        # Freight invoice origin (e.g. "FROM ST. MARTIN TO GRENADA") overrides country_origin
        if freight_inv.get('origin_country'):
            country_origin = freight_inv['origin_country']

    # Fallback: extract man_reg from email.txt if not found in BL/declaration
    if not getattr(args, 'man_reg', ''):
        email_txt = os.path.join(output_dir, 'email.txt')
        if os.path.isfile(email_txt):
            try:
                with open(email_txt, 'r', encoding='utf-8', errors='replace') as f:
                    email_body = f.read()
                mr = re.search(
                    r'(?:MNF|MAN(?:IFEST)?)\s*(?:(?:REG(?:ISTRY)?|NUM(?:BER)?))?\s*#?\s*:?\s*(\d{4})\s*[-/\s]\s*(\d+)',
                    email_body, re.IGNORECASE)
                if mr:
                    args.man_reg = f"{mr.group(1)} {mr.group(2)}"
                    logger.info(f"man_reg from email.txt: {args.man_reg}")
            except Exception:
                pass

    # Collect OCR quality warnings
    ocr_warnings = []
    for r in results:
        inv_data = getattr(r, 'invoice_data', {}) or {}
        ocr_q = inv_data.get('ocr_quality', {})
        if ocr_q and ocr_q.get('rating') in ('poor', 'unusable'):
            ocr_warnings.append(
                f"{getattr(r, 'pdf_file', '?')}: OCR quality {ocr_q.get('rating').upper()} "
                f"(score {ocr_q.get('score', 0)}/100)"
            )

    _email_location, _email_office = _resolve_location_office(args, manifest_meta)
    params = {
        'waybill': bl_number,
        'consignee_name': consignee_name,
        'consignee_code': consignee_code,
        'consignee_address': consignee_address,
        'total_invoices': total_invoices or len(results),
        'expected_entries': len(results),
        'packages': email_packages,
        'weight': email_weight,
        'country_origin': country_origin,
        'freight': str(email_freight),
        'man_reg': getattr(args, 'man_reg', ''),
        'attachment_paths': _dedupe_attachments(all_attachments),
        'location': _email_location,
        'office': _email_office,
    }

    if ocr_warnings:
        params['ocr_warnings'] = ocr_warnings
        params['notes'] = (
            'WARNING: Some pages have poor OCR quality. '
            'Check the "OCR Notes" sheet in the combined XLSX for details. '
            'Consider asking sender to rescan and resubmit.'
        )

    params_path = os.path.join(output_dir, '_email_params.json')
    with open(params_path, 'w') as f:
        json.dump(params, f, indent=2)
    print(f"    Email params saved: {params_path}")
    return params_path


def _record_send_history(args, params: dict, email_draft: dict) -> None:
    """Record a successful send in data/send_history.json (best-effort)."""
    try:
        send_history.record_send(
            waybill=params.get('waybill', ''),
            subject=email_draft.get('subject', ''),
            source_input=getattr(args, '_source_input', '') or '',
            source_mode=getattr(args, '_source_mode', '') or '',
            output_dir=getattr(args, '_output_dir', '') or '',
            params=params,
            attachments=email_draft.get('attachments', []),
        )
    except Exception as e:
        logger.warning(f"send history recording failed: {e}")


def _send_proposed_fixes_sidecar(output_dir: str) -> bool:
    """Send the Proposed Fixes email from ``_proposed_fixes_params.json``.

    Mirrors the sidecar handoff in ``send_shipment_email.py`` so the
    legacy in-process ``--send-email`` path (_send_bl_email /
    _send_batch_email / _send_email_from_params) also delivers the
    reviewer-facing email when the pipeline detected uncertain invoices.

    Returns True if the sidecar was sent (or absent — nothing to do);
    returns False only when a sidecar exists and the send failed.
    Failures are logged but do not raise — they must never take down
    the main shipment email flow.
    """
    if not output_dir:
        return True
    fixes_params_path = os.path.join(output_dir, '_proposed_fixes_params.json')
    if not os.path.exists(fixes_params_path):
        return True
    try:
        from workflow.email import compose_proposed_fixes_email, send_email as do_send_email
        from core.config import get_config
        with open(fixes_params_path) as f:
            fparams = json.load(f)
        draft = compose_proposed_fixes_email(
            waybill=fparams.get('waybill', 'UNKNOWN'),
            subject=fparams.get('subject', ''),
            body=fparams.get('body', ''),
            attachment_paths=fparams.get('attachment_paths', []),
        )
        cfg = get_config()
        recipient = getattr(cfg, 'email_fixes_recipient', None) or cfg.email_sender
        sent = do_send_email(
            subject=draft['subject'],
            body=draft['body'],
            attachments=draft['attachments'],
            recipient=recipient,
        )
        if sent:
            print(
                f"\nProposed Fixes email sent to {recipient}: "
                f"{draft['subject']} ({len(draft['attachments'])} attachments)"
            )
        else:
            print(f"\nProposed Fixes email FAILED: {draft['subject']}")
        return sent
    except Exception as e:
        logger.warning(f"proposed fixes send failed: {e}")
        return False


def _send_bl_email(args, results: list, bl_alloc, all_attachments: list,
                   manifest_meta: dict = None, total_invoices: int = 0) -> bool:
    """Legacy: send BL email directly (used when --send-email is passed from CLI)."""
    from workflow.email import compose_email, send_email as do_send_email

    output_dir = getattr(args, '_output_dir', '')
    params_path = _save_email_params(args, results, bl_alloc, all_attachments,
                                     manifest_meta, output_dir,
                                     total_invoices=total_invoices)
    with open(params_path) as f:
        params = json.load(f)

    email_draft = compose_email(
        waybill=params['waybill'],
        consignee_name=params['consignee_name'],
        consignee_code=params['consignee_code'],
        consignee_address=params['consignee_address'],
        total_invoices=params['total_invoices'],
        packages=params['packages'],
        weight=params['weight'],
        country_origin=params['country_origin'],
        freight=params['freight'],
        man_reg=params['man_reg'],
        attachment_paths=params['attachment_paths'],
        location=params['location'],
        office=params['office'],
        expected_entries=params.get('expected_entries', 0),
    )

    email_sent = do_send_email(
        subject=email_draft['subject'],
        body=email_draft['body'],
        attachments=email_draft['attachments'],
    )

    if email_sent:
        print(f"\nEmail sent: {email_draft['subject']} "
              f"({len(email_draft['attachments'])} attachments)")
        _record_send_history(args, params, email_draft)
    else:
        print(f"\nEmail FAILED to send")

    # Send the Proposed Fixes sidecar email (if any) to the reviewer mailbox
    _send_proposed_fixes_sidecar(output_dir)

    return email_sent


def _save_batch_email_params(args, results: list, all_attachments: list,
                             manifest_meta: dict = None, output_dir: str = '') -> str:
    """Compute batch email params and save to _email_params.json. Returns path."""
    countries = [r.supplier_info.get('country', 'US') for r in results]
    country_origin = max(set(countries), key=countries.count)

    email_freight = sum(r.freight for r in results)
    email_packages = str(len(results))
    # Compute minimum weight from total item quantities (ASYCUDA minimum: 0.01 kg per item)
    total_qty = sum(
        sum(int(m.get('quantity', 0) or 0) for m in r.matched_items)
        for r in results
    )
    email_weight = str(round(total_qty * 0.01, 2)) if total_qty > 0 else '0'
    consignee_address = ''
    consignee_name = _extract_consignee(args) or args.consignee

    # Use consignee_code from matched rule (config/document_types.json), fall back to CLI arg
    matched_rule = getattr(args, '_matched_consignee_rule', None)
    match_type = getattr(args, '_consignee_match_type', None)
    consignee_code = (matched_rule or {}).get('consignee_code', '') or args.consignee_code
    if matched_rule and matched_rule.get('consignee_address'):
        consignee_address = matched_rule['consignee_address']
    # Supplier match with no consignee extracted — use rule's consignee name as fallback
    if matched_rule and not consignee_name:
        consignee_name = matched_rule.get('consignee_name', '') or matched_rule.get('match', '').title() or ''
        if consignee_name:
            logger.info(f"Using consignee from rule fallback: {consignee_name}")

    # Fallback: look up consignee code from shipment_rules.yaml
    if not consignee_code and consignee_name:
        sr_code, sr_address = _lookup_consignee_code(consignee_name)
        if sr_code:
            consignee_code = sr_code
            if sr_address and not consignee_address:
                consignee_address = sr_address

    import re as _re_wb
    waybill = args.bl or args.waybill or ''
    if not waybill and results:
        # Fallback: use invoice number as waybill (for PDFs without declarations)
        first_inv_num = getattr(results[0], 'invoice_num', '') or ''
        if first_inv_num and first_inv_num not in ('unknown', 'combined'):
            waybill = _re_wb.sub(r'[<>:"/\\|?*]', '_', first_inv_num)
    if not waybill:
        waybill = 'Next Shipment'

    if manifest_meta:
        if manifest_meta.get('waybill') and _looks_like_bl_number(manifest_meta['waybill']):
            waybill = manifest_meta['waybill']
        if manifest_meta.get('man_reg'):
            args.man_reg = manifest_meta['man_reg']
        _placeholders = {'SAME AS CONSIGNEE', 'SAME AS SHIPPER', 'SAME AS ABOVE',
                         'AS PER SHIPPER', 'TO ORDER', 'TO THE ORDER OF'}
        manifest_consignee = (manifest_meta.get('consignee_name') or '').strip()
        if manifest_consignee and manifest_consignee.upper() not in _placeholders:
            consignee_name = manifest_consignee
        if manifest_meta.get('consignee_address'):
            consignee_address = manifest_meta['consignee_address']
        if manifest_meta.get('packages'):
            email_packages = manifest_meta['packages']
        if manifest_meta.get('weight'):
            email_weight = manifest_meta['weight']
        if manifest_meta.get('freight'):
            email_freight = float(manifest_meta['freight'])

    # Freight invoice takes precedence over all other freight sources
    # Use the freight line item (excludes landing charges), fall back to total
    freight_inv = getattr(args, '_freight_invoice_data', {})
    if freight_inv:
        if freight_inv.get('freight'):
            email_freight = float(freight_inv['freight'])
        elif freight_inv.get('total'):
            email_freight = float(freight_inv['total'])
        if freight_inv.get('packages'):
            email_packages = freight_inv['packages']
        if freight_inv.get('origin_country'):
            country_origin = freight_inv['origin_country']

    # Fallback: extract man_reg from email.txt if not found in BL/declaration
    if not getattr(args, 'man_reg', ''):
        email_txt = os.path.join(output_dir, 'email.txt')
        if os.path.isfile(email_txt):
            try:
                with open(email_txt, 'r', encoding='utf-8', errors='replace') as f:
                    email_body = f.read()
                mr = re.search(
                    r'(?:MNF|MAN(?:IFEST)?)\s*(?:(?:REG(?:ISTRY)?|NUM(?:BER)?))?\s*#?\s*:?\s*(\d{4})\s*[-/\s]\s*(\d+)',
                    email_body, re.IGNORECASE)
                if mr:
                    args.man_reg = f"{mr.group(1)} {mr.group(2)}"
                    logger.info(f"man_reg from email.txt: {args.man_reg}")
            except Exception:
                pass

    params = {
        'waybill': waybill,
        'consignee_name': consignee_name,
        'consignee_code': consignee_code,
        'consignee_address': consignee_address,
        'total_invoices': len(results),
        'expected_entries': len(results),
        'packages': email_packages,
        'weight': email_weight,
        'country_origin': country_origin,
        'freight': str(email_freight),
        'man_reg': getattr(args, 'man_reg', ''),
        'attachment_paths': _dedupe_attachments(all_attachments),
        'location': _resolve_location_office(args, manifest_meta)[0],
        'office': _resolve_location_office(args, manifest_meta)[1],
    }

    params_path = os.path.join(output_dir, '_email_params.json')
    with open(params_path, 'w') as f:
        json.dump(params, f, indent=2)
    print(f"    Email params saved: {params_path}")
    return params_path


def _send_batch_email(args, results: list, all_attachments: list,
                      manifest_meta: dict = None) -> bool:
    """Legacy: send batch email directly (used when --send-email is passed from CLI)."""
    from workflow.email import compose_email, send_email as do_send_email

    output_dir = getattr(args, '_output_dir', '')
    params_path = _save_batch_email_params(args, results, all_attachments,
                                           manifest_meta, output_dir)
    with open(params_path) as f:
        params = json.load(f)

    email_draft = compose_email(**{k: v for k, v in params.items()
                                   if k != 'attachment_paths'},
                                attachment_paths=params['attachment_paths'])

    email_sent = do_send_email(
        subject=email_draft['subject'],
        body=email_draft['body'],
        attachments=email_draft['attachments'],
    )

    if email_sent:
        print(f"\nEmail sent: {email_draft['subject']} "
              f"({len(email_draft['attachments'])} attachments)")
        _record_send_history(args, params, email_draft)
    else:
        print(f"\nEmail FAILED to send")

    # Send the Proposed Fixes sidecar email (if any) to the reviewer mailbox
    _send_proposed_fixes_sidecar(output_dir)

    return email_sent


def _send_email_from_params(params_path: str, args=None) -> bool:
    """Send email from a saved _email_params.json file."""
    from workflow.email import compose_email, send_email as do_send_email

    with open(params_path) as f:
        params = json.load(f)

    email_draft = compose_email(**{k: v for k, v in params.items()
                                   if k != 'attachment_paths'},
                                attachment_paths=params['attachment_paths'])

    email_sent = do_send_email(
        subject=email_draft['subject'],
        body=email_draft['body'],
        attachments=email_draft['attachments'],
    )

    if email_sent:
        print(f"    Email sent: {email_draft['subject']} "
              f"({len(email_draft['attachments'])} attachments)")
        if args is not None:
            _record_send_history(args, params, email_draft)
    else:
        print(f"    Email FAILED: {email_draft['subject']}")

    return email_sent


def _build_report(results: list, bl_alloc, email_sent: bool, failures: list = None) -> dict:
    """Build JSON report for TypeScript consumption."""
    invoices = []
    for r in results:
        invoices.append({
            'invoice_num': r.invoice_num,
            'supplier': r.supplier_info.get('name', ''),
            'format': r.format_name,
            'items': len(r.matched_items),
            'matched': r.matched_count,
            'classified': r.classified_count,
            'total': r.invoice_data.get('invoice_total', 0),
            'freight': r.freight,
            'packages': r.packages,
            'xlsx_path': r.xlsx_path,
            'pdf_path': r.pdf_output_path,
        })

    report = {
        'status': 'success',
        'mode': 'bl' if bl_alloc else 'batch',
        'invoice_count': len(results),
        'invoices': invoices,
        'email_sent': email_sent,
    }

    if bl_alloc:
        report['bl'] = {
            'bl_number': bl_alloc.bl_data.get('bl_number', ''),
            'freight': bl_alloc.freight,
            'packages': bl_alloc.packages,
            'weight': bl_alloc.weight,
            'insurance': bl_alloc.insurance,
            'bl_pdf': bl_alloc.bl_output_path,
        }

    if failures:
        report['failures'] = failures
        report['failed_count'] = len(failures)

    return report


def run_send_email(args) -> dict:
    """
    Send-email-only mode: compose and send email via workflow/email.py.

    This is the SINGLE code path for ALL email sending in the system.
    Called by TypeScript after pipeline processing (single invoice or BL).
    """
    from workflow.email import compose_email, send_email as do_send_email

    # Parse attachment paths from comma-separated string
    attachment_paths = []
    if args.attachments:
        attachment_paths = [p.strip() for p in args.attachments.split(',') if p.strip()]

    email_draft = compose_email(
        waybill=args.waybill or 'Next Shipment',
        consignee_name=args.consignee,
        consignee_code=args.consignee_code,
        total_invoices=args.total_invoices,
        packages=args.packages or '1',
        weight=args.weight or '0',
        country_origin=args.country_origin or 'US',
        freight=args.freight or '0',
        man_reg=args.man_reg,
        attachment_paths=attachment_paths,
        location=args.location,
        office=args.office,
    )

    email_sent = do_send_email(
        subject=email_draft['subject'],
        body=email_draft['body'],
        attachments=email_draft['attachments'],
    )

    if email_sent:
        print(f"Email sent: {email_draft['subject']} "
              f"({len(email_draft['attachments'])} attachments)")
    else:
        print(f"Email FAILED to send")

    report = {
        'status': 'success' if email_sent else 'error',
        'email_sent': email_sent,
        'subject': email_draft['subject'],
        'attachments': len(email_draft['attachments']),
    }

    if args.json_output:
        print(f"\nREPORT:JSON:{json.dumps(report)}")

    return report


def _reprocess_history_entry(entry: dict) -> Optional[dict]:
    """
    Re-run the pipeline against the source_input of a prior send.

    Reuses the original output_dir so XLSX/PDF artifacts are refreshed in place.
    Returns the new params dict (same shape as _email_params.json), or None if
    the source is no longer available.
    """
    source_input = entry.get('source_input', '')
    output_dir = entry.get('output_dir', '')
    waybill = entry.get('waybill', '')

    if not source_input or not os.path.exists(source_input):
        logger.error(
            f"resend {waybill}: source input missing ({source_input!r}); "
            f"cannot reprocess"
        )
        return None

    # Build a synthetic args namespace that mirrors the original invocation.
    fake = argparse.Namespace(
        input_dir=None,
        input=None,
        send_email_only=False,
        output_dir=output_dir or os.path.join(BASE_DIR, 'workspace', 'shipments'),
        output=None,
        bl=None,
        po_file=None,
        doc_type='auto',
        send_email=False,         # do NOT send during the dry re-run
        waybill=None,
        consignee='',
        consignee_code='',
        man_reg=None,
        location='WebSource',
        office='',
        attachments=None,
        total_invoices=1,
        packages=None,
        weight=None,
        country_origin=None,
        freight=None,
        config=None,
        json_output=False,
        verbose=False,
        resend=None,
        resend_stale=False,
    )
    if os.path.isdir(source_input):
        fake.input_dir = source_input
    else:
        fake.input = source_input

    # Run detect_mode + pipeline. This produces fresh _email_params.json.
    mode = detect_mode(fake)
    fake._source_input = source_input
    fake._source_mode = mode
    if mode == 'bl':
        run_bl_mode(fake)
    elif mode == 'batch':
        run_batch_mode(fake)
    else:
        run_single_mode(fake)

    # Read the freshly-written params.
    new_params_path = os.path.join(
        getattr(fake, '_output_dir', '') or fake.output_dir,
        '_email_params.json',
    )
    if not os.path.isfile(new_params_path):
        logger.error(f"resend {waybill}: no _email_params.json at {new_params_path}")
        return None
    with open(new_params_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def run_resend(args) -> dict:
    """
    Resend previously-sent shipments whose re-run would differ from what was
    sent. Entry points:
        --resend <waybill>   -- unconditionally reprocess and resend one waybill
        --resend-stale       -- scan history, reprocess each entry, resend if
                                the re-run produces different params
    """
    from workflow.email import compose_email, send_email as do_send_email

    targets: list = []
    if args.resend:
        entry = send_history.find_by_waybill(args.resend)
        if not entry:
            print(f"ERROR: no send history entry for waybill {args.resend!r}")
            return {'status': 'error', 'resent': 0}
        targets.append(('force', entry))
    else:
        for entry in send_history.all_entries():
            targets.append(('stale', entry))

    if not targets:
        print("No history entries to consider.")
        return {'status': 'success', 'resent': 0, 'checked': 0}

    resent = 0
    unchanged = 0
    skipped = 0
    for reason, entry in targets:
        waybill = entry.get('waybill', '(unknown)')
        print(f"\n[resend:{reason}] {waybill}  source={entry.get('source_input','')}")
        new_params = _reprocess_history_entry(entry)
        if new_params is None:
            skipped += 1
            continue

        new_hash = send_history.params_hash(new_params)
        old_hash = entry.get('params_hash', '')
        if reason == 'stale' and new_hash == old_hash:
            print(f"    unchanged (params_hash match) — not resending")
            unchanged += 1
            continue

        _COMPOSE_KEYS = {
            'waybill', 'consignee_name', 'consignee_code', 'consignee_address',
            'total_invoices', 'packages', 'weight', 'country_origin', 'freight',
            'man_reg', 'location', 'office', 'expected_entries',
        }
        email_draft = compose_email(
            **{k: v for k, v in new_params.items() if k in _COMPOSE_KEYS},
            attachment_paths=new_params.get('attachment_paths', []),
        )
        ok = do_send_email(
            subject=email_draft['subject'],
            body=email_draft['body'],
            attachments=email_draft['attachments'],
        )
        if ok:
            print(f"    RESENT: {email_draft['subject']} "
                  f"({len(email_draft['attachments'])} attachments)")
            # Refresh the history entry with the new params/hash.
            fake = argparse.Namespace(
                _source_input=entry.get('source_input', ''),
                _source_mode=entry.get('source_mode', ''),
                _output_dir=entry.get('output_dir', ''),
            )
            _record_send_history(fake, new_params, email_draft)
            resent += 1
        else:
            print(f"    RESEND FAILED: {email_draft['subject']}")
            skipped += 1

    print(f"\nResend summary: resent={resent}, unchanged={unchanged}, "
          f"skipped={skipped}, total={len(targets)}")
    report = {
        'status': 'success',
        'resent': resent,
        'unchanged': unchanged,
        'skipped': skipped,
        'checked': len(targets),
    }
    if args.json_output:
        print(f"\nREPORT:JSON:{json.dumps(report)}")
    return report


def parse_args():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description='Unified pipeline — auto-detects BL batch vs single invoice'
    )

    # Input (mutually exclusive: folder, single file, send-email-only, or resend)
    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument('--input-dir', '-d',
                             help='Directory containing PDF invoices (+ optional PO XLSX + BL PDF)')
    input_group.add_argument('--input', '-i',
                             help='Single PDF invoice file')
    input_group.add_argument('--send-email-only', action='store_true',
                             help='Send email only (no processing) — requires --waybill and --attachments')
    input_group.add_argument('--resend', metavar='WAYBILL',
                             help='Reprocess and unconditionally resend the email '
                                  'for a given waybill from data/send_history.json')
    input_group.add_argument('--resend-stale', action='store_true',
                             help='Scan send_history.json, reprocess every entry, '
                                  'and resend only those whose params would differ '
                                  '(e.g. after a supplier DB correction)')

    # Output
    parser.add_argument('--output-dir',
                        default=os.path.join(BASE_DIR, 'workspace', 'shipments'),
                        help='Directory for output XLSX files')
    parser.add_argument('--output', '-o',
                        help='Output XLSX path (single mode only)')

    # BL-specific
    parser.add_argument('--bl',
                        help='Bill of Lading number (auto-detected from PDF if not specified)')
    parser.add_argument('--po-file',
                        help='Path to PO XLSX file (auto-detected if not specified)')
    parser.add_argument('--doc-type', default='auto',
                        help='CARICOM document type (e.g., 7400-000, 4000-000). '
                             'Default: auto (resolved from consignee via config/document_types.json)')

    # Email parameters (used by all modes when --send-email or --send-email-only)
    parser.add_argument('--send-email', action='store_true',
                        help='Send email after processing')
    parser.add_argument('--waybill',
                        help='Waybill/BL number for email subject')
    parser.add_argument('--consignee', default='',
                        help='Consignee name')
    parser.add_argument('--consignee-code', default='',
                        help='Consignee code')
    parser.add_argument('--man-reg',
                        help='Manifest registration (e.g. "2026 148")')
    parser.add_argument('--location', default='WebSource',
                        help='Location of goods code')
    parser.add_argument('--office', default='',
                        help='Office code (defaults to Customs Office from Simplified Declaration)')
    parser.add_argument('--attachments',
                        help='Comma-separated attachment file paths (for --send-email-only)')
    parser.add_argument('--total-invoices', type=int, default=1,
                        help='Total number of invoices (for email)')
    parser.add_argument('--packages',
                        help='Package count (for email)')
    parser.add_argument('--weight',
                        help='Weight in kg (for email)')
    parser.add_argument('--country-origin',
                        help='Country of origin (for email)')
    parser.add_argument('--freight',
                        help='Freight amount (for email)')

    # Pipeline config
    parser.add_argument('--config', '-c',
                        help='Pipeline config YAML (single mode)')
    parser.add_argument('--json-output', action='store_true',
                        help='Emit JSON report for programmatic consumption')
    parser.add_argument('--verbose', '-v', action='store_true',
                        help='Enable verbose logging')

    return parser.parse_args()


def main():
    args = parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Send-email-only mode — no pipeline processing
    if args.send_email_only:
        run_send_email(args)
        return

    # Resend mode — reprocess prior sends from data/send_history.json
    if args.resend or args.resend_stale:
        init_resolver(BASE_DIR)
        run_resend(args)
        return

    # Initialize the resolver with project base directory
    init_resolver(BASE_DIR)

    # Capture source for send_history recording, then run pipeline.
    args._source_input = args.input or args.input_dir or ''
    mode = detect_mode(args)
    args._source_mode = mode
    logger.info(f"Detected mode: {mode}")

    if mode == 'bl':
        run_bl_mode(args)
    elif mode == 'batch':
        run_batch_mode(args)
    else:
        run_single_mode(args)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"FATAL ERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
