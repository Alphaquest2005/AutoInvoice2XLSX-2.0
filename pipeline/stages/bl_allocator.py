"""
Bill of Lading allocation stage.

Parses BL PDF → matches invoices to shipments → allocates packages/weight
→ updates XLSX files with per-invoice package counts.

Used after invoice_processor has generated initial XLSX files (Packages=1).
"""

import logging
import os
import re
import shutil
import sys
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Set

logger = logging.getLogger(__name__)

# Ensure pipeline directory is on path for imports
PIPELINE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PIPELINE_DIR not in sys.path:
    sys.path.insert(0, PIPELINE_DIR)

from bl_parser import parse_bl_pdf, match_invoice_to_bl
from bl_xlsx_generator import update_xlsx_packages


@dataclass
class BLAllocation:
    """Result of BL parsing and package allocation."""
    bl_data: Dict                    # full parsed BL data
    freight: float                   # total_usd - landing
    packages: str                    # grand total packages (string for email)
    weight: str                      # grand total weight_kg (string for email)
    insurance: float                 # insurance premium from BL
    per_invoice: List[Dict]          # [{packages, weight_kg, matched}] per invoice
    bl_output_path: Optional[str]    # path to renamed BL PDF in output dir
    allocated_packages: List[int]    # exclusive package counts per invoice


def allocate_bl_packages(
    bl_pdf_path: Optional[str],
    invoice_results: List,           # List[InvoiceResult] from invoice_processor
    output_dir: str,
    bl_number: str = '',
    supplementary_bl_paths: Optional[List[str]] = None,
) -> Optional[BLAllocation]:
    """
    Parse BL PDF, match invoices to shipments, allocate packages, update XLSX.

    When multiple BL-classified files exist (e.g. an actual BL + a Caricom
    declaration), the primary BL supplies freight/consignee/waybill while
    supplementary files supply package counts and shipment-level detail that
    the primary may lack.

    Args:
        bl_pdf_path: Path to primary Bill of Lading PDF (None if no BL found)
        invoice_results: List of InvoiceResult from invoice_processor
        output_dir: Directory where XLSX files are stored
        bl_number: BL number for file naming (auto-detected from PDF if empty)
        supplementary_bl_paths: Additional BL-classified files to merge data from

    Returns:
        BLAllocation with all allocation data, or None if no BL PDF
    """
    if not bl_pdf_path:
        print("\n[BL] No Bill of Lading PDF found — Packages remain at default (1)")
        return None

    print(f"\n[BL] Parsing Bill of Lading: {os.path.basename(bl_pdf_path)}")
    bl_data = parse_bl_pdf(bl_pdf_path)

    # Extract cost breakdown
    cost = bl_data.get('cost_breakdown', {})
    bl_freight = cost.get('total_usd', 0) - cost.get('landing', 0)
    insurance = cost.get('insurance_premium', 0)
    gt = bl_data.get('grand_total', {})
    # Use piece count when available (e.g. "165 PIECE(S)") — packages may be
    # the container count (1) rather than the actual number of pieces shipped.
    bl_piece_count = gt.get('pieces', 0) or gt.get('packages', 1)
    bl_packages = str(bl_piece_count)
    bl_weight = str(gt.get('weight_kg', 0))

    # ── Supplement from other BL-classified files ──
    # When the final BL isn't available, multiple partial documents (actual BL,
    # Caricom declaration, packing list) may each carry different data.  Merge
    # the best values: highest package count, heaviest weight, extra shipments.
    if supplementary_bl_paths and bl_piece_count <= len(invoice_results):
        for sup_path in supplementary_bl_paths:
            if not os.path.exists(sup_path):
                continue
            try:
                sup_data = parse_bl_pdf(sup_path)
                sup_gt = sup_data.get('grand_total', {})
                sup_pkgs = sup_gt.get('pieces', 0) or sup_gt.get('packages', 0)
                sup_weight = sup_gt.get('weight_kg', 0)
                if sup_pkgs > bl_piece_count:
                    logger.info(f"Supplemented packages from {os.path.basename(sup_path)}: "
                                f"{bl_piece_count} -> {sup_pkgs}")
                    bl_piece_count = sup_pkgs
                    bl_packages = str(sup_pkgs)
                if sup_weight > float(bl_weight or 0):
                    logger.info(f"Supplemented weight from {os.path.basename(sup_path)}: "
                                f"{bl_weight} -> {sup_weight}")
                    bl_weight = str(sup_weight)
                # Merge shipments for better invoice matching
                if sup_data.get('shipments'):
                    bl_data.setdefault('shipments', []).extend(sup_data['shipments'])
                    logger.info(f"Merged {len(sup_data['shipments'])} shipment(s) from "
                                f"{os.path.basename(sup_path)}")
            except Exception as e:
                logger.debug(f"Supplementary BL parse failed for {sup_path}: {e}")

    # Auto-detect BL number from parsed data if not provided
    if not bl_number:
        bl_number = bl_data.get('bl_number', 'UNKNOWN')

    print(f"    Shipments: {len(bl_data.get('shipments', []))}")
    print(f"    Total USD: ${cost.get('total_usd', 0):.2f}  "
          f"Landing: ${cost.get('landing', 0):.2f}  "
          f"Insurance: ${insurance:.2f}")
    print(f"    Email Freight: ${bl_freight:.2f}  "
          f"Packages: {bl_packages}  Weight: {bl_weight} KG")

    # ── Match each invoice to BL shipments ──
    bl_matches: List[Dict] = []
    for result in invoice_results:
        invoice_num = result.invoice_num
        file_ref = os.path.splitext(result.pdf_file)[0]
        file_ref = re.sub(r'\s*\(\d+\)$', '', file_ref)

        fallback_refs = []
        if file_ref != invoice_num:
            fallback_refs.append(file_ref)

        bl_match = match_invoice_to_bl(invoice_num, fallback_refs, bl_data)
        if bl_match['matched']:
            print(f"    {invoice_num}: pkgs={bl_match['packages']} "
                  f"weight={bl_match['weight_kg']}kg")
        bl_matches.append(bl_match)

    # ── Elimination matching: unclaimed BL shipments → unmatched invoices ──
    claimed_indices: Set[int] = set()
    for m in bl_matches:
        claimed_indices.update(m.get('matched_indices', set()))

    all_indices = set(range(len(bl_data.get('shipments', []))))
    unclaimed = all_indices - claimed_indices
    unmatched_inv = [(i, result.pdf_file) for i, (m, result) in
                     enumerate(zip(bl_matches, invoice_results))
                     if not m['matched']]

    if unmatched_inv and unclaimed:
        shipments = bl_data['shipments']
        elim_pkgs = sum(shipments[i]['packages'] for i in unclaimed)
        elim_kg = sum(shipments[i]['weight_kg'] for i in unclaimed)
        elim_refs = []
        for i in sorted(unclaimed):
            elim_refs.extend(shipments[i].get('invoice_refs', []))

        if len(unmatched_inv) == 1:
            idx, pdf = unmatched_inv[0]
            bl_matches[idx] = {
                'packages': elim_pkgs or 1,
                'weight_kg': elim_kg,
                'matched': True,
                'matched_indices': unclaimed,
            }
            print(f"    Elimination: {pdf} → pkgs={elim_pkgs} "
                  f"weight={elim_kg}kg "
                  f"({len(unclaimed)} unclaimed shipment(s): "
                  f"refs {elim_refs})")
        else:
            print(f"\n    WARNING: {len(unmatched_inv)} unmatched invoices, "
                  f"{len(unclaimed)} unclaimed BL shipments — "
                  f"cannot resolve by elimination")

    # ── Exclusive package allocation (no double-counting) ──
    # Sort by invoice value descending so the highest-value invoice claims
    # shared shipments first (e.g. if invoices A=$3500 and B=$94 both match
    # shipments 6,7,9 then A should get the packages, not B).
    assigned_shipments: Set[int] = set()
    bl_shipments = bl_data.get('shipments', [])
    allocated_packages: List[int] = [0] * len(bl_matches)

    # Build priority order: highest invoice value first
    priority_order = sorted(
        range(len(bl_matches)),
        key=lambda i: _get_invoice_total(invoice_results[i]),
        reverse=True,
    )

    for i in priority_order:
        bl_match = bl_matches[i]
        matched_indices = bl_match.get('matched_indices', set())
        pkgs = 0
        for idx in sorted(matched_indices):
            if idx not in assigned_shipments:
                assigned_shipments.add(idx)
                pkgs += bl_shipments[idx]['packages']
        if not matched_indices and not bl_match.get('matched'):
            pkgs = 1
        allocated_packages[i] = pkgs

    # ── Enforce minimum 1 package per entry ──
    bl_grand_pkgs = bl_piece_count or 1
    zero_pkg_indices = [i for i, p in enumerate(allocated_packages) if p == 0]

    if zero_pkg_indices:
        new_results, allocated_packages = _enforce_min_packages(
            invoice_results, allocated_packages, bl_grand_pkgs, output_dir)
        # Update the original list in-place so callers see the changes
        if len(new_results) != len(invoice_results):
            invoice_results.clear()
            invoice_results.extend(new_results)

    # ── Redistribute packages proportionally by value ──
    allocated_packages = _redistribute_packages(
        invoice_results, allocated_packages, bl_grand_pkgs)

    # ── Distribute BL freight proportionally by invoice value ──
    # Use invoice_total first; fall back to item cost sum if total is 0 (e.g. Walmart)
    inv_totals = [_get_invoice_total(r) for r in invoice_results]
    if sum(inv_totals) == 0:
        inv_totals = [
            sum((m.get('total_cost', 0) or 0) for m in r.matched_items)
            for r in invoice_results
        ]
    grand_inv_total = sum(inv_totals) or 1  # avoid div by zero
    allocated_freight = [round(bl_freight * (t / grand_inv_total), 2) for t in inv_totals]
    allocated_insurance = [round(insurance * (t / grand_inv_total), 2) for t in inv_totals]

    # ── Update existing XLSX files with BL package counts + freight ──
    print(f"\n[BL] Updating XLSX packages from BL allocation:\n")
    for i, result in enumerate(invoice_results):
        pkgs = allocated_packages[i]
        if not os.path.exists(result.xlsx_path):
            print(f"    {result.invoice_num}.xlsx  SKIPPED (file not found, will use combined XLSX)")
            result.packages = pkgs
            continue
        update_xlsx_packages(result.xlsx_path, pkgs,
                             freight=allocated_freight[i],
                             insurance=allocated_insurance[i])
        result.packages = pkgs
        print(f"    {result.invoice_num}.xlsx  Packages: 1 -> {pkgs}")

    pkg_total = sum(allocated_packages)
    print(f"\n    Package allocation: {pkg_total} "
          f"(BL grand total: {bl_grand_pkgs})")

    # ── Copy BL PDF to output with standardized name ──
    bl_output_path = None
    if bl_number and bl_number != 'UNKNOWN':
        bl_output_name = f"{bl_number}-BL.pdf"
        bl_output_path = os.path.join(output_dir, bl_output_name)
        if os.path.abspath(bl_pdf_path) != os.path.abspath(bl_output_path):
            shutil.copy2(bl_pdf_path, bl_output_path)
            print(f"\n    BL PDF copied: {bl_output_name}")
        else:
            print(f"\n    BL PDF already in place: {bl_output_name}")

    return BLAllocation(
        bl_data=bl_data,
        freight=bl_freight,
        packages=bl_packages,
        weight=bl_weight,
        insurance=insurance,
        per_invoice=bl_matches,
        bl_output_path=bl_output_path,
        allocated_packages=allocated_packages,
    )


def _get_invoice_total(result) -> float:
    """Get invoice total value from an InvoiceResult."""
    if hasattr(result, 'invoice_data') and result.invoice_data:
        return float(result.invoice_data.get('invoice_total', 0) or 0)
    return 0.0


def _enforce_min_packages(
    invoice_results: List,
    allocated_packages: List[int],
    bl_total_packages: int,
    output_dir: str,
) -> tuple:
    """
    Ensure every entry has at least 1 package.

    Strategy:
      1. If BL has enough packages (>= number of entries), donate 1 from the
         largest-package invoice to each zero-package invoice. No combining.
      2. Only combine invoices if BL total packages < number of entries
         (impossible to give each entry 1 package otherwise).

    Returns:
        (updated_invoice_results, updated_allocated_packages)
    """
    n_entries = len(invoice_results)
    zero_count = sum(1 for p in allocated_packages if p == 0)

    if zero_count == 0:
        return invoice_results, allocated_packages

    print(f"\n[BL] Package enforcement: {zero_count} entries have 0 packages "
          f"(BL total: {bl_total_packages}, entries: {n_entries})")

    # ── Case 1: Enough packages — redistribute without combining ──
    if bl_total_packages >= n_entries:
        print(f"    Sufficient packages ({bl_total_packages} >= {n_entries} entries)"
              f" — donating from largest donors")
        for i in range(len(allocated_packages)):
            if allocated_packages[i] == 0:
                # Find the invoice with the most packages to donate from
                donor_idx = max(range(len(allocated_packages)),
                                key=lambda j: allocated_packages[j])
                if allocated_packages[donor_idx] > 1:
                    allocated_packages[donor_idx] -= 1
                    allocated_packages[i] = 1
                    print(f"    {invoice_results[i].invoice_num}: "
                          f"0 → 1 (donated from "
                          f"{invoice_results[donor_idx].invoice_num}, "
                          f"now {allocated_packages[donor_idx]})")
                else:
                    # All donors have 1 — just force 1 (total will be checked
                    # by _redistribute_packages later)
                    allocated_packages[i] = 1
                    print(f"    {invoice_results[i].invoice_num}: "
                          f"0 → 1 (forced minimum)")
        return invoice_results, allocated_packages

    # ── Case 2: Not enough packages — must combine to reduce entries ──
    print(f"    Insufficient packages ({bl_total_packages} < {n_entries} entries)"
          f" — combining invoices")

    # Group invoices by supplier name
    supplier_groups: Dict[str, List[int]] = {}
    for i, result in enumerate(invoice_results):
        supplier = result.supplier_info.get('name', '') if hasattr(result, 'supplier_info') else ''
        supplier_groups.setdefault(supplier, []).append(i)

    merge_plan: List[List[int]] = []
    merged_indices: Set[int] = set()

    # Phase 1: Combine same-supplier invoices (prioritise groups with 0-pkg members)
    for supplier, indices in supplier_groups.items():
        if len(indices) < 2:
            continue
        has_zero = any(allocated_packages[i] == 0 for i in indices)
        if has_zero:
            merge_plan.append(sorted(indices))
            merged_indices.update(indices)
            print(f"    Combining {len(indices)} {supplier} invoices: "
                  f"{[invoice_results[i].invoice_num for i in indices]}")

    # Phase 2: If still too many entries, combine smallest values
    remaining_entries = (n_entries - sum(len(g) - 1 for g in merge_plan))
    while remaining_entries > bl_total_packages:
        unmerged = [i for i in range(n_entries) if i not in merged_indices]
        if len(unmerged) < 2:
            break
        unmerged.sort(key=lambda i: _get_invoice_total(invoice_results[i]))
        pair = [unmerged[0], unmerged[1]]
        merge_plan.append(pair)
        merged_indices.update(pair)
        remaining_entries -= 1
        print(f"    Combining smallest-value entries: "
              f"{[invoice_results[i].invoice_num for i in pair]}")

    if not merge_plan:
        for i in range(len(allocated_packages)):
            if allocated_packages[i] == 0:
                allocated_packages[i] = 1
                print(f"    Forced {invoice_results[i].invoice_num} to 1 package (minimum)")
        return invoice_results, allocated_packages

    # Execute merges
    try:
        import openpyxl
        from copy import copy as copy_style
    except ImportError:
        logger.warning("openpyxl not available — forcing 1 package minimum")
        for i in range(len(allocated_packages)):
            if allocated_packages[i] == 0:
                allocated_packages[i] = 1
        return invoice_results, allocated_packages

    indices_to_remove: Set[int] = set()
    for group in merge_plan:
        if len(group) < 2:
            continue

        primary_idx = group[0]
        primary = invoice_results[primary_idx]
        invoice_nums = [invoice_results[i].invoice_num for i in group]

        combined_name = f"{invoice_nums[0]}-combined.xlsx"
        combined_path = os.path.join(output_dir, combined_name)

        try:
            _combine_xlsx_with_formulas(
                [invoice_results[i].xlsx_path for i in group],
                combined_path)
        except Exception as e:
            logger.warning(f"Failed to combine {invoice_nums}: {e}")
            continue

        print(f"    Combined XLSX: {combined_name}")

        combined_pkgs = sum(allocated_packages[i] for i in group)
        combined_freight = sum(invoice_results[i].freight for i in group)
        combined_total = sum(_get_invoice_total(invoice_results[i]) for i in group)
        all_pdf_paths = [invoice_results[i].pdf_output_path for i in group]

        primary.xlsx_path = combined_path
        primary.invoice_num = invoice_nums[0]
        primary.freight = combined_freight
        primary._combined_pdf_paths = all_pdf_paths
        if hasattr(primary, 'invoice_data') and primary.invoice_data:
            primary.invoice_data['invoice_total'] = combined_total
        allocated_packages[primary_idx] = combined_pkgs

        for i in group[1:]:
            indices_to_remove.add(i)

    if indices_to_remove:
        new_results = []
        new_packages = []
        for i in range(n_entries):
            if i not in indices_to_remove:
                new_results.append(invoice_results[i])
                new_packages.append(allocated_packages[i])
        invoice_results = new_results
        allocated_packages = new_packages
        print(f"    After combining: {len(invoice_results)} entries "
              f"(was {n_entries})")

    return invoice_results, allocated_packages


def _redistribute_packages(
    invoice_results: List,
    allocated_packages: List[int],
    bl_total_packages: int,
) -> List[int]:
    """
    Ensure allocated packages sum to BL total.

    If total already matches, keep allocations as-is.
    If not, use BL-matched allocations as the base and adjust:
      - Deficit (sum < BL total): add extras to BL-matched invoices,
        proportionally by their existing allocation.
      - Surplus (sum > BL total): shouldn't happen (exclusive assignment),
        but handled by trimming from largest.

    Every entry keeps at least 1 package.
    """
    n = len(invoice_results)
    if n == 0:
        return allocated_packages

    current_total = sum(allocated_packages)

    # Already balanced — nothing to do
    if current_total == bl_total_packages:
        print(f"\n[BL] Package allocation balanced: {current_total} = BL total")
        return allocated_packages

    # Safety: if BL has fewer packages than entries, each gets 1
    if bl_total_packages <= n:
        result = [1] * n
        print(f"\n[BL] Package redistribution: BL total ({bl_total_packages}) "
              f"<= entries ({n}), each gets 1")
        return result

    result = list(allocated_packages)  # copy

    if current_total < bl_total_packages:
        # Deficit — distribute extras to invoices that already have packages,
        # proportionally by their current allocation
        extras = bl_total_packages - current_total
        # Sort by current allocation descending (largest gets extras first)
        candidates = [(i, result[i]) for i in range(n) if result[i] > 1]
        if not candidates:
            # All have 1 — distribute to highest-value invoices
            candidates = [(i, result[i]) for i in range(n)]

        total_weight = sum(c[1] for c in candidates) or 1
        proportions = [(c[0], (c[1] / total_weight) * extras) for c in candidates]
        floors = [(idx, int(p)) for idx, p in proportions]

        for idx, f in floors:
            result[idx] += f

        remainder = extras - sum(f for _, f in floors)
        # Distribute remainder to candidates with largest fractional parts
        fractionals = [(idx, (c[1] / total_weight) * extras - int((c[1] / total_weight) * extras))
                       for idx, c in zip([c[0] for c in candidates], candidates)]
        fractionals.sort(key=lambda x: x[1], reverse=True)
        for j in range(remainder):
            if j < len(fractionals):
                result[fractionals[j][0]] += 1

    elif current_total > bl_total_packages:
        # Surplus — trim from largest allocations
        surplus = current_total - bl_total_packages
        while surplus > 0:
            max_idx = max(range(n), key=lambda i: result[i])
            if result[max_idx] <= 1:
                break  # can't trim below 1
            trim = min(result[max_idx] - 1, surplus)
            result[max_idx] -= trim
            surplus -= trim

    # Log changes
    changed = any(result[i] != allocated_packages[i] for i in range(n))
    if changed:
        print(f"\n[BL] Package redistribution ({current_total} → {bl_total_packages}):")
        for i, r in enumerate(invoice_results):
            old = allocated_packages[i]
            new = result[i]
            if old != new:
                print(f"    {r.invoice_num:<20} {old} → {new} packages")
            else:
                print(f"    {r.invoice_num:<20} {new} packages")

    return result


def _combine_xlsx_with_formulas(file_paths: List[str], output_path: str) -> None:
    """
    Combine multiple invoice XLSX files into one entry with proper formulas.

    Each invoice keeps its own verification section:
      - Data rows (with metadata in first data row: InvoiceTotal, Freight, etc.)
      - Subtotal formula (sum of that invoice's P cells)
      - Adjustments formula (T+U+V-W from that invoice's first data row)
      - Net Total formula (Subtotal + Adjustments)
      - Variance Check formula (InvoiceTotal - Net Total)
    Invoices are separated by a single blank row.
    """
    import openpyxl
    from copy import copy as copy_style

    COL_P = 16       # TotalCost
    COL_S = 19       # InvoiceTotal
    COL_T = 20       # Freight
    COL_U = 21       # Insurance
    COL_V = 22       # Tax/OtherCost
    COL_W = 23       # Deduction

    FORMULA_LABELS = {'Subtotal', 'Adjustments', 'Net Total', 'Variance Check',
                      'Subtotal Grouped', 'Subtotal Details', 'Group Verification',
                      'Grand Subtotal', 'Grand Adjustments', 'Grand Net Total',
                      'Grand Invoice Total', 'Grand Variance'}

    # Load all source workbooks (with formulas, not data_only)
    sources = []
    for fp in file_paths:
        wb = openpyxl.load_workbook(fp)
        sources.append(wb)

    first_ws = sources[0].active
    max_col = first_ws.max_column

    # Create output workbook
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active

    # Copy header row from first file
    for c in range(1, max_col + 1):
        src = first_ws.cell(1, c)
        dst = out_ws.cell(1, c)
        dst.value = src.value
        if src.has_style:
            dst.font = copy_style(src.font)
            dst.border = copy_style(src.border)
            dst.fill = copy_style(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy_style(src.alignment)

    # Copy column widths from first file
    for c in range(1, max_col + 1):
        letter = openpyxl.utils.get_column_letter(c)
        if first_ws.column_dimensions[letter].width:
            out_ws.column_dimensions[letter].width = first_ws.column_dimensions[letter].width

    out_row = 2  # start writing data at row 2
    bold_font = openpyxl.styles.Font(bold=True)
    currency_fmt = '#,##0.00'

    # Track per-invoice rows for grand total formulas
    invoice_sections = []  # [{first_data_row, sub_row, net_row}]

    for file_idx, wb in enumerate(sources):
        ws = wb.active

        # Add blank separator row between invoices (not before the first)
        if file_idx > 0:
            out_row += 1  # leave one blank row

        first_data_row = out_row  # track where this invoice's data starts

        # Copy data rows (skip header=1, skip formula rows)
        for r in range(2, ws.max_row + 1):
            tc = ws.cell(r, COL_P).value
            # Stop at formula rows (Subtotal, Net Total, etc.)
            if isinstance(tc, str) and tc.startswith('='):
                break
            # Also stop at label rows
            label = ws.cell(r, 12).value
            if isinstance(label, str) and label.strip() in FORMULA_LABELS:
                break

            for c in range(1, max_col + 1):
                src = ws.cell(r, c)
                dst = out_ws.cell(out_row, c)
                dst.value = src.value
                if src.has_style:
                    dst.font = copy_style(src.font)
                    dst.border = copy_style(src.border)
                    dst.fill = copy_style(src.fill)
                    dst.number_format = src.number_format
                    dst.alignment = copy_style(src.alignment)

            # Clear packages (col X=24) for non-first invoices — the combined
            # entry's packages are set once on row 2 by update_xlsx_packages()
            if file_idx > 0:
                out_ws.cell(out_row, 24).value = None

            out_row += 1

        last_data_row = out_row - 1

        if last_data_row < first_data_row:
            continue  # empty invoice, skip

        # Write per-invoice formula rows
        # Subtotal: sum of this invoice's P cells
        p_refs = '+'.join(f'P{r}' for r in range(first_data_row, last_data_row + 1))

        sub_row = out_row
        out_ws.cell(sub_row, 12).value = 'Subtotal'
        out_ws.cell(sub_row, 12).font = bold_font
        out_ws.cell(sub_row, COL_P).value = f'={p_refs}'
        out_ws.cell(sub_row, COL_P).number_format = currency_fmt

        # Adjustments: T+U+V-W from this invoice's first data row
        adj_row = sub_row + 1
        out_ws.cell(adj_row, 12).value = 'Adjustments'
        out_ws.cell(adj_row, 12).font = bold_font
        out_ws.cell(adj_row, COL_P).value = (
            f'=(T{first_data_row}+U{first_data_row}'
            f'+V{first_data_row}-W{first_data_row})')
        out_ws.cell(adj_row, COL_P).number_format = currency_fmt

        # Net Total: Subtotal + Adjustments
        net_row = adj_row + 1
        out_ws.cell(net_row, 12).value = 'Net Total'
        out_ws.cell(net_row, 12).font = bold_font
        out_ws.cell(net_row, COL_P).value = f'=(P{sub_row}+P{adj_row})'
        out_ws.cell(net_row, COL_P).number_format = currency_fmt

        # Variance Check: InvoiceTotal(S) - Net Total
        var_row = net_row + 1
        out_ws.cell(var_row, 12).value = 'Variance Check'
        out_ws.cell(var_row, 12).font = bold_font
        out_ws.cell(var_row, COL_P).value = f'=(S{first_data_row}-P{net_row})'
        out_ws.cell(var_row, COL_P).number_format = currency_fmt

        invoice_sections.append({
            'first_data_row': first_data_row,
            'sub_row': sub_row,
            'net_row': net_row,
        })

        out_row = var_row + 1

    # ── Grand Total section (only for multi-invoice combined files) ──
    if len(invoice_sections) > 1:
        out_row += 1  # blank separator

        # Grand Subtotal: sum of all per-invoice Subtotal rows
        grand_sub_refs = '+'.join(f'P{s["sub_row"]}' for s in invoice_sections)
        grand_sub_row = out_row
        out_ws.cell(grand_sub_row, 12).value = 'Grand Subtotal'
        out_ws.cell(grand_sub_row, 12).font = bold_font
        out_ws.cell(grand_sub_row, COL_P).value = f'={grand_sub_refs}'
        out_ws.cell(grand_sub_row, COL_P).number_format = currency_fmt

        # Grand Adjustments: sum of all invoices' T+U+V-W
        adj_parts = []
        for s in invoice_sections:
            r = s['first_data_row']
            adj_parts.append(f'(T{r}+U{r}+V{r}-W{r})')
        grand_adj_row = grand_sub_row + 1
        out_ws.cell(grand_adj_row, 12).value = 'Grand Adjustments'
        out_ws.cell(grand_adj_row, 12).font = bold_font
        out_ws.cell(grand_adj_row, COL_P).value = f'={"+".join(adj_parts)}'
        out_ws.cell(grand_adj_row, COL_P).number_format = currency_fmt

        # Grand Net Total
        grand_net_row = grand_adj_row + 1
        out_ws.cell(grand_net_row, 12).value = 'Grand Net Total'
        out_ws.cell(grand_net_row, 12).font = bold_font
        out_ws.cell(grand_net_row, COL_P).value = f'=(P{grand_sub_row}+P{grand_adj_row})'
        out_ws.cell(grand_net_row, COL_P).number_format = currency_fmt

        # Grand Invoice Total: sum of all S values
        s_refs = '+'.join(f'S{s["first_data_row"]}' for s in invoice_sections)
        grand_inv_row = grand_net_row + 1
        out_ws.cell(grand_inv_row, 12).value = 'Grand Invoice Total'
        out_ws.cell(grand_inv_row, 12).font = bold_font
        out_ws.cell(grand_inv_row, COL_P).value = f'={s_refs}'
        out_ws.cell(grand_inv_row, COL_P).number_format = currency_fmt

        # Grand Variance Check: Grand Invoice Total - Grand Net Total
        grand_var_row = grand_inv_row + 1
        out_ws.cell(grand_var_row, 12).value = 'Grand Variance'
        out_ws.cell(grand_var_row, 12).font = bold_font
        out_ws.cell(grand_var_row, COL_P).value = f'=(P{grand_inv_row}-P{grand_net_row})'
        out_ws.cell(grand_var_row, COL_P).number_format = currency_fmt

    # Close source workbooks
    for wb in sources:
        wb.close()

    out_wb.save(output_path)
