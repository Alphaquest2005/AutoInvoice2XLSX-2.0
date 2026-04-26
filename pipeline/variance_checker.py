#!/usr/bin/env python3
"""
Stage 7: Variance Checker
Validates formulas and checks variance equals $0.00.
"""

import os
from typing import Dict

try:
    import openpyxl
except ImportError:
    openpyxl = None

from pipeline.config_loader import (
    load_library_enums,
    load_validation_tolerances,
    load_xlsx_labels,
)

# ── SSOT-loaded module constants ────────────────────────────────────────
_LABELS = load_xlsx_labels()
_LABEL_VARIANCE_CHECK = _LABELS["totals"]["VARIANCE_CHECK"]
_LABEL_GROUP_VERIFICATION = _LABELS["totals"]["GROUP_VERIFICATION"]

_FORMULA_ERROR_TOKENS = tuple(load_library_enums()["excel"]["formula_errors"])

_VC = load_validation_tolerances()["variance_checker"]
_COL_VARIANCE_VALUE = _VC["variance_value_col"]
_COL_PER_ROW_VARIANCE = _VC["per_row_variance_col"]
_COL_LABEL_SCAN_MAX = _VC["label_scan_max_col"]
_COL_FORMULA_SCAN_MAX = _VC["formula_error_scan_max_col"]
_FORMULA_ERR_MAX_REPORT = _VC["formula_error_max_reported"]
_PER_ROW_VAR_MAX_REPORT = _VC["non_zero_variance_max_reported"]
_VARIANCE_EPSILON = _VC["variance_epsilon_usd"]
_PER_ROW_FIRST_DATA_ROW = _VC["per_row_variance_first_data_row"]

# Error-payload keys for run()'s return dict. These match the upstream
# pipeline consumer's contract and are therefore magic-ok.
_STATUS_SUCCESS = "success"                          # magic-ok: result dict contract
_STATUS_ERROR = "error"                              # magic-ok: result dict contract
_STATUS_VALIDATION_ERROR = "validation_error"        # magic-ok: result dict contract
_ERR_NO_OPENPYXL = "openpyxl not installed"          # magic-ok: error payload text
_ERR_FILE_NOT_FOUND_TPL = "File not found: {path}"   # magic-ok: error payload text


def run(input_path: str, output_path: str = None, config: Dict = None, context: Dict = None) -> Dict:
    """Validate Excel formulas and variance checks."""
    if not openpyxl:
        return {"status": _STATUS_ERROR, "error": _ERR_NO_OPENPYXL}

    if not input_path or not os.path.exists(input_path):
        return {
            "status": _STATUS_ERROR,
            "error": _ERR_FILE_NOT_FOUND_TPL.format(path=input_path),
        }

    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active

    result = {
        "status": _STATUS_SUCCESS,
        "valid": True,
        "errors": [],
        "warnings": [],
        "checks": {},
    }

    # Check VARIANCE CHECK total.
    variance = find_labeled_value(ws, _LABEL_VARIANCE_CHECK, _COL_VARIANCE_VALUE)
    result["checks"]["variance_check"] = variance
    if variance is not None and abs(float(variance)) > _VARIANCE_EPSILON:
        result["valid"] = False
        result["errors"].append(
            f"{_LABEL_VARIANCE_CHECK} = ${variance}, expected $0.00"
        )

    # Check GROUP VERIFICATION total.
    group_var = find_labeled_value(
        ws, _LABEL_GROUP_VERIFICATION, _COL_VARIANCE_VALUE
    )
    result["checks"]["group_verification"] = group_var
    if group_var is not None and abs(float(group_var)) > _VARIANCE_EPSILON:
        result["valid"] = False
        result["errors"].append(
            f"{_LABEL_GROUP_VERIFICATION} = ${group_var}, expected $0.00"
        )

    # Scan for formula errors.
    formula_errors = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, min(ws.max_column + 1, _COL_FORMULA_SCAN_MAX)):
            cell = ws.cell(row=row, column=col)
            if cell.value and str(cell.value) in _FORMULA_ERROR_TOKENS:
                col_letter = openpyxl.utils.get_column_letter(col)
                formula_errors.append(f"{cell.value} at {col_letter}{row}")

    result["checks"]["formula_errors"] = len(formula_errors)
    if formula_errors:
        result["valid"] = False
        result["errors"].extend(formula_errors[:_FORMULA_ERR_MAX_REPORT])

    # Check per-row Variance column for non-zero values.
    non_zero_variance = []
    for row in range(_PER_ROW_FIRST_DATA_ROW, ws.max_row + 1):
        cell = ws.cell(row=row, column=_COL_PER_ROW_VARIANCE)
        if cell.value is not None:
            try:
                val = float(cell.value)
                if abs(val) > _VARIANCE_EPSILON:
                    non_zero_variance.append(f"Row {row}: ${val}")
            except (ValueError, TypeError):
                pass

    if non_zero_variance:
        result["warnings"].extend(
            f"Non-zero variance: {v}"
            for v in non_zero_variance[:_PER_ROW_VAR_MAX_REPORT]
        )

    if not result["valid"]:
        result["status"] = _STATUS_VALIDATION_ERROR

    return result


def find_labeled_value(ws, label: str, value_col: int):
    """Find a labeled row and return value from specified column."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, _COL_LABEL_SCAN_MAX):
            cell = ws.cell(row=row, column=col)
            if cell.value and label.upper() in str(cell.value).upper():
                val = ws.cell(row=row, column=value_col).value
                if val is not None:
                    try:
                        return float(val)
                    except (ValueError, TypeError):
                        return None
    return None
