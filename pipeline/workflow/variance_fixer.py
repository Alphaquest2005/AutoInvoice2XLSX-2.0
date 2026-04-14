"""
LLM-based variance fixing for XLSX files.

Targeted fix: only modifies cell values in existing XLSX.
Does NOT re-run OCR, parsing, or classification.
"""

import logging
import os
from typing import Dict, Optional

logger = logging.getLogger(__name__)

# Prompts are constants - no variance from prompt construction
SYSTEM_PROMPT = """You are an invoice processing assistant. Your task is to fix variance issues in XLSX files.

IMPORTANT: The XLSX has two types of data rows:
1. GROUP rows (blue background) like "BAGS & LUGGAGE (1 items)" - these are GROUP TOTALS, do NOT include in sum
2. DETAIL rows (no background) - these ARE the individual items

The variance = Invoice Total - Net Total (where Net Total = Subtotal + Adjustments)
- Negative variance: items sum to LESS than invoice total (missing items or prices too low)
- Positive variance: items sum to MORE than invoice total (duplicate items or prices too high)

To fix variance, you can ONLY:
1. Adjust NUMERIC values (prices, quantities) on existing rows
2. If you cannot identify the exact fix, say so — do NOT add new rows

CRITICAL CONSTRAINTS:
- You MUST NOT add new rows (add_items). The XLSX structure is fixed.
- You MUST NOT change descriptions or text — only numbers.
- Calculate current sum of detail items
- Calculate what the sum SHOULD be (= Invoice Total - Adjustments)
- Propose changes that fix the variance

Respond with JSON only:
{
  "analysis": "Brief explanation of the variance cause and your fix strategy",
  "current_sum": <number>,
  "target_sum": <number>,
  "fixes": [
    {"row": N, "column": "total_cost", "new_value": Y, "reason": "..."},
    ...
  ],
  "expected_new_variance": 0.00
}

If you cannot determine the exact fix, return an empty fixes array and explain why in the analysis."""


def fix_variance(
    xlsx_path: str,
    invoice_text: str,
    current_variance: float,
    honest_mode: bool = True,
) -> Dict:
    """
    Use LLM to analyze and fix variance in the XLSX file.

    This is a TARGETED fix - it only modifies cell values in the existing XLSX.
    It does NOT re-run OCR, parsing, classification, or email composition.

    Args:
        xlsx_path: Path to the XLSX file to fix
        invoice_text: Pre-extracted invoice text (not re-extracted)
        current_variance: Current variance amount
        honest_mode: When True (default) skip LLM cell-scaling entirely and
            absorb the residual in the ADJUSTMENTS row via force_adjustment.
            This is the "honest" path: no hallucinated per-item numbers, the
            variance is recorded visibly instead of being hidden inside
            scaled item totals.  Pass False only when a spec explicitly
            opts in to LLM scaling (spec.allow_llm_fix=true).

    Returns:
        dict with success, new_variance, fixes_applied, analysis, error
    """
    from core.config import get_config

    cfg = get_config()

    # Honest mode: skip LLM entirely — absorb residual via ADJUSTMENTS.
    # This keeps per-item P/O cells untouched (true to source) and records
    # any gap visibly in the ADJUSTMENTS row where a reviewer can see it.
    if honest_mode:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            logger.info(
                f"variance_fixer (honest_mode): absorbing ${current_variance:.2f} "
                f"into ADJUSTMENTS row — no per-item scaling"
            )
            _force_adjustment(ws, current_variance, cfg)
            _update_variance_row(ws, 0.00, cfg)
            wb.save(xlsx_path)
            return {
                'success': True,
                'fixes_applied': 0,
                'new_variance': 0.00,
                'analysis': (
                    f"Honest mode: ${current_variance:.2f} absorbed into "
                    f"ADJUSTMENTS row (visible, traceable)."
                ),
            }
        except Exception as e:
            logger.error(f"Honest variance fix failed: {e}", exc_info=True)
            return {'success': False, 'error': str(e)}

    from core.llm_client import get_llm_client

    llm = get_llm_client()

    try:
        import openpyxl

        # Load with data_only=True for reading computed values (formulas → numbers)
        wb_data = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws_data = wb_data.active

        # Build XLSX summary from computed values
        xlsx_summary = _build_xlsx_summary(ws_data, cfg)
        invoice_total = xlsx_summary['invoice_total']
        total_cost_sum = xlsx_summary['total_cost_sum']

        # Read raw adjustment components (freight/ins/tax/deduction) from row 2
        # so we can tell the LLM the correct *items* target:
        #     target_items_sum = invoice_total - adjustments
        # Without this the LLM was told "Sum = Invoice Total", which caused it
        # to target e.g. $67.80 instead of $63.37 (dropping a $4.43 tax into
        # the items sum and effectively losing the tax).
        def _num_cell(v):
            try:
                return float(v) if v is not None else 0.0
            except (TypeError, ValueError):
                return 0.0
        _freight = _num_cell(ws_data.cell(row=2, column=20).value)   # T
        _insurance = _num_cell(ws_data.cell(row=2, column=21).value) # U
        _tax = _num_cell(ws_data.cell(row=2, column=22).value)       # V
        _deduction = _num_cell(ws_data.cell(row=2, column=23).value) # W
        adjustments = _freight + _insurance + _tax - _deduction
        try:
            inv_total_num = float(invoice_total) if invoice_total else 0.0
        except (TypeError, ValueError):
            inv_total_num = 0.0
        target_items_sum = round(inv_total_num - adjustments, 2)

        wb_data.close()

        # Load again without data_only for writing (preserves formulas)
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        # Truncate invoice text
        if len(invoice_text) > 8000:
            invoice_text = invoice_text[:8000] + "\n... [truncated]"

        user_message = f"""{xlsx_summary['text']}

Invoice Total (S2): ${invoice_total}
Adjustments (Freight ${_freight:.2f} + Insurance ${_insurance:.2f} + Tax ${_tax:.2f} - Deduction ${_deduction:.2f}) = ${adjustments:.2f}
Current Sum of Individual Items: ${total_cost_sum:.2f}
Current Variance: ${current_variance:.2f}
Target: Sum of detail items MUST equal ${target_items_sum:.2f}
  (derivation: Invoice Total ${inv_total_num:.2f} − Adjustments ${adjustments:.2f} = ${target_items_sum:.2f})

Invoice PDF text (for reference):
{invoice_text}

Analyze the data and provide fixes to bring the items sum to ${target_items_sum:.2f}. Remember:
- Only modify NUMERIC values on DETAIL ITEM rows (not group rows or summary rows)
- Do NOT touch the Adjustments / Freight / Tax values — those are separate
- Do NOT propose add_items — the XLSX structure is fixed
- If the fix is unclear, return empty fixes array"""

        # Use cache key based on xlsx content + variance (deterministic).
        # Include target_items_sum so cache busts when the prompt's target
        # changes (the LLM used to be told the wrong target; a stale cache
        # entry would otherwise keep returning fixes aimed at the old target).
        cache_extra = (
            f"variance:{current_variance:.2f}:total:{total_cost_sum:.2f}"
            f":target:{target_items_sum:.2f}:v2"
        )

        fixes_data = llm.call_json(
            user_message=user_message,
            system_prompt=SYSTEM_PROMPT,
            cache_key_extra=cache_extra,
        )

        if not fixes_data:
            return {'success': False, 'error': 'LLM returned no valid JSON response'}

        logger.info(f"LLM analysis: {fixes_data.get('analysis', 'No analysis')[:80]}...")

        # Apply fixes to XLSX
        fixes_applied = _apply_fixes(ws, fixes_data, cfg)

        # Propagate detail-row changes up to their group headers so grouped
        # subtotals (=P{group_header_row}) stay consistent with the new detail
        # sums. Without this, SUBTOTAL (GROUPED) still references the pre-fix
        # header value and NET TOTAL in Excel remains wrong even though our
        # in-memory variance calculation is right.
        if fixes_applied:
            updated = _update_group_header_totals(ws, cfg)
            if updated:
                logger.info(f"Updated {updated} group header total(s) to match fixed details")

        # Recalculate and update variance
        new_variance = _recalculate_variance(ws, invoice_total, cfg)

        # Force adjustment if LLM fix was incomplete. We use a tight 0.02
        # tolerance here (not cfg.variance_threshold which is 0.50) because
        # the downstream combined-block check also uses 0.02 — anything above
        # that bubbles up as an unfixed variance blocker. Small residuals
        # (e.g. 16¢ rounding errors when the LLM patches one row) should be
        # absorbed by the ADJUSTMENTS formula so the block is fully balanced.
        if abs(new_variance) >= 0.02:
            logger.info(f"LLM fix incomplete - forcing adjustment for ${new_variance:.2f}")
            _force_adjustment(ws, new_variance, cfg)
            new_variance = 0.00
            _update_variance_row(ws, 0.00, cfg)

        wb.save(xlsx_path)

        return {
            'success': True,
            'fixes_applied': fixes_applied,
            'new_variance': round(new_variance, 2),
            'analysis': fixes_data.get('analysis', ''),
        }

    except Exception as e:
        logger.error(f"Variance fix failed: {e}", exc_info=True)
        return {'success': False, 'error': str(e)}


def _get_desc_column(ws) -> int:
    """Detect the description column from the XLSX header row.
    bl_xlsx_generator uses column J (10) = 'Supplier Item Description'.
    Falls back to checking column L (12) for legacy xlsx_generator format.
    """
    for col in (10, 12):
        val = ws.cell(row=1, column=col).value
        if val and 'desc' in str(val).lower():
            return col
    return 10  # Default to J


def _build_xlsx_summary(ws, cfg) -> Dict:
    """Build text summary of XLSX content for the LLM prompt."""
    lines = ["Current XLSX content:"]

    # Detect description column from headers (not hardcoded)
    desc_col = _get_desc_column(ws)

    # Headers
    headers = []
    for col in range(1, min(ws.max_column + 1, 20)):
        val = ws.cell(row=1, column=col).value
        if val:
            headers.append(str(val))
    lines.append(f"Columns: {', '.join(headers)}")

    # Detect group rows by fill color (D9E1F2 = grouped format blue)
    def _is_group_row(row):
        fill = ws.cell(row=row, column=1).fill
        return (fill and fill.start_color and
                'D9E1F2' in str(getattr(fill.start_color, 'rgb', '') or ''))

    # Parse rows
    items = []
    group_rows = []
    summary_rows = []
    total_cost_sum = 0

    for row in range(2, ws.max_row + 1):
        desc = ws.cell(row=row, column=desc_col).value
        qty = ws.cell(row=row, column=cfg.col_quantity).value
        unit_cost = ws.cell(row=row, column=cfg.col_unit_cost).value
        total_cost = ws.cell(row=row, column=cfg.col_total_cost).value

        if not desc:
            continue

        desc_str = str(desc)
        desc_upper = desc_str.upper()

        if any(x in desc_upper for x in ['SUBTOTAL', 'VARIANCE', 'NET TOTAL', 'GROUP VERIFICATION',
                                          'ADJUSTMENTS', 'INVOICE TOTAL', 'TOTAL INTERNAL',
                                          'TOTAL INSURANCE', 'TOTAL OTHER', 'TOTAL DEDUCTION']):
            summary_rows.append({'row': row, 'description': desc_str[:60], 'total_cost': total_cost})
        elif _is_group_row(row):
            group_rows.append({'row': row, 'description': desc_str[:60], 'total_cost': total_cost})
        elif '(' in desc_str and ')' in desc_str and 'items' in desc_str.lower():
            group_rows.append({'row': row, 'description': desc_str[:60], 'total_cost': total_cost})
        else:
            items.append({'row': row, 'description': desc_str[:60], 'qty': qty, 'unit_cost': unit_cost, 'total_cost': total_cost})
            if total_cost and isinstance(total_cost, (int, float)):
                total_cost_sum += float(total_cost)

    lines.append(f"\nGroup Rows ({len(group_rows)}):")
    for gr in group_rows[:10]:
        lines.append(f"  Row {gr['row']}: {gr['description'][:40]} | Total: ${gr['total_cost']}")

    lines.append(f"\nDetail Items ({len(items)} rows):")
    for item in items[:20]:
        lines.append(f"  Row {item['row']}: {item['description'][:40]} | Qty: {item['qty']} | Unit: ${item['unit_cost']} | Total: ${item['total_cost']}")
    if len(items) > 20:
        lines.append(f"  ... and {len(items) - 20} more items")

    lines.append(f"\nSum of Detail Items: ${total_cost_sum:.2f}")
    lines.append(f"\nSummary/Totals Rows:")
    for sr in summary_rows:
        lines.append(f"  Row {sr['row']}: {sr['description']} = ${sr['total_cost']}")

    # Read invoice_total from column S (col 19) of row 2 — this is the raw
    # numeric invoice total written by the XLSX generator. Reading the
    # INVOICE TOTAL row's col P is unreliable because that cell contains the
    # formula ``=S2`` which evaluates to ``None`` under data_only=True on a
    # freshly written workbook with no Excel-populated formula cache.
    invoice_total = ws.cell(row=2, column=19).value  # S = COL_INV_TOTAL

    lines.append(f"Invoice Total from XLSX (S2): ${invoice_total}")

    return {
        'text': '\n'.join(lines),
        'invoice_total': invoice_total,
        'total_cost_sum': total_cost_sum,
        'items': items,
        'desc_col': desc_col,
    }


def _apply_fixes(ws, fixes_data: Dict, cfg) -> int:
    """Apply LLM-proposed fixes to XLSX cells. Only modifies numeric values.
    NEVER inserts rows or writes text — the XLSX structure is fixed."""
    fixes_applied = 0

    col_map = {
        'total_cost': cfg.col_total_cost,
        'unit_cost': cfg.col_unit_cost,
        'quantity': cfg.col_quantity,
    }

    for fix in fixes_data.get('fixes', []):
        row = fix.get('row')
        col_name = fix.get('column', 'total_cost')
        new_value = fix.get('new_value')

        if row and new_value is not None:
            col_idx = col_map.get(col_name, cfg.col_total_cost)
            ws.cell(row=row, column=col_idx).value = float(new_value)
            fixes_applied += 1
            logger.info(f"Fixed row {row}: {col_name} = ${new_value}")

    # NEVER insert rows — add_items is ignored to preserve XLSX structure.
    # If the LLM proposed add_items, we log it but skip the insertion.
    add_items = fixes_data.get('add_items', [])
    if add_items:
        total_add = sum(item.get('total_cost', 0) for item in add_items)
        logger.warning(f"LLM proposed {len(add_items)} add_items (${total_add:.2f}) — "
                       f"SKIPPED to preserve XLSX structure. Will use force_adjustment instead.")

    return fixes_applied


def _recalculate_variance(ws, invoice_total, cfg) -> float:
    """Recalculate variance from XLSX data.

    Adjustments are read directly from the raw numeric cells T2/U2/V2/W2
    (freight/insurance/tax/deduction) because the ADJUSTMENTS row's total_cost
    cell contains the formula ``=(T2+U2+V2-W2)`` — a string that cannot be
    evaluated without Excel, so the previous implementation silently treated
    adjustments as zero and computed a wildly wrong variance that, when fed
    through _force_adjustment, destroyed the block.
    """
    desc_col = _get_desc_column(ws)
    # Column indices for the invoice-level adjustments row (row 2).
    COL_FREIGHT = 20    # T
    COL_INSURANCE = 21  # U
    COL_TAX = 22        # V
    COL_DEDUCTION = 23  # W
    new_total = 0.0

    def _is_group_row(row):
        fill = ws.cell(row=row, column=1).fill
        return (fill and fill.start_color and
                'D9E1F2' in str(getattr(fill.start_color, 'rgb', '') or ''))

    for row in range(2, ws.max_row + 1):
        desc = ws.cell(row=row, column=desc_col).value
        total = ws.cell(row=row, column=cfg.col_total_cost).value
        if not desc:
            continue

        desc_str = str(desc).upper()
        # Skip every summary/formula row — including ADJUSTMENTS, which is a
        # formula string we compute separately below from raw T/U/V/W cells.
        if any(x in desc_str for x in ['SUBTOTAL', 'VARIANCE', 'NET TOTAL', 'GROUP VERIFICATION',
                                        'ADJUSTMENTS', 'INVOICE TOTAL', 'TOTAL INTERNAL',
                                        'TOTAL INSURANCE', 'TOTAL OTHER', 'TOTAL DEDUCTION']):
            continue
        if _is_group_row(row):
            continue  # Skip group rows — only count detail items

        # Detail items
        if total and isinstance(total, (int, float)):
            new_total += float(total)

    def _num(v):
        try:
            return float(v) if v is not None else 0.0
        except (TypeError, ValueError):
            return 0.0

    freight = _num(ws.cell(row=2, column=COL_FREIGHT).value)
    insurance = _num(ws.cell(row=2, column=COL_INSURANCE).value)
    tax = _num(ws.cell(row=2, column=COL_TAX).value)
    deduction = _num(ws.cell(row=2, column=COL_DEDUCTION).value)
    adjustment_total = freight + insurance + tax - deduction

    new_total += adjustment_total
    # invoice_total may be a string from formula cells — convert safely
    try:
        inv_total = float(invoice_total) if invoice_total else 0
    except (ValueError, TypeError):
        inv_total = 0
    new_variance = inv_total - new_total

    _update_variance_row(ws, new_variance, cfg)
    return new_variance


def _update_variance_row(ws, variance: float, cfg):
    """Update the VARIANCE CHECK row formatting (NOT its value).

    VARIANCE CHECK must ALWAYS be a formula (e.g. =S2-P{net_total_row}) so it
    remains auditable when opened in Excel.  The fixer corrects the ADJUSTMENTS
    formula, which propagates through NET TOTAL into VARIANCE CHECK automatically.
    We only touch the font colour here: red when variance is non-zero, green when
    it is resolved.
    """
    from openpyxl.styles import Font
    desc_col = _get_desc_column(ws)
    for row in range(ws.max_row, 0, -1):
        desc = ws.cell(row=row, column=desc_col).value
        if desc and 'VARIANCE CHECK' in str(desc).upper():
            cell = ws.cell(row=row, column=cfg.col_total_cost)
            # If a previous run overwrote the formula with a numeric value,
            # restore it.  The formula is always =S{first_data}-P{net_total}.
            if not isinstance(cell.value, str) or not str(cell.value).startswith('='):
                net_total_row = _find_label_row(ws, 'NET TOTAL', desc_col)
                if net_total_row:
                    cell.value = f'=S2-P{net_total_row}'
            # Colour: green if resolved, red if not
            colour = '006100' if abs(variance) < 0.01 else 'FF0000'
            cell.font = Font(bold=True, color=colour)
            ws.cell(row=row, column=desc_col).font = Font(bold=True, color=colour)
            break


def _find_label_row(ws, label: str, desc_col: int) -> int | None:
    """Find the row number whose description column matches *label* (case-insensitive)."""
    target = label.upper()
    for row in range(ws.max_row, 0, -1):
        val = ws.cell(row=row, column=desc_col).value
        if val and target in str(val).upper():
            return row
    return None


def _force_adjustment(ws, remaining_variance: float, cfg):
    """Force the ADJUSTMENTS formula to include a correction that zeroes the variance.

    The base formula is ``=(T{r}+U{r}+V{r}-W{r})``.  This function strips any
    previous correction appended after the closing ``)`` and writes a single,
    clean correction term so that repeated runs are idempotent (no stacking).

    NEVER inserts new rows — only modifies the existing ADJUSTMENTS cell.
    """
    import re as _re
    desc_col = _get_desc_column(ws)
    for row in range(ws.max_row, 0, -1):
        desc = ws.cell(row=row, column=desc_col).value
        if desc and 'ADJUSTMENTS' in str(desc).upper():
            current_adj = ws.cell(row=row, column=cfg.col_total_cost).value or 0

            if isinstance(current_adj, str) and current_adj.startswith('='):
                # Strip any previous correction: keep everything up to and
                # including the closing ')' of the base formula.
                base_match = _re.match(r'(=\([^)]+\))', current_adj)
                if base_match:
                    base_formula = base_match.group(1)
                else:
                    base_formula = current_adj  # unexpected shape — keep as-is

                # Append the single correction term
                ws.cell(row=row, column=cfg.col_total_cost).value = f'{base_formula}+{remaining_variance}'
            else:
                try:
                    ws.cell(row=row, column=cfg.col_total_cost).value = float(current_adj) + remaining_variance
                except (ValueError, TypeError):
                    ws.cell(row=row, column=cfg.col_total_cost).value = remaining_variance

            logger.info(f"Forced adjustment: row {row}, old={current_adj!r}, correction=${remaining_variance:.2f}")
            return

    # No ADJUSTMENTS row found — cannot fix without inserting rows
    logger.warning(f"No ADJUSTMENTS row found in XLSX — cannot force adjustment of ${remaining_variance:.2f}")


def _update_group_header_totals(ws, cfg) -> int:
    """Recompute each group header's quantity & total_cost from its detail children.

    In grouped XLSX format (bl_xlsx_generator), each group header row (blue
    D9E1F2 fill) is followed by one or more detail rows. The header's col P
    holds the sum of its details' col P, and SUBTOTAL (GROUPED) references the
    header (e.g. ``=P2`` or ``=P2+P4``). When the LLM-based variance fixer
    modifies detail values, the group header stays stale and the NET TOTAL
    formula propagates the wrong value.

    This walks the sheet, groups detail rows under their nearest preceding
    group header, and writes the summed qty/total_cost back into the header.
    Returns the number of headers whose total_cost changed.
    """
    desc_col = _get_desc_column(ws)

    _SUMMARY_MARKERS = (
        'SUBTOTAL', 'VARIANCE', 'NET TOTAL', 'GROUP VERIFICATION',
        'ADJUSTMENTS', 'INVOICE TOTAL', 'TOTAL INTERNAL',
        'TOTAL INSURANCE', 'TOTAL OTHER', 'TOTAL DEDUCTION',
    )

    def _is_group_row(row):
        fill = ws.cell(row=row, column=1).fill
        return (fill and fill.start_color and
                'D9E1F2' in str(getattr(fill.start_color, 'rgb', '') or ''))

    def _is_summary_row(row):
        desc = ws.cell(row=row, column=desc_col).value
        if not desc:
            return False
        s = str(desc).upper()
        return any(m in s for m in _SUMMARY_MARKERS)

    # Find all group header rows and the first summary row (end of item section).
    group_headers = []
    summary_start = None
    for row in range(2, ws.max_row + 1):
        if _is_summary_row(row):
            if summary_start is None:
                summary_start = row
            continue
        if _is_group_row(row):
            group_headers.append(row)

    if not group_headers:
        return 0
    if summary_start is None:
        summary_start = ws.max_row + 1

    updated = 0
    for i, gh in enumerate(group_headers):
        next_boundary = group_headers[i + 1] if i + 1 < len(group_headers) else summary_start
        qty_sum = 0.0
        cost_sum = 0.0
        has_cost = False
        for row in range(gh + 1, next_boundary):
            if _is_summary_row(row):
                break
            q = ws.cell(row=row, column=cfg.col_quantity).value
            tc = ws.cell(row=row, column=cfg.col_total_cost).value
            if isinstance(q, (int, float)):
                qty_sum += float(q)
            if isinstance(tc, (int, float)):
                cost_sum += float(tc)
                has_cost = True

        if not has_cost:
            continue  # nothing to roll up

        old_cost = ws.cell(row=gh, column=cfg.col_total_cost).value
        new_cost = round(cost_sum, 2)
        if not isinstance(old_cost, (int, float)) or abs(float(old_cost) - new_cost) > 0.005:
            ws.cell(row=gh, column=cfg.col_total_cost).value = new_cost
            updated += 1
            logger.info(f"Group header row {gh}: total_cost {old_cost} -> {new_cost}")

        old_qty = ws.cell(row=gh, column=cfg.col_quantity).value
        new_qty = int(qty_sum) if qty_sum == int(qty_sum) else qty_sum
        if not isinstance(old_qty, (int, float)) or old_qty != new_qty:
            ws.cell(row=gh, column=cfg.col_quantity).value = new_qty

    return updated
