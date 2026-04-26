#!/usr/bin/env python3
"""
XLSX Combiner - Combine multiple invoice XLSX files into one.

Creates a combined file with:
- One header row at the top
- Each invoice's FULL data including verification totals
- Blank row between invoices
- Grand total with variance check at the end
"""

import argparse
import json
import os
import re
from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

from pipeline.config_loader import load_xlsx_labels, load_validation_tolerances


def copy_cell_style(src_cell, dest_cell):
    """Copy all styling from source cell to destination cell."""
    if src_cell.has_style:
        dest_cell.font = copy(src_cell.font)
        dest_cell.border = copy(src_cell.border)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.number_format = src_cell.number_format
        dest_cell.protection = copy(src_cell.protection)
        dest_cell.alignment = copy(src_cell.alignment)


def _shift_formula(formula: str, row_offset: int) -> str:
    """Shift all row references in an Excel formula by row_offset.

    E.g. '=O2*K2' with offset 5 → '=O7*K7'
         '=P2+P15' with offset 5 → '=P7+P20'
         '=SUM(P2:P19)-P21' with offset 5 → '=SUM(P7:P24)-P26'
    """
    if not formula or not formula.startswith('='):
        return formula

    def replace_ref(m):
        col = m.group(1)
        row_num = int(m.group(2))
        return f'{col}{row_num + row_offset}'

    # Match column letter(s) followed by row number, but not inside quotes
    return re.sub(r'([A-Z]{1,3})(\d+)', replace_ref, formula)   # magic-ok: Excel cell-reference regex (column-letters + row-number)


# Columns (bl_xlsx_generator layout, see config/columns.yaml):
#   P (16) = total_cost per row / sums
#   S (19) = invoice total (only on first data row)
#   T (20) = freight, U (21) = insurance, V (22) = tax, W (23) = deduction
_COL_DESC_J = 10      # magic-ok: columns.yaml description column for bl_xlsx_generator
_COL_DESC_L = 12      # magic-ok: columns.yaml description column for legacy xlsx_generator
_COL_TOTAL_COST = 16  # magic-ok: columns.yaml TotalCost column index
_COL_INV_TOTAL = 19   # magic-ok: columns.yaml InvoiceTotal column index
_COL_FREIGHT = 20     # magic-ok: columns.yaml Freight column index
_COL_INSURANCE = 21   # magic-ok: columns.yaml Insurance column index
_COL_TAX = 22         # magic-ok: columns.yaml Tax/OtherCost column index
_COL_DEDUCTION = 23   # magic-ok: columns.yaml Deduction column index
_COL_A = 1            # magic-ok: openpyxl column-A fill-color probe

# SSOT-loaded module constants.
_LABELS = load_xlsx_labels()
_SUMMARY_MARKERS = tuple(_LABELS["summary_row_markers"])
_GROUP_FILL_RGB = "D9E1F2"   # magic-ok: openpyxl ARGB for grouped-row fill (bl_xlsx_generator style)

_VAR_GREEN_THRESHOLD = load_validation_tolerances()["variance_fixer"]["variance_row_green_threshold"]


def _is_group_header_row(ws, row: int) -> bool:
    """A group header is filled with 'D9E1F2' in bl_xlsx_generator's grouped
    output. Its col-P value is the sum of its child detail rows, so if both
    header and children are counted the detail total is double-counted."""
    fill = ws.cell(row=row, column=_COL_A).fill
    return bool(
        fill and fill.start_color
        and _GROUP_FILL_RGB in str(getattr(fill.start_color, 'rgb', '') or '')
    )


def _is_summary_row(ws, row: int) -> bool:
    """Rows whose description column (J=10 or L=12) matches any summary marker."""
    for col in (_COL_DESC_J, _COL_DESC_L):
        desc = ws.cell(row=row, column=col).value
        if desc and isinstance(desc, str):
            up = desc.upper()
            if any(m in up for m in _SUMMARY_MARKERS):
                return True
    return False


def _parse_adjustment_correction(formula_str) -> float:
    """Return the trailing numeric correction appended to an ADJUSTMENTS formula.

    The variance_fixer writes corrections as ``=(T{r}+U{r}+V{r}-W{r})+12.73``
    or ``...+-4.43``. We parse whatever follows the closing parenthesis so the
    combined summary picks up the same net total Excel will show on open.
    """
    if not isinstance(formula_str, str) or not formula_str.startswith('='):
        return 0.0
    m = re.match(r'=\([^)]+\)\s*([+\-].+)$', formula_str)   # magic-ok: ADJUSTMENTS-formula trailing-correction parser
    if not m:
        return 0.0
    try:
        return float(m.group(1))
    except (TypeError, ValueError):
        return 0.0


def _compute_block_totals(ws) -> Dict[str, float]:
    """Compute (invoice_total, net_total, variance) from raw cells of a
    single per-invoice XLSX.

    Mirrors ``variance_fixer._recalculate_variance`` so what the combined
    summary displays matches what Excel will evaluate for the per-block
    VARIANCE CHECK formulas. Without this, the old fallback double-counted
    grouped header rows and ignored the ADJUSTMENTS correction, producing
    wildly wrong per-invoice variance numbers in the grand summary block.
    """
    detail_sum = 0.0
    adjustment_correction = 0.0
    invoice_total_s = 0.0
    # Note: ADJUSTMENTS correction can appear on any row, so we hunt for it
    # across the whole sheet rather than assuming row 2. Also note that in
    # bl_xlsx_generator's *grouped* output, column S is written only on the
    # group-header row (blue fill), so S must be collected even on rows that
    # we otherwise skip from the detail sum.
    for r in range(2, ws.max_row + 1):
        # Invoice total lives in col S on the first data or group-header row.
        s_val = ws.cell(row=r, column=_COL_INV_TOTAL).value
        if isinstance(s_val, (int, float)) and s_val and invoice_total_s == 0.0:
            invoice_total_s = float(s_val)

        if _is_summary_row(ws, r):
            # ADJUSTMENTS row may carry a trailing +correction in its formula
            for col in (_COL_DESC_J, _COL_DESC_L):
                desc = ws.cell(row=r, column=col).value
                if desc and isinstance(desc, str) and 'ADJUSTMENTS' in desc.upper():   # magic-ok: xlsx_labels.totals.ADJUSTMENTS label substring
                    adjustment_correction += _parse_adjustment_correction(
                        ws.cell(row=r, column=_COL_TOTAL_COST).value
                    )
                    break
            continue
        if _is_group_header_row(ws, r):
            continue
        # Detail row: sum col P if numeric
        tc = ws.cell(row=r, column=_COL_TOTAL_COST).value
        if isinstance(tc, (int, float)):
            detail_sum += float(tc)

    def _num(v):
        try:
            return float(v) if v is not None else 0.0
        except (TypeError, ValueError):
            return 0.0

    freight = _num(ws.cell(row=2, column=_COL_FREIGHT).value)
    insurance = _num(ws.cell(row=2, column=_COL_INSURANCE).value)
    tax = _num(ws.cell(row=2, column=_COL_TAX).value)
    deduction = _num(ws.cell(row=2, column=_COL_DEDUCTION).value)
    adjustments = freight + insurance + tax - deduction + adjustment_correction

    net_total = round(detail_sum + adjustments, 2)
    variance = round(invoice_total_s - net_total, 2)
    return {
        'invoice_total': round(invoice_total_s, 2),
        'net_total': net_total,
        'variance': variance,
    }


def combine_xlsx_files(file_paths: List[str], output_path: str,
                       ocr_notes: List[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    Combine multiple XLSX invoice files into one.

    Args:
        file_paths: List of paths to XLSX files to combine
        output_path: Path for the combined output file
        ocr_notes: Optional list of dicts with OCR quality data per invoice:
                   [{'pdf_file': str, 'score': int, 'rating': str,
                     'details': str, 'raw_text': str}]

    Returns:
        Dict with status and details
    """
    # Deduplicate by absolute path
    seen = set()
    unique_paths = []
    for p in file_paths:
        ap = os.path.abspath(p)
        if ap not in seen:
            seen.add(ap)
            unique_paths.append(p)
    file_paths = unique_paths

    if len(file_paths) < 2:   # magic-ok: require at least 2 files for combining (name says "combine")
        return {'status': 'error', 'error': 'Need at least 2 files to combine'}   # magic-ok: error-payload message surfaced to caller

    # Debug: log files being combined
    print(f"    [combiner] Combining {len(file_paths)} files:", flush=True)
    for fp in file_paths:
        try:
            wb_dbg = load_workbook(fp)
            ws_dbg = wb_dbg.active
            print(f"      {os.path.basename(fp)}: {ws_dbg.max_row} rows, {ws_dbg.max_column} cols, sheet={ws_dbg.title}", flush=True)
            # Check first data row structure
            j2 = ws_dbg.cell(2, _COL_DESC_J).value
            l2 = ws_dbg.cell(2, _COL_DESC_L).value
            fill2 = ws_dbg.cell(2, _COL_A).fill
            is_grp = fill2 and fill2.start_color and _GROUP_FILL_RGB in str(fill2.start_color.rgb or '')
            print(f"        Row2: J={str(j2)[:30] if j2 else None!r} L={str(l2)[:30] if l2 else None!r} grouped={is_grp}", flush=True)   # magic-ok: debug print truncation width
            wb_dbg.close()
        except Exception as e:
            print(f"      {os.path.basename(fp)}: ERROR {e}", flush=True)

    # Create output workbook
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Combined Invoices'   # magic-ok: openpyxl sheet title for the combined-invoices output

    # Track totals for grand summary
    invoice_summaries = []
    grand_total_cost = 0.0
    grand_total_variance = 0.0

    current_row = 1
    header_written = False
    max_col = 17   # magic-ok: bl_xlsx_generator writes 17 columns (A..Q) per invoice; overridden from file

    # Styles for separator and grand summary
    separator_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')   # magic-ok: openpyxl light-grey separator fill
    summary_font = Font(bold=True, size=11)   # magic-ok: openpyxl summary-row font size
    error_font = Font(bold=True, color='FF0000')   # magic-ok: openpyxl ARGB (red) for error
    ok_font = Font(bold=True, color='008000')   # magic-ok: openpyxl ARGB (green) for OK

    for file_idx, file_path in enumerate(file_paths):
        if not os.path.exists(file_path):
            return {'status': 'error', 'error': f'File not found: {file_path}'}   # magic-ok: error-payload key surfaced to caller

        # First pass: compute block totals from raw cells.
        #
        # We deliberately do NOT rely on openpyxl's ``data_only=True`` cached
        # formula values here because openpyxl cannot evaluate formulas — it
        # only returns what Excel already wrote into the cache. Our per-invoice
        # XLSXs are produced by openpyxl and never opened by Excel, so every
        # formula cell (NET TOTAL, VARIANCE CHECK, ADJUSTMENTS, …) reads as
        # None. Instead, reconstruct the numbers from raw detail cells using
        # the same algorithm as ``variance_fixer._recalculate_variance`` so the
        # GRAND SUMMARY matches what Excel will show when the combined file
        # is opened.
        try:
            wb_data = load_workbook(file_path)
            ws_data = wb_data.active
        except Exception as e:
            return {'status': 'error', 'error': f'Failed to open {file_path}: {str(e)}'}   # magic-ok: error-payload key surfaced to caller

        # Invoice number lives in column C of the first data row
        invoice_num = ''
        for r in range(2, ws_data.max_row + 1):
            inv_cell = ws_data.cell(row=r, column=3).value   # magic-ok: columns.yaml InvoiceNumber column C (3)
            if inv_cell and str(inv_cell).strip():
                invoice_num = str(inv_cell).strip()
                break

        block = _compute_block_totals(ws_data)
        net_total = block['net_total']
        variance = block['variance']

        invoice_summaries.append({
            'file': os.path.basename(file_path),
            'invoice_num': invoice_num,
            'total': net_total,
            'variance': variance,
        })
        grand_total_cost += net_total
        grand_total_variance += variance
        wb_data.close()

        # Second pass: read with formulas preserved
        try:
            wb = load_workbook(file_path)
            ws = wb.active
        except Exception as e:
            return {'status': 'error', 'error': f'Failed to open {file_path}: {str(e)}'}   # magic-ok: error-payload key surfaced to caller

        max_col = max(max_col, ws.max_column)

        # Determine row range to copy
        start_row = 1 if not header_written else 2  # Skip header for subsequent files
        end_row = ws.max_row

        # Calculate row offset for formula adjustment
        # Source rows start at start_row, destination rows start at current_row
        row_offset = current_row - start_row

        # Add blank separator between invoices (not before first)
        if header_written:
            for c in range(1, max_col + 1):
                cell = out_ws.cell(row=current_row, column=c)
                cell.fill = separator_fill
            current_row += 1
            # Recalculate offset after separator
            row_offset = current_row - start_row

        # Copy rows from this file, adjusting formula row references
        for r in range(start_row, end_row + 1):
            for c in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=r, column=c)
                dest_cell = out_ws.cell(row=current_row, column=c)

                value = src_cell.value
                # Adjust formula row references
                if isinstance(value, str) and value.startswith('='):
                    value = _shift_formula(value, row_offset)

                dest_cell.value = value
                copy_cell_style(src_cell, dest_cell)

            current_row += 1

        if not header_written:
            header_written = True

        wb.close()

    # Add grand summary section
    current_row += 1  # Blank row

    # Separator before grand summary
    for c in range(1, max_col + 1):
        out_ws.cell(row=current_row, column=c).fill = PatternFill(
            start_color='4472C4', end_color='4472C4', fill_type='solid'   # magic-ok: openpyxl ARGB (dark-blue) summary-separator fill
        )
    current_row += 1

    # Grand Summary Header
    header_cell = out_ws.cell(row=current_row, column=_COL_DESC_L, value='═══ COMBINED GRAND SUMMARY ═══')   # magic-ok: grand-summary section title
    header_cell.font = Font(bold=True, size=12, color='FFFFFF')   # magic-ok: openpyxl header font (size 12, white)
    header_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')   # magic-ok: openpyxl ARGB (dark-blue) header fill
    for c in range(1, max_col + 1):
        out_ws.cell(row=current_row, column=c).fill = PatternFill(
            start_color='4472C4', end_color='4472C4', fill_type='solid'   # magic-ok: openpyxl ARGB (dark-blue) header fill
        )
    current_row += 1

    # List each invoice with its totals
    for summary in invoice_summaries:
        label = f"Invoice {summary['invoice_num'] or summary['file']}"
        out_ws.cell(row=current_row, column=_COL_DESC_L, value=label)
        out_ws.cell(row=current_row, column=_COL_DESC_L).font = Font(bold=True)

        # Net total
        total_cell = out_ws.cell(row=current_row, column=_COL_TOTAL_COST, value=summary['total'])
        total_cell.number_format = '"$"#,##0.00'   # magic-ok: Excel currency number_format

        # Variance
        if abs(summary['variance']) > _VAR_GREEN_THRESHOLD:
            var_cell = out_ws.cell(row=current_row, column=17, value=summary['variance'])   # magic-ok: column Q (17) holds variance in combined output
            var_cell.number_format = '"$"#,##0.00'   # magic-ok: Excel currency number_format
            var_cell.font = error_font

        current_row += 1

    current_row += 1  # Blank row

    # Grand totals row
    out_ws.cell(row=current_row, column=_COL_DESC_L, value='GRAND TOTAL')   # magic-ok: grand-totals row label
    out_ws.cell(row=current_row, column=_COL_DESC_L).font = summary_font
    grand_cell = out_ws.cell(row=current_row, column=_COL_TOTAL_COST, value=grand_total_cost)
    grand_cell.number_format = '"$"#,##0.00'   # magic-ok: Excel currency number_format
    grand_cell.font = summary_font
    current_row += 1

    # Grand variance check
    out_ws.cell(row=current_row, column=_COL_DESC_L, value='GRAND VARIANCE CHECK')   # magic-ok: grand-variance row label
    var_font = error_font if abs(grand_total_variance) > _VAR_GREEN_THRESHOLD else ok_font
    out_ws.cell(row=current_row, column=_COL_DESC_L).font = var_font
    grand_var_cell = out_ws.cell(row=current_row, column=_COL_TOTAL_COST, value=grand_total_variance)
    grand_var_cell.number_format = '"$"#,##0.00'   # magic-ok: Excel currency number_format
    grand_var_cell.font = var_font
    current_row += 1

    # File count note
    note_cell = out_ws.cell(row=current_row, column=_COL_DESC_L, value=f'Combined {len(file_paths)} invoice files')
    note_cell.font = Font(italic=True, color='666666')   # magic-ok: openpyxl ARGB (mid-grey) for italic note

    # Copy column widths from first file
    try:
        first_wb = load_workbook(file_paths[0])
        first_ws = first_wb.active
        for col_idx in range(1, first_ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if first_ws.column_dimensions[col_letter].width:
                out_ws.column_dimensions[col_letter].width = first_ws.column_dimensions[col_letter].width
        first_wb.close()
    except:
        # Fallback column widths (openpyxl character units — display-only policy).
        col_widths = {   # magic-ok: openpyxl column-width fallback table (display-only)
            1: 10, 2: 10, 3: 12, 4: 12, 5: 25, 6: 12, 7: 10, 8: 10, 9: 15,   # magic-ok: openpyxl column widths (display)
            10: 40, 11: 10, 12: 45, 13: 8, 14: 8, 15: 12, 16: 14, 17: 14     # magic-ok: openpyxl column widths (display)
        }
        for col, width in col_widths.items():
            out_ws.column_dimensions[get_column_letter(col)].width = width

    # ─── OCR Notes Sheet ──────────────────────────────────────
    # When OCR quality data is provided, add a sheet showing raw OCR text
    # per invoice so users can see extraction quality and identify rescans needed.
    any_poor = False
    if ocr_notes:
        ocr_ws = out_wb.create_sheet('OCR Notes')   # magic-ok: openpyxl sheet name for the OCR-quality addendum
        ocr_header_font = Font(bold=True, size=11, color='FFFFFF')   # magic-ok: openpyxl OCR-header font (size 11, white)
        ocr_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')   # magic-ok: openpyxl ARGB (dark-blue) OCR-header fill
        poor_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')   # magic-ok: openpyxl ARGB (light-red) "poor" row fill
        poor_font = Font(bold=True, color='9C0006')   # magic-ok: openpyxl ARGB (dark-red) "poor" row text
        fair_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')   # magic-ok: openpyxl ARGB (light-yellow) "fair" row fill
        good_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')   # magic-ok: openpyxl ARGB (light-green) "good" row fill
        wrap_align = Alignment(wrap_text=True, vertical='top')   # magic-ok: openpyxl vertical-align enum (stdlib value)

        # Headers
        headers = ['PDF File', 'Quality Score', 'Rating', 'Details', 'OCR Text (first 2000 chars)']   # magic-ok: OCR-notes sheet column headers (display-only)
        for c, h in enumerate(headers, 1):
            cell = ocr_ws.cell(row=1, column=c, value=h)
            cell.font = ocr_header_font
            cell.fill = ocr_header_fill

        ocr_ws.column_dimensions['A'].width = 25   # magic-ok: openpyxl column width for OCR sheet (display)
        ocr_ws.column_dimensions['B'].width = 14   # magic-ok: openpyxl column width for OCR sheet (display)
        ocr_ws.column_dimensions['C'].width = 10   # magic-ok: openpyxl column width for OCR sheet (display)
        ocr_ws.column_dimensions['D'].width = 60   # magic-ok: openpyxl column width for OCR sheet (display)
        ocr_ws.column_dimensions['E'].width = 100  # magic-ok: openpyxl column width for OCR sheet (display)

        for idx, note in enumerate(ocr_notes, 2):
            rating = note.get('rating', 'unknown')   # magic-ok: default rating when OCR meta missing
            score = note.get('score', 0)

            ocr_ws.cell(row=idx, column=1, value=note.get('pdf_file', ''))
            score_cell = ocr_ws.cell(row=idx, column=2, value=f"{score}/100")   # magic-ok: "X/100" score display format
            rating_cell = ocr_ws.cell(row=idx, column=3, value=rating.upper())   # magic-ok: column C (3) = Rating in OCR Notes sheet
            ocr_ws.cell(row=idx, column=4, value=note.get('details', ''))   # magic-ok: column D (4) = Details in OCR Notes sheet

            raw_text = note.get('raw_text', '')[:2000]   # magic-ok: OCR-text display truncation width
            text_cell = ocr_ws.cell(row=idx, column=5, value=raw_text)   # magic-ok: column E (5) = OCR Text in OCR Notes sheet
            text_cell.alignment = wrap_align

            # Color-code by rating
            if rating == 'poor' or rating == 'unusable':   # magic-ok: OCR rating enum values
                any_poor = True
                for c in range(1, 6):   # magic-ok: columns A..E of OCR sheet (5 cols)
                    ocr_ws.cell(row=idx, column=c).fill = poor_fill
                rating_cell.font = poor_font
            elif rating == 'fair':   # magic-ok: OCR rating enum value
                for c in range(1, 6):   # magic-ok: columns A..E of OCR sheet (5 cols)
                    ocr_ws.cell(row=idx, column=c).fill = fair_fill
            elif rating == 'good':   # magic-ok: OCR rating enum value
                for c in range(1, 6):   # magic-ok: columns A..E of OCR sheet (5 cols)
                    ocr_ws.cell(row=idx, column=c).fill = good_fill

        # Add warning row if any poor quality
        if any_poor:
            warn_row = len(ocr_notes) + 3   # magic-ok: 1 header row + 1 blank row + 1 offset
            warn_cell = ocr_ws.cell(row=warn_row, column=1,
                                     value='WARNING: OCR text quality is too bad for some pages. '   # magic-ok: user-visible warning banner
                                           'Consider rescanning and resubmitting the email.')
            warn_cell.font = Font(bold=True, color='FF0000', size=12)   # magic-ok: openpyxl ARGB (red) size-12 warning font
            ocr_ws.merge_cells(start_row=warn_row, start_column=1,
                               end_row=warn_row, end_column=5)   # magic-ok: merge all 5 columns of warning banner

    # Save
    out_wb.save(output_path)

    return {
        'status': 'success',   # magic-ok: result-payload status value surfaced to caller
        'output': output_path,
        'files_combined': len(file_paths),
        'grand_total': grand_total_cost,
        'grand_variance': grand_total_variance,
        'any_poor_ocr': any_poor,
    }


def main():
    parser = argparse.ArgumentParser(description='Combine multiple XLSX invoice files')
    parser.add_argument('--files', required=True, help='JSON array of file paths')
    parser.add_argument('--output', required=True, help='Output file path')

    args = parser.parse_args()

    try:
        file_paths = json.loads(args.files)
    except json.JSONDecodeError as e:
        print(f'Error parsing files JSON: {e}', flush=True)
        exit(1)

    result = combine_xlsx_files(file_paths, args.output)

    if result['status'] == 'success':   # magic-ok: result-payload status value check
        print(f"Combined {result['files_combined']} files into {result['output']}", flush=True)
        print(f"Grand Total: ${result['grand_total']:.2f}", flush=True)
        print(f"Grand Variance: ${result['grand_variance']:.2f}", flush=True)
        exit(0)
    else:
        print(f"Error: {result['error']}", flush=True)
        exit(1)


if __name__ == '__main__':
    main()
