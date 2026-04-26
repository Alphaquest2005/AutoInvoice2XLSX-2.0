#!/usr/bin/env python3
"""
Stage 6: XLSX Generator
Generates Excel output per columns.yaml specification.
37 columns (A-AK), group/detail rows, totals section.

All user-visible labels, HS category mappings, supplier-code aliases,
tolerances, filenames, styles, and fallbacks are sourced from
``config/`` via ``pipeline.config_loader``. There are no policy strings
or numeric thresholds hardcoded in this module — change wording or
rates by editing YAML, not Python.
"""

from __future__ import annotations

import json
import os
import re
from pathlib import Path
from typing import Any, Dict, List

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    openpyxl = None
try:
    import yaml
except ImportError:
    yaml = None

from pipeline.config_loader import (
    load_columns,
    load_document_types,
    load_file_paths,
    load_hs_categories,
    load_invoice_formats,
    load_issue_types,
    load_library_enums,
    load_patterns,
    load_validation_tolerances,
    load_xlsx_labels,
)

# ── Cached config views (module-level — avoid reloading in hot loops) ──
_COLS_CFG = load_columns()
_LABELS = load_xlsx_labels()
_FP = load_file_paths()
_TOL = load_validation_tolerances()
_DT_CFG = load_document_types()
_ISSUE = load_issue_types()
_ENUMS = load_library_enums()
_PATTERNS = load_patterns()

# Status tokens (controlled vocabulary from issue_types.yaml).
_STATUS_SUCCESS = _ISSUE["status"]["SUCCESS"]
_STATUS_ERROR = _ISSUE["status"]["ERROR"]

# openpyxl library enum: fill_type=solid.
_FILL_SOLID = _ENUMS["openpyxl"]["fill_type"]["SOLID"]

# Regex pattern for filename version-suffix detection.
_VERSION_SUFFIX_RE = _PATTERNS["filename_version_suffix"]

# Error-message catalog for run() return payloads.
_ERRS = _LABELS["errors"]

# Column letter → 1-based index from columns.yaml (single source of truth).
_COL_INDEX: Dict[str, int] = {
    letter: spec["index"] for letter, spec in _COLS_CFG["columns"].items()
}
_NUM_COLS: int = max(_COL_INDEX.values())

# Named column indices used throughout this module.
_COL_DOC_TYPE = _COL_INDEX["A"]
_COL_INVOICE_NUM = _COL_INDEX["C"]
_COL_DATE = _COL_INDEX["D"]
_COL_CATEGORY = _COL_INDEX["E"]
_COL_TARIFF = _COL_INDEX["F"]
_COL_PO_ITEM_NUM = _COL_INDEX["G"]
_COL_PO_ITEM_DESC = _COL_INDEX["H"]
_COL_SUPPLIER_ITEM_NUM = _COL_INDEX["I"]
_COL_SUPPLIER_ITEM_DESC = _COL_INDEX["J"]
_COL_QTY = _COL_INDEX["K"]
_COL_UOM = _COL_INDEX["M"]
_COL_CURRENCY = _COL_INDEX["N"]
_COL_COST = _COL_INDEX["O"]
_COL_TOTAL_COST = _COL_INDEX["P"]
_COL_TOTAL = _COL_INDEX["Q"]
_COL_TOTALCOST_VS_TOTAL = _COL_INDEX["R"]
_COL_INVOICE_TOTAL = _COL_INDEX["S"]
_COL_FREIGHT = _COL_INDEX["T"]
_COL_INSURANCE = _COL_INDEX["U"]
_COL_OTHER_COST = _COL_INDEX["V"]
_COL_DEDUCTION = _COL_INDEX["W"]
_COL_SUPPLIER_CODE = _COL_INDEX["Z"]
_COL_SUPPLIER_NAME = _COL_INDEX["AA"]
_COL_SUPPLIER_ADDR = _COL_INDEX["AB"]
_COL_COUNTRY = _COL_INDEX["AC"]
_COL_GROUPBY = _COL_INDEX["AK"]

_FIRST_DATA_ROW = 2  # row 1 = header; row 2 = first data row per columns.yaml

# Columns that receive currency number_format on group rows.
_CURRENCY_COLS_GROUP: List[int] = [
    _COL_INDEX[letter] for letter in _COLS_CFG["currency"]["all_rows"]
]
# Columns that receive currency number_format on detail rows.
_CURRENCY_COLS_DETAIL: List[int] = [
    _COL_INDEX[letter] for letter in _COLS_CFG["currency"]["detail_only"]
]
_CURRENCY_FMT: str = _COLS_CFG["currency"]["format"]

# Currency code for column N (columns.yaml is SSOT).
_CURRENCY_CODE: str = _COLS_CFG["columns"]["N"]["value"]


# ── Helpers ────────────────────────────────────────────────────────────


def load_columns_spec() -> List[Dict]:
    """Ordered list of {index, name, header} dicts from columns.yaml.
    Retained for external tooling that imports this helper."""
    columns = _COLS_CFG["columns"]
    result = []
    for _letter, col_def in sorted(columns.items(), key=lambda x: x[1].get("index", 0)):
        result.append(
            {
                "index": col_def["index"],
                "name": col_def.get("name", ""),
                "header": col_def.get("header", col_def.get("name", "")),
            }
        )
    return result


def load_formula_templates() -> Dict[str, str]:
    """Load formula_templates from config/columns.yaml."""
    templates = _COLS_CFG.get("formula_templates", {})
    return {k: v.get("pattern", "") for k, v in templates.items()}


def _resolve_supplier_code(supplier: str) -> str:
    """Resolve short supplier code via config aliases, falling back to first token."""
    if not supplier:
        return ""
    aliases = load_invoice_formats().get("supplier_code_aliases", {})
    upper = supplier.upper()
    for needle, code in aliases.items():
        if needle in upper:
            return code
    supplier_str = str(supplier)
    tokens = supplier_str.split()
    return tokens[0] if tokens else ""


def _strip_intermediate_suffixes(stem: str) -> str:
    """Remove pipeline-intermediate stage suffixes from a filename stem."""
    for suffix in _FP["intermediate_suffixes"]:
        if stem.endswith(suffix):
            return stem[: -len(suffix)]
    return stem


def get_next_version_filename(output_path: str, input_path: str | None = None) -> str:
    """Auto-increment version number in filename based on sibling files."""
    output_path = os.path.abspath(output_path)
    output_dir = os.path.dirname(output_path)
    ext = os.path.splitext(output_path)[1] or _FP["extensions"]["xlsx"]

    if input_path:
        input_base = os.path.splitext(os.path.basename(input_path))[0]
        base_stem = _strip_intermediate_suffixes(input_base)
    else:
        name_without_ext = os.path.splitext(os.path.basename(output_path))[0]
        m = re.match(_VERSION_SUFFIX_RE, name_without_ext)
        base_stem = m.group(1) if m else name_without_ext

    existing_versions: List[int] = []
    if os.path.exists(output_dir):
        scan_pattern = re.compile(
            rf"^{re.escape(base_stem)}(?:_v(\d+))?{re.escape(ext)}$"
        )
        for filename in os.listdir(output_dir):
            match = scan_pattern.match(filename)
            if match:
                version_num = int(match.group(1)) if match.group(1) else 1
                existing_versions.append(version_num)

    if not existing_versions:
        new_filename = f"{base_stem}{ext}"
    else:
        next_version = max(existing_versions) + 1
        new_filename = f"{base_stem}_v{next_version}{ext}"

    return os.path.join(output_dir, new_filename)


# ── Main entry point ───────────────────────────────────────────────────


def run(
    input_path: str,
    output_path: str,
    config: Dict | None = None,
    context: Dict | None = None,
) -> Dict:
    """Generate Excel output from grouped data."""
    if not openpyxl:
        return {
            "status": _STATUS_ERROR,
            "error": _ERRS["openpyxl_missing_install_hint"],
        }

    if not input_path or not os.path.exists(input_path):
        return {"status": _STATUS_ERROR, "error": f"Input not found: {input_path}"}

    # Auto-increment version number in output filename
    original_input = context.get("input_file") if context else None
    output_path = get_next_version_filename(output_path, original_input or input_path)

    with open(input_path) as f:
        data = json.load(f)

    groups = data.get("groups", [])
    metadata = data.get("invoice_metadata", {})

    # Document type: context > grouped.json metadata > config default.
    # Hardcoding here would silently override the consignee-rule
    # resolution done by run.py::resolve_doc_type (e.g. Budget Marine → 7400-000).
    document_type = (
        (context or {}).get("document_type")
        or metadata.get("document_type")
        or _DT_CFG["default"]
    )

    if not groups:
        return {"status": _STATUS_ERROR, "error": _ERRS["no_groups"]}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _LABELS["defaults"]["sheet_title"]

    # Styles (all sourced from columns.yaml)
    styles = _COLS_CFG["styles"]
    sih = styles["single_invoice_header"]
    sig = styles["single_invoice_group"]
    sid = styles["single_invoice_detail"]
    svr = styles["verify"]
    svar = styles["variance"]
    sts = styles["totals_section"]

    header_fill = PatternFill(
        start_color=sih["fill_color"], end_color=sih["fill_color"], fill_type=_FILL_SOLID
    )
    header_font = Font(
        bold=sih["font_bold"], size=sih["font_size"], color=sih["font_color"]
    )
    group_fill = PatternFill(
        start_color=sig["fill_color"], end_color=sig["fill_color"], fill_type=_FILL_SOLID
    )
    group_font = Font(
        bold=sig["font_bold"], size=sig["font_size"], color=sig["font_color"]
    )
    detail_font = Font(size=sid["font_size"])
    bold_font = Font(bold=True)
    verify_font = Font(bold=svr["font_bold"], color=svr["font_color"])
    variance_font = Font(bold=svar["font_bold"], color=svar["font_color"])
    totals_fill = PatternFill(
        start_color=sts["fill_color"], end_color=sts["fill_color"], fill_type=_FILL_SOLID
    )

    # Headers from columns.yaml (specification-driven — no hardcoded fallback).
    for col_def in load_columns_spec() or []:
        cell = ws.cell(row=1, column=col_def["index"], value=col_def["header"])
        cell.font = header_font
        cell.fill = header_fill

    po_number = metadata.get("po_number", "")
    is_invoice = not po_number or po_number in _TOL["invalid_po_sentinels"]

    row_num = _FIRST_DATA_ROW
    group_row_nums: List[int] = []
    detail_row_nums: List[int] = []
    computed_group_total = 0.0
    computed_detail_total = 0.0

    category_map = load_hs_categories()["categories"]
    category_fallback = _LABELS["defaults"]["category_fallback"]
    group_label_tpl = _LABELS["defaults"]["group_label_count_tpl"]
    detail_indent = _LABELS["defaults"]["detail_row_indent"]

    for group in groups:
        tariff = group["tariff_code"]
        raw_category = group.get("category", category_fallback)
        from bl_xlsx_generator import _normalize_date, get_cet_category

        cet_desc = get_cet_category(tariff)
        category = category_map.get(tariff, cet_desc or str(raw_category))
        item_count = group["item_count"]
        category_label = group_label_tpl.format(category=category, n=item_count)

        sum_qty = group["sum_quantity"]
        sum_cost = group["sum_total_cost"]
        avg_cost = group["average_unit_cost"]

        group_row_nums.append(row_num)
        computed_group_total += sum_cost

        # Only FIRST group row per invoice gets Document Type
        if row_num == _FIRST_DATA_ROW:
            ws.cell(row_num, column=_COL_DOC_TYPE, value=document_type)
        else:
            ws.cell(row_num, column=_COL_DOC_TYPE, value=None)

        ws.cell(row_num, column=_COL_INVOICE_NUM, value=metadata.get("invoice_number", ""))
        ws.cell(row_num, column=_COL_DATE, value=_normalize_date(metadata.get("date", "")))
        ws.cell(row_num, column=_COL_CATEGORY, value=category)
        ws.cell(row_num, column=_COL_TARIFF, value=tariff)

        if not is_invoice:
            ws.cell(row_num, column=_COL_PO_ITEM_NUM, value=tariff)
            ws.cell(row_num, column=_COL_PO_ITEM_DESC, value=category_label)
        else:
            ws.cell(row_num, column=_COL_PO_ITEM_NUM, value=None)
            ws.cell(row_num, column=_COL_PO_ITEM_DESC, value=None)

        ws.cell(row_num, column=_COL_SUPPLIER_ITEM_NUM, value=tariff)
        ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC, value=category_label)
        ws.cell(row_num, column=_COL_QTY, value=sum_qty)
        ws.cell(row_num, column=_COL_UOM, value=category_label)
        ws.cell(row_num, column=_COL_CURRENCY, value=_CURRENCY_CODE)
        ws.cell(row_num, column=_COL_COST, value=avg_cost)
        ws.cell(row_num, column=_COL_TOTAL_COST, value=sum_cost)

        # Q=Stat Value, R=Variance as Excel formulas (formulas not hardcodes)
        ws.cell(row_num, column=_COL_TOTAL).value = f"=O{row_num}*K{row_num}"
        ws.cell(row_num, column=_COL_TOTALCOST_VS_TOTAL).value = (
            f"=P{row_num}-Q{row_num}"
        )
        ws.cell(row_num, column=_COL_GROUPBY, value=tariff)

        if row_num == _FIRST_DATA_ROW:
            total = metadata.get("total")
            if total:
                ws.cell(row_num, column=_COL_INVOICE_TOTAL, value=total)
            freight = metadata.get("freight", 0)
            if freight:
                ws.cell(row_num, column=_COL_FREIGHT, value=freight)
            # Column U: insurance OR customer-induced credits (x -1)
            insurance = metadata.get("insurance", 0)
            credits = metadata.get("credits", 0)
            col_u_value = insurance if insurance else (-credits if credits else 0)
            if col_u_value:
                ws.cell(row_num, column=_COL_INSURANCE, value=col_u_value)
            # V=Other Cost (tax, fees, service charges)
            other_cost = metadata.get("other_cost", 0) or metadata.get("tax", 0)
            if other_cost:
                ws.cell(row_num, column=_COL_OTHER_COST, value=other_cost)
            # W: Total Deduction = discount + free_shipping (both supplier-induced)
            discount = metadata.get("discount", 0) or 0
            free_shipping = metadata.get("free_shipping", 0) or 0
            total_deduction = discount + free_shipping
            if total_deduction:
                ws.cell(row_num, column=_COL_DEDUCTION, value=total_deduction)

            supplier = metadata.get("supplier", "")
            supplier_code = _resolve_supplier_code(supplier)
            if supplier_code:
                ws.cell(row_num, column=_COL_SUPPLIER_CODE, value=supplier_code)
            if supplier:
                ws.cell(row_num, column=_COL_SUPPLIER_NAME, value=supplier)
            sup_addr = metadata.get("supplier_address", "")
            if sup_addr:
                ws.cell(row_num, column=_COL_SUPPLIER_ADDR, value=sup_addr)
            country = metadata.get("country_code", "")
            if country:
                ws.cell(row_num, column=_COL_COUNTRY, value=country)

        for col in range(1, _NUM_COLS + 1):
            cell = ws.cell(row_num, column=col)
            cell.fill = group_fill
            cell.font = group_font
        for col in _CURRENCY_COLS_GROUP:
            ws.cell(row_num, column=col).number_format = _CURRENCY_FMT
        row_num += 1

        for item in group["items"]:
            detail_row_nums.append(row_num)
            # Quantity defaults to 1 when the format parser didn't map a
            # quantity field (most multiline receipts — AURORA, etc.).
            # qty=0 with a positive unit_cost/total_cost produces
            # Q=O*K=0 and spurious variance in R (TotalCost vs Total).
            item_qty = item.get("quantity", 1)
            item_cost = item.get("unit_cost", 0)
            item_total = item.get("total_cost", 0)
            # Back-derive unit_cost when only total_cost was extracted.
            if not item_cost and item_total and item_qty:
                item_cost = round(item_total / item_qty, 2)
            if item.get("billable", True):
                computed_detail_total += item_total

            ws.cell(row_num, column=_COL_TARIFF, value=tariff)
            ws.cell(
                row_num,
                column=_COL_SUPPLIER_ITEM_NUM,
                value=item.get("supplier_item") or item.get("sku", ""),
            )
            ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC, value=item.get("description", ""))
            ws.cell(row_num, column=_COL_QTY, value=item_qty)
            ws.cell(
                row_num,
                column=_COL_UOM,
                value=f"{detail_indent}{item.get('description', '')}",
            )
            ws.cell(row_num, column=_COL_CURRENCY, value=_CURRENCY_CODE)
            ws.cell(row_num, column=_COL_COST, value=item_cost)
            ws.cell(row_num, column=_COL_TOTAL_COST, value=item_total)

            ws.cell(row_num, column=_COL_TOTAL).value = f"=O{row_num}*K{row_num}"
            ws.cell(row_num, column=_COL_TOTALCOST_VS_TOTAL).value = (
                f"=P{row_num}-Q{row_num}"
            )

            for col in range(1, _NUM_COLS + 1):
                ws.cell(row_num, column=col).font = detail_font

            for col in _CURRENCY_COLS_DETAIL:
                ws.cell(row_num, column=col).number_format = _CURRENCY_FMT

            row_num += 1

    row_num += 1

    # --- Extract metadata values for cells S2, T2, U2, V2, W2 ---
    # These are the SAME values written to row 2 above. Python mirrors
    # what the Excel formulas will compute so we can report values in
    # the return dict and set font colors.
    invoice_total = metadata.get("total", 0) or 0       # -> S2
    freight = metadata.get("freight", 0) or 0           # -> T2
    insurance = (
        metadata.get("insurance", 0) or metadata.get("credits", 0) or 0
    )  # -> U2
    other_cost = metadata.get("other_cost", 0) or metadata.get("tax", 0) or 0  # -> V2
    discount = metadata.get("discount", 0) or 0
    free_shipping = metadata.get("free_shipping", 0) or 0
    total_deduction = discount + free_shipping          # -> W2

    def fill_summary_row(r: int) -> None:
        for c in range(1, _NUM_COLS + 1):
            ws.cell(r, column=c).fill = totals_fill

    group_verification = round(computed_group_total - computed_detail_total, 2)
    adjustments = freight + insurance + other_cost - total_deduction
    net_total = computed_group_total + adjustments
    variance_check = round(invoice_total - net_total, 2)

    formulas = load_formula_templates()
    totals_labels = _LABELS["totals"]

    # --- Totals Section (all formulas from spec, labels from config) ---

    # SUBTOTAL (GROUPED) — formula sums all group row P values.
    ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC, value=totals_labels["SUBTOTAL_GROUPED"])
    ws.cell(row_num, column=_COL_QTY, value=sum(g["sum_quantity"] for g in groups))
    group_p_refs = "+".join(f"P{r}" for r in group_row_nums)
    ws.cell(row_num, column=_COL_TOTAL_COST).value = (
        f"={group_p_refs}" if group_p_refs else 0
    )
    ws.cell(row_num, column=_COL_TOTAL_COST).number_format = _CURRENCY_FMT
    ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC).font = bold_font
    fill_summary_row(row_num)
    subtotal_grouped_row = row_num
    row_num += 1

    # SUBTOTAL (DETAILS) — formula sums all detail row P values.
    ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC, value=totals_labels["SUBTOTAL_DETAILS"])
    detail_p_refs = "+".join(f"P{r}" for r in detail_row_nums)
    ws.cell(row_num, column=_COL_TOTAL_COST).value = (
        f"={detail_p_refs}" if detail_p_refs else 0
    )
    ws.cell(row_num, column=_COL_TOTAL_COST).number_format = _CURRENCY_FMT
    ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC).font = bold_font
    fill_summary_row(row_num)
    subtotal_details_row = row_num
    row_num += 1

    # GROUP VERIFICATION — spec pattern from columns.yaml formula_templates.
    # All formula placeholders resolved via str.format() so the literal
    # placeholder names live only in columns.yaml, not in this module.
    ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC, value=totals_labels["GROUP_VERIFICATION"])
    formula_args = {
        "first_row": _FIRST_DATA_ROW,
        "subtotal_grouped_row": subtotal_grouped_row,
        "subtotal_details_row": subtotal_details_row,
    }
    gv_formula = formulas["group_verification"].format(**formula_args)
    ws.cell(row_num, column=_COL_TOTAL_COST).value = gv_formula
    ws.cell(row_num, column=_COL_TOTAL_COST).number_format = _CURRENCY_FMT
    ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC).font = verify_font
    ws.cell(row_num, column=_COL_TOTAL_COST).font = verify_font
    fill_summary_row(row_num)
    row_num += 1

    # ADJUSTMENTS — spec pattern from columns.yaml formula_templates.
    ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC, value=totals_labels["ADJUSTMENTS"])
    adj_formula = formulas["adjustments"].format(**formula_args)
    ws.cell(row_num, column=_COL_TOTAL_COST).value = adj_formula
    ws.cell(row_num, column=_COL_TOTAL_COST).number_format = _CURRENCY_FMT
    ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC).font = bold_font
    fill_summary_row(row_num)
    adjustments_row = row_num
    formula_args["adjustments_row"] = adjustments_row
    row_num += 1

    # NET TOTAL — spec pattern from columns.yaml formula_templates.
    ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC, value=totals_labels["NET_TOTAL"])
    nt_formula = formulas["net_total"].format(**formula_args)
    ws.cell(row_num, column=_COL_TOTAL_COST).value = nt_formula
    ws.cell(row_num, column=_COL_TOTAL_COST).number_format = _CURRENCY_FMT
    ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC).font = bold_font
    fill_summary_row(row_num)
    net_total_row = row_num
    formula_args["net_total_row"] = net_total_row
    row_num += 1

    # VARIANCE CHECK — spec pattern from columns.yaml formula_templates.
    ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC, value=totals_labels["VARIANCE_CHECK"])
    vc_formula = formulas["variance_check"].format(**formula_args)
    ws.cell(row_num, column=_COL_TOTAL_COST).value = vc_formula
    ws.cell(row_num, column=_COL_TOTAL_COST).number_format = _CURRENCY_FMT
    if abs(variance_check) > _TOL["variance_display_epsilon"]:
        ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC).font = variance_font
        ws.cell(row_num, column=_COL_TOTAL_COST).font = variance_font
    else:
        ws.cell(row_num, column=_COL_SUPPLIER_ITEM_DESC).font = verify_font
        ws.cell(row_num, column=_COL_TOTAL_COST).font = verify_font
    fill_summary_row(row_num)
    row_num += 1

    wb.save(output_path)

    return {
        "status": _STATUS_SUCCESS,
        "output": output_path,
        "total_groups": len(groups),
        "total_rows": row_num,
        "group_verification": group_verification,
        "variance_check": variance_check,
        "invoice_total": invoice_total,
        "freight": freight,
        "insurance": insurance,
        "other_cost": other_cost,
        "net_total": net_total,
    }


def run_split_declarations(
    input_path: str,
    output_dir: str,
    config: Dict | None = None,
    context: Dict | None = None,
) -> Dict:
    """
    Generate MULTIPLE Excel files - one per tariff code group.

    Each group becomes its own simplified declaration (XLSX).
    Used when one invoice needs to be split into multiple declarations.
    """
    if not openpyxl:
        return {"status": _STATUS_ERROR, "error": _ERRS["openpyxl_missing"]}

    if not input_path or not os.path.exists(input_path):
        return {"status": _STATUS_ERROR, "error": f"Input not found: {input_path}"}

    with open(input_path) as f:
        data = json.load(f)

    groups = data.get("groups", [])
    metadata = data.get("invoice_metadata", {})

    if not groups:
        return {"status": _STATUS_ERROR, "error": _ERRS["no_groups"]}

    os.makedirs(output_dir, exist_ok=True)

    defaults = _LABELS["defaults"]
    invoice_number = metadata.get("invoice_number", defaults["invoice_number_fallback"])
    # Clean invoice number for filename
    clean_inv = "".join(
        c if c.isalnum() or c in "._-" else "_" for c in str(invoice_number)
    )

    outputs: List[Dict[str, Any]] = []

    for _i, group in enumerate(groups, 1):
        tariff_code = group.get("tariff_code", defaults["tariff_fallback"])
        category = group.get("category", defaults["category_fallback"])

        # Create single-group data for this declaration
        single_group_data = {
            "invoice_metadata": metadata,
            "groups": [group],
            "total_groups": 1,
            "total_items": group.get("item_count", len(group.get("items", []))),
        }

        # Generate filename: invoice_tariffcode.xlsx
        output_filename = defaults["split_filename_tpl"].format(
            invoice=clean_inv, tariff=tariff_code
        )
        output_path = os.path.join(output_dir, output_filename)

        # Write temp JSON and run normal generator
        import tempfile

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=_FP["extensions"]["json"], delete=False
        ) as tmp:
            json.dump(single_group_data, tmp)
            tmp_path = tmp.name

        try:
            split_context = {
                "input_file": defaults["split_input_hint_tpl"].format(
                    invoice=clean_inv, tariff=tariff_code
                )
            }
            result = run(tmp_path, output_path, config, split_context)
            if result.get("status") == _STATUS_SUCCESS:
                actual_path = result.get("output", output_path)
                outputs.append(
                    {
                        "path": actual_path,
                        "tariff_code": tariff_code,
                        "category": category,
                        "item_count": group.get("item_count", 0),
                        "total_cost": group.get("sum_total_cost", 0),
                    }
                )
        finally:
            os.unlink(tmp_path)

    return {
        "status": _STATUS_SUCCESS,
        "outputs": outputs,
        "total_declarations": len(outputs),
        "invoice_number": invoice_number,
    }
