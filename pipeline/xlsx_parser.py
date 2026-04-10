#!/usr/bin/env python3
"""
XLSX Parser - Extract data from existing XLSX files for reprocessing.

This module parses processed XLSX invoice files to extract:
- Invoice metadata (from first group row)
- Line items (from detail rows)

The extracted data can be fed back into the grouping and generation stages
to fix issues or regenerate the file with updates.
"""

import json
import os
from typing import Any, Dict, List, Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None


def run(input_path: str, output_path: str, config: Dict = None, context: Dict = None) -> Dict:
    """
    Extract invoice data from an existing XLSX file.

    Args:
        input_path: Path to the XLSX file
        output_path: Path to write extracted JSON

    Returns:
        Dict with status, items, and metadata
    """
    if not openpyxl:
        return {'status': 'error', 'error': 'openpyxl not installed'}

    if not input_path or not os.path.exists(input_path):
        return {'status': 'error', 'error': f'File not found: {input_path}'}

    if not input_path.lower().endswith(('.xlsx', '.xls')):
        return {'status': 'error', 'error': f'Not an XLSX file: {input_path}'}

    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
        ws = wb.active

        # Extract metadata from first data row (row 2)
        metadata = extract_metadata(ws)

        # Extract all items (detail rows have indented descriptions or blank GroupBy)
        items = extract_items(ws)

        result = {
            'status': 'success',
            'invoice_metadata': metadata,
            'items': items,
            'total_items': len(items),
            'source_file': input_path,
        }

        # Write output JSON if path provided
        if output_path:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            with open(output_path, 'w') as f:
                json.dump(result, f, indent=2)

        return result

    except Exception as e:
        return {'status': 'error', 'error': str(e)}


def extract_metadata(ws) -> Dict:
    """Extract invoice metadata from the XLSX."""
    metadata = {}

    # Row 2 is typically the first group row with invoice-level data
    # Column mappings (1-indexed):
    # C (3) = Invoice #
    # D (4) = Date
    # S (19) = Invoice Total
    # T (20) = Freight
    # U (21) = Insurance
    # V (22) = Other Cost
    # W (23) = Deductions/Discount
    # Z (26) = Supplier Code
    # AA (27) = Supplier Name
    # AB (28) = Supplier Address
    # AC (29) = Country Code

    row = 2

    metadata['invoice_number'] = ws.cell(row, 3).value
    metadata['date'] = ws.cell(row, 4).value
    metadata['total'] = safe_float(ws.cell(row, 19).value)
    metadata['freight'] = safe_float(ws.cell(row, 20).value)
    metadata['insurance'] = safe_float(ws.cell(row, 21).value)
    metadata['other_cost'] = safe_float(ws.cell(row, 22).value)
    metadata['discount'] = safe_float(ws.cell(row, 23).value)
    metadata['supplier_code'] = ws.cell(row, 26).value
    metadata['supplier'] = ws.cell(row, 27).value
    metadata['supplier_address'] = ws.cell(row, 28).value
    metadata['country_code'] = ws.cell(row, 29).value

    # Clean up None values
    metadata = {k: v for k, v in metadata.items() if v is not None}

    return metadata


def extract_items(ws) -> List[Dict]:
    """
    Extract line items from the XLSX.

    Detail rows are identified by:
    - Column AH (34) is blank (no GroupBy value)
    - Column L (12) has indented description (starts with spaces)
    - Column A (1) is blank (no Document Type)
    """
    items = []

    for row in range(2, ws.max_row + 1):
        # Check if this is a detail row (blank GroupBy column AK)
        group_by = ws.cell(row, 34).value
        doc_type = ws.cell(row, 1).value
        description = ws.cell(row, 12).value

        # Skip header row, group rows (have GroupBy), and totals section
        if group_by:
            continue  # This is a group row

        if description and isinstance(description, str):
            # Check if it's a totals row
            desc_upper = description.upper().strip()
            if any(label in desc_upper for label in [
                'SUBTOTAL', 'VERIFICATION', 'ADJUSTMENTS',
                'NET TOTAL', 'VARIANCE', 'GRAND'
            ]):
                continue  # Skip totals section

            # This is likely a detail row
            # Get the tariff code from column F (might need to look at parent group)
            tariff = ws.cell(row, 6).value

            # If tariff is blank on detail row, find it from the nearest group row above
            if not tariff:
                tariff = find_parent_tariff(ws, row)

            sku = ws.cell(row, 9).value  # Column I - Supplier Item #

            item = {
                'index': len(items),
                'sku': sku,
                'description': description.strip() if description else '',
                'quantity': safe_float(ws.cell(row, 11).value),  # Column K
                'unit_cost': safe_float(ws.cell(row, 15).value),  # Column O
                'total_cost': safe_float(ws.cell(row, 16).value),  # Column P
                'classification': {
                    'code': str(tariff) if tariff else 'UNKNOWN',
                    'confidence': 1.0 if tariff else 0.0,
                }
            }

            # Detect bundle items by SKU prefix (same logic as item_parser.py)
            bundle_info = detect_bundle(sku, desc_upper)
            if bundle_info:
                item['is_bundle'] = True
                item['bundle_type'] = bundle_info['type']
                item['billable'] = bundle_info['billable']

            # Only add if it has meaningful data
            if item['description'] and (item['quantity'] or item['total_cost']):
                items.append(item)

    return items


def detect_bundle(sku: str, desc_upper: str) -> Optional[Dict]:
    """
    Detect if an item is a bundled item (set, display, tester).

    Bundled items are:
    - ST-*: Starter kits/sets - contain same items as individual SKUs, NOT billable
    - DP-*: Display units - promotional displays
    - TST-*: Tester sets - samples for display
    - T-*: Individual testers

    Returns bundle info dict or None if not a bundle.
    """
    if not sku:
        return None

    sku_upper = str(sku).upper()

    # Check SKU prefixes for bundle types
    bundle_prefixes = {
        'ST-': {'type': 'starter_kit', 'billable': False},  # Starter kits duplicate individual items
        'DP-': {'type': 'display', 'billable': True},       # Displays are usually $0 anyway
        'TST-': {'type': 'tester_set', 'billable': True},   # Tester sets are usually $0
        'T-': {'type': 'tester', 'billable': True},         # Individual testers
    }

    for prefix, info in bundle_prefixes.items():
        if sku_upper.startswith(prefix):
            return info.copy()

    # Check description patterns for sets/kits not caught by SKU prefix
    if desc_upper:
        set_patterns = ['SET FOR ', 'KIT FOR ', 'STARTER KIT', 'DISPLAY SET']
        for pattern in set_patterns:
            if pattern in desc_upper:
                return {'type': 'set', 'billable': False}

    return None


def find_parent_tariff(ws, detail_row: int) -> Optional[str]:
    """Find the tariff code from the nearest group row above this detail row."""
    for row in range(detail_row - 1, 1, -1):
        group_by = ws.cell(row, 34).value  # Column AK
        if group_by:
            return str(group_by)
        tariff = ws.cell(row, 6).value  # Column F
        if tariff and str(tariff).isdigit() and len(str(tariff)) == 8:
            return str(tariff)
    return None


def safe_float(value) -> float:
    """Safely convert value to float."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        result = run(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
        print(json.dumps(result, indent=2))
