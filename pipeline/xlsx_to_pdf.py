#!/usr/bin/env python3
"""Convert an XLSX worksheet to a landscape PDF preserving colours and layout.

The PDF is sized dynamically so every column fits on a single page at a
readable font size.  Cell background colours and bold fonts are preserved.
Formulas are evaluated in-process so computed values appear instead of raw
formula strings.

Usage:
    from pipeline.xlsx_to_pdf import convert_xlsx_to_pdf
    pdf_path = convert_xlsx_to_pdf("input.xlsx")           # -> "input-WorkSheet.pdf"
    pdf_path = convert_xlsx_to_pdf("input.xlsx", "out.pdf") # explicit path
"""

import os
import re
from typing import Dict, Optional, Tuple

import fitz  # PyMuPDF
import openpyxl
from openpyxl.utils import get_column_letter


# ── Simple formula evaluator ────────────────────────────────────────

def _col_letter_to_idx(letter: str) -> int:
    """Convert column letter(s) to 0-based index. A=0, B=1, ..., AK=36."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result - 1


def _parse_cell_ref(ref: str) -> Tuple[int, int]:
    """Parse 'P2' -> (row_0based, col_0based)."""
    m = re.match(r'([A-Z]+)(\d+)', ref.upper())
    if not m:
        raise ValueError(f"Bad cell ref: {ref}")
    return int(m.group(2)) - 1, _col_letter_to_idx(m.group(1))


def _evaluate_formulas(ws) -> Dict[Tuple[int, int], float]:
    """Evaluate all formula cells in the worksheet.

    Handles common patterns:
      =O2*K2, =P2-Q2, =P2+P4+P6, =SUM(P2:P11)-P12,
      =(T2+U2+V2-W2), =S2, =S2-P14
    """
    total_rows = ws.max_row
    total_cols = ws.max_column

    # First pass: collect raw numeric values
    values: Dict[Tuple[int, int], Optional[float]] = {}
    formulas: Dict[Tuple[int, int], str] = {}

    for r in range(1, total_rows + 1):
        for c in range(1, total_cols + 1):
            val = ws.cell(row=r, column=c).value
            key = (r - 1, c - 1)
            if val is None:
                values[key] = None
            elif isinstance(val, (int, float)):
                values[key] = float(val)
            elif isinstance(val, str) and val.startswith("="):
                formulas[key] = val
                values[key] = None
            else:
                values[key] = None  # text cells

    def _get(ref_str: str) -> float:
        """Resolve a cell reference to its numeric value."""
        r, c = _parse_cell_ref(ref_str)
        v = values.get((r, c))
        return v if v is not None else 0.0

    def _eval_formula(formula: str) -> Optional[float]:
        """Evaluate a single formula string."""
        f = formula.lstrip("=").strip()
        # Remove outer parens: (T2+U2+V2-W2) -> T2+U2+V2-W2
        if f.startswith("(") and f.endswith(")"):
            f = f[1:-1]

        # SUM(range)-cell: =SUM(P2:P11)-P12
        m = re.match(
            r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)\s*([+-])\s*([A-Z]+\d+)$',
            f, re.IGNORECASE,
        )
        if m:
            col_start = _col_letter_to_idx(m.group(1))
            row_start = int(m.group(2)) - 1
            col_end = _col_letter_to_idx(m.group(3))
            row_end = int(m.group(4)) - 1
            total = 0.0
            for rr in range(row_start, row_end + 1):
                for cc in range(col_start, col_end + 1):
                    v = values.get((rr, cc))
                    if v is not None:
                        total += v
            op = m.group(5)
            adj = _get(m.group(6))
            return total + adj if op == "+" else total - adj

        # SUM(range): =SUM(P2:P11)
        m = re.match(
            r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$', f, re.IGNORECASE,
        )
        if m:
            col_start = _col_letter_to_idx(m.group(1))
            row_start = int(m.group(2)) - 1
            col_end = _col_letter_to_idx(m.group(3))
            row_end = int(m.group(4)) - 1
            total = 0.0
            for rr in range(row_start, row_end + 1):
                for cc in range(col_start, col_end + 1):
                    v = values.get((rr, cc))
                    if v is not None:
                        total += v
            return total

        # Simple cell ref: =S2
        if re.match(r'^[A-Z]+\d+$', f, re.IGNORECASE):
            return _get(f)

        # Binary op: =O2*K2, =P2-Q2, =P2+P12
        m = re.match(r'^([A-Z]+\d+)\s*([+\-*/])\s*([A-Z]+\d+)$', f, re.IGNORECASE)
        if m:
            a, op, b = _get(m.group(1)), m.group(2), _get(m.group(3))
            if op == "+":
                return a + b
            elif op == "-":
                return a - b
            elif op == "*":
                return a * b
            elif op == "/" and b != 0:
                return a / b
            return 0.0

        # Multi-term addition/subtraction: P2+P4+P6+P8+P10 or T2+U2+V2-W2
        refs = re.findall(r'([+-]?)([A-Z]+\d+)', f, re.IGNORECASE)
        if refs and len(refs) >= 2:
            # Check that the full string is just refs with +/-
            rebuilt = ""
            for sign, ref in refs:
                rebuilt += sign + ref
            # Handle first ref without sign (implicit +)
            if not refs[0][0]:
                rebuilt = refs[0][1] + rebuilt[len(refs[0][1]):]
            if rebuilt.replace(" ", "") == f.replace(" ", ""):
                total = 0.0
                for sign, ref in refs:
                    v = _get(ref)
                    if sign == "-":
                        total -= v
                    else:
                        total += v
                return total

        return None  # can't evaluate

    # Iterative evaluation (formulas may depend on other formulas)
    max_passes = 5
    for _ in range(max_passes):
        remaining = {}
        for key, formula in formulas.items():
            result = _eval_formula(formula)
            if result is not None:
                values[key] = result
            else:
                remaining[key] = formula
        formulas = remaining
        if not formulas:
            break

    return values


def _hex_to_rgb(hex_str: str):
    """Convert an openpyxl theme/hex colour to (r, g, b) floats 0-1."""
    if not hex_str or hex_str == "00000000" or len(hex_str) < 6:
        return None
    h = hex_str[-6:]
    try:
        r = int(h[0:2], 16) / 255.0
        g = int(h[2:4], 16) / 255.0
        b = int(h[4:6], 16) / 255.0
        return (r, g, b)
    except ValueError:
        return None


def _format_number(val: float) -> str:
    """Format a number for display."""
    if val == int(val) and abs(val) < 1e12:
        return str(int(val))
    return f"{val:,.2f}"


def convert_xlsx_to_pdf(
    xlsx_path: str,
    output_path: Optional[str] = None,
    font_size: float = 8.0,
    padding: float = 4,
    min_col_width: float = 28,
    max_col_width: float = 280,
    max_desc_width: float = 360,
) -> str:
    """Convert an XLSX file to a single-page landscape PDF.

    Args:
        xlsx_path: Path to the .xlsx file.
        output_path: Output PDF path.  Defaults to same directory,
            ``{stem}-WorkSheet.pdf``.
        font_size: Base font size in points.
        padding: Cell padding in points.
        min_col_width: Minimum column width.
        max_col_width: Maximum column width for most columns.
        max_desc_width: Maximum width for description columns (J).

    Returns:
        Absolute path to the generated PDF.
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    total_rows = ws.max_row
    total_cols = ws.max_column

    # ── Evaluate all formulas ─────────────────────────────────
    computed = _evaluate_formulas(ws)

    # ── Collect cell data ─────────────────────────────────────
    fontname = "helv"
    bold_fontname = "hebo"

    grid = []  # grid[row][col] = (display_text, fill_rgb, is_bold, font_rgb)
    for r in range(1, total_rows + 1):
        row_data = []
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            key = (r - 1, c - 1)

            # Display value
            raw = cell.value
            if raw is None:
                display = ""
            elif isinstance(raw, str) and raw.startswith("="):
                # Use computed value
                cv = computed.get(key)
                if cv is not None:
                    display = _format_number(cv)
                else:
                    display = ""  # hide unresolvable formulas
            elif isinstance(raw, (int, float)):
                display = _format_number(float(raw))
            else:
                display = str(raw)

            # Fill colour
            fill_rgb = None
            if cell.fill and cell.fill.fgColor and cell.fill.fill_type == "solid":
                rgb_val = cell.fill.fgColor.rgb
                if isinstance(rgb_val, str):
                    fill_rgb = _hex_to_rgb(rgb_val)

            # Font
            is_bold = bool(cell.font and cell.font.bold)
            font_rgb = None
            if cell.font and cell.font.color and cell.font.color.rgb:
                rgb_val = cell.font.color.rgb
                if isinstance(rgb_val, str):
                    font_rgb = _hex_to_rgb(rgb_val)

            row_data.append((display, fill_rgb, is_bold, font_rgb))
        grid.append(row_data)

    # ── Determine column widths ───────────────────────────────
    col_widths = []
    for c in range(total_cols):
        max_w = min_col_width
        for r in range(len(grid)):
            text = grid[r][c][0]
            if not text:
                continue
            is_bold = grid[r][c][2]
            fn = bold_fontname if is_bold else fontname
            tw = fitz.get_text_length(text, fontname=fn, fontsize=font_size)
            max_w = max(max_w, tw + padding * 2)

        # Cap width
        cap = max_desc_width if c == 9 else max_col_width  # col J = index 9
        col_widths.append(min(max_w, cap))

    row_height = font_size + padding * 2 + 2

    # ── Page dimensions ───────────────────────────────────────
    margin = 24
    table_width = sum(col_widths)
    table_height = total_rows * row_height
    page_width = table_width + margin * 2
    page_height = table_height + margin * 2

    # Minimum page height for readability
    page_height = max(page_height, 600)

    # ── Create PDF ────────────────────────────────────────────
    doc = fitz.open()
    page = doc.new_page(width=page_width, height=page_height)

    # Draw cells
    y = margin
    for r in range(len(grid)):
        x = margin
        for c in range(total_cols):
            display, fill_rgb, is_bold, font_rgb = grid[r][c]
            cw = col_widths[c]

            rect = fitz.Rect(x, y, x + cw, y + row_height)

            # Fill
            if fill_rgb:
                page.draw_rect(rect, color=None, fill=fill_rgb)

            # Border (thin grey)
            page.draw_rect(rect, color=(0.75, 0.75, 0.75), width=0.3)

            # Text
            if display:
                fn = bold_fontname if is_bold else fontname
                fc = font_rgb or (0, 0, 0)

                # Truncate text if too wide for cell
                tw = fitz.get_text_length(display, fontname=fn, fontsize=font_size)
                if tw > cw - padding * 2:
                    while len(display) > 1:
                        display = display[:-1]
                        tw = fitz.get_text_length(
                            display + "…", fontname=fn, fontsize=font_size
                        )
                        if tw <= cw - padding * 2:
                            display += "…"
                            break

                text_y = y + padding + font_size
                page.insert_text(
                    fitz.Point(x + padding, text_y),
                    display,
                    fontname=fn,
                    fontsize=font_size,
                    color=fc,
                )

            x += cw
        y += row_height

    # ── Save ──────────────────────────────────────────────────
    if not output_path:
        stem = os.path.splitext(xlsx_path)[0]
        output_path = stem + "-WorkSheet.pdf"

    doc.save(output_path)
    doc.close()
    wb.close()

    return os.path.abspath(output_path)


def generate_worksheet_pdf(xlsx_path: str, waybill: str) -> str:
    """Generate a worksheet PDF named ``{waybill}-WorkSheet.pdf`` next to the XLSX.

    This is the main entry point called by the email sender.
    """
    out_dir = os.path.dirname(xlsx_path)
    pdf_name = f"{waybill}-WorkSheet.pdf"
    pdf_path = os.path.join(out_dir, pdf_name)
    return convert_xlsx_to_pdf(xlsx_path, pdf_path)


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python xlsx_to_pdf.py <input.xlsx> [output.pdf]")
        sys.exit(1)
    out = convert_xlsx_to_pdf(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
    print(f"Generated: {out}")
