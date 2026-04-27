"""
Post-generation XLSX validation and LLM-assisted auto-fix.

Detects and fixes:
  1. Missing tariff codes (col F empty/placeholder)
  2. Invalid tariff codes (not 8-digit or not in CET valid codes)
  3. Zero-value items (qty>0 but $0 price, not backordered)
  4. Variance (InvoiceTotal != calculated net total)

Usage:
    from xlsx_validator import validate_and_fix
    summary = validate_and_fix(results, base_dir)
"""

import json
import logging
import os
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)

# ── Column constants (match bl_xlsx_generator.py / columns.yaml) ───────────
COL_TARIFF = 6         # magic-ok: columns.yaml TariffCode column F (6)
COL_SUPP_DESC = 10     # magic-ok: columns.yaml Supplier Item Description column J (10)
COL_QTY = 11           # magic-ok: columns.yaml Quantity column K (11)
COL_UNIT_PRICE = 12    # magic-ok: columns.yaml Per Unit column L (12)
COL_COST = 15          # magic-ok: columns.yaml Cost column O (15)
COL_TOTAL_COST = 16    # magic-ok: columns.yaml Total Cost column P (16)
COL_INV_TOTAL = 19     # magic-ok: columns.yaml InvoiceTotal column S (19)
COL_FREIGHT = 20       # magic-ok: columns.yaml Freight column T (20)
COL_INSURANCE = 21     # magic-ok: columns.yaml Insurance column U (21)
COL_TAX = 22           # magic-ok: columns.yaml Tax column V (22)
COL_DEDUCTION = 23     # magic-ok: columns.yaml Deduction column W (23)
COL_PACKAGES = 24      # magic-ok: columns.yaml Packages column X (24)
COL_GROUPBY = 37       # magic-ok: columns.yaml GroupBy column AK (37)

# Issue type constants (internal enum values — used as dict keys and compared
# as strings throughout; keep in sync with downstream consumers).
MISSING_TARIFF = 'MISSING_TARIFF'    # magic-ok: internal issue-type enum
INVALID_TARIFF = 'INVALID_TARIFF'    # magic-ok: internal issue-type enum
ZERO_VALUE = 'ZERO_VALUE'            # magic-ok: internal issue-type enum
VARIANCE = 'VARIANCE'                # magic-ok: internal issue-type enum
PACKAGE_MISMATCH = 'PACKAGE_MISMATCH' # magic-ok: internal issue-type enum
DUPLICATE_FILE = 'DUPLICATE_FILE'    # magic-ok: internal issue-type enum
COLUMN_SPEC = 'COLUMN_SPEC'          # magic-ok: internal issue-type enum

# ── Tariff LLM prompt ─────────────────────────────────────────────────────
# Loaded from prompts/xlsx_validator_tariff_system.txt so long LLM-instruction
# text stays out of Python source (same pattern as variance_fixer).
from pathlib import Path as _Path
_PROMPT_DIR = _Path(__file__).resolve().parent.parent / "prompts"   # magic-ok: top-level prompts/ directory (sibling of pipeline/)
TARIFF_SYSTEM_PROMPT = (_PROMPT_DIR / "xlsx_validator_tariff_system.txt").read_text(encoding="utf-8")   # magic-ok: filename of LLM system prompt


# ── Label sets used by _detect_issues to skip formula/summary rows ─────────
# Labels are the exact strings written into the description column of
# generated XLSX files — bl_xlsx_generator uses the title-case set, the
# current xlsx_generator uses the upper-case set. Combined (multi-invoice)
# files contain both, plus the GRAND/Grand variants.
_FORMULA_LABELS = {
    'Subtotal', 'Adjustments', 'Net Total', 'Variance Check',                  # magic-ok: xlsx_labels combiner row labels (title-case)
    'Subtotal Grouped', 'Subtotal Details', 'Group Verification',              # magic-ok: xlsx_labels combiner row labels (title-case)
    'Grand Subtotal', 'Grand Adjustments', 'Grand Net Total',                  # magic-ok: xlsx_labels combiner grand-row labels
    'Grand Invoice Total', 'Grand Variance',                                   # magic-ok: xlsx_labels combiner grand-row labels
    'SUBTOTAL (GROUPED)', 'SUBTOTAL (DETAILS)', 'GROUP VERIFICATION',          # magic-ok: xlsx_labels totals-section upper-case labels
    'ADJUSTMENTS', 'NET TOTAL', 'INVOICE TOTAL', 'VARIANCE CHECK',             # magic-ok: xlsx_labels totals-section upper-case labels
    'TOTAL INTERNAL FREIGHT', 'TOTAL INSURANCE',                               # magic-ok: xlsx_labels totals-section upper-case labels
    'TOTAL OTHER COST', 'TOTAL DEDUCTION',                                     # magic-ok: xlsx_labels totals-section upper-case labels
    'GRAND SUBTOTAL (GROUPED)', 'GRAND SUBTOTAL (DETAILS)',                    # magic-ok: combined-XLSX grand-section upper-case labels
    'GRAND ADJUSTMENTS', 'GRAND NET TOTAL', 'GRAND INVOICE TOTAL',             # magic-ok: combined-XLSX grand-section upper-case labels
    'GRAND VARIANCE CHECK', 'GRAND VARIANCE',                                  # magic-ok: combined-XLSX grand-section upper-case labels
}

# Duty estimation section labels — these are informational rows, not item data.
_DUTY_ESTIMATION_PREFIXES = (
    'DUTY ESTIMATION', 'CIF ', 'CET ', 'CSC ', 'VAT ',                         # magic-ok: xlsx_labels duty-section row-label prefixes
    'ESTIMATED TOTAL', 'CLIENT DECLARED', 'DUTY VARIANCE',                     # magic-ok: xlsx_labels duty-section row-label prefixes
    'IMPLIED CET', 'IMPLIED CIF', 'IMPLIED ITEMS',                             # magic-ok: xlsx_labels duty-section reverse-calc row-label prefixes
    'Effective Duty', 'CET MISMATCH', 'ITEM ALLOCATION', '\u26a0',              # magic-ok: xlsx_labels duty-section row-label prefixes (warn sign)
    '\u2514',                                                                  # magic-ok: └ duty-breakdown sub-row prefix
    'REFERENCE SUBTOTAL', 'REFERENCE FREIGHT', 'REFERENCE INSURANCE',          # magic-ok: xlsx_labels reference-section row-label prefixes
    'REFERENCE OTHER COST', 'REFERENCE DEDUCTION', 'REFERENCE ADJUSTMENTS',    # magic-ok: xlsx_labels reference-section row-label prefixes
    'REFERENCE NET TOTAL', 'COMBINED ', 'FULL INVOICE', 'FULL ADJUSTMENTS',    # magic-ok: xlsx_labels reference-section row-label prefixes
    'Items on other',                                                          # magic-ok: xlsx_labels reference.default_label prefix
)


def _num(v):
    """Safe numeric conversion."""
    return v if isinstance(v, (int, float)) else 0


# ── Main entry point ───────────────────────────────────────────────────────

def validate_and_fix(results: list, base_dir: str = '.', bl_alloc=None,
                     manifest_meta: dict = None) -> dict:
    """
    Validate all generated XLSX files and auto-fix issues via LLM.

    Args:
        results: List of InvoiceResult objects from the pipeline
        base_dir: Project base directory (for loading CET codes, etc.)
        bl_alloc: BLAllocation object (has .packages as total BL packages)
        manifest_meta: Manifest metadata dict (has 'packages' key)

    Returns:
        Summary dict: {total_issues, fixed, unfixed, per_file: [...]}
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not available — skipping XLSX validation")
        return {'total_issues': 0, 'fixed': 0, 'unfixed': 0, 'per_file': []}

    if not results:
        return {'total_issues': 0, 'fixed': 0, 'unfixed': 0, 'per_file': []}

    all_file_results = []
    total_issues = 0
    total_fixed = 0
    total_unfixed = 0

    for r in results:
        xlsx_path = r.xlsx_path
        if not xlsx_path or not os.path.exists(xlsx_path):
            continue

        fname = os.path.basename(xlsx_path)
        supplier = r.supplier_info.get('name', '') if hasattr(r, 'supplier_info') else ''

        try:
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
        except Exception as e:
            logger.warning(f"Cannot open {fname}: {e}")
            continue

        # Phase 1: Detect issues
        issues = _detect_issues(ws, base_dir)

        # For combined multi-invoice XLSX, each block was variance-fixed
        # individually in invoice_processor before combining. The grand
        # variance computed by _detect_issues is not a per-invoice target
        # that variance_fixer can act on, and aggregating has known edge
        # cases with the ADJUSTMENTS-correction parser. Replace the global
        # variance check with a per-block check that only flags blocks
        # with residual variance.
        if getattr(r, '_combined_pdf_paths', None):
            issues = [i for i in issues if i['type'] != VARIANCE]
            block_variance_issues = _detect_combined_block_variance(ws)
            issues.extend(block_variance_issues)

        if not issues:
            all_file_results.append({
                'file': fname,
                'supplier': supplier,
                'issues': [],
                'fixes': [],
                'fixed': 0,
                'unfixed': 0,
            })
            wb.close()
            continue

        total_issues += len(issues)
        fixes = []

        # Phase 2a: Fix tariff issues
        tariff_issues = [i for i in issues if i['type'] in (MISSING_TARIFF, INVALID_TARIFF)]
        if tariff_issues:
            tariff_fixes = _fix_tariff_issues(ws, tariff_issues, r, base_dir)
            fixes.extend(tariff_fixes)
            wb.save(xlsx_path)

        wb.close()

        # Phase 2b: Fix variance issues (uses its own wb open/save)
        # Skip for combined multi-invoice XLSX: variance_fixer is designed
        # for single-invoice files and will scribble corrections into the
        # wrong block. Per-invoice variance fixing already ran in
        # invoice_processor before the files were combined.
        is_combined = bool(getattr(r, '_combined_pdf_paths', None))
        # Variance is left visible — do NOT auto-absorb into ADJUSTMENTS.
        # The broker needs to see the real discrepancy.

        # Phase 3: Re-validate
        try:
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            remaining = _detect_issues(ws, base_dir)
            if getattr(r, '_combined_pdf_paths', None):
                remaining = [i for i in remaining if i['type'] != VARIANCE]
                remaining.extend(_detect_combined_block_variance(ws))
            wb.close()
        except Exception:
            remaining = issues  # couldn't re-check

        fixed_count = len(issues) - len(remaining)
        unfixed_count = len(remaining)
        total_fixed += fixed_count
        total_unfixed += unfixed_count

        all_file_results.append({
            'file': fname,
            'supplier': supplier,
            'issues': issues,
            'fixes': fixes,
            'remaining': remaining,
            'fixed': fixed_count,
            'unfixed': unfixed_count,
        })

    # Phase 4: Cross-file package total check
    pkg_check = _check_package_totals(results, bl_alloc, manifest_meta)

    # Phase 4b: Duplicate file check
    dup_check = _check_duplicates(results)

    # Phase 5: Print report
    _print_report(all_file_results, pkg_check, dup_check)

    dup_issues = len(dup_check.get('duplicates', []))
    return {
        'total_issues': total_issues + (1 if pkg_check.get('mismatch') else 0) + dup_issues,
        'fixed': total_fixed,
        'unfixed': total_unfixed + (1 if pkg_check.get('mismatch') else 0) + dup_issues,
        'per_file': all_file_results,
        'package_check': pkg_check,
        'duplicate_check': dup_check,
    }


# ── Issue detection ────────────────────────────────────────────────────────

def _detect_issues(ws, base_dir: str) -> list:
    """
    Scan XLSX data rows for all issue types.
    Skips formula/summary rows (supports combined XLSX with per-invoice sections).
    """
    issues = []
    cet_codes = _load_cet_codes(base_dir)

    FORMULA_LABELS = _FORMULA_LABELS
    DUTY_ESTIMATION_PREFIXES = _DUTY_ESTIMATION_PREFIXES

    for row in range(2, ws.max_row + 1):
        # Skip formula rows and label rows (combined XLSX has these between invoices)
        tc_val = ws.cell(row, COL_TOTAL_COST).value
        if isinstance(tc_val, str) and tc_val.startswith('='):
            continue
        # Check both column J (COL_SUPP_DESC) and column L for summary labels
        label_j = ws.cell(row, COL_SUPP_DESC).value
        label_l = ws.cell(row, COL_UNIT_PRICE).value
        if isinstance(label_j, str) and label_j.strip() in FORMULA_LABELS:
            continue
        if isinstance(label_l, str) and label_l.strip() in FORMULA_LABELS:
            continue
        # Skip invoice-level metadata rows (notes, totals appended by parser)
        if isinstance(label_j, str) and label_j.strip().startswith('INVOICE NOTES'):   # magic-ok: xlsx_labels metadata row prefix
            continue
        # Skip duty estimation section rows (informational, not item data)
        if isinstance(label_j, str) and any(label_j.strip().startswith(p) for p in DUTY_ESTIMATION_PREFIXES):
            continue

        tariff = str(ws.cell(row, COL_TARIFF).value or '').strip()
        desc = str(ws.cell(row, COL_SUPP_DESC).value or '').strip()
        qty = _num(ws.cell(row, COL_QTY).value)
        unit_price = _num(ws.cell(row, COL_COST).value)
        total_cost = _num(ws.cell(row, COL_TOTAL_COST).value)

        # Skip rows with no description (empty/padding rows)
        if not desc:
            continue

        # 1. MISSING_TARIFF
        if not tariff or tariff == 'None' or tariff == '00000000':   # magic-ok: placeholder-tariff sentinel values
            issues.append({
                'type': MISSING_TARIFF,
                'row': row,
                'current': tariff,
                'description': desc,
                'qty': qty,
                'total_cost': total_cost,
                'why': f'No tariff code assigned. Customs requires a valid 8-digit HS code.',
                'how': f'Classify "{desc[:60]}" using CARICOM CET schedule based on material and function.',   # magic-ok: LLM-prompt description truncation width
            })

        # 2. INVALID_TARIFF (only if not already flagged as missing)
        elif len(tariff) != 8 or not tariff.isdigit():   # magic-ok: HS/CET tariff codes are always 8 digits
            issues.append({
                'type': INVALID_TARIFF,
                'row': row,
                'current': tariff,
                'description': desc,
                'qty': qty,
                'total_cost': total_cost,
                'why': f'Tariff code must be exactly 8 digits. Got: "{tariff}" ({len(tariff)} chars).',
                'how': f'Provide the full 8-digit CARICOM CET end-node code for "{desc[:60]}".',   # magic-ok: LLM-prompt description truncation width
            })
        elif cet_codes and tariff not in cet_codes:
            issues.append({
                'type': INVALID_TARIFF,
                'row': row,
                'current': tariff,
                'description': desc,
                'qty': qty,
                'total_cost': total_cost,
                'why': f'Code {tariff} not in CARICOM CET valid codes list (not an end-node code).',
                'how': f'Find the correct end-node under heading {tariff[:4]}.{tariff[4:6]}.',   # magic-ok: HS heading (4-digit) and subheading (6-digit) slice widths
            })

        # 3. ZERO_VALUE (qty > 0 but price = 0, not backordered)
        if qty > 0 and total_cost == 0 and unit_price == 0:
            # Skip items that look like free/promo/backordered
            desc_lower = desc.lower()
            if not any(kw in desc_lower for kw in ('free', 'promo', 'sample', 'backorder', 'b/o', 'cancel')):   # magic-ok: zero-price-exemption keyword list (backordered/promo/free items)
                issues.append({
                    'type': ZERO_VALUE,
                    'row': row,
                    'description': desc,
                    'qty': qty,
                    'why': f'Item has qty={qty} but $0.00 unit price and $0.00 total. Likely OCR missed the price.',
                    'how': f'Look for "{desc[:40]}" price in the invoice text. Set unit_price and total_cost = qty * unit_price.',   # magic-ok: LLM-prompt description truncation width
                })

    # 4. COLUMN_SPEC: columns.yaml conformance check
    # Per columns.yaml:
    #   Col L (12, Per Unit): group = "${category_name} (N items)", detail = "    ${description}" — always text
    #   Col O (15, Cost):     group = total_cost/sum_quantity, detail = unit_price — always numeric
    for row in range(2, ws.max_row + 1):
        tc_val = ws.cell(row, COL_TOTAL_COST).value
        if isinstance(tc_val, str) and tc_val.startswith('='):
            continue
        label = ws.cell(row, COL_UNIT_PRICE).value
        if isinstance(label, str) and label.strip() in FORMULA_LABELS:
            continue
        desc = str(ws.cell(row, COL_SUPP_DESC).value or '').strip()
        if not desc:
            continue

        per_unit_val = ws.cell(row, COL_UNIT_PRICE).value  # Col L
        cost_val = ws.cell(row, COL_COST).value             # Col O

        # Per Unit (col L) must be text, never numeric
        if isinstance(per_unit_val, (int, float)):
            issues.append({
                'type': COLUMN_SPEC,
                'row': row,
                'column': 'L (Per Unit)',   # magic-ok: report-label for column-spec violation
                'expected': 'text (description)',   # magic-ok: report-label expected-kind
                'actual': f'numeric ({per_unit_val})',
                'why': f'Col L (Per Unit) must be text per columns.yaml: group="category (N items)", detail="    description". Got numeric value {per_unit_val}.',
                'how': 'Fix bl_xlsx_generator.py to write description text to col L, not unit_price.',   # magic-ok: remediation instruction
            })

        # Cost (col O) must be numeric, never text (unless empty)
        if cost_val is not None and isinstance(cost_val, str) and cost_val.strip():
            issues.append({
                'type': COLUMN_SPEC,
                'row': row,
                'column': 'O (Cost)',   # magic-ok: report-label for column-spec violation
                'expected': 'numeric (unit price)',   # magic-ok: report-label expected-kind
                'actual': f'text ("{cost_val[:30]}")',   # magic-ok: display truncation width
                'why': f'Col O (Cost) must be numeric per columns.yaml: group=total_cost/sum_quantity, detail=unit_price. Got text "{cost_val[:30]}".',   # magic-ok: display truncation width
                'how': 'Fix bl_xlsx_generator.py to write numeric unit_price to col O.',   # magic-ok: remediation instruction
            })

    # 5. VARIANCE (invoice-level check)
    # For combined XLSX, sum metadata (S/T/U/V/W) from all invoice sections
    # BL freight is stored in col 38 (BLFreight), NOT in col T, so no
    # exclusion needed — col T only contains invoice-level freight.
    inv_total = 0.0
    freight = 0.0
    insurance = 0.0
    tax = 0.0
    deduction = 0.0
    for row in range(2, ws.max_row + 1):
        s_val = _num(ws.cell(row, COL_INV_TOTAL).value)
        if s_val:
            inv_total += s_val
            freight += _num(ws.cell(row, COL_FREIGHT).value)
            insurance += _num(ws.cell(row, COL_INSURANCE).value)
            tax += _num(ws.cell(row, COL_TAX).value)
            deduction += _num(ws.cell(row, COL_DEDUCTION).value)

    # Detect grouped XLSX: if there's a "SUBTOTAL (GROUPED)" label in column J,
    # then group rows already contain totals and detail rows should be excluded.
    # Plain "SUBTOTAL" does NOT indicate grouped mode — non-grouped XLSX files
    # also have a plain SUBTOTAL summary row.
    is_grouped = False
    for row in range(2, ws.max_row + 1):
        j_val = str(ws.cell(row, COL_SUPP_DESC).value or '')
        if 'SUBTOTAL (GROUPED)' in j_val:   # magic-ok: xlsx_labels grouped-mode sentinel label
            is_grouped = True
            break

    sum_items = 0.0
    import re as _re
    _group_re = _re.compile(r'\(\d+\s+items?\)$')   # magic-ok: group-header suffix pattern "(N items)"
    def _is_summary_row(row):
        """Return True if row is a formula label, duty estimation, or empty label."""
        j_val = str(ws.cell(row, COL_SUPP_DESC).value or '').strip()
        if not j_val:
            return True
        if j_val in FORMULA_LABELS:
            return True
        if any(j_val.startswith(p) for p in DUTY_ESTIMATION_PREFIXES):
            return True
        return False

    if is_grouped:
        # Grouped mode: only sum group-header rows.
        # Group headers end with "(N items)" in column J (e.g. "slippers (27 items)").
        # Detail rows are sub-items whose totals are already included in the
        # group header's P value — skip them to avoid double-counting.
        for row in range(2, ws.max_row + 1):
            tc = ws.cell(row, COL_TOTAL_COST).value
            if isinstance(tc, str) and tc.startswith('='):
                continue
            if _is_summary_row(row):
                continue
            # Group header: description ends with "(N items)"
            j_val = str(ws.cell(row, COL_SUPP_DESC).value or '').strip()
            if _group_re.search(j_val):
                if isinstance(tc, (int, float)):
                    sum_items += tc
            # else: detail row — skip (already counted in group total)
    else:
        for row in range(2, ws.max_row + 1):
            tc = ws.cell(row, COL_TOTAL_COST).value
            if isinstance(tc, str) and tc.startswith('='):
                continue  # skip formula rows (combined XLSX has per-invoice formulas)
            if _is_summary_row(row):
                continue
            if isinstance(tc, (int, float)):
                sum_items += tc

    adjustments = freight + insurance + tax - deduction

    # The variance fixer writes a correction into the ADJUSTMENTS formula:
    #   =(T2+U2+V2-W2)+<correction>
    # Since openpyxl reads formula strings (not computed values), we must
    # parse the correction term and include it in the adjustments total.
    # This keeps the validator in sync with the formula that Excel evaluates.
    for row in range(2, ws.max_row + 1):
        j_val = str(ws.cell(row, COL_SUPP_DESC).value or '').strip()
        if j_val in ('ADJUSTMENTS', 'Adjustments'):   # magic-ok: xlsx_labels ADJUSTMENTS row label (both cases)
            adj_formula = ws.cell(row, COL_TOTAL_COST).value
            if isinstance(adj_formula, str) and adj_formula.startswith('='):
                # Extract correction after the base formula's closing ')'
                # e.g. "=(T2+U2+V2-W2)+-30.84" → correction = -30.84
                paren_idx = adj_formula.rfind(')')
                if paren_idx >= 0 and paren_idx < len(adj_formula) - 1:
                    tail = adj_formula[paren_idx + 1:]
                    # tail is like "+{value}" or "+-{value}"
                    # Evaluate by summing all numeric terms
                    for num_match in _re.finditer(r'[+\-]?\s*\d+\.?\d*', tail):   # magic-ok: numeric-term regex for ADJUSTMENTS correction tail
                        try:
                            adjustments += float(num_match.group().replace(' ', ''))
                        except ValueError:
                            pass
            break

    net_total = round(sum_items + adjustments, 2)
    variance = round(inv_total - net_total, 2) if inv_total else 0

    if abs(variance) > 0.01:   # magic-ok: cent-level rounding tolerance for invoice variance
        issues.append({
            'type': VARIANCE,
            'variance': variance,
            'invoice_total': inv_total,
            'sum_items': sum_items,
            'freight': freight,
            'insurance': insurance,
            'tax': tax,
            'deduction': deduction,
            'why': f'Variance ${variance:+.2f}. InvoiceTotal(${inv_total:.2f}) != NetTotal(${net_total:.2f}).',
            'how': 'Check invoice text for unextracted freight, tax, fees, or misread item prices.',   # magic-ok: remediation instruction
        })

    return issues


_COMBINED_FORMULA_LABELS = {
    'SUBTOTAL (GROUPED)', 'SUBTOTAL (DETAILS)', 'GROUP VERIFICATION',          # magic-ok: xlsx_labels combined-XLSX block summary row labels
    'ADJUSTMENTS', 'NET TOTAL', 'INVOICE TOTAL', 'VARIANCE CHECK',             # magic-ok: xlsx_labels combined-XLSX block summary row labels
    'TOTAL INTERNAL FREIGHT', 'TOTAL INSURANCE',                               # magic-ok: xlsx_labels combined-XLSX block summary row labels
    'TOTAL OTHER COST', 'TOTAL DEDUCTION',                                     # magic-ok: xlsx_labels combined-XLSX block summary row labels
    'GRAND SUBTOTAL (GROUPED)', 'GRAND SUBTOTAL (DETAILS)',                    # magic-ok: xlsx_labels combined-XLSX grand-section labels
    'GRAND ADJUSTMENTS', 'GRAND NET TOTAL', 'GRAND INVOICE TOTAL',             # magic-ok: xlsx_labels combined-XLSX grand-section labels
    'GRAND VARIANCE CHECK', 'GRAND VARIANCE',                                  # magic-ok: xlsx_labels combined-XLSX grand-section labels
}


def _detect_combined_block_variance(ws) -> list:
    """
    Per-block variance check for combined multi-invoice XLSX.

    Uses each block's ADJUSTMENTS row as an anchor: its formula
    ``=(T{r}+U{r}+V{r}-W{r})[+correction]`` references the block's header
    row ``r``. Between the current block's header and the next block's
    header (or the ADJUSTMENTS row for the last block) we sum the item
    rows in col P, skipping formula rows, blank rows, and known summary
    labels. Compare against S{r} (invoice total) minus the adjustments
    (including any variance-fixer correction term).

    Blocks that already balance — because invoice_processor ran
    variance_fixer per-invoice before combining — produce no issue.
    variance_fixer must NOT be re-invoked on combined files because its
    search-from-bottom logic scribbles corrections into the wrong block.
    """
    import re as _re

    issues = []
    # Find block anchors via ADJUSTMENTS rows. Each has a formula whose
    # argument is the block's header row number.
    adj_re = _re.compile(r'^=\(T(\d+)\+U\d+\+V\d+-W\d+\)')   # magic-ok: ADJUSTMENTS base-formula anchor pattern (columns T/U/V/W)
    blocks = []  # list of (header_row, adjustments_row)
    for row in range(2, ws.max_row + 1):
        j_val = str(ws.cell(row, COL_SUPP_DESC).value or '').strip().upper()
        if j_val == 'ADJUSTMENTS':   # magic-ok: xlsx_labels ADJUSTMENTS row label (upper-case)
            tc = ws.cell(row, COL_TOTAL_COST).value
            if isinstance(tc, str):
                m = adj_re.match(tc)
                if m:
                    header_row = int(m.group(1))
                    blocks.append((header_row, row))

    if len(blocks) < 2:
        return issues  # not a multi-block combined file

    for block_idx, (header, adj_row) in enumerate(blocks, 1):
        s_val = _num(ws.cell(header, COL_INV_TOTAL).value)
        freight = _num(ws.cell(header, COL_FREIGHT).value)
        insurance = _num(ws.cell(header, COL_INSURANCE).value)
        tax = _num(ws.cell(header, COL_TAX).value)
        deduction = _num(ws.cell(header, COL_DEDUCTION).value)
        adjustments = freight + insurance + tax - deduction

        # Parse any variance-fixer correction term on the ADJUSTMENTS formula
        adj_formula = ws.cell(adj_row, COL_TOTAL_COST).value
        if isinstance(adj_formula, str) and adj_formula.startswith('='):
            paren_idx = adj_formula.rfind(')')
            if paren_idx >= 0 and paren_idx < len(adj_formula) - 1:
                tail = adj_formula[paren_idx + 1:]
                for m in _re.finditer(r'[+\-]?\s*\d+\.?\d*', tail):   # magic-ok: numeric-term regex for correction tail
                    try:
                        adjustments += float(m.group().replace(' ', ''))
                    except ValueError:
                        pass

        # Sum item costs for this block. The generator produces grouped
        # XLSX where row `header` is a group header and subsequent rows
        # are detail rows belonging to that header (detail P values are
        # already rolled up into the header's P). The SUBTOTAL (GROUPED)
        # formula is the source of truth — e.g. "=P2" or "=P124+P126" —
        # so we parse it and sum the referenced P cells.
        sum_items = 0.0
        grouped_ref_re = _re.compile(r'P(\d+)')   # magic-ok: SUBTOTAL(GROUPED) formula cell-ref pattern (col P)
        found_grouped = False
        for row in range(header, adj_row):
            j_val = str(ws.cell(row, COL_SUPP_DESC).value or '').strip().upper()
            if j_val == 'SUBTOTAL (GROUPED)':   # magic-ok: xlsx_labels grouped-mode subtotal label
                grouped_formula = ws.cell(row, COL_TOTAL_COST).value
                if isinstance(grouped_formula, str) and grouped_formula.startswith('='):
                    for m in grouped_ref_re.finditer(grouped_formula):
                        ref_row = int(m.group(1))
                        ref_val = ws.cell(ref_row, COL_TOTAL_COST).value
                        if isinstance(ref_val, (int, float)):
                            sum_items += ref_val
                found_grouped = True
                break
        if not found_grouped:
            # Ungrouped block — sum non-formula item rows directly.
            for row in range(header, adj_row):
                j_val = str(ws.cell(row, COL_SUPP_DESC).value or '').strip()
                if j_val.upper() in _COMBINED_FORMULA_LABELS:
                    continue
                p_val = ws.cell(row, COL_TOTAL_COST).value
                if isinstance(p_val, str) and p_val.startswith('='):
                    continue
                if isinstance(p_val, (int, float)):
                    sum_items += p_val

        net_total = round(sum_items + adjustments, 2)
        variance = round(s_val - net_total, 2)
        # Match invoice_processor's $0.02 tolerance (covers cent rounding
        # between grouped subtotal and detail sum).
        if abs(variance) > 0.02:   # magic-ok: 2-cent tolerance matching invoice_processor (covers grouped-subtotal rounding)
            issues.append({
                'type': VARIANCE,
                'variance': variance,
                'invoice_total': s_val,
                'sum_items': sum_items,
                'freight': freight,
                'insurance': insurance,
                'tax': tax,
                'deduction': deduction,
                'block': block_idx,
                'block_header_row': header,
                'block_adjustments_row': adj_row,
                'why': (f'Block {block_idx} (row {header}) variance ${variance:+.2f}. '
                        f'InvoiceTotal(${s_val:.2f}) != NetTotal(${net_total:.2f}).'),
                'how': ('Open source PDF for this invoice and rerun pipeline — '   # magic-ok: remediation instruction
                        'per-invoice variance_fixer should resolve before combining.'),   # magic-ok: remediation instruction
            })

    return issues


# ── Tariff fix ─────────────────────────────────────────────────────────────

def _fix_tariff_issues(ws, issues: list, result, base_dir: str) -> list:
    """
    Fix tariff issues in two passes:
      1. Try classifier.validate_and_correct_code() for invalid codes
      2. LLM call for remaining unresolved items

    Returns list of fix records: {row, old, new, method, description}
    """
    fixes = []
    unresolved = []

    # Pass 1: Try CET auto-correction for INVALID codes
    try:
        from classifier import validate_and_correct_code
    except ImportError:
        validate_and_correct_code = None

    for issue in issues:
        if issue['type'] == INVALID_TARIFF and validate_and_correct_code:
            current = issue['current']
            corrected = validate_and_correct_code(current, base_dir)
            if corrected != current and corrected and corrected != 'UNKNOWN':   # magic-ok: classifier sentinel value for unclassified codes
                # Apply CET correction
                ws.cell(row=issue['row'], column=COL_TARIFF).value = corrected
                ws.cell(row=issue['row'], column=COL_GROUPBY).value = corrected
                fixes.append({
                    'type': INVALID_TARIFF,
                    'row': issue['row'],
                    'old': current,
                    'new': corrected,
                    'method': 'CET auto-correction',   # magic-ok: fix-method label
                    'description': issue['description'][:50],   # magic-ok: display truncation width
                })
                continue

        # Still unresolved
        unresolved.append(issue)

    if not unresolved:
        return fixes

    # Pass 2: LLM classification for remaining items
    llm_fixes = _llm_classify_items(ws, unresolved, result, base_dir)
    fixes.extend(llm_fixes)

    return fixes


def _llm_classify_items(ws, unresolved: list, result, base_dir: str) -> list:
    """Call LLM to classify unresolved tariff items. Validate codes before applying."""
    try:
        from core.llm_client import get_llm_client
        llm = get_llm_client()
    except Exception as e:
        logger.warning(f"LLM client unavailable for tariff fix: {e}")
        return []

    cet_codes = _load_cet_codes(base_dir)

    # Collect already-classified items as reference examples
    classified_items = []
    for row in range(2, ws.max_row + 1):
        tc_val = ws.cell(row, COL_TOTAL_COST).value
        if isinstance(tc_val, str) and tc_val.startswith('='):
            continue  # skip formula rows (combined XLSX)
        tariff = str(ws.cell(row, COL_TARIFF).value or '').strip()
        desc = str(ws.cell(row, COL_SUPP_DESC).value or '').strip()
        if tariff and len(tariff) == 8 and tariff.isdigit() and desc:   # magic-ok: HS/CET tariff codes are always 8 digits
            # Only include if it's actually valid (in CET or we don't have CET list)
            if not cet_codes or tariff in cet_codes:
                classified_items.append({'row': row, 'description': desc[:60], 'code': tariff})   # magic-ok: LLM-prompt description truncation width

    # Build prompt
    supplier = ''
    invoice_num = ''
    invoice_text = ''
    if hasattr(result, 'supplier_info'):
        supplier = result.supplier_info.get('name', '')
    if hasattr(result, 'invoice_num'):
        invoice_num = result.invoice_num
    if hasattr(result, 'invoice_data'):
        invoice_text = result.invoice_data.get('raw_text', '')[:3000]   # magic-ok: LLM-prompt raw-text truncation width (tokens budget)

    user_lines = [f"Invoice: {invoice_num} from {supplier}", ""]
    user_lines.append("Items needing tariff classification:")   # magic-ok: LLM-prompt section header
    for issue in unresolved:
        user_lines.append(f"  Row {issue['row']}: \"{issue['description'][:60]}\"  qty={issue.get('qty', '?')}  ${issue.get('total_cost', 0):.2f}")   # magic-ok: LLM-prompt description truncation width
        user_lines.append(f"    WHY: {issue['why']}")
        user_lines.append(f"    HOW: {issue['how']}")
        user_lines.append("")

    if classified_items:
        user_lines.append("Already classified items from this invoice (for reference):")   # magic-ok: LLM-prompt section header
        for ci in classified_items[:15]:   # magic-ok: max reference examples shown to LLM
            user_lines.append(f"  Row {ci['row']}: \"{ci['description']}\" -> {ci['code']}")
        user_lines.append("")

    if invoice_text:
        user_lines.append("Invoice text (for additional context):")   # magic-ok: LLM-prompt section header
        user_lines.append(invoice_text)

    user_message = '\n'.join(user_lines)

    # LLM call
    cache_extra = f"tariff_fix:{invoice_num}:{len(unresolved)}"
    try:
        response = llm.call_json(
            user_message=user_message,
            system_prompt=TARIFF_SYSTEM_PROMPT,
            cache_key_extra=cache_extra,
        )
    except Exception as e:
        logger.warning(f"LLM tariff classification call failed: {e}")
        return []

    if not response or 'classifications' not in response:   # magic-ok: LLM response JSON key
        logger.warning("LLM returned no classifications")
        return []

    # Apply validated fixes
    fixes = []
    unresolved_rows = {i['row']: i for i in unresolved}

    for cls in response.get('classifications', []):
        row = cls.get('row')
        code = str(cls.get('code', '')).strip()
        reasoning = cls.get('reasoning', '')

        if row not in unresolved_rows:
            continue

        # Validate: must be 8 digits
        if len(code) != 8 or not code.isdigit():   # magic-ok: HS/CET tariff codes are always 8 digits
            logger.warning(f"LLM returned invalid code '{code}' for row {row}, skipping")
            continue

        # Validate against CET if available
        if cet_codes and code not in cet_codes:
            # Try suffix correction
            fallback = code[:6] + '00'   # magic-ok: 6-digit HS subheading + '00' end-node suffix fallback
            if fallback in cet_codes:
                logger.info(f"LLM code {code} not in CET, using fallback {fallback}")
                code = fallback
            else:
                logger.warning(f"LLM code '{code}' not in CET valid codes for row {row}, skipping")
                continue

        issue = unresolved_rows[row]
        ws.cell(row=row, column=COL_TARIFF).value = code
        ws.cell(row=row, column=COL_GROUPBY).value = code
        fixes.append({
            'type': issue['type'],
            'row': row,
            'old': issue['current'],
            'new': code,
            'method': f'LLM: {reasoning[:60]}',   # magic-ok: fix-method display truncation width
            'description': issue['description'][:50],   # magic-ok: display truncation width
        })

    return fixes


# ── Variance fix ───────────────────────────────────────────────────────────

def _fix_variance_issues(xlsx_path: str, result, variance: float) -> Optional[dict]:
    """
    Delegate variance fixing to the existing variance_fixer module.
    Also handles zero-value items (they contribute to variance).
    """
    try:
        from workflow.variance_fixer import fix_variance
    except ImportError:
        logger.warning("variance_fixer module not available")
        return None

    invoice_text = ''
    if hasattr(result, 'invoice_data'):
        invoice_text = result.invoice_data.get('raw_text', '')

    if not invoice_text:
        logger.warning("No invoice text available for variance fix")
        return None

    try:
        fix_result = fix_variance(
            xlsx_path=xlsx_path,
            invoice_text=invoice_text,
            current_variance=variance,
        )
        if fix_result.get('success'):
            return {
                'type': VARIANCE,
                'old_variance': variance,
                'new_variance': fix_result.get('new_variance', 0),
                'fixes_applied': fix_result.get('fixes_applied', 0),
                'method': f"variance_fixer: {fix_result.get('analysis', '')[:60]}",   # magic-ok: display truncation width for variance-fix method label
            }
        else:
            logger.warning(f"Variance fix failed: {fix_result.get('error', 'unknown')}")   # magic-ok: default-value sentinel for logging
            return None
    except Exception as e:
        logger.warning(f"Variance fix error: {e}")
        return None


# ── Duplicate file check ──────────────────────────────────────────────────

def _check_duplicates(results: list) -> dict:
    """
    Check for duplicate XLSX and PDF files in the output.

    Detects:
      1. Duplicate filenames — multiple invoices producing the same output file
      2. Duplicate content — different source PDFs producing XLSX with the same
         invoice number, supplier, and matching item data (overwrite / double-count)
      3. Stale files — leftover XLSX/PDF from a previous run

    Returns:
        {duplicates: [...], content_duplicates: [...], stale: [...], has_issues: bool}
    """
    try:
        import openpyxl
    except ImportError:
        return {'duplicates': [], 'content_duplicates': [], 'stale': [],
                'has_issues': False}

    # ── 1. Duplicate filenames ──
    seen_xlsx = {}
    seen_pdf = {}
    output_dir = None

    for r in results:
        src_pdf = getattr(r, 'pdf_file', '') or ''
        xlsx_path = r.xlsx_path
        pdf_out = getattr(r, 'pdf_output_path', '') or ''

        if xlsx_path:
            if output_dir is None:
                output_dir = os.path.dirname(xlsx_path)
            basename = os.path.basename(xlsx_path)
            seen_xlsx.setdefault(basename, []).append(src_pdf)

        if pdf_out and os.path.exists(pdf_out):
            basename = os.path.basename(pdf_out)
            seen_pdf.setdefault(basename, []).append(src_pdf)

    duplicates = []
    for fname, sources in seen_xlsx.items():
        if len(sources) > 1:
            duplicates.append({'filename': fname, 'type': 'xlsx', 'sources': sources})   # magic-ok: file-kind label
    for fname, sources in seen_pdf.items():
        if len(sources) > 1:
            duplicates.append({'filename': fname, 'type': 'pdf', 'sources': sources})   # magic-ok: file-kind label

    # ── 2. Duplicate content ──
    # Detect duplicate invoices/documents (same invoice number + supplier).
    # This does NOT flag product overlap between different invoices.
    content_duplicates = []
    file_fingerprints = []  # [(xlsx_basename, src_pdf, inv_num, supplier, total)]

    for r in results:
        xlsx_path = r.xlsx_path
        if not xlsx_path or not os.path.exists(xlsx_path):
            continue
        src_pdf = getattr(r, 'pdf_file', '') or ''
        basename = os.path.basename(xlsx_path)

        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            inv_num = ws.cell(row=2, column=3).value    # magic-ok: columns.yaml Supplier Invoice# column C (3)
            supplier = ws.cell(row=2, column=27).value   # magic-ok: columns.yaml Supplier Name column AA (27)
            inv_total = _num(ws.cell(row=2, column=COL_INV_TOTAL).value)
            wb.close()

            file_fingerprints.append((basename, src_pdf, str(inv_num),
                                      str(supplier), inv_total))
        except Exception as e:
            logger.debug(f"Duplicate content check skipped for {basename}: {e}")

    # Compare all pairs — only flag same invoice number from same supplier
    for i in range(len(file_fingerprints)):
        for j in range(i + 1, len(file_fingerprints)):
            f1 = file_fingerprints[i]
            f2 = file_fingerprints[j]
            name1, src1, inv1, sup1, tot1 = f1
            name2, src2, inv2, sup2, tot2 = f2

            # Same invoice number from same supplier = duplicate document
            if inv1 and inv2 and inv1 == inv2 and sup1 == sup2:
                content_duplicates.append({
                    'type': 'same_invoice_number',   # magic-ok: duplicate-detection category
                    'invoice_num': inv1,
                    'supplier': sup1,
                    'files': [name1, name2],
                    'sources': [src1, src2],
                })

    # ── 2b. Auto-resolve content duplicates ──
    # When two source PDFs produce identical content and one filename prefix is
    # contained in the other (e.g. "bmgrn INV_..." vs "bmgrn tbm INV_..."),
    # the shorter-prefix version is the consignee-specific copy — keep it and
    # remove the other.
    resolved_dups = []
    for cd in content_duplicates:
        if len(cd.get('sources', [])) != 2:
            continue
        src_a, src_b = cd['sources']
        file_a, file_b = cd['files']
        # Compare the prefix before the invoice number (e.g. "bmgrn " vs "bmgrn tbm ")
        prefix_a = src_a.split('INV')[0].strip().lower() if 'INV' in src_a else src_a.lower()   # magic-ok: invoice-number marker in source PDF filename
        prefix_b = src_b.split('INV')[0].strip().lower() if 'INV' in src_b else src_b.lower()   # magic-ok: invoice-number marker in source PDF filename

        remove_file = None
        keep_src = None
        if prefix_a and prefix_b and prefix_a != prefix_b:
            if prefix_a in prefix_b:
                # prefix_a is shorter / base variant — keep it, remove b
                remove_file = file_b
                keep_src = src_a
            elif prefix_b in prefix_a:
                remove_file = file_a
                keep_src = src_b

        if remove_file:
            # Remove the duplicate from results
            for idx, r in enumerate(results):
                if r.xlsx_path and os.path.basename(r.xlsx_path) == remove_file:
                    # Delete the duplicate XLSX and PDF from output
                    if os.path.exists(r.xlsx_path):
                        os.remove(r.xlsx_path)
                        logger.info(f"Auto-resolved duplicate: removed {remove_file} (keeping {keep_src})")
                    pdf_out = getattr(r, 'pdf_output_path', '')
                    if pdf_out and os.path.exists(pdf_out):
                        os.remove(pdf_out)
                    results.pop(idx)
                    break
            resolved_dups.append(cd)
            print(f"    AUTO-RESOLVED duplicate: removed {remove_file} (keeping {keep_src})")

    # Remove resolved duplicates from the issue list
    for rd in resolved_dups:
        content_duplicates.remove(rd)

    # ── 3. Stale files ──
    stale = []
    if output_dir and os.path.isdir(output_dir):
        expected_files = set()
        for r in results:
            if r.xlsx_path:
                expected_files.add(os.path.basename(r.xlsx_path))
            pdf_out = getattr(r, 'pdf_output_path', '') or ''
            if pdf_out:
                expected_files.add(os.path.basename(pdf_out))
            # Combined entries stash per-invoice source PDFs here; they are
            # produced by the current run even though the individual XLSX
            # files were merged away.
            combined_pdfs = getattr(r, '_combined_pdf_paths', None) or []
            for cp in combined_pdfs:
                expected_files.add(os.path.basename(cp))
        expected_files.add('_email_params.json')   # magic-ok: filename of per-run email params sidecar

        for f in os.listdir(output_dir):
            fp = os.path.join(output_dir, f)
            if not os.path.isfile(fp):
                continue
            if f in expected_files:
                continue
            if f.endswith('-BL.pdf'):   # magic-ok: BL PDF filename suffix (from bl_generator)
                continue
            if f.endswith('.xlsx') or f.endswith('.pdf'):   # magic-ok: pipeline output file extensions
                stale.append(f)

    return {
        'duplicates': duplicates,
        'content_duplicates': content_duplicates,
        'stale': stale,
        'has_issues': bool(duplicates or content_duplicates or stale),
    }


# ── Package total check ───────────────────────────────────────────────────

def _check_package_totals(results: list, bl_alloc=None, manifest_meta: dict = None) -> dict:
    """
    Verify that the sum of Packages (col X) across all XLSX files matches the
    expected total from BL or manifest.

    Priority: BL packages → manifest packages → 1 per entry (default).

    Returns:
        {expected, actual, source, per_file: [{file, packages}], mismatch: bool}
    """
    try:
        import openpyxl
    except ImportError:
        return {}

    # Determine expected total and source
    expected = None
    source = 'default (1 per entry)'   # magic-ok: expected-packages source label

    if bl_alloc and getattr(bl_alloc, 'packages', None):
        try:
            expected = int(bl_alloc.packages)
            source = 'BL'   # magic-ok: expected-packages source label
        except (ValueError, TypeError):
            pass

    if expected is None and manifest_meta and manifest_meta.get('packages'):
        try:
            expected = int(manifest_meta['packages'])
            source = 'manifest'   # magic-ok: expected-packages source label
        except (ValueError, TypeError):
            pass

    if expected is None:
        # Default: 1 package per XLSX entry
        expected = len([r for r in results if r.xlsx_path and os.path.exists(r.xlsx_path)])
        source = 'default (1 per entry)'   # magic-ok: expected-packages source label

    # Read actual packages from each XLSX col X (row 2)
    per_file = []
    actual_total = 0
    for r in results:
        xlsx_path = r.xlsx_path
        if not xlsx_path or not os.path.exists(xlsx_path):
            continue
        fname = os.path.basename(xlsx_path)
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            pkg_val = ws.cell(row=2, column=COL_PACKAGES).value
            wb.close()
            pkg = int(pkg_val) if pkg_val is not None else 1
        except Exception:
            pkg = 1
        per_file.append({'file': fname, 'packages': pkg})
        actual_total += pkg

    mismatch = actual_total != expected

    return {
        'expected': expected,
        'actual': actual_total,
        'source': source,
        'per_file': per_file,
        'mismatch': mismatch,
    }


# ── Report ─────────────────────────────────────────────────────────────────

def _print_report(all_results: list, pkg_check: dict = None,
                   dup_check: dict = None) -> None:
    """Print comprehensive validation report with before/after states."""
    if not all_results:
        return

    total_files = len(all_results)
    total_issues = sum(len(r.get('issues', [])) for r in all_results)
    total_fixed = sum(r.get('fixed', 0) for r in all_results)
    total_unfixed = sum(r.get('unfixed', 0) for r in all_results)
    clean_files = sum(1 for r in all_results if not r.get('issues'))

    print()
    print("XLSX VALIDATION")   # magic-ok: report section header
    print("-" * 80)   # magic-ok: report rule width (characters)

    for fr in all_results:
        fname = fr['file']
        supplier = fr.get('supplier', '')
        issues = fr.get('issues', [])
        fixes = fr.get('fixes', [])
        remaining = fr.get('remaining', [])

        label = f"  {fname}"
        if supplier:
            label += f" ({supplier})"

        if not issues:
            print(f"{label}")
            print(f"    No issues found.")
            print()
            continue

        print(f"{label}")

        # Build fix lookup for display
        fix_by_row = {}
        for f in fixes:
            if isinstance(f, dict) and 'row' in f:   # magic-ok: fix dict 'row' key
                fix_by_row[f['row']] = f

        remaining_types = set()
        for ri in remaining:
            remaining_types.add((ri.get('type'), ri.get('row', 0)))

        for issue in issues:
            itype = issue['type']
            row = issue.get('row', 0)

            if itype in (MISSING_TARIFF, INVALID_TARIFF):
                desc = issue.get('description', '')[:30]   # magic-ok: report display truncation width
                current = issue.get('current', '?')
                fix = fix_by_row.get(row)

                if fix:
                    new_code = fix.get('new', '?')
                    method = fix.get('method', '')
                    print(f"    Row {row:>3}  {itype:<16} \"{desc}\" [{current}] -> {new_code} ({method})")
                else:
                    is_remaining = (itype, row) in remaining_types
                    status = "UNFIXED" if is_remaining else "FIXED"   # magic-ok: report status labels
                    print(f"    Row {row:>3}  {itype:<16} \"{desc}\" [{current}]  {status}")

            elif itype == ZERO_VALUE:
                desc = issue.get('description', '')[:30]   # magic-ok: report display truncation width
                qty = issue.get('qty', 0)
                is_remaining = (itype, row) in remaining_types
                status = "UNFIXED" if is_remaining else "FIXED"   # magic-ok: report status labels
                print(f"    Row {row:>3}  {itype:<16} \"{desc}\" qty={qty} $0.00  {status}")

            elif itype == COLUMN_SPEC:
                col_name = issue.get('column', '?')
                expected = issue.get('expected', '?')
                actual = issue.get('actual', '?')
                print(f"    Row {row:>3}  {itype:<16} {col_name}: expected {expected}, got {actual}")

            elif itype == VARIANCE:
                old_var = issue.get('variance', 0)
                # Find variance fix in fixes list
                var_fix = next((f for f in fixes if isinstance(f, dict) and f.get('type') == VARIANCE), None)
                if var_fix:
                    new_var = var_fix.get('new_variance', 0)
                    print(f"    {itype:<20} ${old_var:+.2f} -> ${new_var:+.2f} ({var_fix.get('method', '')[:40]})")   # magic-ok: report display truncation width
                else:
                    has_remaining_var = any(ri.get('type') == VARIANCE for ri in remaining)
                    status = "UNFIXED" if has_remaining_var else "FIXED"   # magic-ok: report status labels
                    print(f"    {itype:<20} ${old_var:+.2f}  {status}")

        fixed_count = fr.get('fixed', 0)
        unfixed_count = fr.get('unfixed', 0)
        print(f"    Result: {len(issues)} issues, {fixed_count} fixed, {unfixed_count} unfixed")
        print()

    # Package total check
    if pkg_check and pkg_check.get('per_file'):
        expected = pkg_check.get('expected', '?')
        actual = pkg_check.get('actual', '?')
        source = pkg_check.get('source', '?')
        mismatch = pkg_check.get('mismatch', False)

        print(f"  PACKAGE CHECK (source: {source})")
        for pf in pkg_check['per_file']:
            print(f"    {pf['file']:<30} {pf['packages']:>3} packages")
        if mismatch:
            print(f"    MISMATCH: XLSX total = {actual}, expected = {expected} ({source})")
        else:
            print(f"    PASS: {actual} packages = {expected} ({source})")
        print()

    # Duplicate / stale file check
    if dup_check and dup_check.get('has_issues'):
        print(f"  DUPLICATE / STALE FILE CHECK")
        for dup in dup_check.get('duplicates', []):
            sources = ', '.join(dup['sources']) if dup['sources'] else 'unknown'   # magic-ok: report display fallback
            print(f"    DUPLICATE {dup['type'].upper()}: {dup['filename']}  (from: {sources})")
        for cd in dup_check.get('content_duplicates', []):
            dtype = cd['type']
            files = ', '.join(cd['files'])
            sources = ', '.join(cd.get('sources', []))
            if dtype == 'same_invoice_number':   # magic-ok: duplicate-detection category key
                print(f"    DUPLICATE CONTENT: same invoice #{cd['invoice_num']} ({cd['supplier']})")
                print(f"      Files: {files}")
                print(f"      Source PDFs: {sources}")
            elif dtype == 'identical_items':   # magic-ok: duplicate-detection category key
                print(f"    DUPLICATE CONTENT: identical items ({cd['item_count']} items)")
                print(f"      Files: {files}")
                print(f"      Source PDFs: {sources}")
            elif dtype == 'high_overlap':   # magic-ok: duplicate-detection category key
                print(f"    DUPLICATE CONTENT: {cd['pct']} item overlap ({cd['overlap']} of {cd['sizes']})")
                print(f"      Files: {files}")
                print(f"      Source PDFs: {sources}")
        for stale_f in dup_check.get('stale', []):
            print(f"    STALE: {stale_f}  (not produced by current run — leftover from previous run?)")
        print()
    elif dup_check:
        print(f"  DUPLICATE CHECK")
        print(f"    PASS: no duplicate or stale XLSX/PDF files")
        print()

    print("-" * 80)   # magic-ok: report rule width (characters)
    if total_issues == 0:
        print(f"  All {total_files} files clean - no issues detected.")
    elif total_unfixed == 0:
        print(f"  {total_files} files checked, {total_issues} issues found, all {total_fixed} fixed.")
    else:
        print(f"  {total_files} files checked, {total_issues} issues found, {total_fixed} fixed, {total_unfixed} unfixed.")
    print()


# ── Shipment Pre-Send Checklist ───────────────────────────────────────────

# Patterns for new checks added 2026-04-25 in response to recurring user reports.
# Each pattern targets a specific, observed pipeline failure mode that was not
# being caught by the existing rule-based checks before email send.
import re as _re

_BL_DOC_SUFFIX_RE = _re.compile(   # magic-ok: BL/waybill document-type suffix detector
    r'[-_ ]+(Declaration|Manifest|WorkSheet|Invoice|Packing(?:\s*List)?)\s*$',
    _re.IGNORECASE,
)
_PLACEHOLDER_SKU_RE = _re.compile(r'^ITEM-\d+$', _re.IGNORECASE)   # magic-ok: auto-generated sequential SKU pattern
_TARIFF_PREFIX_RE = _re.compile(r'^\d{8}\s')   # magic-ok: 8-digit tariff prefix at start of description
_GENERIC_PURCHASE_RE = _re.compile(   # magic-ok: placeholder-description sentinel pattern
    r'^(amazon\.com|walmart|target|shein|temu|ebay|aliexpress|alibaba)[\s\.,]*purchase\s*$',
    _re.IGNORECASE,
)
_PAYMENT_KEYWORDS = ('paypal', 'credit card', 'payment method',                 # magic-ok: payment-method sentinel substrings
                     'visa ending', 'mastercard', 'amex', 'gift card balance')  # magic-ok: payment-method sentinel substrings
_GARBAGE_REPEAT_RE = _re.compile(r'^(.)\1{2,}$')   # magic-ok: same-char-repeated description (e.g. 'eee', 'aaaa')
_MAN_REG_RE = _re.compile(r'^\d{4}[\s/]+\d+$')   # magic-ok: man_reg "YYYY NN" or "YYYY/NN" format
_COUNTRY_CODE_RE = _re.compile(r'^[A-Z]{2}$')   # magic-ok: ISO-3166 alpha-2 country code

_DESC_MIN_CHARS = 5                       # magic-ok: shortest plausible item description
_REPEATED_ITEM_PRICE_TOLERANCE = 0.01     # magic-ok: $0.01 unit-cost equality tolerance
_GENERIC_TLD_PURCHASE_FALLBACK = 'amazon.com purchase'   # magic-ok: most common stub description string


def _checklist_email_params_extra(email_params: dict, fail_fn) -> None:
    """Extra rule-based checks that operate purely on email_params dict.

    Detects:
      - BL number with leftover document-type suffix (defense-in-depth for
        the auto-strip in run.py:_auto_detect_bl_number)
      - expected_entries vs xlsx attachment count mismatch
      - WorkSheet PDFs included as attachments
      - Stale tmp / rerun paths in attachment_paths
      - Duplicate attachment basenames
      - Missing XLSX/PDF pairing per invoice
      - country_origin not 2-letter ISO
      - man_reg blank or wrong format
    """
    waybill = str(email_params.get('waybill', '')).strip()
    if waybill and _BL_DOC_SUFFIX_RE.search(waybill):
        fail_fn('bl_has_doc_suffix', 'warn',   # magic-ok: check key + severity
                f'Waybill "{waybill}" still ends with a document-type suffix '
                f'(e.g. -Declaration). The BL auto-detect should have stripped this.',
                'waybill', waybill,   # magic-ok: email-params field key
                'Update _email_params.json field "waybill" to drop the suffix.')   # magic-ok: remediation instruction

    attachments = email_params.get('attachment_paths', []) or []
    basenames = [os.path.basename(p) for p in attachments]

    # XLSX attachment count vs expected_entries (per feedback_expected_entries.md)
    xlsx_atts = [b for b in basenames if b.lower().endswith('.xlsx')]
    expected_entries = email_params.get('expected_entries')
    try:
        ee_int = int(expected_entries) if expected_entries not in (None, '') else None
    except (TypeError, ValueError):
        ee_int = None
    if ee_int is not None and xlsx_atts and ee_int != len(xlsx_atts):
        fail_fn('expected_entries_mismatch', 'warn',   # magic-ok: check key + severity
                f'expected_entries={ee_int} but {len(xlsx_atts)} XLSX attachment(s) found. '
                f'Expected Entries should equal the number of declarations / XLSX sheets sent.',
                'expected_entries', str(ee_int),   # magic-ok: email-params field key
                'Set expected_entries to len([a for a in attachment_paths if a.endswith(".xlsx")]).')   # magic-ok: remediation instruction

    # WorkSheet PDFs as attachments — they're auto-generated, not source invoices.
    ws_atts = [b for b in basenames if 'worksheet' in b.lower() and b.lower().endswith('.pdf')]
    if ws_atts:
        fail_fn('worksheet_pdf_in_attachments', 'warn',   # magic-ok: check key + severity
                f'{len(ws_atts)} WorkSheet PDF(s) included as attachments: {", ".join(ws_atts[:3])}.',
                'attachment_paths', str(len(ws_atts)),   # magic-ok: email-params field key
                'WorkSheet PDFs are auto-generated by send_shipment_email.py at send time. '   # magic-ok: remediation instruction
                'They should not appear in attachment_paths during pipeline output.')   # magic-ok: remediation instruction

    # Stale tmp / rerun paths
    stale_path_markers = ('_rerun_tmp', '/tmp/', '\\Temp\\', '/var/folders/')   # magic-ok: tmp-path markers
    stale = [p for p in attachments if any(m in p for m in stale_path_markers)]
    if stale:
        fail_fn('stale_tmp_path', 'block',   # magic-ok: check key + severity
                f'{len(stale)} attachment path(s) point at a temp / rerun directory.',
                'attachment_paths', os.path.basename(stale[0]),   # magic-ok: email-params field key
                'Re-resolve attachment paths to point at the real output directory.')   # magic-ok: remediation instruction

    # Duplicate basenames
    seen = {}
    dupes = []
    for b in basenames:
        seen[b] = seen.get(b, 0) + 1
    for b, n in seen.items():
        if n > 1:
            dupes.append(f'{b} ×{n}')
    if dupes:
        fail_fn('duplicate_attachment_basenames', 'block',   # magic-ok: check key + severity
                f'{len(dupes)} duplicate attachment basename(s): {", ".join(dupes[:5])}.',
                'attachment_paths', dupes[0],   # magic-ok: email-params field key
                'Remove duplicates from attachment_paths.')   # magic-ok: remediation instruction

    # XLSX without matching PDF (per source invoice — paired by basename stem).
    # Skip:
    #  - combined_*.xlsx / BL_combined.xlsx (not 1:1 with a single source PDF)
    #  - per-declaration {waybill}.xlsx (one source receipt PDF feeds N waybill
    #    XLSX via reverse-calc allocation; carrier-prefix HAWB/TSCW/HBL/MEDU
    #    is the signal)
    pdf_stems = {os.path.splitext(b)[0] for b in basenames if b.lower().endswith('.pdf')}
    _WAYBILL_STEM_RE = _re.compile(r'^(HAWB|TSCW|HBL|MEDU|MAEU|HDMU|COSU|TS)\d+',   # magic-ok: per-declaration XLSX-stem carrier prefixes
                                    _re.IGNORECASE)
    orphan_xlsx = []
    for x in xlsx_atts:
        stem = os.path.splitext(x)[0]
        if stem.endswith('_combined') or stem.lower().endswith('-combined'):
            continue
        if _WAYBILL_STEM_RE.match(stem):
            continue
        if stem not in pdf_stems:
            orphan_xlsx.append(x)
    if orphan_xlsx:
        fail_fn('xlsx_without_source_pdf', 'warn',   # magic-ok: check key + severity
                f'{len(orphan_xlsx)} XLSX attachment(s) have no matching source PDF: {", ".join(orphan_xlsx[:3])}.',
                'attachment_paths', orphan_xlsx[0],   # magic-ok: email-params field key
                'Each per-invoice XLSX should ship with the source PDF it was generated from.')   # magic-ok: remediation instruction

    # country_origin format
    country_origin = str(email_params.get('country_origin', '')).strip()
    if country_origin and not _COUNTRY_CODE_RE.match(country_origin):
        fail_fn('country_origin_invalid', 'warn',   # magic-ok: check key + severity
                f'country_origin "{country_origin}" is not a 2-letter ISO code (e.g. US, CN).',
                'country_origin', country_origin,   # magic-ok: email-params field key
                'Set country_origin to the 2-letter ISO-3166 code of the supplier country.')   # magic-ok: remediation instruction

    # man_reg format ("YYYY NN")
    man_reg = str(email_params.get('man_reg', '')).strip()
    if not man_reg:
        fail_fn('man_reg_blank', 'warn',   # magic-ok: check key + severity
                'man_reg is blank. Manifest registration is required for ASYCUDA filing.',
                'man_reg', '',   # magic-ok: email-params field key
                'Set man_reg to "YYYY NN" (e.g. "2024 19").')   # magic-ok: remediation instruction
    elif not _MAN_REG_RE.match(man_reg):
        fail_fn('man_reg_invalid', 'warn',   # magic-ok: check key + severity
                f'man_reg "{man_reg}" does not match expected "YYYY NN" format.',
                'man_reg', man_reg,   # magic-ok: email-params field key
                'Use "YYYY NN" format (year + slot number).')   # magic-ok: remediation instruction


def _inspect_xlsx_items(xlsx_path: str) -> list:
    """Walk a per-invoice XLSX and return a list of item-level findings.

    Each finding: {row, col, kind, severity, detail, value}.
    Kinds: placeholder_sku, description_starts_with_tariff,
           description_too_short, payment_line_as_item,
           generic_purchase_description, quantity_zero_with_cost,
           garbage_description_pattern, duplicate_items_in_xlsx,
           negative_invoice_total.
    """
    findings = []
    try:
        import openpyxl
    except ImportError:
        return findings
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        logger.debug(f"_inspect_xlsx_items could not open {xlsx_path}: {e}")   # magic-ok: debug log message
        return findings

    # Row 2 holds InvoiceTotal in column S
    try:
        inv_total_cell = ws.cell(row=2, column=COL_INV_TOTAL).value   # magic-ok: row 2 = first data row in per-invoice XLSX
        if inv_total_cell is not None:
            try:
                if float(inv_total_cell) <= 0:
                    findings.append({
                        'row': 2, 'col': 'S', 'kind': 'negative_invoice_total',
                        'severity': 'block',   # magic-ok: severity label
                        'detail': f'InvoiceTotal is {inv_total_cell} (must be > 0)',
                        'value': str(inv_total_cell),
                    })
            except (TypeError, ValueError):
                pass
    except Exception:
        pass

    # Walk rows. Detail rows = COL_TARIFF (F) is empty/None and COL_GROUPBY (AK)
    # is empty (group rows have a tariff in F). Skip totals rows (description in
    # _FORMULA_LABELS or matches duty-section prefix).
    item_signatures = {}  # (description, unit_cost) -> list of rows for dup detection

    for row in range(2, ws.max_row + 1):   # magic-ok: data rows start at row 2
        tariff_v = ws.cell(row=row, column=COL_TARIFF).value
        desc_v = ws.cell(row=row, column=COL_SUPP_DESC).value
        sku_v = ws.cell(row=row, column=9).value   # magic-ok: SupplierItemNumber column I (9)
        qty_v = ws.cell(row=row, column=COL_QTY).value
        cost_v = ws.cell(row=row, column=COL_COST).value
        total_v = ws.cell(row=row, column=COL_TOTAL_COST).value

        desc = str(desc_v).strip() if desc_v else ''
        sku = str(sku_v).strip() if sku_v else ''

        # Skip totals/duty/reference rows by description label
        if desc in _FORMULA_LABELS:
            continue
        if any(desc.startswith(p) for p in _DUTY_ESTIMATION_PREFIXES):
            continue

        # An item row is any row whose K (qty), O (cost), or P (total) holds a
        # numeric value — covers both detail rows and single-item-per-group
        # rows (where tariff in F and item data in I/J/K are merged).
        has_item_data = False
        for v in (qty_v, cost_v, total_v):
            if v is None:
                continue
            try:
                if float(v) != 0:
                    has_item_data = True
                    break
            except (TypeError, ValueError):
                continue
        if not has_item_data:
            continue

        # On group rows the SKU column (I) holds the tariff code, not a real
        # SKU — suppress placeholder-SKU & garbage checks against it.
        sku_is_tariff = bool(tariff_v) and str(sku) == str(tariff_v)

        # Numeric coercion
        try:
            qty = float(qty_v) if qty_v is not None else 0
        except (TypeError, ValueError):
            qty = 0
        try:
            unit_cost = float(cost_v) if cost_v is not None else 0
        except (TypeError, ValueError):
            unit_cost = 0
        try:
            total_cost = float(total_v) if total_v is not None else 0
        except (TypeError, ValueError):
            total_cost = 0

        # ── Placeholder SKU (auto-generated ITEM-001 etc.)
        if sku and not sku_is_tariff and _PLACEHOLDER_SKU_RE.match(sku):
            findings.append({
                'row': row, 'col': 'I', 'kind': 'placeholder_sku',
                'severity': 'warn',   # magic-ok: severity label
                'detail': f'SKU "{sku}" looks auto-generated (format extraction likely failed)',
                'value': sku,
            })

        # ── Payment-method line ingested as an item
        desc_l = desc.lower()
        if sku.upper() == 'PAYMENT' or any(kw in desc_l for kw in _PAYMENT_KEYWORDS):
            findings.append({
                'row': row, 'col': 'J', 'kind': 'payment_line_as_item',
                'severity': 'block',   # magic-ok: severity label
                'detail': f'Payment-method line treated as item: SKU="{sku}", desc="{desc[:60]}"',
                'value': desc[:60],   # magic-ok: display truncation width
            })

        # ── Description starts with 8-digit tariff prefix (e.g. WorkSheet ingest)
        if _TARIFF_PREFIX_RE.match(desc):
            findings.append({
                'row': row, 'col': 'J', 'kind': 'description_starts_with_tariff',
                'severity': 'warn',   # magic-ok: severity label
                'detail': f'Description begins with 8-digit tariff: "{desc[:60]}"',
                'value': desc[:60],   # magic-ok: display truncation width
            })

        # ── Generic placeholder description
        if _GENERIC_PURCHASE_RE.match(desc) or desc.lower() == _GENERIC_TLD_PURCHASE_FALLBACK:
            findings.append({
                'row': row, 'col': 'J', 'kind': 'generic_purchase_description',
                'severity': 'warn',   # magic-ok: severity label
                'detail': f'Description is a generic placeholder ("{desc}") — real product name was not extracted',
                'value': desc,
            })

        # ── Description too short
        if desc and len(desc) < _DESC_MIN_CHARS:
            findings.append({
                'row': row, 'col': 'J', 'kind': 'description_too_short',
                'severity': 'warn',   # magic-ok: severity label
                'detail': f'Description "{desc}" is shorter than {_DESC_MIN_CHARS} chars',
                'value': desc,
            })

        # ── Garbage description pattern
        if desc and _GARBAGE_REPEAT_RE.match(desc):
            findings.append({
                'row': row, 'col': 'J', 'kind': 'garbage_description_pattern',
                'severity': 'warn',   # magic-ok: severity label
                'detail': f'Description "{desc}" looks like keymash / OCR garbage',
                'value': desc,
            })

        # ── Quantity zero but total_cost non-zero
        if qty == 0 and total_cost > 0:
            findings.append({
                'row': row, 'col': 'K', 'kind': 'quantity_zero_with_cost',
                'severity': 'warn',   # magic-ok: severity label
                'detail': f'Quantity=0 but TotalCost=${total_cost:.2f} — extraction broke',
                'value': '0',
            })

        # ── Track for in-XLSX dup detection
        if desc and unit_cost > 0:
            sig = (desc.lower(), round(unit_cost, 2))   # magic-ok: 2-decimal price-rounding for dup detection
            item_signatures.setdefault(sig, []).append(row)

    try:
        wb.close()
    except Exception:
        pass

    # Same description + same unit_cost appearing 2+ times in one XLSX
    for (desc_l, price), rows in item_signatures.items():
        if len(rows) > 1:
            findings.append({
                'row': rows[0], 'col': 'J', 'kind': 'duplicate_items_in_xlsx',
                'severity': 'warn',   # magic-ok: severity label
                'detail': f'Same description+price appears on rows {rows} (desc="{desc_l[:40]}" @ ${price})',
                'value': f'{len(rows)} copies',
            })

    return findings


def _checklist_xlsx_inspection(email_params: dict, fail_fn) -> None:
    """Iterate XLSX attachments and surface item-level findings via fail_fn.

    Also checks each XLSX's sidecar .meta.json for format_source=='llm_auto'
    (no YAML format match — extraction is brittle).
    """
    attachments = email_params.get('attachment_paths', []) or []
    xlsx_paths = [p for p in attachments if p.lower().endswith('.xlsx')]

    # Aggregate findings by kind for compact reporting
    kind_counter = {}   # kind -> [(file, row, detail), ...]
    for xp in xlsx_paths:
        if not os.path.exists(xp):
            continue
        # Only inspect per-invoice XLSX, skip the BL combined XLSX which has a
        # different layout (per-block totals, no single InvoiceTotal in row 2).
        base = os.path.basename(xp)
        is_combined = base.lower().endswith('_combined.xlsx') or '-combined' in base.lower()

        # Sidecar .meta.json check — applies to per-invoice XLSX only
        meta_path = xp[:-5] + '.meta.json'   # magic-ok: replace .xlsx (5 chars) with .meta.json
        if not is_combined and os.path.exists(meta_path):
            try:
                with open(meta_path, 'r', encoding='utf-8') as fh:
                    meta = json.load(fh)
                if meta.get('format_source') == 'llm_auto':   # magic-ok: format_source sentinel
                    kind_counter.setdefault('llm_auto_format_used', []).append(
                        (base, 0, f'format_name={meta.get("format_name", "?")}')
                    )
            except (OSError, json.JSONDecodeError):
                pass

        if is_combined:
            # Combined XLSX has different row layout — skip per-row item checks.
            continue

        for f in _inspect_xlsx_items(xp):
            kind_counter.setdefault(f['kind'], []).append(
                (base, f['row'], f['detail'])
            )

    # Map kind -> (default severity, hint)
    KIND_DEFAULTS = {
        'placeholder_sku': ('warn',
            'Format YAML is not extracting real SKUs — fall-back assigned ITEM-NNN. '
            'Check the format spec for this supplier (config/formats/<name>.yaml).'),
        'description_starts_with_tariff': ('warn',
            'A WorkSheet PDF was likely processed as a source invoice. WorkSheets '
            'concatenate tariff codes and item descriptions; they are not invoice data.'),
        'description_too_short': ('warn',
            'Item description is too short to be a real product name. '
            'Re-check the format YAML and source PDF text extraction.'),
        'payment_line_as_item': ('block',
            'Payment-method/Paypal lines must not be ingested as items. '
            'Add the payment-line marker to the format YAML skip rules.'),
        'generic_purchase_description': ('warn',
            'Description is a marketplace-name placeholder, not the actual product. '
            'Improve the format YAML so it captures the real product title from the PDF.'),
        'quantity_zero_with_cost': ('warn',
            'Quantity=0 with non-zero TotalCost indicates failed quantity extraction. '
            'Check the format YAML quantity capture group.'),
        'garbage_description_pattern': ('warn',
            'Description is OCR garbage / keymash. '
            'Try a higher-resolution OCR pass on this page or update the format YAML.'),
        'duplicate_items_in_xlsx': ('warn',
            'Same item appears multiple times in one XLSX. May be legitimate (two of the '
            'same product) or duplicate extraction. Verify against source PDF.'),
        'negative_invoice_total': ('block',
            'InvoiceTotal must be > 0. Re-check the source invoice OCR.'),
        'llm_auto_format_used': ('warn',
            'No format YAML matched — pipeline auto-generated extraction via LLM. '
            'Promote the auto-generated spec into a real config/formats/*.yaml so future '
            'invoices use deterministic extraction.'),
    }

    for kind, rows in kind_counter.items():
        sev, hint = KIND_DEFAULTS.get(kind, ('warn', ''))
        # Deduplicate file/detail tuples to keep message terse
        sample = rows[:5]   # magic-ok: max 5 examples in message body
        sample_str = '; '.join(
            f'{f}{":r" + str(r) if r else ""} — {d}'   # magic-ok: file:row separator
            for (f, r, d) in sample
        )
        fail_fn(f'item_{kind}', sev,   # magic-ok: check key prefix
                f'{len(rows)} occurrence(s) of "{kind}" across XLSX attachment(s).',
                f'xlsx_items.{kind}', f'{len(rows)} hits',   # magic-ok: validation field key
                f'{hint}\nExamples: {sample_str}')


def _llm_review_email_params(email_params: dict, validation: dict, fail_fn) -> None:
    """
    Use LLM to review email params for obvious data quality issues that
    rule-based checks might miss (e.g. invoice text in address fields,
    garbled OCR in consignee name, nonsensical values).
    """
    try:
        from core.llm_client import get_llm_client
        client = get_llm_client()
    except Exception:
        return  # LLM not available, skip

    # Build a concise summary of what the email will contain
    review_data = {
        'waybill': email_params.get('waybill', ''),
        'consignee_name': email_params.get('consignee_name', ''),
        'consignee_address': str(email_params.get('consignee_address', ''))[:300],   # magic-ok: LLM-prompt address truncation width
        'consignee_code': email_params.get('consignee_code', ''),
        'packages': email_params.get('packages', ''),
        'weight': email_params.get('weight', ''),
        'freight': email_params.get('freight', ''),
        'total_invoices': email_params.get('total_invoices', ''),
        'man_reg': email_params.get('man_reg', ''),
    }
    # Add attachment filenames
    att_paths = email_params.get('attachment_paths', [])
    review_data['attachment_files'] = [os.path.basename(p) for p in att_paths]

    system_prompt = (_PROMPT_DIR / "xlsx_validator_email_review_system.txt").read_text(encoding="utf-8")   # magic-ok: filename of LLM email-review system prompt

    user_msg = f"Review this shipment email data:\n{json.dumps(review_data, indent=2)}"

    response = client.call(user_msg, system_prompt=system_prompt, max_tokens=500)   # magic-ok: LLM token budget for email review
    if not response:
        return

    # Parse LLM response
    try:
        # Extract JSON from response (may have markdown fencing)
        json_str = response.strip()
        if '```' in json_str:   # magic-ok: markdown fence marker
            json_str = json_str.split('```')[1]   # magic-ok: markdown fence split index
            if json_str.startswith('json'):   # magic-ok: markdown fence language tag
                json_str = json_str[4:]   # magic-ok: len('json') strip offset
            json_str = json_str.strip()
        result = json.loads(json_str)

        if not result.get('ok', True):
            for issue in result.get('issues', []):
                fail_fn(
                    f'llm_review_{issue.get("field", "unknown")}',   # magic-ok: fail-event key default
                    issue.get('severity', 'warn'),   # magic-ok: severity default ('warn')
                    f'[LLM Review] {issue.get("message", "Data quality issue")}',   # magic-ok: fail-message default
                    issue.get('field', ''),
                    str(review_data.get(issue.get('field', ''), '')),
                    'Review the source PDFs and correct the email params.'   # magic-ok: remediation instruction
                )
    except (json.JSONDecodeError, KeyError, TypeError):
        logger.debug(f"Could not parse LLM review response: {response[:200]}")   # magic-ok: debug-log truncation width


def shipment_checklist(email_params: dict, validation: dict = None) -> dict:
    """
    Final gate before sending a shipment email.
    Returns {passed: bool, failures: [{check, severity, message, field, value, fix_hint}]}

    Severity levels:
      - 'block'   : email MUST NOT be sent until fixed
      - 'warn'    : email can be sent but issue should be flagged
    """
    failures = []

    def fail(check, severity, message, field='', value='', fix_hint=''):
        failures.append({
            'check': check,
            'severity': severity,
            'message': message,
            'field': field,
            'value': str(value),
            'fix_hint': fix_hint,
        })

    # ── 1. Waybill / BL Number ──
    waybill = str(email_params.get('waybill', '')).strip()
    if not waybill:
        fail('waybill_missing', 'block',   # magic-ok: check key + severity
             'Waybill/BL number is empty. Every shipment must have a BL number.',   # magic-ok: fail message
             'waybill', waybill,   # magic-ok: email-params field key
             'Read the BL PDF to extract the BL number. Look for "B/L No" or "BILL OF LADING".')   # magic-ok: remediation instruction
    elif len(waybill) < 6:   # magic-ok: min BL-number length threshold
        fail('waybill_short', 'warn',   # magic-ok: check key + severity
             f'Waybill "{waybill}" is unusually short ({len(waybill)} chars). Typical BL numbers are 10+ chars.',
             'waybill', waybill,   # magic-ok: email-params field key
             'Verify BL number against the BL PDF.')   # magic-ok: remediation instruction

    # ── 2. Weight ──
    weight_str = str(email_params.get('weight', '0')).strip()   # magic-ok: default weight string '0' when missing
    try:
        weight = float(weight_str)
    except ValueError:
        weight = 0
    if weight <= 0:
        fail('weight_zero', 'block',   # magic-ok: check key + severity
             f'Weight is {weight_str} kg. A shipment cannot weigh 0 kg.',
             'weight', weight_str,   # magic-ok: email-params field key
             'Read the BL PDF and look for GRAND TOTAL line with weight in kg. '   # magic-ok: remediation instruction
             'Update _email_params.json field "weight" with the correct value.')   # magic-ok: remediation instruction

    # ── 3. Packages ──
    pkg_str = str(email_params.get('packages', '0')).strip()   # magic-ok: default package string '0' when missing
    try:
        packages = int(pkg_str)
    except ValueError:
        packages = 0
    if packages <= 0:
        fail('packages_zero', 'block',   # magic-ok: check key + severity
             f'Packages is {pkg_str}. A shipment must have at least 1 package.',
             'packages', pkg_str,   # magic-ok: email-params field key
             'Read the BL PDF GRAND TOTAL line for the package count. '   # magic-ok: remediation instruction
             'Update _email_params.json field "packages" with the correct value.')   # magic-ok: remediation instruction

    # ── 4. Freight ──
    freight_str = str(email_params.get('freight', '0')).strip()   # magic-ok: default freight string '0' when missing
    try:
        freight = float(freight_str)
    except ValueError:
        freight = 0
    if freight <= 0:
        fail('freight_zero', 'warn',   # magic-ok: check key + severity
             f'Freight is ${freight_str}. Ocean freight is typically > $0.',
             'freight', freight_str,   # magic-ok: email-params field key
             'Check the BL cost breakdown for freight charges. '   # magic-ok: remediation instruction
             'If freight is genuinely $0 (e.g. prepaid/included), this may be OK.')   # magic-ok: remediation instruction

    # ── 5. Consignee ──
    consignee_name = str(email_params.get('consignee_name', '')).strip()
    CONSIGNEE_JUNK = ('unknown', 'consignee name not found', 'none',                # magic-ok: consignee-placeholder sentinel values
                      'charges will be billed to consignee',                        # magic-ok: consignee-placeholder sentinel values
                      'to order', 'to the order of', 'same as above')               # magic-ok: consignee-placeholder sentinel values
    if not consignee_name or consignee_name.lower() in CONSIGNEE_JUNK:
        fail('consignee_missing', 'block',   # magic-ok: check key + severity
             f'Consignee name is missing or placeholder: "{consignee_name}".',
             'consignee_name', consignee_name,   # magic-ok: email-params field key
             'Read the BL PDF or invoice headers to find the consignee/ship-to name.')   # magic-ok: remediation instruction

    # Consignee name sanity: should be a short company name, not invoice text
    if consignee_name and len(consignee_name) > 60:   # magic-ok: consignee-name suspicious-length threshold
        fail('consignee_name_garbage', 'block',   # magic-ok: check key + severity
             f'Consignee name is suspiciously long ({len(consignee_name)} chars) — '
             f'likely contains invoice text instead of a company name.',
             'consignee_name', consignee_name[:80] + '...',   # magic-ok: email-params field key + display truncation width
             'Read the BL PDF to find the correct consignee name. '   # magic-ok: remediation instruction
             'Update _email_params.json field "consignee_name".')   # magic-ok: remediation instruction

    # Consignee address sanity: should be a short address, not invoice dump
    consignee_address = str(email_params.get('consignee_address', '')).strip()
    if consignee_address and len(consignee_address) > 200:   # magic-ok: consignee-address suspicious-length threshold
        fail('consignee_address_garbage', 'block',   # magic-ok: check key + severity
             f'Consignee address is suspiciously long ({len(consignee_address)} chars) — '
             f'likely contains invoice text instead of an address.',
             'consignee_address', consignee_address[:80] + '...',   # magic-ok: email-params field key + display truncation width
             'Read the BL PDF to find the correct consignee address. '   # magic-ok: remediation instruction
             'Update _email_params.json field "consignee_address".')   # magic-ok: remediation instruction


    # ── 6. Invoice count ──
    total_invoices = email_params.get('total_invoices', 0)
    if not total_invoices or int(total_invoices) < 1:
        fail('no_invoices', 'block',   # magic-ok: check key + severity
             f'Total invoices is {total_invoices}. At least 1 invoice is required.',
             'total_invoices', str(total_invoices),   # magic-ok: email-params field key
             'Check the pipeline output for failed invoice processing.')   # magic-ok: remediation instruction

    # ── 7. Attachments ──
    attachments = email_params.get('attachment_paths', [])
    if not attachments:
        fail('no_attachments', 'block',   # magic-ok: check key + severity
             'No attachment files found. Email must have XLSX/PDF attachments.',   # magic-ok: fail message
             'attachment_paths', '[]',   # magic-ok: email-params field key + empty-list display
             'Check output directory for generated XLSX and PDF files.')   # magic-ok: remediation instruction
    else:
        def _path_exists(p):
            """Check if path exists, handling Windows paths on WSL."""
            if os.path.exists(p):
                return True
            # On WSL, Windows paths like C:\... need translation to /mnt/c/...
            if len(p) >= 3 and p[1] == ':' and p[2] in ('\\', '/'):   # magic-ok: Windows drive-letter path shape (e.g. "C:\")
                wsl_path = '/mnt/' + p[0].lower() + '/' + p[3:].replace('\\', '/')   # magic-ok: WSL mount prefix + drive-letter offset
                return os.path.exists(wsl_path)
            return False
        missing = [p for p in attachments if not _path_exists(p)]
        if missing:
            fail('missing_attachments', 'block',   # magic-ok: check key + severity
                 f'{len(missing)} attachment file(s) not found on disk.',
                 'attachment_paths', f'{len(missing)} missing',   # magic-ok: email-params field key
                 f'Missing files: {", ".join(os.path.basename(p) for p in missing[:5])}')   # magic-ok: max missing-file names shown

    # ── 8. Unfixed validation issues ──
    if validation:
        unfixed = validation.get('unfixed', 0)
        total_issues = validation.get('total_issues', 0)
        if unfixed > 0:
            # Count blocking vs warning issues
            tariff_unfixed = 0
            variance_unfixed = 0
            for pf in validation.get('per_file', []):
                for rem in pf.get('remaining', []):
                    if rem.get('type') in ('MISSING_TARIFF', 'INVALID_TARIFF'):   # magic-ok: issue-type enum values (string-compared downstream)
                        tariff_unfixed += 1
                    elif rem.get('type') == 'VARIANCE':   # magic-ok: issue-type enum value
                        variance_unfixed += 1

            if tariff_unfixed > 0:
                # Collect details for LLM context
                tariff_details = []
                for pf in validation.get('per_file', []):
                    for rem in pf.get('remaining', []):
                        if rem.get('type') in ('MISSING_TARIFF', 'INVALID_TARIFF'):   # magic-ok: issue-type enum values
                            tariff_details.append(
                                f'  {pf["file"]} row {rem["row"]}: "{rem.get("description", "")[:50]}" '   # magic-ok: description display truncation width
                                f'[current: {rem.get("current", "none")}]')   # magic-ok: display fallback for missing 'current' value
                detail_str = '\n'.join(tariff_details[:20])  # magic-ok: cap at 20 tariff-detail lines for readability
                fail('unfixed_tariff', 'block',   # magic-ok: check key + severity
                     f'{tariff_unfixed} item(s) have invalid/missing tariff codes.',
                     'validation.tariff', str(tariff_unfixed),   # magic-ok: validation-summary field key
                     f'For each item below, use lookup_tariff with the item description to find '
                     f'the correct 8-digit CET end-node code. Then edit the XLSX file: set column F '
                     f'(TariffCode) and column AK (GroupBy) to the new code. After fixing, run '
                     f'validate_xlsx to confirm.\n{detail_str}')

            if variance_unfixed > 0:
                var_details = []
                # Per-block variance issues on a combined multi-invoice XLSX
                # are downgraded to a warning: invoice_processor already ran
                # variance_fixer per-invoice before combining, so any
                # residual variance is an OCR / source-data quality issue
                # that the operator must inspect directly in the combined
                # XLSX. Blocking the email would just prevent them from
                # seeing the problematic attachments.
                has_per_block = False
                for pf in validation.get('per_file', []):
                    for rem in pf.get('remaining', []):
                        if rem.get('type') == 'VARIANCE':   # magic-ok: issue-type enum value
                            if rem.get('block'):
                                has_per_block = True
                            var_details.append(
                                f'  {pf["file"]}: {rem.get("why", "")}')
                detail_str = '\n'.join(var_details[:10])   # magic-ok: cap at 10 variance-detail lines
                severity = 'warn' if has_per_block else 'block'   # magic-ok: severity labels
                fail('unfixed_variance', severity,   # magic-ok: check key
                     f'{variance_unfixed} block(s)/file(s) have unresolved variance.',
                     'validation.variance', str(variance_unfixed),   # magic-ok: validation-summary field key
                     f'For each file: read_file the XLSX to see line items and totals (row 2 cols '
                     f'S=InvoiceTotal, T=Freight, U=Insurance, V=Tax, W=Deduction). Read the source '
                     f'invoice PDF to find unextracted charges. Use edit_file to correct values. '
                     f'Variance = InvoiceTotal - (SumItems + Freight + Insurance + Tax - Deduction). '
                     f'After fixing, run validate_xlsx to confirm.\n{detail_str}')

        # Package mismatch from validation
        pkg_check = validation.get('package_check', {})
        if pkg_check.get('mismatch') and pkg_check.get('source') != 'default (1 per entry)':   # magic-ok: expected-packages source label (skip defaulted)
            per_file_str = '\n'.join(
                f'  {pf["file"]}: {pf["packages"]} pkg'
                for pf in pkg_check.get('per_file', [])[:15]   # magic-ok: cap per-file package-list at 15 entries
            )
            fail('package_mismatch', 'block',   # magic-ok: check key + severity
                 f'Package count mismatch: XLSX total={pkg_check.get("actual")}, '
                 f'expected={pkg_check.get("expected")} (from {pkg_check.get("source")}).',
                 'packages', str(pkg_check.get('actual')),   # magic-ok: email-params field key
                 f'Read the BL PDF to find the correct package count per invoice. Each XLSX has a '
                 f'Packages value in column X row 2. The sum across all XLSX files must equal the '
                 f'BL total. Edit the XLSX files to set the correct package count per invoice.\n'
                 f'{per_file_str}')

        # Duplicate content
        dup_check = validation.get('duplicate_check', {})
        if dup_check.get('has_issues'):
            content_dups = dup_check.get('content_duplicates', [])
            dup_count = len(content_dups)
            if dup_count > 0:
                dup_details = []
                for cd in content_dups:
                    files = ', '.join(cd.get('files', []))
                    sources = ', '.join(cd.get('sources', []))
                    dup_details.append(f'  {cd["type"]}: {files} (from PDFs: {sources})')
                detail_str = '\n'.join(dup_details[:10])   # magic-ok: cap at 10 duplicate-detail lines
                fail('duplicate_content', 'block',   # magic-ok: check key + severity
                     f'{dup_count} duplicate content issue(s) detected across XLSX files.',
                     'validation.duplicates', str(dup_count),   # magic-ok: validation-summary field key
                     f'Compare the listed duplicate XLSX files. If they have the same invoice number '
                     f'and items, delete the one with fewer items or the one from the wrong format. '
                     f'Use list_files to see the output directory, read_file to compare contents, '
                     f'and delete the duplicate file. Then update _email_params.json to remove it '
                     f'from attachment_paths.\n{detail_str}')

    # ── 9. Extra rule-based email_params checks (added 2026-04-25) ──
    try:
        _checklist_email_params_extra(email_params, fail)
    except Exception as e:
        logger.warning(f"Extra email_params checks failed (non-blocking): {e}")

    # ── 10. XLSX item-level inspection (added 2026-04-25) ──
    try:
        _checklist_xlsx_inspection(email_params, fail)
    except Exception as e:
        logger.warning(f"XLSX item inspection failed (non-blocking): {e}")

    # ── 11. LLM sanity review of email params ──
    try:
        _llm_review_email_params(email_params, validation, fail)
    except Exception as e:
        logger.warning(f"LLM email review failed (non-blocking): {e}")

    # ── Summary ──
    has_blockers = any(f['severity'] == 'block' for f in failures)   # magic-ok: severity value 'block'
    passed = not has_blockers

    # Print checklist report
    if failures:
        print()
        print("SHIPMENT CHECKLIST")   # magic-ok: report section header
        print("-" * 80)   # magic-ok: report rule width (characters)
        for f in failures:
            icon = "BLOCK" if f['severity'] == 'block' else " WARN"   # magic-ok: report severity icons + severity value
            print(f"  [{icon}] {f['check']}: {f['message']}")
        print("-" * 80)   # magic-ok: report rule width (characters)
        if passed:
            print(f"  PASSED with {len(failures)} warning(s). Email can be sent.")
        else:
            blockers = sum(1 for f in failures if f['severity'] == 'block')   # magic-ok: severity value 'block'
            print(f"  FAILED: {blockers} blocker(s). Email will NOT be sent until fixed.")
        print()

    return {
        'passed': passed,
        'failures': failures,
        'blocker_count': sum(1 for f in failures if f['severity'] == 'block'),   # magic-ok: severity value 'block'
        'warning_count': sum(1 for f in failures if f['severity'] == 'warn'),   # magic-ok: severity value 'warn'
    }


# ── Helpers ────────────────────────────────────────────────────────────────

def _load_cet_codes(base_dir: str) -> set:
    """Load CET valid codes. Try classifier module first, then direct file read."""
    try:
        from classifier import _load_cet_valid_codes
        return _load_cet_valid_codes(base_dir)
    except ImportError:
        pass

    # Fallback: direct file read
    codes = set()
    cet_path = os.path.join(base_dir, 'data', 'cet_valid_codes.txt')   # magic-ok: CET valid-codes data file path
    try:
        if os.path.exists(cet_path):
            with open(cet_path, 'r', encoding='utf-8') as f:
                for line in f:
                    code = line.strip().split('\t')[0]
                    if len(code) == 8 and code.isdigit():   # magic-ok: HS/CET tariff codes are always 8 digits
                        codes.add(code)
    except Exception as e:
        logger.warning(f"Failed to load CET codes: {e}")
    return codes
