"""Audit VARIANCE CHECK cell across every XLSX in the regression corpus.

Per docs/V2_REQUIREMENTS.md R3.1.3 / L2, every XLSX must carry the VARIANCE
CHECK as an Excel *formula* (starts with ``=``) that evaluates to $0.00.
This script classifies the current state of every .xlsx under
``workspace/output/downloads-regression-emails/`` into four buckets and
reports counts + the file paths in each bucket.

Run:
    .venv/bin/python scripts/audit_variance_check.py
"""

from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parent.parent
CORPUS = REPO_ROOT / "workspace" / "output" / "downloads-regression-emails"

COL_P_TOTAL_COST = 16  # column P — holds the VARIANCE CHECK formula
COL_J_LABEL = 10  # column J — holds row labels per invariants.py:196
VARIANCE_LABEL = "VARIANCE CHECK"


def _find_variance_rows(ws: Worksheet) -> list[int]:
    """Return all rows whose column-J label contains VARIANCE CHECK."""
    target = VARIANCE_LABEL.upper()
    rows: list[int] = []
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=COL_J_LABEL).value
        if isinstance(val, str) and target in val.upper():
            rows.append(row)
    return rows


def classify(xlsx_path: Path) -> tuple[str, str]:
    """Return (bucket, detail) for one XLSX file.

    Buckets:
      formula_zero        — formula text + cached value == 0 (or very near)
      formula_nonzero     — formula text + non-zero cached value
      value_zero          — literal 0 (historical corruption)
      value_nonzero       — literal non-zero value
      no_variance_row     — no VARIANCE CHECK label found (aggregate sheets)
      error               — file could not be opened
    """
    try:
        wb = load_workbook(xlsx_path, data_only=False)
        wb_vals = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        return "error", f"open failed: {e}"

    try:
        worst_bucket = "formula_zero"
        worst_detail = ""
        found_any_row = False

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws_v = wb_vals[sheet_name]
            rows = _find_variance_rows(ws)
            if not rows:
                continue
            for row in rows:
                found_any_row = True
                raw = ws.cell(row=row, column=COL_P_TOTAL_COST).value
                cached = ws_v.cell(row=row, column=COL_P_TOTAL_COST).value

                is_formula = isinstance(raw, str) and raw.startswith("=")
                near_zero = cached is None or (
                    isinstance(cached, (int, float)) and abs(cached) < 0.01
                )

                if is_formula and near_zero:
                    bucket = "formula_zero"
                elif is_formula and not near_zero:
                    bucket = "formula_nonzero"
                elif not is_formula and near_zero:
                    bucket = "value_zero"
                else:
                    bucket = "value_nonzero"

                # Escalate to worst across sheets + rows.
                severity = {
                    "formula_zero": 0,
                    "value_zero": 1,
                    "formula_nonzero": 2,
                    "value_nonzero": 3,
                }
                if severity[bucket] > severity[worst_bucket]:
                    worst_bucket = bucket
                    worst_detail = f"sheet={sheet_name} row={row} raw={raw!r} cached={cached!r}"

        if not found_any_row:
            return "no_variance_row", ""
        return worst_bucket, worst_detail
    finally:
        wb.close()
        wb_vals.close()


def main() -> int:
    if not CORPUS.is_dir():
        print(f"corpus not found: {CORPUS}", file=sys.stderr)
        return 2

    xlsx_files = sorted(CORPUS.rglob("*.xlsx"))
    if not xlsx_files:
        print(f"no .xlsx files under {CORPUS}", file=sys.stderr)
        return 2

    buckets: dict[str, list[tuple[Path, str]]] = defaultdict(list)
    for path in xlsx_files:
        bucket, detail = classify(path)
        buckets[bucket].append((path, detail))

    print(f"Audited {len(xlsx_files)} XLSX files under {CORPUS.relative_to(REPO_ROOT)}\n")

    order = [
        "formula_zero",
        "value_zero",
        "formula_nonzero",
        "value_nonzero",
        "no_variance_row",
        "error",
    ]
    for bucket in order:
        entries = buckets.get(bucket, [])
        print(f"{bucket}: {len(entries)}")
    print()

    # Detail listings for the buckets that need attention.
    for bucket in ("value_zero", "value_nonzero", "formula_nonzero", "error"):
        entries = buckets.get(bucket, [])
        if not entries:
            continue
        print(f"── {bucket} ─" + "─" * 60)
        for path, detail in entries:
            rel = path.relative_to(REPO_ROOT)
            line = f"  {rel}"
            if detail:
                line += f"  | {detail}"
            print(line)
        print()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
