"""Arithmetically evaluate the VARIANCE CHECK formula in each XLSX.

Reading cached values via ``data_only=True`` yields ``None`` when Excel has
never opened the file (which is the case for pipeline-generated files that
were never touched by an Excel client). So we cannot rely on the cache to
know whether the formulas *actually* evaluate to zero.

This script parses the minimal arithmetic of the VARIANCE CHECK formula by
substituting referenced cells' resolved numeric values and computing the
result in Python. Only the specific formula patterns produced by the
pipeline are supported (``=S2-P15`` / ``=Sn-Pn`` / ``=S{r}-{col}{r}``):
if a formula doesn't match the expected shapes we report ``unrecognised``
rather than silently passing.

Usage:
    .venv/bin/python scripts/evaluate_variance.py
"""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parent.parent
CORPUS = REPO_ROOT / "workspace" / "output" / "downloads-regression-emails"

COL_P_TOTAL_COST = 16
COL_J_LABEL = 10
VARIANCE_LABEL = "VARIANCE CHECK"

# e.g. "=S2-P15", "=S2-P15-P13"; we accept any sequence of +/- cell refs.
CELL_TERM = re.compile(r"([+-]?)\s*\$?([A-Z]+)\$?(\d+)")


def _find_variance_rows(ws: Worksheet) -> list[int]:
    rows: list[int] = []
    target = VARIANCE_LABEL.upper()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=COL_J_LABEL).value
        if isinstance(v, str) and target in v.upper():
            rows.append(r)
    return rows


def _resolve(ws: Worksheet, col: str, row: int, depth: int = 0) -> float | None:
    """Resolve a cell reference to a numeric value, recursing through formulas.

    Returns None when the cell can't be resolved to a number (text, cycle,
    unsupported formula).
    """
    if depth > 15:
        return None
    try:
        c = column_index_from_string(col)
    except Exception:
        return None
    val = ws.cell(row=row, column=c).value
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str) and val.startswith("="):
        return _eval_formula(ws, val, depth + 1)
    # Try to coerce numeric-looking text.
    try:
        return float(str(val).replace(",", "").replace("$", "").strip())
    except ValueError:
        return None


def _eval_formula(ws: Worksheet, formula: str, depth: int = 0) -> float | None:
    """Minimal evaluator for the +/-/cellref/SUM forms the pipeline emits."""
    body = formula.lstrip("=").strip()
    if not body:
        return None

    # Handle SUM(range) by expanding the range and summing.
    sum_match = re.fullmatch(
        r"SUM\(\s*([A-Z]+)(\d+)\s*:\s*([A-Z]+)(\d+)\s*\)(?:\s*-\s*(.+))?", body
    )
    if sum_match:
        col1, r1, col2, r2, tail = sum_match.groups()
        c1 = column_index_from_string(col1)
        c2 = column_index_from_string(col2)
        total = 0.0
        for c in range(min(c1, c2), max(c1, c2) + 1):
            for r in range(int(r1), int(r2) + 1):
                col_letter = ws.cell(row=r, column=c).column_letter
                v = _resolve(ws, col_letter, r, depth + 1)
                if v is None:
                    return None
                total += v
        if tail:
            tail_val = _eval_formula(ws, "=" + tail, depth + 1)
            if tail_val is None:
                return None
            total -= tail_val
        return total

    # Drop surrounding parens.
    while body.startswith("(") and body.endswith(")"):
        body = body[1:-1].strip()

    # Linear +/- of cell refs.
    total = 0.0
    pos = 0
    seen_any = False
    while pos < len(body):
        m = CELL_TERM.match(body, pos)
        if not m:
            # Ignore whitespace.
            if body[pos].isspace():
                pos += 1
                continue
            return None
        seen_any = True
        sign_s, col, row_s = m.groups()
        sign = -1.0 if sign_s == "-" else 1.0
        v = _resolve(ws, col, int(row_s), depth + 1)
        if v is None:
            return None
        total += sign * v
        pos = m.end()
    return total if seen_any else None


def main() -> int:
    files = sorted(CORPUS.rglob("*.xlsx"))
    if not files:
        print(f"no .xlsx under {CORPUS}")
        return 2

    buckets: dict[str, list[tuple[Path, str]]] = defaultdict(list)
    for path in files:
        try:
            wb = load_workbook(path, data_only=False)
        except Exception as e:
            buckets["error"].append((path, f"open failed: {e}"))
            continue
        worst = "zero"  # best-case default
        worst_detail = ""
        found = False
        try:
            for name in wb.sheetnames:
                ws = wb[name]
                for r in _find_variance_rows(ws):
                    found = True
                    raw = ws.cell(row=r, column=COL_P_TOTAL_COST).value
                    if not (isinstance(raw, str) and raw.startswith("=")):
                        if worst in ("zero", "nonzero"):
                            worst = "non_formula"
                            worst_detail = f"sheet={name} row={r} raw={raw!r}"
                        continue
                    val = _eval_formula(ws, raw)
                    if val is None:
                        if worst == "zero":
                            worst = "unrecognised"
                            worst_detail = f"sheet={name} row={r} formula={raw!r}"
                    elif abs(val) < 0.01:
                        pass  # still zero
                    else:
                        worst = "nonzero"
                        worst_detail = f"sheet={name} row={r} formula={raw!r} → {val:.4f}"
        finally:
            wb.close()
        if not found:
            buckets["no_variance_row"].append((path, ""))
        else:
            buckets[worst].append((path, worst_detail))

    print(f"Evaluated {len(files)} XLSX files arithmetically\n")
    for bucket in (
        "zero",
        "nonzero",
        "non_formula",
        "unrecognised",
        "no_variance_row",
        "error",
    ):
        entries = buckets.get(bucket, [])
        print(f"  {bucket:20s} {len(entries)}")
    print()

    for bucket in ("nonzero", "non_formula", "unrecognised", "error"):
        entries = buckets.get(bucket, [])
        if not entries:
            continue
        print(f"── {bucket} " + "─" * 60)
        for path, detail in entries:
            rel = path.relative_to(REPO_ROOT)
            print(f"  {rel} | {detail}")
        print()

    return 0 if not (buckets.get("nonzero") or buckets.get("error")) else 1


if __name__ == "__main__":
    raise SystemExit(main())
