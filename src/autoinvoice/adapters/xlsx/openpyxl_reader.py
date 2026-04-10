"""XLSX reader adapter using openpyxl."""

from __future__ import annotations

from typing import Any

from openpyxl import load_workbook


class OpenpyxlXlsxReader:
    """Reads XLSX workbooks and returns structured data."""

    def read_workbook(self, xlsx_path: str) -> dict[str, Any]:
        """Read all sheets, cells, and metadata from an XLSX file.

        Args:
            xlsx_path: Path to the XLSX file.

        Returns:
            Dictionary with ``sheets`` key containing per-sheet data.
        """
        wb = load_workbook(xlsx_path, data_only=True)
        result: dict[str, Any] = {"sheets": {}}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))

            result["sheets"][sheet_name] = {
                "rows": rows,
                "row_count": ws.max_row,
                "column_count": ws.max_column,
            }

        return result
