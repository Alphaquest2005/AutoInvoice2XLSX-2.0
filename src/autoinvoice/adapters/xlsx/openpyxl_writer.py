"""XLSX writer adapter using openpyxl."""

from __future__ import annotations

from decimal import Decimal
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from autoinvoice.domain.models.grouping import GroupedInvoice
    from autoinvoice.domain.models.xlsx_spec import ColumnDef, SheetSpec


class OpenpyxlXlsxWriter:
    """Generates and validates XLSX workbooks using openpyxl."""

    # ------------------------------------------------------------------
    # Generation
    # ------------------------------------------------------------------

    def generate(
        self,
        invoices: list[GroupedInvoice],
        output_path: str,
        column_spec: SheetSpec,
    ) -> str:
        """Generate an XLSX file from grouped invoices.

        Args:
            invoices: Grouped invoices to write.
            output_path: Destination file path.
            column_spec: Column and sheet layout specification.

        Returns:
            Path to the generated XLSX file.
        """
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = column_spec.sheet_name

        columns = column_spec.columns

        # -- Header row --
        self._write_header_row(ws, columns, column_spec.start_row)

        current_row = column_spec.start_row + 1

        # -- Invoice data rows --
        for invoice in invoices:
            for group in invoice.groups:
                # Group header row
                self._write_group_header_row(ws, columns, current_row, group, invoice)
                current_row += 1

                # Detail rows
                for classification in group.items:
                    self._write_detail_row(ws, columns, current_row, classification)
                    current_row += 1

        # -- Column widths --
        for col_def in columns:
            col_letter = get_column_letter(col_def.index + 1)
            ws.column_dimensions[col_letter].width = col_def.width

        wb.save(output_path)
        return output_path

    # ------------------------------------------------------------------
    # Validation
    # ------------------------------------------------------------------

    def validate(self, xlsx_path: str) -> dict[str, Any]:
        """Validate an XLSX file for structural correctness.

        Args:
            xlsx_path: Path to the XLSX file to validate.

        Returns:
            Dictionary with ``valid``, ``row_count``, and ``variance`` keys.
        """
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        if ws is None:
            return {"valid": False, "row_count": 0, "column_count": 0, "variance": 0.0}

        row_count = ws.max_row
        col_count = ws.max_column

        # Try to compute a basic variance by summing any numeric values in
        # the last column that appear to be totals.  This is a heuristic
        # check; real validation would be format-aware.
        items_total = Decimal("0")
        for row_idx in range(2, row_count + 1):
            for col_idx in range(1, col_count + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(val, (int, float)) and col_idx == col_count:
                    items_total += Decimal(str(val))

        variance = float(items_total)  # no invoice total in standalone file

        return {
            "valid": row_count > 1,
            "row_count": row_count,
            "column_count": col_count,
            "variance": variance,
        }

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _write_header_row(
        self,
        ws: Any,
        columns: tuple[ColumnDef, ...],
        row: int,
    ) -> None:
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", wrap_text=True)

        for col_def in columns:
            cell = ws.cell(row=row, column=col_def.index + 1, value=col_def.name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def _write_group_header_row(
        self,
        ws: Any,
        columns: tuple[ColumnDef, ...],
        row: int,
        group: Any,
        invoice: GroupedInvoice,
    ) -> None:
        group_font = Font(bold=True, size=11)
        group_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for col_def in columns:
            cell = ws.cell(row=row, column=col_def.index + 1)
            cell.font = group_font
            cell.fill = group_fill

            name_lower = col_def.name.lower()
            if "tariff" in name_lower or "code" in name_lower:
                cell.value = group.tariff_code.code
            elif "category" in name_lower:
                cell.value = group.category
            elif "quantity" in name_lower or "qty" in name_lower:
                cell.value = float(group.sum_quantity)
            elif "total" in name_lower:
                cell.value = float(group.sum_total_cost)
                if col_def.data_type == "currency":
                    cell.number_format = "#,##0.00"
            elif "invoice" in name_lower and "number" in name_lower:
                cell.value = invoice.metadata.invoice_number
            elif "supplier" in name_lower:
                cell.value = invoice.metadata.supplier_name

    def _write_detail_row(
        self,
        ws: Any,
        columns: tuple[ColumnDef, ...],
        row: int,
        classification: Any,
    ) -> None:
        item = classification.item
        for col_def in columns:
            cell = ws.cell(row=row, column=col_def.index + 1)
            name_lower = col_def.name.lower()

            if "description" in name_lower:
                cell.value = item.description
            elif "quantity" in name_lower or "qty" in name_lower:
                cell.value = float(item.quantity)
            elif "unit" in name_lower and "cost" in name_lower:
                cell.value = float(item.unit_cost)
                if col_def.data_type == "currency":
                    cell.number_format = "#,##0.00"
            elif "total" in name_lower:
                cell.value = float(item.total_cost)
                if col_def.data_type == "currency":
                    cell.number_format = "#,##0.00"
            elif "sku" in name_lower:
                cell.value = item.sku
            elif "tariff" in name_lower or "code" in name_lower:
                cell.value = classification.tariff_code.code
            elif "confidence" in name_lower:
                cell.value = classification.confidence
            elif "source" in name_lower:
                cell.value = classification.source
