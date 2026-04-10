"""Tests for the openpyxl XLSX writer adapter."""

from __future__ import annotations

from decimal import Decimal
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from autoinvoice.adapters.xlsx.openpyxl_writer import OpenpyxlXlsxWriter
from autoinvoice.domain.models.classification import Classification, TariffCode
from autoinvoice.domain.models.grouping import GroupedInvoice, ItemGroup
from autoinvoice.domain.models.invoice import InvoiceItem, InvoiceMetadata
from autoinvoice.domain.models.xlsx_spec import ColumnDef, SheetSpec

if TYPE_CHECKING:
    from pathlib import Path

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_column_spec() -> SheetSpec:
    """Build a minimal SheetSpec for testing."""
    return SheetSpec(
        columns=(
            ColumnDef(name="Tariff Code", index=0, data_type="string", width=14),
            ColumnDef(name="Description", index=1, data_type="string", width=30),
            ColumnDef(name="Quantity", index=2, data_type="number", width=10),
            ColumnDef(name="Unit Cost", index=3, data_type="currency", width=12),
            ColumnDef(name="Total Cost", index=4, data_type="currency", width=12),
        ),
        sheet_name="Invoice",
        start_row=1,
    )


def _make_item(
    description: str = "Widget",
    quantity: str = "10",
    unit_cost: str = "5.00",
    total_cost: str = "50.00",
    sku: str = "W-001",
) -> InvoiceItem:
    return InvoiceItem(
        description=description,
        quantity=Decimal(quantity),
        unit_cost=Decimal(unit_cost),
        total_cost=Decimal(total_cost),
        sku=sku,
    )


def _make_classification(
    item: InvoiceItem | None = None,
    tariff_code: str = "84719000",
    confidence: float = 0.95,
) -> Classification:
    if item is None:
        item = _make_item()
    return Classification(
        item=item,
        tariff_code=TariffCode(tariff_code),
        confidence=confidence,
        source="rules",
        category="PRODUCTS",
    )


def _make_group(
    tariff_code: str = "84719000",
    items: tuple[Classification, ...] | None = None,
) -> ItemGroup:
    if items is None:
        items = (_make_classification(tariff_code=tariff_code),)
    sum_qty = sum(c.item.quantity for c in items)
    sum_total = sum(c.item.total_cost for c in items)
    avg_unit = sum_total / sum_qty if sum_qty else Decimal("0")
    return ItemGroup(
        tariff_code=TariffCode(tariff_code),
        category="PRODUCTS",
        items=items,
        sum_quantity=sum_qty,
        sum_total_cost=sum_total,
        average_unit_cost=avg_unit,
    )


def _make_metadata(
    invoice_number: str = "INV-001",
    invoice_total: str = "150.00",
) -> InvoiceMetadata:
    return InvoiceMetadata(
        invoice_number=invoice_number,
        invoice_date="2026-01-15",
        supplier_name="Acme Corp",
        supplier_code="ACME",
        invoice_total=Decimal(invoice_total),
    )


def _make_grouped_invoice(
    groups: tuple[ItemGroup, ...] | None = None,
    invoice_number: str = "INV-001",
) -> GroupedInvoice:
    if groups is None:
        groups = (_make_group(),)
    total_items = sum(len(g.items) for g in groups)
    return GroupedInvoice(
        metadata=_make_metadata(invoice_number=invoice_number),
        groups=groups,
        total_items=total_items,
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestOpenpyxlXlsxWriterGenerate:
    """Tests for OpenpyxlXlsxWriter.generate."""

    def test_generate_creates_xlsx_file(self, tmp_path: Path) -> None:
        writer = OpenpyxlXlsxWriter()
        out = tmp_path / "output.xlsx"
        spec = _make_column_spec()

        writer.generate([_make_grouped_invoice()], str(out), spec)

        assert out.exists()

    def test_generate_header_row(self, tmp_path: Path) -> None:
        writer = OpenpyxlXlsxWriter()
        out = tmp_path / "output.xlsx"
        spec = _make_column_spec()

        writer.generate([_make_grouped_invoice()], str(out), spec)

        wb = load_workbook(str(out))
        ws = wb.active
        header_values = [ws.cell(row=1, column=i + 1).value for i in range(5)]
        assert header_values == [
            "Tariff Code",
            "Description",
            "Quantity",
            "Unit Cost",
            "Total Cost",
        ]

    def test_generate_group_and_detail_rows(self, tmp_path: Path) -> None:
        writer = OpenpyxlXlsxWriter()
        out = tmp_path / "output.xlsx"
        spec = _make_column_spec()

        item1 = _make_item(description="Widget A", total_cost="50.00")
        item2 = _make_item(
            description="Widget B", quantity="5", unit_cost="10.00", total_cost="50.00"
        )
        cls1 = _make_classification(item=item1)
        cls2 = _make_classification(item=item2)
        group = _make_group(items=(cls1, cls2))
        invoice = _make_grouped_invoice(groups=(group,))

        writer.generate([invoice], str(out), spec)

        wb = load_workbook(str(out))
        ws = wb.active

        # Row 1 = header, Row 2 = group header, Row 3-4 = detail
        assert ws.max_row == 4

        # Group header has tariff code
        assert ws.cell(row=2, column=1).value == "84719000"

        # Detail rows have descriptions
        assert ws.cell(row=3, column=2).value == "Widget A"
        assert ws.cell(row=4, column=2).value == "Widget B"

    def test_generate_returns_output_path(self, tmp_path: Path) -> None:
        writer = OpenpyxlXlsxWriter()
        out = tmp_path / "output.xlsx"
        spec = _make_column_spec()

        result = writer.generate([_make_grouped_invoice()], str(out), spec)

        assert result == str(out)

    def test_generate_empty_invoice(self, tmp_path: Path) -> None:
        """Invoice with no groups produces header-only file."""
        writer = OpenpyxlXlsxWriter()
        out = tmp_path / "output.xlsx"
        spec = _make_column_spec()
        invoice = _make_grouped_invoice(groups=())

        writer.generate([invoice], str(out), spec)

        wb = load_workbook(str(out))
        ws = wb.active
        # Only the header row
        assert ws.max_row == 1

    def test_generate_multiple_groups(self, tmp_path: Path) -> None:
        """Three groups each with 2 items -> 1 header + 3*(1 group + 2 detail) = 10 rows."""
        writer = OpenpyxlXlsxWriter()
        out = tmp_path / "output.xlsx"
        spec = _make_column_spec()

        groups = []
        for code in ("84719000", "39269090", "73181500"):
            i1 = _make_item(description=f"Item {code} A")
            i2 = _make_item(description=f"Item {code} B")
            c1 = _make_classification(item=i1, tariff_code=code)
            c2 = _make_classification(item=i2, tariff_code=code)
            groups.append(_make_group(tariff_code=code, items=(c1, c2)))

        invoice = _make_grouped_invoice(groups=tuple(groups))
        writer.generate([invoice], str(out), spec)

        wb = load_workbook(str(out))
        ws = wb.active
        # 1 header + 3 groups * (1 group_header + 2 detail) = 10
        assert ws.max_row == 10


class TestOpenpyxlXlsxWriterValidate:
    """Tests for OpenpyxlXlsxWriter.validate."""

    def test_validate_valid_xlsx(self, tmp_path: Path) -> None:
        writer = OpenpyxlXlsxWriter()
        out = tmp_path / "output.xlsx"
        spec = _make_column_spec()

        writer.generate([_make_grouped_invoice()], str(out), spec)
        result = writer.validate(str(out))

        assert result["valid"] is True
        assert result["row_count"] == 3  # header + group + detail
        assert "variance" in result
