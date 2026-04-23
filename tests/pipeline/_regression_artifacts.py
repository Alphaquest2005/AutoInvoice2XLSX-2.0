"""Regression-artifact snapshot + diff helpers.

Purpose
-------
Joseph's bug-intake pattern is: run the Downloads folder through the pipeline,
open an XLSX in Excel, spot something weird, describe it. Previously we had no
automated way to say "has this file's output actually *changed* since last run,
or does it just *look* wrong now?".

This module turns every XLSX + ``_email_params.json`` produced by the
Downloads regression test into a stable, canonical snapshot (timestamps,
openpyxl's cosmetic defaults, and cell styles stripped) and compares it to a
golden baseline stored under ``tests/regression_artifacts/downloads/``.

Run semantics
-------------
- **Baseline missing** — current snapshot is written as the new baseline and
  the test passes (first run / new PDF).
- **Baseline present and snapshot matches** — test passes silently.
- **Baseline present and snapshot differs** — the current snapshot is copied
  to ``<baseline_dir>/<stem>/current/`` for side-by-side diffing and the test
  is annotated with the changed keys. The test still passes — artifact drift
  is reported for review, not treated as hard failure, because variance fixes
  legitimately change totals. Promotion to baseline is explicit, via
  ``AUTOINVOICE_UPDATE_GOLDENS=1``.
- **``AUTOINVOICE_UPDATE_GOLDENS=1``** — any snapshot mismatch overwrites the
  baseline and the diff is preserved as ``<stem>.last_diff.json``.

Canonicalisation rules
----------------------
For XLSX we extract: ``{sheet_name -> {"A1": {"value": ..., "formula": ...},
...}}``. Only cells with a value or formula are emitted. Formulas are captured
as their string form (leading ``=`` preserved); values are captured as
``repr()`` of the typed value (so ``1.0`` and ``1`` do not collide).

For ``_email_params.json`` we sort keys and drop any ``attachment_paths``
entries that point outside ``workspace/`` (those are tmp_path paths that rotate
every run).
"""

from __future__ import annotations

import hashlib
import json
import os
import shutil
from pathlib import Path
from typing import Any, Dict

from openpyxl import load_workbook

_REPO_ROOT = Path(__file__).resolve().parents[2]
_BASELINE_DIR = _REPO_ROOT / "tests" / "regression_artifacts" / "downloads"

# When set, diffs overwrite baselines and the old baseline is preserved as
# ``<stem>.last_diff.json`` for audit.
UPDATE_GOLDENS = os.environ.get("AUTOINVOICE_UPDATE_GOLDENS", "").strip() == "1"


def _canonicalise_xlsx(path: Path) -> Dict[str, Any]:
    """Return a deterministic, style-free representation of *path*.

    Only cell values and formulas are retained. Sheets are emitted in natural
    order.
    """
    out: Dict[str, Any] = {"__type__": "xlsx", "sheets": {}}
    wb = load_workbook(str(path), data_only=False)
    try:
        for sheet in wb.worksheets:
            cells: Dict[str, Dict[str, Any]] = {}
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    raw = cell.value
                    if isinstance(raw, str) and raw.startswith("="):
                        cells[cell.coordinate] = {"formula": raw}
                    else:
                        cells[cell.coordinate] = {"value": repr(raw)}
            out["sheets"][sheet.title] = cells
    finally:
        wb.close()
    return out


def _canonicalise_email_params(path: Path) -> Dict[str, Any]:
    """Normalise ``_email_params.json`` so tmp_path-rotating attachments and
    ordering noise don't show up as false-positive diffs.
    """
    data = json.loads(path.read_text(encoding="utf-8"))
    attachments = data.get("attachment_paths") or []
    # Keep only basenames — tmp_path directories rotate every run.
    data["attachment_paths"] = sorted(Path(p).name for p in attachments)
    return {"__type__": "email_params", "data": data}


def _hash_snapshot(snapshot: Dict[str, Any]) -> str:
    encoded = json.dumps(snapshot, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _build_snapshot(output_dir: Path) -> Dict[str, Any]:
    """Collect + canonicalise every XLSX and email-params file under *output_dir*."""
    snapshot: Dict[str, Any] = {}
    for xlsx in sorted(output_dir.rglob("*.xlsx")):
        rel = xlsx.relative_to(output_dir).as_posix()
        try:
            snapshot[rel] = _canonicalise_xlsx(xlsx)
        except Exception as e:  # noqa: BLE001
            snapshot[rel] = {"__type__": "xlsx", "error": str(e)}
    for params in sorted(output_dir.rglob("_email_params.json")):
        rel = params.relative_to(output_dir).as_posix()
        try:
            snapshot[rel] = _canonicalise_email_params(params)
        except Exception as e:  # noqa: BLE001
            snapshot[rel] = {"__type__": "email_params", "error": str(e)}
    return snapshot


def _diff_snapshots(old: Dict[str, Any], new: Dict[str, Any]) -> list[str]:
    """Return a flat list of human-readable differences between *old* and *new*."""
    diffs: list[str] = []
    old_keys = set(old)
    new_keys = set(new)
    for k in sorted(old_keys - new_keys):
        diffs.append(f"REMOVED: {k}")
    for k in sorted(new_keys - old_keys):
        diffs.append(f"ADDED: {k}")
    for k in sorted(old_keys & new_keys):
        if old[k] == new[k]:
            continue
        # Drill one level into XLSX cells to give actionable info.
        o, n = old[k], new[k]
        if isinstance(o, dict) and isinstance(n, dict) and o.get("__type__") == "xlsx":
            for sheet in sorted(set(o.get("sheets", {})) | set(n.get("sheets", {}))):
                o_cells = o.get("sheets", {}).get(sheet, {})
                n_cells = n.get("sheets", {}).get(sheet, {})
                if o_cells == n_cells:
                    continue
                for coord in sorted(set(o_cells) | set(n_cells)):
                    if o_cells.get(coord) != n_cells.get(coord):
                        diffs.append(
                            f"CHANGED: {k}::{sheet}!{coord}  "
                            f"{o_cells.get(coord)!r} → {n_cells.get(coord)!r}"
                        )
        else:
            diffs.append(f"CHANGED: {k} (top-level)")
    return diffs


def snapshot_and_compare(
    pdf_stem: str,
    output_dir: Path,
) -> Dict[str, Any]:
    """Snapshot *output_dir* and diff against the baseline for *pdf_stem*.

    Returns a dict with keys:
      - ``status``  : ``"new"`` | ``"match"`` | ``"drift"``
      - ``hash``    : sha256 of the current snapshot
      - ``diffs``   : list of diff strings (empty on match/new)
      - ``baseline_path`` : path to the baseline file
      - ``current_dir``   : when drift, path where current artifacts are preserved
    """
    _BASELINE_DIR.mkdir(parents=True, exist_ok=True)
    baseline_path = _BASELINE_DIR / f"{pdf_stem}.baseline.json"
    snapshot = _build_snapshot(output_dir)
    current_hash = _hash_snapshot(snapshot)

    result: Dict[str, Any] = {
        "hash": current_hash,
        "baseline_path": str(baseline_path),
        "diffs": [],
        "current_dir": None,
    }

    if not baseline_path.exists():
        baseline_path.write_text(
            json.dumps(snapshot, indent=2, sort_keys=True, default=str),
            encoding="utf-8",
        )
        result["status"] = "new"
        return result

    old_snapshot = json.loads(baseline_path.read_text(encoding="utf-8"))
    old_hash = _hash_snapshot(old_snapshot)
    if old_hash == current_hash:
        result["status"] = "match"
        return result

    diffs = _diff_snapshots(old_snapshot, snapshot)
    result["diffs"] = diffs

    # Preserve the current XLSX + email params side-by-side so the user can
    # open both in Excel. ``current/`` rotates every run; a ``last_diff``
    # copy stays around when UPDATE_GOLDENS promotes the change.
    current_dir = _BASELINE_DIR / pdf_stem / "current"
    if current_dir.exists():
        shutil.rmtree(current_dir)
    current_dir.mkdir(parents=True, exist_ok=True)
    for src in output_dir.rglob("*"):
        if not src.is_file():
            continue
        if src.suffix.lower() not in {".xlsx", ".json"}:
            continue
        if src.suffix.lower() == ".json" and src.name != "_email_params.json":
            continue
        rel = src.relative_to(output_dir)
        dest = current_dir / rel
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dest)
    result["current_dir"] = str(current_dir)

    if UPDATE_GOLDENS:
        # Archive the diff for audit, then promote.
        (_BASELINE_DIR / f"{pdf_stem}.last_diff.json").write_text(
            json.dumps({"diffs": diffs, "old_hash": old_hash, "new_hash": current_hash},
                       indent=2, default=str),
            encoding="utf-8",
        )
        baseline_path.write_text(
            json.dumps(snapshot, indent=2, sort_keys=True, default=str),
            encoding="utf-8",
        )
        result["status"] = "promoted"
        return result

    result["status"] = "drift"
    return result


__all__ = ["snapshot_and_compare", "UPDATE_GOLDENS"]
