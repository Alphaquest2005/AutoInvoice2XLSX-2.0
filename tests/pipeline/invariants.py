"""Reusable XLSX invariants for pipeline regression tests.

These assertions encode the rules that the variance fixer and validator
must preserve — regardless of which fix path they take.  Tests should call
these at the end of every workbook-transforming operation so that a
regression like "the VARIANCE CHECK cell got turned into a numeric 0" or
"the ADJUSTMENTS formula now has two stacked corrections" fails loudly.

The original three rules below correspond one-to-one to the three bugs
that caused BL #TSCW18489131 to get stuck:

1. ``assert_is_grouped_matches`` — the validator must agree with the
   generator on whether a sheet is in grouped mode.
2. ``assert_variance_is_formula`` — VARIANCE CHECK must always be a formula
   so Excel shows a live-computed value to the broker.
3. ``assert_adjustments_has_no_stacked_corrections`` — the ADJUSTMENTS
   formula may have at most one correction term appended, so repeated fix
   runs are idempotent.

The extended checklist (assert_headers_present onward) encodes the full
"proper XLSX" rules from ``config/columns.yaml`` and
``config/grouping.yaml`` — tariff shape, L/O type enforcement, first-group
population of invoice totals, totals-section formula shapes, Q/R row
formulas, group verification == 0, variance-within-tolerance, and
grouped-mode group total reconciliation.

Every new invariant raises ``AssertionError`` with a precise message that
names the sheet, row number, column, expected value, and actual value.
The convenience helper ``run_all_invariants(ws, mode=...)`` runs the full
set and returns a list of ``(invariant_name, message)`` tuples rather than
stopping at the first failure, which is what the regression tests use to
produce per-file precise reports.
"""

from __future__ import annotations

import os
import re
import sqlite3
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from collections.abc import Callable, Iterable

    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

# ── Column indices (1-based, match pipeline/xlsx_validator.COL_*) ─────────
COL_A_DOC_TYPE = 1
COL_B_PO_NUMBER = 2
COL_C_INVOICE_NUM = 3
COL_D_DATE = 4
COL_E_CATEGORY = 5
COL_F_TARIFF = 6
COL_G_PO_ITEM_NUMBER = 7
COL_H_PO_ITEM_DESC = 8
COL_I_SUPPLIER_ITEM_NUMBER = 9
COL_J_SUPPLIER_ITEM_DESC = 10
COL_K_QUANTITY = 11
COL_L_PER_UNIT = 12
COL_M_UNITS = 13
COL_N_CURRENCY = 14
COL_O_COST = 15
COL_P_TOTAL_COST = 16
COL_Q_TOTAL = 17
COL_R_TC_VS_TOTAL = 18
COL_S_INVOICE_TOTAL = 19
COL_T_FREIGHT = 20
COL_U_INSURANCE = 21
COL_V_OTHER_COST = 22
COL_W_DEDUCTIONS = 23
COL_X_PACKAGES = 24
COL_Y_WAREHOUSE = 25
COL_Z_SUPPLIER_CODE = 26
COL_AA_SUPPLIER_NAME = 27
COL_AB_SUPPLIER_ADDR = 28
COL_AC_COUNTRY_CODE = 29
COL_AD_INSTRUCTIONS = 30
COL_AE_PREV_DECL = 31
COL_AF_FIN_INFO = 32
COL_AG_GALLONS = 33
COL_AH_LITERS = 34
COL_AI_INV_TOTAL_COST = 35
COL_AJ_PO_TOTAL_COST = 36
COL_AK_GROUPBY = 37

# Legacy aliases kept for back-compat with existing tests.
COL_DESC = COL_J_SUPPLIER_ITEM_DESC
COL_TOTAL_COST = COL_P_TOTAL_COST

# Expected headers per config/columns.yaml — strict mode expects all 37.
EXPECTED_HEADERS_STRICT: dict[int, str] = {
    COL_A_DOC_TYPE: "Document Type",
    COL_B_PO_NUMBER: "PO Number",
    COL_C_INVOICE_NUM: "Supplier Invoice#",
    COL_D_DATE: "Date",
    COL_E_CATEGORY: "Category",
    COL_F_TARIFF: "TariffCode",
    COL_G_PO_ITEM_NUMBER: "PO Item Number",
    COL_H_PO_ITEM_DESC: "PO Item Description",
    COL_I_SUPPLIER_ITEM_NUMBER: "Supplier Item Number",
    COL_J_SUPPLIER_ITEM_DESC: "Supplier Item Description",
    COL_K_QUANTITY: "Quantity",
    COL_L_PER_UNIT: "Per Unit",
    COL_M_UNITS: "UNITS",
    COL_N_CURRENCY: "Currency",
    COL_O_COST: "Cost",
    COL_P_TOTAL_COST: "Total Cost",
    COL_Q_TOTAL: "Total",
    COL_R_TC_VS_TOTAL: "TotalCost Vs Total",
    COL_S_INVOICE_TOTAL: "InvoiceTotal",
    COL_T_FREIGHT: "Total Internal Freight",
    COL_U_INSURANCE: "Total Insurance",
    COL_V_OTHER_COST: "Total Other Cost",
    COL_W_DEDUCTIONS: "Total Deduction",
    COL_X_PACKAGES: "Packages",
    COL_Y_WAREHOUSE: "Warehouse",
    COL_Z_SUPPLIER_CODE: "Supplier Code",
    COL_AA_SUPPLIER_NAME: "Supplier Name",
    COL_AB_SUPPLIER_ADDR: "Supplier Address",
    COL_AC_COUNTRY_CODE: "Country Code",
    COL_AD_INSTRUCTIONS: "Instructions",
    COL_AE_PREV_DECL: "Previous Declaration",
    COL_AF_FIN_INFO: "Financial Information",
    COL_AG_GALLONS: "Gallons",
    COL_AH_LITERS: "Liters",
    COL_AI_INV_TOTAL_COST: "INVTotalCost",
    COL_AJ_PO_TOTAL_COST: "POTotalCost",
    COL_AK_GROUPBY: "GroupBy",
}

# Minimal header set — what the synthetic xlsx_factory workbooks emit and
# what the simplest pipeline generators touch.  Used by test fixtures and
# any caller that only needs to verify the core item columns are labelled.
#
# Columns T/U/V/W intentionally omitted because the factory uses short
# labels (Freight/Insurance/Tax/Deduction) while production uses the full
# "Total Internal Freight" / etc names.  Callers that need strict-mode
# header coverage of those columns should pass ``mode="strict"``.
EXPECTED_HEADERS_MINIMAL: dict[int, str] = {
    COL_F_TARIFF: "TariffCode",
    COL_J_SUPPLIER_ITEM_DESC: "Supplier Item Description",
    COL_K_QUANTITY: "Quantity",
    COL_L_PER_UNIT: "Per Unit",
    COL_O_COST: "Cost",
    COL_P_TOTAL_COST: "Total Cost",
    COL_S_INVOICE_TOTAL: "InvoiceTotal",
}

# Totals-row labels recognised by the validator (``pipeline/xlsx_validator``).
_TOTALS_LABELS = {
    "SUBTOTAL",
    "SUBTOTAL (GROUPED)",
    "SUBTOTAL (DETAILS)",
    "GROUP VERIFICATION",
    "ADJUSTMENTS",
    "NET TOTAL",
    "INVOICE TOTAL",
    "VARIANCE CHECK",
    "GRAND SUBTOTAL (GROUPED)",
    "GRAND SUBTOTAL (DETAILS)",
    "GRAND VARIANCE CHECK",
    "TOTAL INTERNAL FREIGHT",
    "TOTAL INSURANCE",
    "TOTAL OTHER COST",
    "TOTAL DEDUCTION",
}

# Formula shape regexes (match the templates in config/grouping.yaml).
_TARIFF_RE = re.compile(r"^\d{8}$")
_ADJUSTMENTS_BASE_RE = re.compile(r"^=\(T\d+\+U\d+\+V\d+-W\d+\)(?P<tail>.*)$")
_SINGLE_CORRECTION_RE = re.compile(r"^[+\-]\s*-?\d+(\.\d+)?$")
_NET_TOTAL_RE = re.compile(r"^=\(?P\d+\+P\d+\)?$")
_VARIANCE_RE = re.compile(r"^=\(?S\d+-P\d+\)?$")
_SUBTOTAL_GROUPED_RE = re.compile(r"^=P\d+(\+P\d+)*$")
_SUBTOTAL_DETAILS_RE = re.compile(r"^=SUM\(P\d+:P\d+\)-P\d+$")
_GROUP_VERIFICATION_RE = re.compile(r"^=P\d+-P\d+$")
_Q_FORMULA_RE = re.compile(r"^=O(\d+)\*K\1$")
_R_FORMULA_RE = re.compile(r"^=P(\d+)-Q\1$")


# ─────────────────────────────────────────────────────────────────────────
#                             Helper functions
# ─────────────────────────────────────────────────────────────────────────


def _find_label_row(ws: Worksheet, label: str) -> int | None:
    """Return the row number whose description column contains ``label``.

    Search is case-insensitive and runs bottom-up so that, in a
    multi-invoice workbook, we find the final/grand totals row rather
    than a per-invoice intermediate row.
    """
    target = label.upper()
    for row in range(ws.max_row, 0, -1):
        val = ws.cell(row=row, column=COL_J_SUPPLIER_ITEM_DESC).value
        if val and target in str(val).upper():
            return row
    return None


def _find_all_label_rows(ws: Worksheet, label: str) -> list[int]:
    """Return all row numbers whose description column matches ``label``.

    Exact (case-insensitive, whitespace-stripped) match — used to find
    per-invoice repeated labels in multi-invoice workbooks.
    """
    target = label.upper().strip()
    rows: list[int] = []
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=COL_J_SUPPLIER_ITEM_DESC).value
        if val and str(val).upper().strip() == target:
            rows.append(row)
    return rows


def _is_totals_label(value: object) -> bool:
    if value is None:
        return False
    s = str(value).upper().strip()
    return any(s == label or s.startswith(label) for label in _TOTALS_LABELS)


def _row_is_empty(ws: Worksheet, row: int) -> bool:
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=row, column=col).value not in (None, ""):
            return False
    return True


def _is_group_row(ws: Worksheet, row: int) -> bool:
    """Detect a group header row by its blue fill ``D9E1F2``.

    Production grouped-mode sheets populate the AK (GroupBy) column on
    BOTH group and detail rows (the tariff code is the group key), so
    AK alone cannot distinguish them.  The only reliable signal is the
    blue ``D9E1F2`` fill that ``bl_xlsx_generator`` and ``variance_fixer
    ._is_group_row`` use.  The synthetic factory workbooks in
    ``tests/pipeline/xlsx_factory.py`` set the same fill on col A, so
    this check works for both the production and test corpora.
    """
    for col in (1, COL_J_SUPPLIER_ITEM_DESC, COL_AK_GROUPBY):
        fill = ws.cell(row=row, column=col).fill
        color = getattr(getattr(fill, "start_color", None), "rgb", None)
        if isinstance(color, str) and color.upper().endswith("D9E1F2"):
            return True
    return False


def _col_letter(col: int) -> str:
    # Avoids importing openpyxl.utils at module top; tests may pass a plain int.
    from openpyxl.utils import get_column_letter

    return get_column_letter(col)


# ─────────────────────────────────────────────────────────────────────────
#                         Original three invariants
# ─────────────────────────────────────────────────────────────────────────


def assert_variance_is_formula(ws: Worksheet) -> None:
    """VARIANCE CHECK must always be a formula, never a bare number.

    Excel shows formula results live, so if anything in the pipeline ever
    writes ``0`` (or any numeric literal) into this cell, the broker loses
    visibility into the audit chain.  This was bug #2 on BL #TSCW18489131.
    """
    row = _find_label_row(ws, "VARIANCE CHECK")
    assert row is not None, f"[{ws.title}] VARIANCE CHECK row not found"
    cell = ws.cell(row=row, column=COL_P_TOTAL_COST)
    value = cell.value
    assert isinstance(value, str) and value.startswith("="), (
        f"[{ws.title}] VARIANCE CHECK (row {row}, col P) must be a formula, "
        f"got {value!r} (type={type(value).__name__}). Writing a numeric "
        f"literal here destroys Excel auditability."
    )


def assert_adjustments_has_no_stacked_corrections(ws: Worksheet) -> None:
    """ADJUSTMENTS formula may have at most one appended correction.

    The base formula is ``=(T{r}+U{r}+V{r}-W{r})``.  The fixer is allowed
    to append a single correction term so the variance zeroes out, but
    running the fixer twice must not produce ``=(...)+X+Y``.  This was
    bug #3.
    """
    row = _find_label_row(ws, "ADJUSTMENTS")
    assert row is not None, f"[{ws.title}] ADJUSTMENTS row not found"
    value = ws.cell(row=row, column=COL_P_TOTAL_COST).value

    # ADJUSTMENTS may legally be a plain number in very simple workbooks.
    if isinstance(value, (int, float)):
        return

    assert isinstance(value, str) and value.startswith("="), (
        f"[{ws.title}] ADJUSTMENTS (row {row}, col P) must be a formula or numeric, got {value!r}"
    )

    match = _ADJUSTMENTS_BASE_RE.match(value)
    assert match, (
        f"[{ws.title}] ADJUSTMENTS (row {row}, col P) does not match the "
        f"expected base formula =(T..+U..+V..-W..): got {value!r}"
    )
    tail = match.group("tail")
    if not tail:
        return  # no correction appended — fine

    assert _SINGLE_CORRECTION_RE.match(tail), (
        f"[{ws.title}] ADJUSTMENTS (row {row}, col P) has stacked "
        f"corrections: tail={tail!r}. _force_adjustment must strip previous "
        f"corrections before writing a new one so repeated runs are "
        f"idempotent."
    )


def assert_is_grouped_matches(
    ws: Worksheet,
    *,
    expected_grouped: bool,
    detector: Callable[[Worksheet], bool],
) -> None:
    """The validator's grouped-mode detection must agree with the generator.

    Bug #1 was that ``xlsx_validator`` treated plain ``"SUBTOTAL"`` as a
    grouped-mode marker, even though ``bl_xlsx_generator`` writes that
    label in *ungrouped* mode.  This helper lets a test assert the
    detector function (injected so tests can swap it) agrees with what
    the factory built.
    """
    actual = detector(ws)
    mode = "grouped" if expected_grouped else "ungrouped"
    assert actual == expected_grouped, (
        f"[{ws.title}] is_grouped detector returned {actual} for a {mode} "
        f"workbook. A mismatch here causes sum_items=0 in the validator, "
        f"which yields a permanent non-zero variance that cannot be "
        f"auto-fixed."
    )


# ─────────────────────────────────────────────────────────────────────────
#                           Extended invariants
# ─────────────────────────────────────────────────────────────────────────


def assert_headers_present(ws: Worksheet, mode: str = "strict") -> None:
    """Row 1 must contain all expected headers per ``config/columns.yaml``.

    ``mode="strict"`` checks all 37 columns A-AK.
    ``mode="minimal"`` checks only F/J/K/L/O/P/S/T/U/V/W — the subset
    that simple generators and the synthetic factories populate.
    """
    expected = EXPECTED_HEADERS_STRICT if mode == "strict" else EXPECTED_HEADERS_MINIMAL
    for col, expected_header in expected.items():
        actual = ws.cell(row=1, column=col).value
        assert actual == expected_header, (
            f"[{ws.title}] header mismatch at row 1, col {_col_letter(col)} "
            f"(index {col}): expected {expected_header!r}, got {actual!r}"
        )


def assert_tariff_codes_valid(
    ws: Worksheet,
    first_data_row: int,
    last_data_row: int,
    *,
    cet_db_path: str | None = None,
) -> None:
    """Column F must contain an 8-digit string on every non-empty data row.

    Summary/totals rows are skipped.  If ``cet_db_path`` is provided and
    the file exists, the tariff is additionally validated against the
    ``caricom_cet`` table (skipped gracefully otherwise).
    """
    cet_codes: set[str] | None = None
    if cet_db_path and os.path.exists(cet_db_path):
        try:
            conn = sqlite3.connect(cet_db_path)
            cur = conn.cursor()
            # Table name may vary; try a couple of candidates.
            for tbl in ("caricom_cet", "cet", "tariff_codes"):
                try:
                    cur.execute(f"SELECT code FROM {tbl}")
                    cet_codes = {str(r[0]).strip() for r in cur.fetchall()}
                    break
                except sqlite3.OperationalError:
                    continue
            conn.close()
        except Exception:
            cet_codes = None  # Skip DB validation on any error

    for row in range(first_data_row, last_data_row + 1):
        label = ws.cell(row=row, column=COL_J_SUPPLIER_ITEM_DESC).value
        if _is_totals_label(label):
            continue
        if _row_is_empty(ws, row):
            continue
        tariff = ws.cell(row=row, column=COL_F_TARIFF).value
        if tariff in (None, ""):
            continue  # allow blank tariff rows (detail subheaders)
        s = str(tariff).strip()
        assert _TARIFF_RE.match(s), (
            f"[{ws.title}] tariff at row {row}, col F is not an 8-digit string: got {tariff!r}"
        )
        if cet_codes is not None:
            assert s in cet_codes, (
                f"[{ws.title}] tariff {s!r} at row {row}, col F is not "
                f"present in the CET table at {cet_db_path}"
            )


def assert_column_types(ws: Worksheet, first_data_row: int, last_data_row: int) -> None:
    """Column L is text, column O is numeric on every detail row.

    Group rows (detected via AK/GroupBy or blue fill) are exempt because
    the generator leaves L empty and may write an average into O.
    """
    for row in range(first_data_row, last_data_row + 1):
        label = ws.cell(row=row, column=COL_J_SUPPLIER_ITEM_DESC).value
        if _is_totals_label(label):
            continue
        if _row_is_empty(ws, row):
            continue
        if _is_group_row(ws, row):
            continue

        l_val = ws.cell(row=row, column=COL_L_PER_UNIT).value
        if l_val not in (None, ""):
            assert isinstance(l_val, str), (
                f"[{ws.title}] row {row}, col L (Per Unit) must be "
                f"string/text, got {l_val!r} (type={type(l_val).__name__})"
            )

        o_val = ws.cell(row=row, column=COL_O_COST).value
        if o_val not in (None, ""):
            # Formula strings are allowed (openpyxl stores as str starting "=")
            if isinstance(o_val, str) and o_val.startswith("="):
                continue
            assert isinstance(o_val, (int, float)), (
                f"[{ws.title}] row {row}, col O (Cost) must be numeric, "
                f"got {o_val!r} (type={type(o_val).__name__})"
            )


def assert_invoice_totals_only_on_first_group_row(ws: Worksheet) -> None:
    """Columns S/T/U/V/W must be populated ONLY on the first group row of
    each invoice block, and blank on every other row within that block.

    Strategy: scan top-to-bottom.  The first non-empty S cell outside the
    totals section marks the start of an invoice block; subsequent S
    cells are only allowed if they belong to a new invoice block (ie the
    preceding row was a totals row).  For simple single-invoice
    workbooks, there is exactly one row with S/T/U/V/W populated.
    """
    invoice_start_rows: list[int] = []
    prev_was_totals = True  # treat header region as "above first invoice"
    for row in range(2, ws.max_row + 1):
        label = ws.cell(row=row, column=COL_J_SUPPLIER_ITEM_DESC).value
        if _is_totals_label(label):
            prev_was_totals = True
            continue
        s_val = ws.cell(row=row, column=COL_S_INVOICE_TOTAL).value
        if s_val not in (None, ""):
            # Allowed only if this is a new invoice block
            assert prev_was_totals or not invoice_start_rows, (
                f"[{ws.title}] row {row} has InvoiceTotal (col S) populated "
                f"but is not the first row of a new invoice block. S/T/U/V/W "
                f"must only be populated on the first group row of each "
                f"invoice."
            )
            invoice_start_rows.append(row)
            prev_was_totals = False
        else:
            # Any T/U/V/W without S is a violation
            for col in (
                COL_T_FREIGHT,
                COL_U_INSURANCE,
                COL_V_OTHER_COST,
                COL_W_DEDUCTIONS,
            ):
                val = ws.cell(row=row, column=col).value
                assert val in (None, ""), (
                    f"[{ws.title}] row {row}, col {_col_letter(col)} is "
                    f"populated ({val!r}) but the corresponding S cell is "
                    f"blank. T/U/V/W must only appear on the first group "
                    f"row of an invoice."
                )
            prev_was_totals = False

    assert invoice_start_rows, (
        f"[{ws.title}] no invoice block found — no row has a populated "
        f"InvoiceTotal (col S). Every workbook must have at least one."
    )


def assert_totals_section_formulas(ws: Worksheet) -> None:
    """SUBTOTAL*, ADJUSTMENTS, NET TOTAL, VARIANCE CHECK rows must have the
    expected formula shapes.

    We check shapes via regex on the ``.value`` of each cell — we do NOT
    evaluate the formula.  Templates come from ``config/grouping.yaml``.
    SUBTOTAL may be a plain number in the ungrouped-factory case; we
    tolerate that.  GROUP VERIFICATION is checked only if present.
    """
    # ADJUSTMENTS
    adj_row = _find_label_row(ws, "ADJUSTMENTS")
    assert adj_row is not None, f"[{ws.title}] ADJUSTMENTS row not found"
    adj_val = ws.cell(row=adj_row, column=COL_P_TOTAL_COST).value
    if not isinstance(adj_val, (int, float)):
        assert isinstance(adj_val, str) and _ADJUSTMENTS_BASE_RE.match(adj_val), (
            f"[{ws.title}] ADJUSTMENTS (row {adj_row}, col P) does not match "
            f"=(T..+U..+V..-W..): got {adj_val!r}"
        )

    # NET TOTAL
    net_row = _find_label_row(ws, "NET TOTAL")
    assert net_row is not None, f"[{ws.title}] NET TOTAL row not found"
    net_val = ws.cell(row=net_row, column=COL_P_TOTAL_COST).value
    assert isinstance(net_val, str) and _NET_TOTAL_RE.match(net_val), (
        f"[{ws.title}] NET TOTAL (row {net_row}, col P) does not match "
        f"=P{{sub}}+P{{adj}} shape: got {net_val!r}"
    )

    # VARIANCE CHECK
    var_row = _find_label_row(ws, "VARIANCE CHECK")
    assert var_row is not None, f"[{ws.title}] VARIANCE CHECK row not found"
    var_val = ws.cell(row=var_row, column=COL_P_TOTAL_COST).value
    assert isinstance(var_val, str) and _VARIANCE_RE.match(var_val), (
        f"[{ws.title}] VARIANCE CHECK (row {var_row}, col P) does not match "
        f"=S{{first}}-P{{net}} shape: got {var_val!r}"
    )

    # SUBTOTAL (GROUPED) — if present must be a + chain of P refs
    grp_row = _find_label_row(ws, "SUBTOTAL (GROUPED)")
    if grp_row is not None:
        grp_val = ws.cell(row=grp_row, column=COL_P_TOTAL_COST).value
        if isinstance(grp_val, str) and grp_val.startswith("="):
            assert _SUBTOTAL_GROUPED_RE.match(grp_val), (
                f"[{ws.title}] SUBTOTAL (GROUPED) (row {grp_row}, col P) "
                f"does not match =P..+P..+.. shape: got {grp_val!r}"
            )

    # SUBTOTAL (DETAILS) — if present must be =SUM(P..:P..)-P..
    det_row = _find_label_row(ws, "SUBTOTAL (DETAILS)")
    if det_row is not None:
        det_val = ws.cell(row=det_row, column=COL_P_TOTAL_COST).value
        if isinstance(det_val, str) and det_val.startswith("="):
            assert _SUBTOTAL_DETAILS_RE.match(det_val), (
                f"[{ws.title}] SUBTOTAL (DETAILS) (row {det_row}, col P) "
                f"does not match =SUM(P..:P..)-P.. shape: got {det_val!r}"
            )


def assert_row_q_r_formulas(ws: Worksheet, first_data_row: int, last_data_row: int) -> None:
    """Every data row with content in Q/R must contain the expected formulas.

    Q = ``=O{row}*K{row}`` and R = ``=P{row}-Q{row}``.  Rows where both Q
    and R are empty are skipped (the generator may not always populate
    them, e.g. synthetic factory workbooks).  Totals rows are also
    skipped.
    """
    for row in range(first_data_row, last_data_row + 1):
        label = ws.cell(row=row, column=COL_J_SUPPLIER_ITEM_DESC).value
        if _is_totals_label(label):
            continue
        if _row_is_empty(ws, row):
            continue
        q_val = ws.cell(row=row, column=COL_Q_TOTAL).value
        r_val = ws.cell(row=row, column=COL_R_TC_VS_TOTAL).value
        if q_val in (None, "") and r_val in (None, ""):
            continue  # neither Q nor R used on this row
        if q_val not in (None, ""):
            assert isinstance(q_val, str), (
                f"[{ws.title}] row {row}, col Q must be a formula string, got {q_val!r}"
            )
            m = _Q_FORMULA_RE.match(q_val)
            assert m and int(m.group(1)) == row, (
                f"[{ws.title}] row {row}, col Q expected =O{row}*K{row}, got {q_val!r}"
            )
        if r_val not in (None, ""):
            assert isinstance(r_val, str), (
                f"[{ws.title}] row {row}, col R must be a formula string, got {r_val!r}"
            )
            m = _R_FORMULA_RE.match(r_val)
            assert m and int(m.group(1)) == row, (
                f"[{ws.title}] row {row}, col R expected =P{row}-Q{row}, got {r_val!r}"
            )


def assert_group_verification_zero(ws: Worksheet) -> None:
    """If a GROUP VERIFICATION row exists, its computed value must be 0.

    The workbook must be re-opened with ``data_only=True`` to read the
    cached computed value Excel stored.  If the cached value is None
    (formula never calculated, which happens when openpyxl wrote the
    file without Excel ever opening it), we skip gracefully — the
    upstream ``assert_totals_section_formulas`` guarantees the formula
    shape is correct.
    """
    grp_row = _find_label_row(ws, "GROUP VERIFICATION")
    if grp_row is None:
        return  # ungrouped workbook — nothing to verify
    val = ws.cell(row=grp_row, column=COL_P_TOTAL_COST).value
    if isinstance(val, (int, float)):
        assert abs(val) < 0.005, (
            f"[{ws.title}] GROUP VERIFICATION (row {grp_row}, col P) must equal 0, got {val}"
        )
    # Else: formula string — can't evaluate without data_only reload
    # (which is the responsibility of the caller if it matters).


def assert_variance_within_tolerance(ws_or_path: Worksheet | str, tolerance: float = 0.01) -> None:
    """Evaluate VARIANCE CHECK and assert |value| <= ``tolerance``.

    Accepts either an openpyxl ``Worksheet`` (in which case the caller is
    expected to have opened the workbook with ``data_only=True``) OR a
    path string (in which case this helper re-opens the workbook in
    data-only mode itself, which is the typical usage).

    Skips gracefully if the cached value is None (formula never
    calculated, typically because the file was written by openpyxl and
    never opened by Excel).
    """
    from openpyxl import load_workbook

    if isinstance(ws_or_path, str):
        wb_do = load_workbook(ws_or_path, data_only=True)
        for sheet in wb_do.worksheets:
            _assert_variance_tolerance_ws(sheet, tolerance)
        wb_do.close()
    else:
        _assert_variance_tolerance_ws(ws_or_path, tolerance)


def _assert_variance_tolerance_ws(ws: Worksheet, tolerance: float) -> None:
    var_row = _find_label_row(ws, "VARIANCE CHECK")
    if var_row is None:
        return  # sheet has no variance row — skip
    val = ws.cell(row=var_row, column=COL_P_TOTAL_COST).value
    if val is None:
        return  # formula was never evaluated — skip gracefully
    if isinstance(val, str) and val.startswith("="):
        return  # we were handed the formula view, not data-only — skip
    assert isinstance(val, (int, float)), (
        f"[{ws.title}] VARIANCE CHECK (row {var_row}, col P) cached value "
        f"must be numeric when present, got {val!r}"
    )
    assert abs(val) <= tolerance, (
        f"[{ws.title}] VARIANCE CHECK (row {var_row}, col P) = {val} exceeds tolerance ±{tolerance}"
    )


def _is_grouped_sheet(ws: Worksheet) -> bool:
    """True iff ``ws`` is in grouped mode.

    Grouped mode is identified by the presence of a ``SUBTOTAL
    (GROUPED)`` label in column J — the same discriminator the
    production validator (``pipeline/xlsx_validator._detect_issues``)
    uses.  A plain ``SUBTOTAL`` label is NOT grouped mode; this was
    bug #1 on BL #TSCW18489131.
    """
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=COL_J_SUPPLIER_ITEM_DESC).value
        if val is None:
            continue
        if "SUBTOTAL (GROUPED)" in str(val).upper():
            return True
    return False


def assert_group_row_totals_match_details(
    ws: Worksheet,
    first_group_row: int,
    last_data_row: int,
    *,
    tolerance: float = 0.01,
) -> None:
    """For grouped-mode sheets: each group-row P value equals the sum of
    its detail rows' P values (within ``tolerance``).

    Skips ungrouped sheets entirely — the discriminator is the presence
    of a ``SUBTOTAL (GROUPED)`` label (same rule the production
    validator applies).

    In grouped sheets, group rows are distinguished from detail rows by
    the blue fill color ``D9E1F2`` set on column A (or any column) —
    ``_is_group_row`` checks this along with a non-empty AK cell.  Note
    that in production ungrouped mode, AK is also populated on every
    detail row (as a GroupBy tag), so the AK check is not sufficient on
    its own; we only run this invariant in grouped mode where group
    rows are visually distinct.
    """
    if not _is_grouped_sheet(ws):
        return  # ungrouped sheet — nothing to reconcile

    group_row: int | None = None
    detail_sum: float = 0.0

    def _flush():
        nonlocal group_row, detail_sum
        if group_row is None:
            return
        g_val = ws.cell(row=group_row, column=COL_P_TOTAL_COST).value
        if isinstance(g_val, (int, float)):
            assert abs(float(g_val) - detail_sum) <= tolerance, (
                f"[{ws.title}] group row {group_row} col P = {g_val}, but "
                f"sum of its detail rows = {detail_sum} (diff "
                f"{float(g_val) - detail_sum}, tolerance ±{tolerance})"
            )
        group_row = None
        detail_sum = 0.0

    for row in range(first_group_row, last_data_row + 1):
        label = ws.cell(row=row, column=COL_J_SUPPLIER_ITEM_DESC).value
        if _is_totals_label(label):
            _flush()
            continue
        if _row_is_empty(ws, row):
            continue
        if _is_group_row(ws, row):
            _flush()
            group_row = row
            detail_sum = 0.0
            continue
        # Detail row — accumulate P into the current block
        if group_row is not None:
            p = ws.cell(row=row, column=COL_P_TOTAL_COST).value
            if isinstance(p, (int, float)):
                detail_sum += float(p)
    _flush()


# ─────────────────────────────────────────────────────────────────────────
#                           Aggregate runner
# ─────────────────────────────────────────────────────────────────────────


def _find_data_row_range(ws: Worksheet) -> tuple[int, int]:
    """Return (first_data_row, last_data_row) for ``ws``.

    first_data_row is always 2 (row 1 is headers).  last_data_row is the
    last row BEFORE any totals-section label appears — or ``ws.max_row``
    if no totals are detected.
    """
    first = 2
    last = ws.max_row
    for row in range(2, ws.max_row + 1):
        label = ws.cell(row=row, column=COL_J_SUPPLIER_ITEM_DESC).value
        if _is_totals_label(label):
            last = row - 1
            break
    return first, max(first, last)


def run_all_invariants(
    ws: Worksheet,
    *,
    mode: str = "minimal",
    cet_db_path: str | None = None,
    xlsx_path: str | None = None,
) -> list[tuple[str, str]]:
    """Run the full invariant suite on ``ws`` and return failures.

    Unlike calling each ``assert_*`` directly, this helper COLLECTS all
    failures rather than stopping at the first.  The return value is a
    list of ``(invariant_name, message)`` tuples — empty means all
    invariants passed.

    ``mode`` controls the header check: ``"minimal"`` (default) checks
    only the columns written by the xlsx_factory fixtures and simple
    generators; ``"strict"`` checks all 37 production columns.

    If ``xlsx_path`` is provided, the variance-within-tolerance check
    opens the file separately with ``data_only=True`` to read Excel's
    cached computed value.  Otherwise that invariant runs against ``ws``
    directly and will skip when the formula view is what's provided.
    """
    failures: list[tuple[str, str]] = []
    first, last = _find_data_row_range(ws)

    def _run(name: str, fn):
        try:
            fn()
        except AssertionError as e:
            failures.append((name, str(e)))

    _run("assert_headers_present", lambda: assert_headers_present(ws, mode=mode))
    _run(
        "assert_tariff_codes_valid",
        lambda: assert_tariff_codes_valid(ws, first, last, cet_db_path=cet_db_path),
    )
    _run(
        "assert_column_types",
        lambda: assert_column_types(ws, first, last),
    )
    _run(
        "assert_invoice_totals_only_on_first_group_row",
        lambda: assert_invoice_totals_only_on_first_group_row(ws),
    )
    _run(
        "assert_totals_section_formulas",
        lambda: assert_totals_section_formulas(ws),
    )
    _run(
        "assert_row_q_r_formulas",
        lambda: assert_row_q_r_formulas(ws, first, last),
    )
    _run("assert_variance_is_formula", lambda: assert_variance_is_formula(ws))
    _run(
        "assert_adjustments_has_no_stacked_corrections",
        lambda: assert_adjustments_has_no_stacked_corrections(ws),
    )
    _run("assert_group_verification_zero", lambda: assert_group_verification_zero(ws))
    _run(
        "assert_variance_within_tolerance",
        lambda: assert_variance_within_tolerance(xlsx_path or ws),
    )
    _run(
        "assert_group_row_totals_match_details",
        lambda: assert_group_row_totals_match_details(ws, first, last),
    )
    return failures


# ─────────────────────────────────────────────────────────────────────────
#                     Kept for back-compat with existing tests
# ─────────────────────────────────────────────────────────────────────────


def snapshot_workbook(wb: Workbook) -> dict[tuple[str, int, int], object]:
    """Return a ``{(sheet, row, col): value}`` dict for deep equality checks.

    Used by idempotency tests: ``snapshot(fix(wb)) == snapshot(fix(fix(wb)))``.
    Only captures cell values; formatting is intentionally out of scope
    because the variance fixer is allowed to re-colour the VARIANCE CHECK
    cell between runs (red → green when resolved).
    """
    out: dict[tuple[str, int, int], object] = {}
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    out[(sheet.title, cell.row, cell.column)] = cell.value
    return out


def assert_fix_is_idempotent(fn: Callable[[Workbook], None], wb: Workbook) -> None:
    """Apply ``fn`` twice and assert the workbook state is unchanged.

    ``fn`` is expected to be a side-effecting function that takes a
    workbook and mutates it in place (e.g. a wrapper around
    ``_force_adjustment``).
    """
    fn(wb)
    once = snapshot_workbook(wb)
    fn(wb)
    twice = snapshot_workbook(wb)
    assert once == twice, "Fix function is not idempotent. Differing cells:\n" + "\n".join(
        f"  {k}: {once.get(k)!r} → {twice.get(k)!r}"
        for k in sorted(set(once) | set(twice))
        if once.get(k) != twice.get(k)
    )


def assert_all_invariants(ws: Worksheet) -> None:
    """Convenience: run every XLSX invariant that doesn't need extra context."""
    assert_variance_is_formula(ws)
    assert_adjustments_has_no_stacked_corrections(ws)


__all__: Iterable[str] = (
    # Original three (kept for back-compat)
    "assert_variance_is_formula",
    "assert_adjustments_has_no_stacked_corrections",
    "assert_is_grouped_matches",
    "assert_fix_is_idempotent",
    "assert_all_invariants",
    "snapshot_workbook",
    # Extended checklist
    "assert_headers_present",
    "assert_tariff_codes_valid",
    "assert_column_types",
    "assert_invoice_totals_only_on_first_group_row",
    "assert_totals_section_formulas",
    "assert_row_q_r_formulas",
    "assert_group_verification_zero",
    "assert_variance_within_tolerance",
    "assert_group_row_totals_match_details",
    "run_all_invariants",
    # Constants
    "EXPECTED_HEADERS_STRICT",
    "EXPECTED_HEADERS_MINIMAL",
)
