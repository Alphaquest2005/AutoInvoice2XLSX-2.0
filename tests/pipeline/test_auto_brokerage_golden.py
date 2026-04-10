"""Cell-level golden comparison for the Auto Brokerage corpus.

``{Downloads}/Auto Brokerage - WebSource/`` contains 47 PDFs, a subset
of which are paired with ``{id}_Invoice.xlsx`` — earlier pipeline
outputs that have been manually reviewed and are now the canonical
"golden" fixtures.  This test re-runs the pipeline on each paired PDF
and diffs the fresh output against its golden cell-by-cell.

**Important conventions**

* ``{id}_Invoice.xlsx`` IS the golden.  ``{id}_Invoice_v2.xlsx`` and
  ``{id}_Invoice_v3.xlsx`` are older experimental cruft and are
  IGNORED entirely.
* Numeric cells compare via ``pytest.approx(abs=0.01)``.
* An exclusion set (``_VOLATILE_FIELDS``) lists cell-level patterns
  that legitimately differ between runs (timestamps, generated IDs).
  It starts empty; add entries as the first run uncovers volatile
  fields, with a comment per entry explaining why.
* **Golden-update mode.** Setting ``AUTOINVOICE_UPDATE_GOLDEN=1`` in
  the environment skips the comparison and instead OVERWRITES each
  golden file with the fresh pipeline output.  This is the user's
  directive: "keep ``{id}_Invoice.xlsx``, just overwrite the files".
"""

from __future__ import annotations

import os
import shutil
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

# ── Paths ──────────────────────────────────────────────────────────────
_REPO_ROOT = Path(__file__).resolve().parents[2]
_PIPELINE_DIR = _REPO_ROOT / "pipeline"
_RUN_PY = _PIPELINE_DIR / "run.py"

_TIMEOUT_SECONDS = 600


def _resolve_downloads_dir() -> Path | None:
    """Locate the Downloads folder (matches file-watcher.ts:56 SSOT)."""
    env = os.environ.get("AUTOINVOICE_DOWNLOADS_DIR")
    if env and Path(env).is_dir():
        return Path(env)
    junction = _REPO_ROOT / "workspace" / "Downloads"
    if junction.is_dir():
        return junction
    fallback = Path("/mnt/d/OneDrive/Clients/WebSource/Downloads")
    if fallback.is_dir():
        return fallback
    return None


_DOWNLOADS = _resolve_downloads_dir()
_AUTO_BROKERAGE = _DOWNLOADS / "Auto Brokerage - WebSource" if _DOWNLOADS is not None else None

if _AUTO_BROKERAGE is None or not _AUTO_BROKERAGE.is_dir():
    pytest.skip(
        "Auto Brokerage - WebSource/ not available (Downloads unresolvable)",
        allow_module_level=True,
    )


def _discover_pairs(root: Path) -> list[tuple[Path, Path]]:
    """Return ``[(pdf, golden_xlsx), ...]`` pairs.

    A pair exists iff ``{id}.pdf`` and ``{id}_Invoice.xlsx`` are both
    present in ``root``.  ``_Invoice_v2.xlsx`` and ``_Invoice_v3.xlsx``
    variants are IGNORED.
    """
    pairs: list[tuple[Path, Path]] = []
    for pdf in sorted(root.glob("*.pdf")):
        stem = pdf.stem
        golden = root / f"{stem}_Invoice.xlsx"
        if golden.is_file():
            pairs.append((pdf, golden))
    # Also try *.PDF (uppercase extension — 8251566282.PDF)
    for pdf in sorted(root.glob("*.PDF")):
        stem = pdf.stem
        golden = root / f"{stem}_Invoice.xlsx"
        if golden.is_file() and not any(p[0] == pdf for p in pairs):
            pairs.append((pdf, golden))
    return pairs


_PAIRS = _discover_pairs(_AUTO_BROKERAGE)

if not _PAIRS:
    pytest.skip(
        f"no (pdf, _Invoice.xlsx) pairs in {_AUTO_BROKERAGE}",
        allow_module_level=True,
    )


# ── Volatile fields exclusion set ──────────────────────────────────────
#
# Each entry is a ``(sheet_title_pattern, row, column_index)`` triple.
# ``sheet_title_pattern`` may be ``*`` to match any sheet, or the exact
# sheet name.  ``row`` and ``column_index`` may be ``None`` to wildcard
# the dimension.
#
# Populate this set only after the first run surfaces a legitimate diff
# that must be tolerated (timestamps, auto-generated invoice numbers,
# etc.).  Leave a comment explaining WHY each entry is here so future
# maintainers can revisit if the underlying volatility is fixed.
_VOLATILE_FIELDS: set[tuple[str, int | None, int | None]] = set()


def _cell_is_volatile(sheet_title: str, row: int, col: int) -> bool:
    for pat, pat_row, pat_col in _VOLATILE_FIELDS:
        if pat != "*" and pat != sheet_title:
            continue
        if pat_row is not None and pat_row != row:
            continue
        if pat_col is not None and pat_col != col:
            continue
        return True
    return False


def _values_match(a: object, b: object) -> bool:
    """Return True if two openpyxl cell values should be considered equal.

    Numeric tolerance ±0.01.  String equality is exact.  ``None`` and
    empty string are treated as equivalent (openpyxl sometimes returns
    one vs the other for a blank cell depending on load mode).
    """
    if (a is None or a == "") and (b is None or b == ""):
        return True
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(float(a) - float(b)) <= 0.01
    return a == b


def _run_pipeline(pdf: Path, out_dir: Path) -> Path:
    """Run ``pipeline/run.py`` on ``pdf`` and return the produced XLSX.

    Raises ``AssertionError`` with context if the pipeline fails or
    produces no XLSX.
    """
    stage = out_dir / "stage"
    stage.mkdir()
    output = out_dir / "out"
    output.mkdir()
    shutil.copy2(pdf, stage / pdf.name)

    proc = subprocess.run(  # noqa: S603 - trusted local path
        [
            sys.executable,
            "-u",
            str(_RUN_PY),
            "--input-dir",
            str(stage),
            "--output-dir",
            str(output),
            "--json-output",
        ],
        cwd=str(_PIPELINE_DIR),
        capture_output=True,
        text=True,
        timeout=_TIMEOUT_SECONDS,
    )
    assert proc.returncode == 0, (
        f"pipeline failed for {pdf.name}: rc={proc.returncode}\nstderr tail: {proc.stderr[-500:]}"
    )

    produced = sorted(list(output.rglob("*.xlsx")) + list(stage.rglob("*.xlsx")))
    assert produced, f"pipeline produced no XLSX for {pdf.name}"
    # Pick the first produced file — for single-PDF runs there should
    # only be one.  If we got more, take the first deterministically.
    return produced[0]


@pytest.mark.integration
@pytest.mark.requires_downloads
@pytest.mark.parametrize(
    ("pdf", "golden"),
    _PAIRS,
    ids=[p.stem for p, _ in _PAIRS],
)
def test_auto_brokerage_golden(pdf: Path, golden: Path, tmp_path: Path) -> None:
    """Run the pipeline on ``pdf`` and diff the output against ``golden``.

    In ``AUTOINVOICE_UPDATE_GOLDEN=1`` mode, overwrite the golden with
    the fresh output instead of diffing.
    """
    fresh_xlsx = _run_pipeline(pdf, tmp_path)

    if os.environ.get("AUTOINVOICE_UPDATE_GOLDEN") == "1":
        shutil.copy2(fresh_xlsx, golden)
        print(f"[update-golden] overwrote {golden}")
        return

    fresh_wb = load_workbook(str(fresh_xlsx))
    golden_wb = load_workbook(str(golden))

    fresh_sheets = {s.title: s for s in fresh_wb.worksheets}
    golden_sheets = {s.title: s for s in golden_wb.worksheets}

    diffs: list[str] = []

    common_sheets = set(fresh_sheets) & set(golden_sheets)
    missing_in_fresh = set(golden_sheets) - set(fresh_sheets)
    missing_in_golden = set(fresh_sheets) - set(golden_sheets)
    for name in sorted(missing_in_fresh):
        diffs.append(f"sheet {name!r} missing from fresh output")
    for name in sorted(missing_in_golden):
        diffs.append(f"sheet {name!r} missing from golden")

    for title in sorted(common_sheets):
        fs = fresh_sheets[title]
        gs = golden_sheets[title]
        max_row = max(fs.max_row, gs.max_row)
        max_col = max(fs.max_column, gs.max_column)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                if _cell_is_volatile(title, r, c):
                    continue
                fv = fs.cell(row=r, column=c).value
                gv = gs.cell(row=r, column=c).value
                if not _values_match(fv, gv):
                    diffs.append(f"[{title}] row {r}, col {c}: fresh={fv!r} golden={gv!r}")

    fresh_wb.close()
    golden_wb.close()

    if diffs:
        lines = [f"[{pdf.name}] {len(diffs)} cell diff(s) vs {golden.name}:"]
        lines.extend(diffs[:50])
        if len(diffs) > 50:
            lines.append(f"... ({len(diffs) - 50} more)")
        pytest.fail("\n".join(lines))
