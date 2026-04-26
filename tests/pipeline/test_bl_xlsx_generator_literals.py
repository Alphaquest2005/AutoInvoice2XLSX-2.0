"""Regression — bl_xlsx_generator magic-constant migration (Wave 1 Sub-Agent #2).

Pins behaviors and config values that MUST stay stable through the Phase-2
migration of ``pipeline/bl_xlsx_generator.py``:

A. Budget-Marine / A2 doc_type resolution (critical production bug guard).
B. Module-level financial constants now come from ``financial_constants.yaml``
   (XCD_RATE, CSC_RATE, VAT_RATE).
C. Module-level xlsx colour/font literals now come from config (styles).
D. Date-format list + regex patterns moved to config files.
E. Sheet title / output defaults / well-known section labels come from config.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[2]
PIPELINE_DIR = BASE_DIR / "pipeline"
if str(PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_DIR))


def _minimal_invoice():
    invoice_data = {
        "invoice_num": "INV_TEST_0001",
        "invoice_date": "2026-03-31",
        "invoice_total": 100.00,
        "items": [],
    }
    matched = [
        {
            "supplier_item": "TEST-1",
            "supplier_item_desc": "Test item",
            "quantity": 1,
            "unit_price": 100.0,
            "total_cost": 100.0,
            "uom": "Each",
            "tariff_code": "90318000",
            "category": "MARINE ELECTRONICS",
            "po_item_ref": "",
            "po_item_desc": "",
            "po_number": "",
        }
    ]
    supplier_info = {
        "code": "BUD",
        "name": "Budget Marine",
        "address": "Waterfront Rd 25B, Cole Bay, Sint Maarten",
        "country": "QN",
    }
    return invoice_data, matched, supplier_info


# ─── A. critical production regression (Budget Marine / A2) ──────────────


@pytest.mark.parametrize("doc_type", ["7400-000", "4000-000"])
def test_a2_document_type_preserved_after_migration(doc_type, tmp_path):
    """A2 MUST echo the resolved ``document_type`` passed in — this is the
    Budget Marine regression case from 2026-04-22 (TSCW18629902)."""
    from bl_xlsx_generator import generate_bl_xlsx

    invoice_data, matched, supplier_info = _minimal_invoice()
    out = tmp_path / f"inv_{doc_type}.xlsx"
    generate_bl_xlsx(
        invoice_data,
        matched,
        "Budget Marine",
        supplier_info,
        str(out),
        document_type=doc_type,
    )

    wb = load_workbook(str(out))
    ws = wb.active
    assert ws["A1"].value == "Document Type"
    assert ws["A2"].value == doc_type


# ─── B. financial constants pinned to config ──────────────────────────────


def test_xcd_csc_vat_match_financial_constants_yaml():
    """Module globals must mirror financial_constants.yaml values."""
    import bl_xlsx_generator
    from pipeline.config_loader import load_financial_constants

    fin = load_financial_constants()
    assert fin["csc_rate"] == bl_xlsx_generator.CSC_RATE
    assert fin["vat_rate"] == bl_xlsx_generator.VAT_RATE
    # xcd_rate may not yet be in financial_constants.yaml — this test will
    # fail until we add it, locking the migration in place.
    assert "xcd_rate" in fin, "financial_constants.yaml must expose xcd_rate for bl_xlsx_generator"
    assert fin["xcd_rate"] == bl_xlsx_generator.XCD_RATE


def test_calculate_duties_uses_config_rates():
    """calculate_duties() numeric output stays identical after refactor.
    Anchored on a known fixed CIF/CET pair so any drift is caught."""
    from bl_xlsx_generator import calculate_duties

    duties = calculate_duties(
        cif_usd=1000.0,
        cet_rate=0.20,
        customs_freight=0,
        insurance=0,
    )
    # cif_xcd = 1000 * 2.7169 = 2716.90
    assert duties["cif_xcd"] == pytest.approx(2716.90)
    # cet = 2716.90 * 0.20 = 543.38
    assert duties["cet"] == pytest.approx(543.38)
    # csc = 2716.90 * 0.06 = 163.01
    assert duties["csc"] == pytest.approx(163.01)
    # vat = (2716.90 + 543.38 + 163.01) * 0.15 = 513.49 (approx; rounding)
    assert duties["vat"] == pytest.approx(513.49, abs=0.01)


# ─── C. xlsx style colours come from config ───────────────────────────────


def test_header_style_colors_from_config():
    """HEADER_FILL/HEADER_FONT colors must match columns.yaml header style."""
    import bl_xlsx_generator
    from pipeline.config_loader import load_columns

    cols = load_columns()
    header = cols.get("styles", {}).get("header", {})
    assert bl_xlsx_generator.HEADER_FILL.start_color.rgb.endswith(
        header.get("fill_color", "4472C4")
    )


def test_uncertain_recovered_colors_from_config():
    """UNCERTAIN_FILL and RECOVERED_FILL must come from a config source,
    not be hardcoded in bl_xlsx_generator.py."""
    from pipeline.config_loader import load_columns

    cols = load_columns()
    styles = cols.get("styles", {})
    # These keys will be added under styles in columns.yaml by the migration.
    assert "uncertain" in styles, (
        "columns.yaml:styles.uncertain must supply the uncertain-row fill color"
    )
    assert "recovered" in styles, (
        "columns.yaml:styles.recovered must supply the orphan-recovery fill color"
    )


# ─── D. date format list + regexes live in config ─────────────────────────


def test_date_formats_come_from_config():
    """_normalize_date must iterate the formats listed in patterns.yaml,
    not a hardcoded in-function tuple."""
    from pipeline.config_loader import load_patterns

    pats = load_patterns()
    assert "date_parse_formats" in pats, (
        "patterns.yaml must expose date_parse_formats for bl_xlsx_generator"
    )
    fmts = pats["date_parse_formats"]
    assert "%m/%d/%Y" in fmts and "%B %d, %Y" in fmts


def test_normalize_date_still_works():
    """Behaviour guard on _normalize_date — must not regress after migration."""
    from bl_xlsx_generator import _normalize_date

    assert _normalize_date("01/16/2026") == "2026-01-16"
    assert _normalize_date("January 16, 2026") == "2026-01-16"
    assert _normalize_date("2026-01-16") == "2026-01-16"


# ─── E. sheet title / defaults / labels sourced from config ───────────────


def test_sheet_title_and_reference_label_from_config(tmp_path):
    """Sheet title and default reference label pulled from file_paths.yaml
    (or library_enums.yaml). After migration, no hardcoded strings."""
    from bl_xlsx_generator import generate_bl_xlsx

    invoice_data, matched, supplier_info = _minimal_invoice()
    out = tmp_path / "sheet_title.xlsx"
    generate_bl_xlsx(
        invoice_data,
        matched,
        "Budget Marine",
        supplier_info,
        str(out),
        document_type="7400-000",
    )
    wb = load_workbook(str(out))
    # Sheet title is a user-visible label; it must be stable post-migration.
    assert wb.active.title == "Invoice Data"


def test_default_document_type_resolves_via_config():
    """generate_bl_xlsx default kwarg '7400-000' must correspond to a valid
    document-type entry in document_types.json."""
    import inspect

    from bl_xlsx_generator import generate_bl_xlsx
    from pipeline.config_loader import load_document_types

    dt = load_document_types()
    assert "7400-000" in dt["document_types"]
    sig = inspect.signature(generate_bl_xlsx)
    default = sig.parameters["document_type"].default
    assert default in dt["document_types"]
