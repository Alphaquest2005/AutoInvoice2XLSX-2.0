"""Per-fixer coverage for ``pipeline.checklist_fixer``.

Each test exercises one fixer end-to-end through the public ``attempt_fixes``
entry-point. The mechanical fixers operate on the ``email_params`` dict
in-place and never call the network. The XLSX fixers mutate openpyxl
workbooks in place at the path supplied by ``attachment_paths``.

Critically: ``consignee_code_missing`` MUST NOT have a registered fixer (per
``feedback_consignee_code_warn_ok.md``).
"""

from __future__ import annotations

import os
import sys

import pytest

_PIPELINE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "pipeline"))
if _PIPELINE_DIR not in sys.path:
    sys.path.insert(0, _PIPELINE_DIR)

from checklist_fixer import _FIXERS, _NO_FIX_KINDS, attempt_fixes  # noqa: E402


def _finding(check, severity="warn"):
    return {
        "check": check,
        "severity": severity,
        "message": f"synthetic {check}",
        "field": "",
        "value": "",
        "fix_hint": "",
    }


# ── Registry guards ────────────────────────────────────────────────


def test_consignee_code_missing_is_NOT_registered():  # noqa: N802
    """`feedback_consignee_code_warn_ok.md`: do NOT silence or auto-source."""
    assert "consignee_code_missing" in _NO_FIX_KINDS
    assert "consignee_code_missing" not in _FIXERS


def test_no_fix_finding_is_neither_fixed_nor_skipped():
    ep = {}
    rep = attempt_fixes(ep, [_finding("consignee_code_missing")])
    assert rep["fixed"] == []
    assert rep["skipped"] == []


def test_unknown_kind_is_skipped_without_crash():
    ep = {}
    rep = attempt_fixes(ep, [_finding("totally_made_up_kind")])
    assert rep["fixed"] == []
    assert "totally_made_up_kind" in rep["skipped"]


# ── email_params (dict) fixers ─────────────────────────────────────


def test_bl_has_doc_suffix_strips_suffix():
    ep = {"waybill": "HAWB9333496-Declaration"}
    rep = attempt_fixes(ep, [_finding("bl_has_doc_suffix")])
    assert ep["waybill"] == "HAWB9333496"
    assert "bl_has_doc_suffix" in rep["fixed"]


def test_stale_tmp_path_resolves_to_local_basename(tmp_path):
    real = tmp_path / "invoice.pdf"
    real.write_bytes(b"%PDF-1.4 fake")
    stale = "/some/legacy/_rerun_tmp/X.out/invoice.pdf"
    ep = {"attachment_paths": [stale]}
    rep = attempt_fixes(ep, [_finding("stale_tmp_path", "block")], output_dir=str(tmp_path))
    assert ep["attachment_paths"] == [str(real)]
    assert "stale_tmp_path" in rep["fixed"]


def test_stale_tmp_path_partial_resolve_does_not_claim_fixed(tmp_path):
    """If at least one stale path can't be re-resolved, leave as residual."""
    ep = {
        "attachment_paths": [
            "/some/legacy/_rerun_tmp/X.out/missing.pdf",  # cannot resolve
        ]
    }
    rep = attempt_fixes(ep, [_finding("stale_tmp_path", "block")], output_dir=str(tmp_path))
    assert "stale_tmp_path" in rep["skipped"]
    assert "stale_tmp_path" not in rep["fixed"]


def test_worksheet_pdf_in_attachments_drops_only_worksheet_pdfs():
    ep = {
        "attachment_paths": [
            "/x/invoice.pdf",
            "/x/HAWB123-Declaration.pdf",
            "/x/items_WorkSheet.pdf",
            "/x/something worksheet.PDF",
            "/x/items.xlsx",
        ]
    }
    rep = attempt_fixes(ep, [_finding("worksheet_pdf_in_attachments")])
    assert ep["attachment_paths"] == [
        "/x/invoice.pdf",
        "/x/HAWB123-Declaration.pdf",
        "/x/items.xlsx",
    ]
    assert "worksheet_pdf_in_attachments" in rep["fixed"]


def test_duplicate_attachment_basenames_dedupes_keep_first():
    ep = {
        "attachment_paths": [
            "/a/file.pdf",
            "/b/file.pdf",
            "/a/items.xlsx",
        ]
    }
    rep = attempt_fixes(ep, [_finding("duplicate_attachment_basenames", "block")])
    assert ep["attachment_paths"] == ["/a/file.pdf", "/a/items.xlsx"]
    assert "duplicate_attachment_basenames" in rep["fixed"]


def test_expected_entries_mismatch_recounts_xlsx_excluding_combined():
    ep = {
        "attachment_paths": [
            "/x/a.xlsx",
            "/x/b.xlsx",
            "/x/HAWB123_combined.xlsx",
            "/x/HAWB123-combined.xlsx",
            "/x/a.pdf",
        ],
        "expected_entries": 99,
    }
    rep = attempt_fixes(ep, [_finding("expected_entries_mismatch")])
    assert ep["expected_entries"] == 2
    assert "expected_entries_mismatch" in rep["fixed"]


def test_country_origin_invalid_uses_results_majority():
    class FakeResult:
        def __init__(self, country):
            self.supplier_info = {"country": country}

    ep = {"country_origin": "XYZ"}
    rep = attempt_fixes(
        ep,
        [_finding("country_origin_invalid")],
        results=[FakeResult("US"), FakeResult("US"), FakeResult("CN")],
    )
    new = ep["country_origin"]
    # Mode is US; CARICOM mapping (if any) returns a 2-letter alpha code
    assert new.isalpha() and len(new) == 2
    assert "country_origin_invalid" in rep["fixed"]


def test_country_origin_invalid_skipped_when_no_results():
    ep = {"country_origin": "XYZ"}
    rep = attempt_fixes(ep, [_finding("country_origin_invalid")])
    assert ep["country_origin"] == "XYZ"
    assert "country_origin_invalid" in rep["skipped"]


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("2024/19", "2024 19"),
        ("2024 / 19", "2024 19"),
        ("2024-19", "2024 19"),
        ("2024-019", "2024 019"),
    ],
)
def test_man_reg_invalid_normalises_to_yyyy_nn(raw, expected):
    ep = {"man_reg": raw}
    rep = attempt_fixes(ep, [_finding("man_reg_invalid")])
    assert ep["man_reg"] == expected
    assert "man_reg_invalid" in rep["fixed"]


def test_man_reg_invalid_returns_residual_when_uncoercible():
    ep = {"man_reg": "gibberish"}
    rep = attempt_fixes(ep, [_finding("man_reg_invalid")])
    assert ep["man_reg"] == "gibberish"
    assert "man_reg_invalid" not in rep["fixed"]


def test_man_reg_blank_is_left_as_residual():
    """Blank manifest reg has no source — Joseph fills it. Don't fabricate."""
    ep = {"man_reg": ""}
    rep = attempt_fixes(ep, [_finding("man_reg_blank")])
    # Either skipped or simply not registered — acceptable, but never fixed.
    assert "man_reg_blank" not in rep["fixed"]


# ── XLSX item fixers ───────────────────────────────────────────────


def _make_per_invoice_xlsx(path: str, rows: list) -> None:
    """Create a per-invoice XLSX matching the bl_xlsx_generator layout.

    ``rows`` is a list of (tariff, sku, description, qty, cost, total) tuples.
    Row 1 is the header (left empty for tests). Row 2 onward holds items.
    InvoiceTotal in S2 is set to sum of total-cost values.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    # Pad row 1 (header)
    ws.cell(row=1, column=1).value = "header"
    inv_total = 0.0
    for i, (tariff, sku, desc, qty, cost, total) in enumerate(rows, start=2):
        if tariff is not None:
            ws.cell(row=i, column=6).value = tariff  # F
        if sku is not None:
            ws.cell(row=i, column=9).value = sku  # I
        if desc is not None:
            ws.cell(row=i, column=10).value = desc  # J
        if qty is not None:
            ws.cell(row=i, column=11).value = qty  # K
        if cost is not None:
            ws.cell(row=i, column=15).value = cost  # O
        if total is not None:
            ws.cell(row=i, column=16).value = total  # P
            try:  # noqa: SIM105
                inv_total += float(total)
            except (TypeError, ValueError):
                pass
    ws.cell(row=2, column=19).value = round(inv_total, 2)  # S2 InvoiceTotal
    wb.save(path)
    wb.close()


def _read_descs_qtys(path: str) -> list:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    out = []
    for row in range(2, ws.max_row + 1):
        desc = ws.cell(row=row, column=10).value
        total = ws.cell(row=row, column=16).value
        out.append((desc, total))
    inv_total = ws.cell(row=2, column=19).value
    wb.close()
    return out, inv_total


def test_payment_line_as_item_drops_row_and_fixes_invoice_total(tmp_path):
    xp = tmp_path / "inv.xlsx"
    _make_per_invoice_xlsx(
        str(xp),
        [
            (None, "SKU1", "Real product A", 1, 10.0, 10.0),
            (None, "PAYMENT", "Visa ending in 4242", 1, 25.0, 25.0),
            (None, "SKU2", "Real product B", 1, 5.0, 5.0),
        ],
    )
    ep = {"attachment_paths": [str(xp)]}
    rep = attempt_fixes(
        ep,
        [_finding("item_payment_line_as_item", "block")],
    )
    assert "item_payment_line_as_item" in rep["fixed"]
    rows, inv_total = _read_descs_qtys(str(xp))
    descs = [d for (d, _) in rows]
    # Payment row removed; only the two real products remain.
    assert "Visa ending in 4242" not in descs
    assert "Real product A" in descs
    assert "Real product B" in descs
    # InvoiceTotal recomputed to sum of remaining P values (10 + 5 = 15).
    assert abs(float(inv_total) - 15.0) < 0.01


def test_description_starts_with_tariff_strips_prefix(tmp_path):
    xp = tmp_path / "inv.xlsx"
    _make_per_invoice_xlsx(
        str(xp),
        [
            (None, "SKU1", "61091000 Cotton T-Shirt Mens", 2, 10.0, 20.0),
            (None, "SKU2", "Plain Description", 1, 5.0, 5.0),
        ],
    )
    ep = {"attachment_paths": [str(xp)]}
    rep = attempt_fixes(
        ep,
        [_finding("item_description_starts_with_tariff")],
    )
    assert "item_description_starts_with_tariff" in rep["fixed"]
    rows, _ = _read_descs_qtys(str(xp))
    descs = [d for (d, _) in rows]
    assert descs[0] == "Cotton T-Shirt Mens"
    assert descs[1] == "Plain Description"


# ── attempt_fixes orchestration ────────────────────────────────────


def test_attempt_fixes_handles_mixed_findings_with_one_fix_per_kind():
    ep = {
        "waybill": "HAWB-Declaration",
        "attachment_paths": ["/a/file.pdf", "/b/file.pdf"],
    }
    rep = attempt_fixes(
        ep,
        [
            _finding("bl_has_doc_suffix"),
            _finding("duplicate_attachment_basenames", "block"),
            _finding("waybill_missing", "block"),  # no mechanical fixer → skipped
        ],
    )
    assert ep["waybill"] == "HAWB"
    assert ep["attachment_paths"] == ["/a/file.pdf"]
    assert set(rep["fixed"]) == {"bl_has_doc_suffix", "duplicate_attachment_basenames"}
    assert "waybill_missing" in rep["skipped"]


def test_attempt_fixes_runs_each_kind_once_even_if_repeated():
    ep = {"waybill": "HAWB-Declaration"}
    rep = attempt_fixes(
        ep,
        [
            _finding("bl_has_doc_suffix"),
            _finding("bl_has_doc_suffix"),
        ],
    )
    assert ep["waybill"] == "HAWB"
    assert rep["fixed"].count("bl_has_doc_suffix") == 1
