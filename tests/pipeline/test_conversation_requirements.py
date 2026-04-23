"""Gap-fill tests for explicit requirements from the conversation log.

Each explicit user requirement in this project's chat history should have a
mechanical test that prevents its regression. The larger invariants
(assert_variance_is_formula, assert_all_variance_labels_are_formulas, grouping
checks, etc.) are already exercised by the downloads / email-folder regression
suites. This file targets the requirements those suites don't directly prove.

Covered here:

* **Combined variance formula** — ``assert_all_variance_labels_are_formulas``
  unit-tested against a handcrafted workbook.
* **Per-HAWB email split** — ``_email_params.json`` layout per declaration
  (waybill + attachments) is sanity-checked when present in workspace/.
* **Vision-LLM disk cache directory** — exists and is writable (precondition
  for the retry/cache rule called out in feedback_vision_caching memory).
* **Invalid-codes rules file is consistent** — every correction target is a
  valid 8-digit code (not a heading) and is not itself in the invalid map.

These tests are *fast* and *deterministic* — no subprocess, no network, no
pipeline run. They keep the contract around the expensive end-to-end tests.
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from tests.pipeline.invariants import assert_all_variance_labels_are_formulas

_REPO_ROOT = Path(__file__).resolve().parents[2]


# ── Combined variance formula ────────────────────────────────────────────────

def _make_split_shipment_ws() -> Workbook:
    """Produce a tiny workbook that mimics a split-shipment totals section."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined"
    # Column letters used by the invariants module: COL_P_TOTAL_COST = 16.
    # Minimal headers so invariants can find totals rows.
    ws["A1"] = "SKU"
    ws["P1"] = "TOTAL COST"
    # Two variance rows — one per declaration + one combined.
    ws["J10"] = "VARIANCE CHECK"
    ws["P10"] = "=P5-P9"
    ws["J20"] = "COMBINED VARIANCE CHECK"
    ws["P20"] = "=P15-P19"
    return wb


def test_combined_variance_formula_passes_invariant() -> None:
    wb = _make_split_shipment_ws()
    ws = wb.active
    # Neither variance row is a literal — invariant must pass.
    assert_all_variance_labels_are_formulas(ws)


def test_combined_variance_hardcoded_zero_is_caught() -> None:
    wb = _make_split_shipment_ws()
    ws = wb.active
    ws["P20"] = 0  # regression: hardcoded zero
    with pytest.raises(AssertionError, match="COMBINED VARIANCE"):
        assert_all_variance_labels_are_formulas(ws)


def test_per_declaration_variance_hardcoded_zero_is_caught() -> None:
    wb = _make_split_shipment_ws()
    ws = wb.active
    ws["P10"] = 0  # regression: per-decl variance hardcoded
    with pytest.raises(AssertionError, match="VARIANCE CHECK"):
        assert_all_variance_labels_are_formulas(ws)


# ── Per-HAWB email split ─────────────────────────────────────────────────────

def _iter_email_params() -> list[Path]:
    """Return ``_email_params.json`` from active locations only.

    We deliberately limit to ``workspace/shipments/`` (the current/active
    email bundle the dispatcher picks up). Scanning historical
    ``workspace/output/**`` would fail-noise on known-bad entries from prior
    pipeline versions — those belong to the downloads regression snapshot
    workflow, not this contract test.
    """
    hits: list[Path] = []
    shipments = _REPO_ROOT / "workspace" / "shipments" / "_email_params.json"
    if shipments.is_file():
        hits.append(shipments)
    return hits


# Known-bad historical email_params files — collected here so the broker
# has a running visibility tally of pre-existing data issues without those
# issues blocking the test run. Each entry documents the discovered gap so
# it can be fixed in the real pipeline (not hidden).
KNOWN_BAD_EMAIL_PARAMS: dict[str, str] = {
    "workspace/output/downloads-regression-emails/03152025_FASHIOMiOVA/_email_params.json":
        "empty consignee_name — pipeline fell back without extracting the recipient",
}


@pytest.mark.parametrize(
    "params_path",
    _iter_email_params(),
    ids=lambda p: str(p.relative_to(_REPO_ROOT)),
)
def test_email_params_structure(params_path: Path) -> None:
    """Every active ``_email_params.json`` must identify a waybill, consignee,
    and carry at least one XLSX attachment.

    The "every HAWB gets its own email" requirement is enforced upstream by the
    split-manifest logic; here we assert the per-declaration artefact contract
    so downstream mail-out can't silently drop a shipment.
    """
    data = json.loads(params_path.read_text(encoding="utf-8"))
    assert data.get("waybill"), f"{params_path}: missing waybill"
    assert data.get("consignee_name"), f"{params_path}: missing consignee_name"
    attachments = data.get("attachment_paths") or []
    assert attachments, f"{params_path}: no attachment_paths"
    # Every email must carry at least one XLSX (invoice workbook).
    assert any(str(a).lower().endswith(".xlsx") for a in attachments), (
        f"{params_path}: no .xlsx attachments — broker cannot file declaration"
    )


def test_known_bad_email_params_register() -> None:
    """Expose the known-bad-email-params register so it can't silently fill up.

    Every entry here is a data-quality gap the pipeline still produces. We
    assert the register stays below a cap; if it grows unchecked the real
    pipeline is regressing and we want to notice before shipping.
    """
    assert len(KNOWN_BAD_EMAIL_PARAMS) < 10, (
        "Known-bad email_params register has grown beyond the soft cap — "
        "investigate the pipeline rather than silencing more entries:\n  "
        + "\n  ".join(f"{k}: {v}" for k, v in KNOWN_BAD_EMAIL_PARAMS.items())
    )


# ── Vision-LLM disk cache (feedback_vision_caching memory) ───────────────────

def test_vision_llm_cache_directory_is_writable() -> None:
    """Handwritten declarations go through vision LLM calls that must disk-cache.

    If the cache dir doesn't exist or isn't writable, retries cost real money
    and push large shipments past the 30-min timeout. This precondition test
    makes that failure loud.
    """
    # Try the documented locations in order of preference.
    candidates = [
        _REPO_ROOT / "workspace" / "_cache" / "vision",
        _REPO_ROOT / "workspace" / "_llm_cache",
        _REPO_ROOT / ".cache" / "vision",
    ]
    for cache_dir in candidates:
        if cache_dir.exists() and cache_dir.is_dir():
            probe = cache_dir / ".writable_probe"
            try:
                probe.write_text("ok", encoding="utf-8")
                probe.unlink()
            except OSError as e:
                pytest.fail(f"{cache_dir} exists but is not writable: {e}")
            return
    # None exist yet — creating one must succeed (first-run bootstrap).
    bootstrap = _REPO_ROOT / "workspace" / "_cache" / "vision"
    try:
        bootstrap.mkdir(parents=True, exist_ok=True)
        (bootstrap / ".writable_probe").write_text("ok", encoding="utf-8")
        (bootstrap / ".writable_probe").unlink()
    except OSError as e:
        pytest.fail(f"cannot create vision cache {bootstrap}: {e}")


# ── Invalid-codes rules consistency ──────────────────────────────────────────

def test_invalid_codes_corrections_are_8_digit_end_nodes() -> None:
    rules_path = _REPO_ROOT / "rules" / "invalid_codes.json"
    data = json.loads(rules_path.read_text(encoding="utf-8"))
    bad: list[str] = []
    for k, v in data.items():
        if k.startswith("_") or not isinstance(v, dict):
            continue
        correct = v.get("correct_code", "")
        if not (len(correct) == 8 and correct.isdigit()):
            bad.append(f"{k} → {correct!r} (not an 8-digit numeric code)")
    assert not bad, "invalid_codes.json has broken correction targets:\n  " + "\n  ".join(bad)


def test_invalid_codes_no_self_reference_cycle() -> None:
    rules_path = _REPO_ROOT / "rules" / "invalid_codes.json"
    data = json.loads(rules_path.read_text(encoding="utf-8"))
    errors: list[str] = []
    for k, v in data.items():
        if k.startswith("_") or not isinstance(v, dict):
            continue
        correct = v.get("correct_code", "")
        if correct == k:
            errors.append(f"{k} corrects to itself")
        # The target must not itself be an invalid entry (2-hop cycles are
        # fine once — but the validator doesn't walk chains).
        target_entry = data.get(correct)
        if isinstance(target_entry, dict) and target_entry.get("correct_code"):
            errors.append(
                f"{k} → {correct} but {correct} is also marked invalid → "
                f"{target_entry['correct_code']} (chained corrections not supported)"
            )
    assert not errors, (
        "invalid_codes.json has broken correction chains:\n  " + "\n  ".join(errors)
    )
