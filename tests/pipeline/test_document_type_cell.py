"""Regression: Document Type cell (A2) must reflect the resolved doc_type.

Client complaint (2026-04-22, TSCW18629902): a Budget Marine BL shipment
came out as IM4 / 4000-000 when it should have been IM7 / 7400-000.

Root cause: config/columns.yaml had ``group_value: "4000-000"`` hardcoded
as a literal, and pipeline/xlsx_generator.py hardcoded the same string
directly. The consignee-rule resolver and the ``document_type`` kwarg
threaded through ``process_single_invoice → generate_bl_xlsx`` were both
silently overridden by the template literal.

Fix: columns.yaml uses ``${document_type}`` and xlsx_generator reads
the value from context / metadata / config default. This test guards
both paths against re-introduction of the bug.
"""

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

# The pipeline modules live in pipeline/; tests import them the same
# way the pipeline does (bare module names, not pipeline.*).
BASE_DIR = Path(__file__).resolve().parents[2]
PIPELINE_DIR = BASE_DIR / "pipeline"
if str(PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_DIR))


def _minimal_invoice():
    """Smallest invoice + single matched item that generate_bl_xlsx accepts."""
    invoice_data = {
        "invoice_num": "INV_TEST_0001",
        "invoice_date": "2026-03-31",
        "invoice_total": 100.00,
        "items": [],
    }
    matched = [
        {
            "supplier_item": "TEST-1",
            "supplier_item_desc": "Test item",
            "quantity": 1,
            "unit_price": 100.0,
            "total_cost": 100.0,
            "uom": "Each",
            "tariff_code": "90318000",
            "category": "MARINE ELECTRONICS",
            "po_item_ref": "",
            "po_item_desc": "",
            "po_number": "",
        }
    ]
    supplier_info = {
        "code": "BUD",
        "name": "Budget Marine",
        "address": "Waterfront Rd 25B, Cole Bay, Sint Maarten",
        "country": "QN",
    }
    return invoice_data, matched, supplier_info


@pytest.mark.parametrize("doc_type", ["7400-000", "4000-000"])
def test_bl_xlsx_writes_resolved_document_type(doc_type, tmp_path):
    """generate_bl_xlsx must write the passed ``document_type`` into A2,
    not a hardcoded default. Budget Marine regression case: '7400-000'."""
    from bl_xlsx_generator import generate_bl_xlsx

    invoice_data, matched, supplier_info = _minimal_invoice()
    out = tmp_path / f"inv_{doc_type}.xlsx"
    generate_bl_xlsx(
        invoice_data,
        matched,
        "Budget Marine",
        supplier_info,
        str(out),
        document_type=doc_type,
    )

    wb = load_workbook(str(out))
    ws = wb.active
    assert ws["A1"].value == "Document Type", (
        f"A1 header changed (got {ws['A1'].value!r}); "
        f"if columns.yaml was reshuffled, update this test."
    )
    assert ws["A2"].value == doc_type, (
        f"Document Type cell A2 should be {doc_type!r} "
        f"(the resolved doc_type passed into generate_bl_xlsx), "
        f"but got {ws['A2'].value!r}. "
        f"Likely cause: config/columns.yaml column A 'group_value' was "
        f"reverted to a hardcoded literal — it must be '${{document_type}}'."
    )


@pytest.mark.parametrize("doc_type", ["7400-000", "4000-000"])
def test_single_xlsx_writes_resolved_document_type(doc_type, tmp_path):
    """xlsx_generator.run (single-invoice path) must write the resolved
    doc_type from context into A2, not the hardcoded '4000-000'."""
    import json

    import xlsx_generator

    grouped = {
        "invoice_metadata": {
            "invoice_number": "INV_TEST_0001",
            "date": "2026-03-31",
            "document_type": doc_type,
        },
        "groups": [
            {
                "tariff_code": "90318000",
                "category": "MARINE ELECTRONICS",
                "item_count": 1,
                "sum_quantity": 1,
                "sum_total_cost": 100.0,
                "average_unit_cost": 100.0,
                "items": [
                    {
                        "supplier_item": "TEST-1",
                        "description": "Test item",
                        "quantity": 1,
                        "unit_price": 100.0,
                        "total_cost": 100.0,
                    }
                ],
            }
        ],
    }
    grouped_path = tmp_path / "grouped.json"
    grouped_path.write_text(json.dumps(grouped))
    out = tmp_path / "inv.xlsx"

    result = xlsx_generator.run(
        str(grouped_path),
        str(out),
        context={"document_type": doc_type, "input_file": str(grouped_path)},
    )
    assert result.get("status") != "error", f"xlsx_generator error: {result}"

    # run() may version-rename the output; follow the path it returned.
    actual_out = result.get("output") or str(out)
    wb = load_workbook(actual_out)
    ws = wb.active
    assert ws["A2"].value == doc_type, (
        f"xlsx_generator.run: Document Type A2 should be {doc_type!r} "
        f"(from context), got {ws['A2'].value!r}. "
        f"Likely cause: xlsx_generator.py reverted to hardcoded '4000-000'."
    )
