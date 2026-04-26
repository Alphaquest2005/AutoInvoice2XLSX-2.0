"""Regression tests over workspace/output/downloads-regression-emails/.

This is the baseline-diffing counterpart to ``test_downloads_regression.py``
but scoped to the per-email shipment corpus under
``workspace/output/downloads-regression-emails/`` (now 105+ folders
including the two TSCW Budget Marine BLs added 2026-04-26).  It exists
so that after pipeline changes we can re-run and spot per-shipment drift
against a frozen golden baseline.

Per-folder flow:

1. Stage only the SOURCE PDFs (exclude generated declarations,
   manifests, generator outputs, ``_email_params*``, ``*.meta.json``,
   and the ``_split_temp`` / ``Unprocessed`` side dirs) into a fresh
   ``tmp_path`` — mirrors ``scripts/rerun_corpus.py``'s
   ``is_source_file`` SSOT.
2. Invoke ``pipeline/run.py --input-dir <stage> --output-dir <out>
   --json-output --no-send-email`` via subprocess.
3. Run the full invariant checklist on every produced XLSX.
4. Apply named-fixture expectations from
   ``tests/fixtures/downloads_emails_regression/expectations.yaml``.
5. Snapshot the XLSX + ``_email_params*.json`` artifacts and diff
   against ``tests/regression_artifacts/downloads_emails/<folder>.baseline.json``.
   Drift is reported (not a hard failure) so legitimate improvements
   surface for review.  Promote with ``AUTOINVOICE_UPDATE_GOLDENS=1``.

Usage:

    # Full corpus (slow — vision API + OCR per PDF):
    source .venv/bin/activate
    pytest tests/pipeline/test_downloads_emails_regression.py -v -m integration

    # Single shipment (fast iteration):
    pytest tests/pipeline/test_downloads_emails_regression.py -v -m integration \\
        -k "TSCW18489131_budget_marine_grenada"

    # Promote a drifted baseline as the new golden:
    AUTOINVOICE_UPDATE_GOLDENS=1 pytest \\
        tests/pipeline/test_downloads_emails_regression.py -v -m integration \\
        -k "<folder>"

Named-fixture sanity:

The two TSCW Budget Marine shipments + the named fixtures above carry
field-level assertions on top of the snapshot diff.  Editing the YAML
expectations file is the authoritative way to add new named regression
cases — the snapshot diff alone catches unexpected drift, but field-
level assertions catch semantic regressions (e.g. doc_type silently
falling back to 4000-000) that might not be flagged by a matching
snapshot.
"""

from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

# tests/_paths is the SSOT bootstrap — it loads config/repo_layout.yaml
# and exposes PIPELINE_DIR + add_pipeline_to_sys_path() so this file
# never has to hardcode 'pipeline' as a directory-name literal.
from tests._paths import (
    PIPELINE_DIR,
    REPO_ROOT,
    add_pipeline_to_sys_path,
)

add_pipeline_to_sys_path()

from config_loader import (  # noqa: E402
    load_file_paths,
    load_library_enums,
    load_pipeline,
    load_xlsx_labels,
)

from tests.pipeline._regression_artifacts import snapshot_and_compare  # noqa: E402
from tests.pipeline.invariants import run_all_invariants  # noqa: E402

# ── Config-loaded constants ─────────────────────────────────────
_FILE_PATHS = load_file_paths()
_LIBRARY = load_library_enums()
_PIPE = load_pipeline()
_XLSX_LABELS = load_xlsx_labels()
_REPO_ROOT = REPO_ROOT  # local alias for path expressions below

_EXT = _FILE_PATHS["extensions"]
_EXT_PDF = _EXT["pdf"]
_EXT_TXT = _EXT["txt"]
_EXT_XLSX = _EXT["xlsx"]
_EXT_JSON = _EXT["json"]
_EXT_META_JSON = _EXT["meta_json"]
_EXT_PAGES_JSON = _EXT["pages_json"]
_EXT_PDF_PAGES_JSON = _EXT["pdf_pages_json"]

_GENERATED_PREFIXES = (
    _FILE_PATHS["generated_file_prefixes"]["email_params"],
    _FILE_PATHS["generated_file_prefixes"]["proposed_fixes_meta"],
    _FILE_PATHS["generated_file_prefixes"]["proposed_fixes_yaml"],
)
_GENERATED_SUFFIXES = {_EXT_XLSX, _EXT_META_JSON, _EXT_PAGES_JSON, _EXT_PDF_PAGES_JSON}

_RUN_PY_BASENAME = _FILE_PATHS["pipeline_script_basenames"]["run"]
_RUN_PY = PIPELINE_DIR / _RUN_PY_BASENAME
_DATA_DIR = _REPO_ROOT / _FILE_PATHS["workspace_dirs"]["data"]
_CET_DB = _DATA_DIR / Path(_FILE_PATHS["references"]["cet_database"]).name
_CORPUS_DIR = (
    _REPO_ROOT
    / _FILE_PATHS["workspace_dirs"]["workspace"]
    / _FILE_PATHS["workspace_dirs"]["workspace_output"]
    / _FILE_PATHS["workspace_dirs"]["downloads_regression_corpus"]
)

_RUN_SUBDIRS = _FILE_PATHS["run_subdirs"]
_SIDE_DIRS_TO_SKIP = {_RUN_SUBDIRS["split_temp"], _RUN_SUBDIRS["unprocessed"]}

_GLOBS = _FILE_PATHS["output_glob_patterns"]
_GLOB_EMAIL_PARAMS = _GLOBS["email_params_files"]
_GLOB_HAWB_XLSX = _GLOBS["hawb_xlsx"]
_GLOB_ANY_XLSX = _GLOBS["any_xlsx"]

_TEST_PATHS = _FILE_PATHS["test_paths"]
_TESTS_DIR_NAME = _TEST_PATHS["tests_dir"]
_FIXTURES_DIR_NAME = _TEST_PATHS["fixtures_dir"]
_DOWNLOADS_FIXTURE_DIR_NAME = _TEST_PATHS["downloads_emails_fixture_dir"]
_DOWNLOADS_EXPECTATIONS_FILE = _TEST_PATHS["downloads_emails_expectations"]

_REG = _PIPE["regression_test_settings"]
_TIMEOUT_SECONDS = _REG["subprocess_timeout_seconds"]
_BANNER_WIDTH = _REG["summary_banner_width"]
_STDERR_TAIL_CHARS = _REG["stderr_tail_chars"]
_SNAPSHOT_DIFF_MAX = _REG["snapshot_diff_max"]
_REPORT_JSON_MARKER = _REG["report_json_marker"]
_INVARIANTS_MODE = _REG["invariants_mode"]
_COMBINED_SUBSTR = _REG["combined_xlsx_substring"]
_CV_COMPARE_EPS = _REG["customs_value_compare_epsilon"]
_STAGE_SUBDIR_NAME = _REG["stage_subdir_name"]
_OUTPUT_SUBDIR_NAME = _REG["output_subdir_name"]
_SNAPSHOT_LABEL_PREFIX = _REG["snapshot_label_prefix"]
_SUMMARY_DIRNAME = _REG["summary_dirname"]
_SUMMARY_FILENAME = _SUMMARY_DIRNAME + _EXT_JSON
_PARAMETRIZE_ARGNAMES_CSV = _REG["parametrize_argnames_csv"]

_EXP_FIELDS = _REG["expectation_field_names"]
_FIELD_EXPECTED_WAYBILLS = _EXP_FIELDS["EXPECTED_WAYBILLS"]
_FIELD_EXPECTED_WAYBILLS_SUBSET = _EXP_FIELDS["EXPECTED_WAYBILLS_SUBSET"]
_FIELD_EXPECTED_CUSTOMS_VALUE = _EXP_FIELDS["EXPECTED_CUSTOMS_VALUE"]
_FIELD_CONSIGNEE = _EXP_FIELDS["CONSIGNEE"]
_FIELD_EXPECTED_DOC_TYPE_ALL = _EXP_FIELDS["EXPECTED_DOC_TYPE_ALL"]
_FIELD_EXPECTED_DOC_TYPE_MAP = _EXP_FIELDS["EXPECTED_DOC_TYPE_MAP"]

_SNAP_STATUS = _PIPE["snapshot_status"]
_STATUS_DRIFT = _SNAP_STATUS["DRIFT"]
_STATUS_NEW = _SNAP_STATUS["NEW"]
_STATUS_PROMOTED = _SNAP_STATUS["PROMOTED"]
_STATUS_MATCH = _SNAP_STATUS["MATCH"]

_RESULT_CATEGORIES = _PIPE["regression_test_result_categories"]
_CAT_ERROR_CRASH = _RESULT_CATEGORIES["ERROR_CRASH"]
_CAT_SKIPPED_NO_PDFS = _RESULT_CATEGORIES["SKIPPED_NO_PDFS"]
_CAT_PROCESSED_PASS = _RESULT_CATEGORIES["PROCESSED_PASS"]
_CAT_PROCESSED_FAIL = _RESULT_CATEGORIES["PROCESSED_FAIL"]

_FAILURE_KINDS = _PIPE["regression_test_failure_kinds"]
_FAIL_TIMEOUT = _FAILURE_KINDS["TIMEOUT"]
_FAIL_NONZERO_EXIT = _FAILURE_KINDS["NONZERO_EXIT"]
_FAIL_STDERR_TAIL = _FAILURE_KINDS["STDERR_TAIL"]
_FAIL_MISSING_REPORT = _FAILURE_KINDS["MISSING_REPORT"]
_FAIL_WAYBILLS_MISMATCH = _FAILURE_KINDS["WAYBILLS_MISMATCH"]
_FAIL_WAYBILLS_MISSING = _FAILURE_KINDS["WAYBILLS_MISSING"]
_FAIL_CONSIGNEE = _FAILURE_KINDS["CONSIGNEE"]
_FAIL_DOC_TYPE = _FAILURE_KINDS["DOC_TYPE"]

_CLI = _PIPE["run_cli_args"]
_CLI_PYTHON_UNBUFFERED = _CLI["python_unbuffered"]
_CLI_INPUT_DIR = _CLI["input_dir"]
_CLI_OUTPUT_DIR = _CLI["output_dir"]
_CLI_JSON_OUTPUT = _CLI["json_output"]
_CLI_NO_SEND_EMAIL = _CLI["no_send_email"]

_PYTEST_SCOPE_SESSION = _LIBRARY["pytest"]["fixture_scope"]["SESSION"]

_LABEL_CLIENT_DECLARED_DUTIES = _XLSX_LABELS["duty"]["client_declared"]

# Per-shipment named-fixture expectations live in YAML so the test
# source carries no domain literals.
_EXPECTATIONS_PATH = (
    _REPO_ROOT
    / _TESTS_DIR_NAME
    / _FIXTURES_DIR_NAME
    / _DOWNLOADS_FIXTURE_DIR_NAME
    / _DOWNLOADS_EXPECTATIONS_FILE
)
with _EXPECTATIONS_PATH.open("r", encoding="utf-8") as _fh:
    _SHIPMENT_EXPECTATIONS: dict[str, dict] = yaml.safe_load(_fh) or {}


def _read_client_declared_duties(xlsx_path: Path) -> float | None:
    """Return the ``CLIENT DECLARED DUTIES`` row value from *xlsx_path*."""
    try:
        wb = load_workbook(str(xlsx_path), data_only=False)
    except Exception:  # noqa: BLE001
        return None
    try:
        target = _LABEL_CLIENT_DECLARED_DUTIES.upper()
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    v = cell.value
                    if not isinstance(v, str):
                        continue
                    if v.strip().upper() != target:
                        continue
                    for other in row[cell.column - 1 :]:
                        if other is cell:
                            continue
                        ov = other.value
                        if isinstance(ov, (int, float)):
                            return float(ov)
                        if isinstance(ov, str):
                            try:
                                return float(ov.replace(",", "").strip())
                            except ValueError:
                                continue
        return None
    finally:
        wb.close()


def _read_doc_type_cell_a2(xlsx_path: Path) -> str | None:
    """Return cell A2 of the first sheet — the document_type written by
    bl_xlsx_generator / xlsx_generator. Used to assert per-XLSX
    expected_doc_type from the YAML expectations."""
    try:
        wb = load_workbook(str(xlsx_path), data_only=True)
    except Exception:  # noqa: BLE001
        return None
    try:
        ws = wb.active
        v = ws["A2"].value
        if v is None:
            return None
        return str(v).strip()
    finally:
        wb.close()


def _is_source_pdf(path: Path) -> bool:
    """True for PDFs the pipeline should re-process from."""
    name = path.name
    if not path.is_file():
        return False
    if path.suffix.lower() != _EXT_PDF:
        return False
    if any(name.startswith(p) for p in _GENERATED_PREFIXES):
        return False
    suffix = "".join(path.suffixes)
    return suffix not in _GENERATED_SUFFIXES


def _discover_corpus_folders() -> list[tuple[str, Path]]:
    """Return ``[(name, path), ...]`` for every folder with at least one
    source PDF."""
    if not _CORPUS_DIR.is_dir():
        return []
    results: list[tuple[str, Path]] = []
    for entry in sorted(_CORPUS_DIR.iterdir()):
        if not entry.is_dir():
            continue
        if entry.name.startswith("."):
            continue
        if entry.name in _SIDE_DIRS_TO_SKIP:
            continue
        pdfs = [p for p in entry.iterdir() if _is_source_pdf(p)]
        if not pdfs:
            continue
        results.append((entry.name, entry))
    return results


_FOLDERS = _discover_corpus_folders()

# Opt-in restriction: when AUTOINVOICE_DOWNLOADS_EMAILS_ONLY is set
# (comma-separated folder names) only those folders run.
_ONLY_ENV = os.environ.get("AUTOINVOICE_DOWNLOADS_EMAILS_ONLY", "").strip()
if _ONLY_ENV:
    _only = {s.strip() for s in _ONLY_ENV.split(",") if s.strip()}
    _FOLDERS = [(n, p) for (n, p) in _FOLDERS if n in _only]

if not _FOLDERS:
    pytest.skip(
        f"no shipment folders with source PDFs under {_CORPUS_DIR.relative_to(_REPO_ROOT)}",
        allow_module_level=True,
    )


# Session-scoped results collector for the summary fixture.
_RESULTS: list[dict] = []


@pytest.fixture(scope=_PYTEST_SCOPE_SESSION, autouse=True)
def _summary_reporter(tmp_path_factory):
    """Emit a session-end summary of per-shipment results + drift list."""
    yield
    if not _RESULTS:
        return
    summary_dir = tmp_path_factory.mktemp(_SUMMARY_DIRNAME)
    summary_path = summary_dir / _SUMMARY_FILENAME
    with summary_path.open("w", encoding="utf-8") as f:
        json.dump(_RESULTS, f, indent=2, default=str)
    drift = [r for r in _RESULTS if r.get("snapshot_status") == _STATUS_DRIFT]
    new = [r for r in _RESULTS if r.get("snapshot_status") == _STATUS_NEW]
    promoted = [r for r in _RESULTS if r.get("snapshot_status") == _STATUS_PROMOTED]
    matched = [r for r in _RESULTS if r.get("snapshot_status") == _STATUS_MATCH]
    with_invariant_fails = [r for r in _RESULTS if r.get("invariant_failure_count")]
    rule = "=" * _BANNER_WIDTH
    print("\n" + rule)
    print(f"Downloads-emails regression summary  ({len(_RESULTS)} shipments)")
    print(rule)
    print(
        f"  snapshot  match={len(matched)}  new={len(new)}  drift={len(drift)}  "
        f"promoted={len(promoted)}"
    )
    print(f"  invariant violations (drift-only): {len(with_invariant_fails)} shipments")
    for r in drift:
        print(
            f"    DRIFT {r['folder']}  ({r.get('snapshot_diff_count', '?')} diffs)  "
            f"→ {r.get('snapshot_current_dir', '?')}"
        )
    for r in with_invariant_fails:
        print(f"    INVARIANT {r['folder']}  ({r['invariant_failure_count']} violations)")
    print(f"\nFull report: {summary_path}")
    print(rule)


def _run_pipeline(stage_dir: Path, output_dir: Path) -> subprocess.CompletedProcess:
    # --no-send-email: regression runs MUST NOT deliver real broker emails.
    return subprocess.run(  # noqa: S603 - trusted local path
        [
            sys.executable,
            _CLI_PYTHON_UNBUFFERED,
            str(_RUN_PY),
            _CLI_INPUT_DIR,
            str(stage_dir),
            _CLI_OUTPUT_DIR,
            str(output_dir),
            _CLI_JSON_OUTPUT,
            _CLI_NO_SEND_EMAIL,
        ],
        cwd=str(PIPELINE_DIR),
        capture_output=True,
        text=True,
        timeout=_TIMEOUT_SECONDS,
    )


def _parse_report(stdout: str) -> dict | None:
    for line in stdout.splitlines():
        stripped = line.strip()
        if stripped.startswith(_REPORT_JSON_MARKER):
            try:
                return json.loads(stripped[len(_REPORT_JSON_MARKER) :])
            except json.JSONDecodeError:
                return None
    return None


def _load_email_params(output_dir: Path) -> list[dict]:
    """Return the contents of every ``_email_params*.json`` produced."""
    out: list[dict] = []
    for p in sorted(output_dir.rglob(_GLOB_EMAIL_PARAMS)):
        try:
            out.append(json.loads(p.read_text(encoding="utf-8")))
        except (OSError, json.JSONDecodeError):
            continue
    return out


def _check_shipment_expectations(
    folder_name: str,
    output_dir: Path,
    report: dict,
) -> list[tuple[str, str]]:
    """Apply the named-fixture sanity checks.  Returns list of failures."""
    exp = _SHIPMENT_EXPECTATIONS.get(folder_name)
    if not exp:
        return []
    failures: list[tuple[str, str]] = []

    produced_waybills: set[str] = set()
    for x in output_dir.rglob(_GLOB_HAWB_XLSX):
        stem = x.stem
        if "-" in stem:
            stem = stem.split("-")[0]
        produced_waybills.add(stem)

    if _FIELD_EXPECTED_WAYBILLS in exp:
        want = set(exp[_FIELD_EXPECTED_WAYBILLS])
        if produced_waybills != want:
            failures.append(
                (
                    _FAIL_WAYBILLS_MISMATCH,
                    f"expected exactly {sorted(want)}, got {sorted(produced_waybills)}",
                )
            )
    if _FIELD_EXPECTED_WAYBILLS_SUBSET in exp:
        want = set(exp[_FIELD_EXPECTED_WAYBILLS_SUBSET])
        missing = want - produced_waybills
        if missing:
            failures.append(
                (
                    _FAIL_WAYBILLS_MISSING,
                    f"expected subset {sorted(want)}, missing {sorted(missing)}",
                )
            )

    expected_cv = exp.get(_FIELD_EXPECTED_CUSTOMS_VALUE) or {}
    if expected_cv:
        for wb, want in expected_cv.items():
            xlsx_paths = list(output_dir.rglob(f"{wb}{_EXT_XLSX}"))
            if not xlsx_paths:
                failures.append(
                    (
                        f"customs_value_ec::{wb}",
                        f"no {wb}{_EXT_XLSX} produced",
                    )
                )
                continue
            xlsx_paths.sort(key=lambda p: _COMBINED_SUBSTR in p.name.lower())
            got = _read_client_declared_duties(xlsx_paths[0])
            if want is None:
                if got is None:
                    failures.append(
                        (
                            f"customs_value_ec::{wb}",
                            f"no {_LABEL_CLIENT_DECLARED_DUTIES} row in XLSX ({xlsx_paths[0].name})",  # noqa: E501
                        )
                    )
                continue
            if got is None or abs(got - float(want)) > _CV_COMPARE_EPS:
                failures.append(
                    (
                        f"customs_value_ec::{wb}",
                        f"expected {want!r}, got {got!r} (xlsx={xlsx_paths[0].name})",
                    )
                )

    if _FIELD_CONSIGNEE in exp:
        params = _load_email_params(output_dir)
        consignees = {p.get("consignee_name") for p in params}
        if exp[_FIELD_CONSIGNEE] not in consignees:
            failures.append(
                (
                    _FAIL_CONSIGNEE,
                    f"expected {exp[_FIELD_CONSIGNEE]!r} in any _email_params, got {consignees}",
                )
            )

    # Per-XLSX doc_type assertion (added 2026-04-26 for the
    # TSCW18489131 / 26006159 rebuild-leak regression).  Two shapes:
    #   expected_doc_type_all : single string applied to every XLSX
    #   expected_doc_type     : per-waybill mapping
    expected_dt_all = exp.get(_FIELD_EXPECTED_DOC_TYPE_ALL)
    if expected_dt_all:
        produced_xlsx = list(output_dir.rglob(_GLOB_ANY_XLSX))
        for xp in produced_xlsx:
            got = _read_doc_type_cell_a2(xp)
            if got != expected_dt_all:
                failures.append(
                    (
                        f"{_FAIL_DOC_TYPE}::{xp.name}",
                        f"expected A2={expected_dt_all!r}, got {got!r}",
                    )
                )
    expected_dt_map = exp.get(_FIELD_EXPECTED_DOC_TYPE_MAP) or {}
    for wb, want_dt in expected_dt_map.items():
        xlsx_paths = list(output_dir.rglob(f"{wb}{_EXT_XLSX}"))
        if not xlsx_paths:
            failures.append(
                (
                    f"{_FAIL_DOC_TYPE}::{wb}",
                    f"no {wb}{_EXT_XLSX} produced",
                )
            )
            continue
        got = _read_doc_type_cell_a2(xlsx_paths[0])
        if got != want_dt:
            failures.append(
                (
                    f"{_FAIL_DOC_TYPE}::{wb}",
                    f"expected A2={want_dt!r}, got {got!r} (xlsx={xlsx_paths[0].name})",
                )
            )

    return failures


@pytest.mark.integration
@pytest.mark.requires_downloads
@pytest.mark.parametrize(
    _PARAMETRIZE_ARGNAMES_CSV,
    _FOLDERS,
    ids=[n for (n, _) in _FOLDERS],
)
def test_downloads_email_folder(
    folder_name: str,
    folder_path: Path,
    tmp_path: Path,
) -> None:
    """Process a single shipment folder and diff against baseline."""
    stage_dir = tmp_path / _STAGE_SUBDIR_NAME
    stage_dir.mkdir()
    output_dir = tmp_path / _OUTPUT_SUBDIR_NAME
    output_dir.mkdir()

    staged = 0
    for item in folder_path.iterdir():
        if _is_source_pdf(item):
            shutil.copy2(item, stage_dir / item.name)
            staged += 1
            sidecar_txt = folder_path / (item.stem + _EXT_TXT)
            if sidecar_txt.exists():
                shutil.copy2(sidecar_txt, stage_dir / sidecar_txt.name)

    result: dict = {
        "folder": folder_name,
        "staged_pdfs": staged,
        "category": _CAT_ERROR_CRASH,
        "failures": [],
    }

    if staged == 0:
        result["category"] = _CAT_SKIPPED_NO_PDFS
        _RESULTS.append(result)
        pytest.skip(f"no source PDFs in {folder_name}")

    try:
        proc = _run_pipeline(stage_dir, output_dir)
    except subprocess.TimeoutExpired as e:
        result["failures"] = [(_FAIL_TIMEOUT, f"subprocess timed out after {_TIMEOUT_SECONDS}s")]
        _RESULTS.append(result)
        pytest.fail(f"[{folder_name}] pipeline timed out: {e}")

    result["returncode"] = proc.returncode
    if proc.returncode != 0:
        result["failures"] = [
            (_FAIL_NONZERO_EXIT, f"returncode={proc.returncode}"),
            (_FAIL_STDERR_TAIL, proc.stderr[-_STDERR_TAIL_CHARS:]),
        ]
        _RESULTS.append(result)
        pytest.fail(
            f"[{folder_name}] pipeline exited with {proc.returncode}.\n"
            f"stderr tail: {proc.stderr[-_STDERR_TAIL_CHARS:]}"
        )

    report = _parse_report(proc.stdout)
    result["report"] = report
    if report is None:
        result["failures"] = [(_FAIL_MISSING_REPORT, f"no {_REPORT_JSON_MARKER} line in stdout")]
        _RESULTS.append(result)
        pytest.fail(f"[{folder_name}] pipeline did not emit {_REPORT_JSON_MARKER}")

    # Invariants on every produced XLSX — recorded as drift, not hard
    # failure.  Many shipments have pre-existing pipeline-level invariant
    # violations (the historical VARIANCE CHECK=0 issue covered by
    # project_variance_check_historical).  We surface them in the
    # summary for review without blocking the regression sweep.
    produced_xlsx = list(output_dir.rglob(_GLOB_ANY_XLSX))
    invariant_failures: list[tuple[str, str]] = []
    for xlsx_path in produced_xlsx:
        try:
            wb = load_workbook(str(xlsx_path))
        except Exception as e:  # noqa: BLE001
            invariant_failures.append((f"load::{xlsx_path.name}", str(e)))
            continue
        for sheet in wb.worksheets:
            failures = run_all_invariants(
                sheet,
                mode=_INVARIANTS_MODE,
                cet_db_path=str(_CET_DB) if _CET_DB.exists() else None,
                xlsx_path=str(xlsx_path),
            )
            for name, msg in failures:
                invariant_failures.append((f"{xlsx_path.name}::{sheet.title}::{name}", msg))
        wb.close()
    if invariant_failures:
        result["invariant_failures"] = invariant_failures
        result["invariant_failure_count"] = len(invariant_failures)

    # Named-fixture semantic checks — these ARE hard failures.
    semantic_failures = _check_shipment_expectations(folder_name, output_dir, report)

    if semantic_failures:
        result["category"] = _CAT_PROCESSED_FAIL
        result["failures"] = semantic_failures
        _RESULTS.append(result)
        lines = [f"[{folder_name}] named-fixture regression failures:"]
        for name, msg in semantic_failures:
            lines.append(f"  - {name}: {msg}")
        pytest.fail("\n".join(lines))

    result["category"] = _CAT_PROCESSED_PASS

    # Snapshot + diff against frozen baseline.
    try:
        snap = snapshot_and_compare(
            _SNAPSHOT_LABEL_PREFIX + folder_name,
            output_dir,
        )
    except Exception as e:  # noqa: BLE001
        result["snapshot_error"] = str(e)
    else:
        result["snapshot_status"] = snap["status"]
        result["snapshot_hash"] = snap["hash"]
        if snap["diffs"]:
            result["snapshot_diffs"] = snap["diffs"][:_SNAPSHOT_DIFF_MAX]
            result["snapshot_diff_count"] = len(snap["diffs"])
            result["snapshot_current_dir"] = snap["current_dir"]

    _RESULTS.append(result)
