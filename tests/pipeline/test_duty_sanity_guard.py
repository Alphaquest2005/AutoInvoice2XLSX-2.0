"""Tests for the CLIENT DECLARED DUTIES reverse-calc display (Fix C-2).

The original Fix C "20% suspiciously low" rule was wrong — a low declared
duty is LEGITIMATE when the waybill contains fewer items (multi-waybill
shipments split items across declarations). The replacement behaviour:

1. Always show CLIENT DECLARED + DUTY VARIANCE + IMPLIED CET (no suppression).
2. Show IMPLIED CIF and IMPLIED ITEMS VALUE (reverse-calc from declared duty).
3. Show ITEM ALLOCATION CHECK warning only when the reverse-calculated items
   value disagrees with the items actually on this sheet by > 35%.

That warning is a diagnostic for mis-allocated items between per-decl XLSX
files — NOT a signal that the declared duty is wrong.
"""

import os
import sys

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "pipeline"))

from bl_xlsx_generator import _write_duty_estimation_section  # noqa: E402


def _fresh_ws():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "seed"
    return ws


def _labels(ws):
    """All non-empty column-J labels."""
    return [row[0] for row in ws.iter_rows(min_col=10, max_col=10, values_only=True) if row[0]]


def _row_value(ws, label_substr: str):
    """Return column-P value for the row whose column-J label contains substr."""
    for row in ws.iter_rows(values_only=False):
        j = row[9].value if len(row) > 9 else None
        if j and label_substr in str(j):
            return row[15].value if len(row) > 15 else None
    return None


# ---------------------------------------------------------------------------
# No false-positive suppression — $8.96 is legitimate for a small-items waybill
# ---------------------------------------------------------------------------

def test_low_declared_duty_does_not_suppress_variance():
    """$8.96 declared vs $86.57 estimated: DUTY VARIANCE and IMPLIED CET
    (when valid) must still be shown — this is a legitimate case of a
    waybill with very few items, not an OCR misread."""
    ws = _fresh_ws()
    matched_items = [{'tariff_code': '62034220', 'total_cost': 100.00}]
    invoice_data = {
        'invoice_total': 100.00,
        '_customs_freight': 0,
        '_customs_insurance': 0,
        '_client_declared_duties': 8.96,  # legitimately low
    }

    _write_duty_estimation_section(ws, matched_items, invoice_data)
    labels = '\n'.join(str(x) for x in _labels(ws))

    # Must NOT have the old false-positive banner
    assert 'SUSPICIOUSLY LOW' not in labels
    # Variance must still be shown
    assert 'DUTY VARIANCE' in labels


def test_client_declared_row_always_present():
    ws = _fresh_ws()
    matched_items = [{'tariff_code': '62034220', 'total_cost': 100.00}]
    invoice_data = {
        'invoice_total': 100.00, '_customs_freight': 0, '_customs_insurance': 0,
        '_client_declared_duties': 8.96,
    }
    _write_duty_estimation_section(ws, matched_items, invoice_data)
    assert _row_value(ws, 'CLIENT DECLARED DUTIES') == 8.96


# ---------------------------------------------------------------------------
# Reverse-calculation rows
# ---------------------------------------------------------------------------

def test_implied_cif_row_rendered():
    """IMPLIED CIF row should appear whenever declared duty > 0."""
    ws = _fresh_ws()
    matched_items = [{'tariff_code': '62034220', 'total_cost': 100.00}]
    invoice_data = {
        'invoice_total': 100.00, '_customs_freight': 0, '_customs_insurance': 0,
        '_client_declared_duties': 8.96,
    }
    _write_duty_estimation_section(ws, matched_items, invoice_data)
    labels = '\n'.join(str(x) for x in _labels(ws))
    assert 'IMPLIED CIF' in labels
    assert 'IMPLIED ITEMS VALUE' in labels


def test_implied_cif_math_at_20pct_cet():
    """With avg CET 20%, composite=0.449. Declared $8.96 → CIF ≈ $19.96 XCD."""
    ws = _fresh_ws()
    matched_items = [{'tariff_code': '62034220', 'total_cost': 100.00}]  # 20% CET
    invoice_data = {
        'invoice_total': 100.00, '_customs_freight': 5.0, '_customs_insurance': 0,
        '_client_declared_duties': 8.96,
    }
    _write_duty_estimation_section(ws, matched_items, invoice_data)
    implied_cif_xcd = _row_value(ws, 'IMPLIED CIF')
    # 8.96 / 0.449 = 19.955...
    assert implied_cif_xcd is not None
    assert 19.5 < implied_cif_xcd < 20.5


def test_implied_items_value_subtracts_freight():
    """IMPLIED ITEMS VALUE = CIF/rate − freight − insurance, floored at 0."""
    ws = _fresh_ws()
    matched_items = [{'tariff_code': '62034220', 'total_cost': 100.00}]
    invoice_data = {
        'invoice_total': 100.00, '_customs_freight': 5.0, '_customs_insurance': 0,
        '_client_declared_duties': 8.96,
    }
    _write_duty_estimation_section(ws, matched_items, invoice_data)
    items_usd = _row_value(ws, 'IMPLIED ITEMS VALUE')
    # CIF USD = 19.96 / 2.7169 = 7.35, minus $5 freight = 2.35
    assert items_usd is not None
    assert 2.0 < items_usd < 2.7


def test_items_value_floored_at_zero():
    """When freight exceeds implied CIF (tiny duty), IMPLIED ITEMS VALUE = 0 (never negative)."""
    ws = _fresh_ws()
    matched_items = [{'tariff_code': '62034220', 'total_cost': 100.00}]
    invoice_data = {
        'invoice_total': 100.00, '_customs_freight': 50.0, '_customs_insurance': 0,
        '_client_declared_duties': 1.0,  # absurdly tiny
    }
    _write_duty_estimation_section(ws, matched_items, invoice_data)
    items_usd = _row_value(ws, 'IMPLIED ITEMS VALUE')
    assert items_usd == 0 or items_usd == 0.0


# ---------------------------------------------------------------------------
# Item allocation check
# ---------------------------------------------------------------------------

def test_allocation_warning_when_items_over_allocated():
    """$8.96 implies ~$2 of items; we have $60 on the sheet → OVER-ALLOCATED warning."""
    ws = _fresh_ws()
    matched_items = [
        {'tariff_code': '62034220', 'total_cost': 41.98},
        {'tariff_code': '65050000', 'total_cost': 12.99},
        {'tariff_code': '00000000', 'total_cost': 5.99},
    ]
    invoice_data = {
        'invoice_total': 60.96, '_customs_freight': 5.0, '_customs_insurance': 0,
        '_client_declared_duties': 8.96,
    }
    _write_duty_estimation_section(ws, matched_items, invoice_data)
    labels = '\n'.join(str(x) for x in _labels(ws))
    assert 'ITEM ALLOCATION CHECK' in labels
    assert 'OVER-ALLOCATED' in labels


def test_no_allocation_warning_when_items_match():
    """Declared $80 on items worth ~$66: within tolerance → no warning."""
    ws = _fresh_ws()
    matched_items = [
        {'tariff_code': '65050000', 'total_cost': 37.98},  # straw hats
        {'tariff_code': '65050000', 'total_cost': 19.99},  # garbage PAYMENT line
    ]
    invoice_data = {
        'invoice_total': 57.97, '_customs_freight': 9.0, '_customs_insurance': 0,
        '_client_declared_duties': 80.06,
    }
    _write_duty_estimation_section(ws, matched_items, invoice_data)
    labels = '\n'.join(str(x) for x in _labels(ws))
    # Within 35% tolerance → no allocation warning
    assert 'ITEM ALLOCATION CHECK' not in labels


def test_allocation_warning_under_allocated():
    """Declared much higher than estimated → UNDER-ALLOCATED warning."""
    ws = _fresh_ws()
    matched_items = [{'tariff_code': '62034220', 'total_cost': 5.00}]
    invoice_data = {
        'invoice_total': 5.00, '_customs_freight': 0, '_customs_insurance': 0,
        '_client_declared_duties': 80.06,
    }
    _write_duty_estimation_section(ws, matched_items, invoice_data)
    labels = '\n'.join(str(x) for x in _labels(ws))
    assert 'ITEM ALLOCATION CHECK' in labels
    assert 'UNDER-ALLOCATED' in labels


def test_no_client_declared_no_reverse_calc():
    """When no client value is set, no CLIENT/VARIANCE/IMPLIED rows."""
    ws = _fresh_ws()
    matched_items = [{'tariff_code': '62034220', 'total_cost': 100.00}]
    invoice_data = {
        'invoice_total': 100.00, '_customs_freight': 0, '_customs_insurance': 0,
    }
    _write_duty_estimation_section(ws, matched_items, invoice_data)
    labels = '\n'.join(str(x) for x in _labels(ws))
    assert 'CLIENT DECLARED DUTIES' not in labels
    assert 'IMPLIED CIF' not in labels
    assert 'IMPLIED ITEMS VALUE' not in labels
