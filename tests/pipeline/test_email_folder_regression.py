"""Regression tests over the workspace/documents/ email folder corpus.

For every subfolder of ``workspace/documents/`` that contains at least
one PDF, this module parametrizes a test that:

1. Stages the PDFs into a fresh ``tmp_path`` (never writing into the
   source folder).
2. Invokes ``pipeline/run.py --input-dir <stage> --output-dir <stage>
   --json-output`` via subprocess — same path as
   ``pipeline/test_email_pipeline.py`` and the Electron app's
   ``runBLPipeline``.
3. Parses the ``REPORT:JSON:`` line from stdout and asserts ``status ==
   'success'``.
4. Opens every ``.xlsx`` produced in ``tmp_path`` and runs the full
   invariant checklist (``run_all_invariants``) in ``minimal`` mode,
   collecting per-file failures with precise row/column messages.

Usage:

    # Activate venv first — the pipeline needs pdfplumber, openpyxl, etc.
    source .venv/bin/activate
    pytest tests/pipeline/test_email_folder_regression.py -v -m integration

If ``workspace/documents/`` contains no PDF-bearing subfolders, the
entire module is skipped with a clear message.
"""

from __future__ import annotations

import json
import shutil
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

from tests.pipeline.invariants import run_all_invariants

# ── Paths ──────────────────────────────────────────────────────────────
_REPO_ROOT = Path(__file__).resolve().parents[2]
_DOCUMENTS_DIR = _REPO_ROOT / "workspace" / "documents"
_PIPELINE_DIR = _REPO_ROOT / "pipeline"
_RUN_PY = _PIPELINE_DIR / "run.py"
_CET_DB = _REPO_ROOT / "data" / "cet.db"

_TIMEOUT_SECONDS = 600  # matches pipeline/test_email_pipeline.py


def _discover_email_folders() -> list[tuple[str, Path]]:
    """Return ``[(name, path), ...]`` for every PDF-containing subfolder."""
    if not _DOCUMENTS_DIR.is_dir():
        return []
    results: list[tuple[str, Path]] = []
    for entry in sorted(_DOCUMENTS_DIR.iterdir()):
        if not entry.is_dir():
            continue
        pdfs = [*entry.glob("*.pdf"), *entry.glob("*.PDF")]
        if not pdfs:
            continue
        results.append((entry.name, entry))
    return results


_EMAIL_FOLDERS = _discover_email_folders()

if not _EMAIL_FOLDERS:
    pytest.skip(
        "no email folders available under workspace/documents/",
        allow_module_level=True,
    )


@pytest.mark.integration
@pytest.mark.parametrize(
    ("folder_name", "folder_path"),
    _EMAIL_FOLDERS,
    ids=[name for name, _ in _EMAIL_FOLDERS],
)
def test_email_folder_regression(
    folder_name: str,
    folder_path: Path,
    tmp_path: Path,
) -> None:
    """Run pipeline on a single email folder and assert invariants."""
    # 1. Stage PDFs into tmp_path (never write into workspace/documents/).
    stage_dir = tmp_path / "stage"
    stage_dir.mkdir()
    for item in folder_path.iterdir():
        if item.is_file():
            shutil.copy2(item, stage_dir / item.name)

    output_dir = tmp_path / "out"
    output_dir.mkdir()

    # 2. Run the pipeline.
    proc = subprocess.run(  # noqa: S603 - trusted local path
        [
            sys.executable,
            "-u",
            str(_RUN_PY),
            "--input-dir",
            str(stage_dir),
            "--output-dir",
            str(output_dir),
            "--json-output",
        ],
        cwd=str(_PIPELINE_DIR),
        capture_output=True,
        text=True,
        timeout=_TIMEOUT_SECONDS,
    )

    assert proc.returncode == 0, (
        f"pipeline/run.py exited with code {proc.returncode} for {folder_name!r}.\n"
        f"stderr (last 1000 chars): {proc.stderr[-1000:]}\n"
        f"stdout (last 1000 chars): {proc.stdout[-1000:]}"
    )

    # 3. Parse REPORT:JSON line from stdout.
    report = None
    for line in proc.stdout.splitlines():
        stripped = line.strip()
        if stripped.startswith("REPORT:JSON:"):
            report = json.loads(stripped[len("REPORT:JSON:") :])
            break

    assert report is not None, (
        f"pipeline did not emit a REPORT:JSON line for {folder_name!r}.\n"
        f"stdout (last 1000 chars): {proc.stdout[-1000:]}"
    )
    assert report.get("status") == "success", (
        f"pipeline report status != 'success' for {folder_name!r}: "
        f"{json.dumps(report, indent=2)[:1500]}"
    )

    # 4. Run full invariant checklist against every produced XLSX.
    # Gather XLSX files from both output_dir (explicit --output-dir) and
    # stage_dir (some generators write alongside the PDFs), but skip the
    # original PO workbooks copied in at step 1.
    original_xlsx_names = {p.name for p in folder_path.iterdir() if p.suffix.lower() == ".xlsx"}
    produced: list[Path] = []
    for d in (output_dir, stage_dir):
        for xlsx in d.rglob("*.xlsx"):
            if xlsx.name not in original_xlsx_names:
                produced.append(xlsx)

    per_file_failures: dict[str, list[tuple[str, str]]] = {}
    for xlsx_path in produced:
        try:
            wb = load_workbook(str(xlsx_path))
        except Exception as e:  # noqa: BLE001 - surface any open error
            per_file_failures[xlsx_path.name] = [("load_workbook", str(e))]
            continue
        for sheet in wb.worksheets:
            failures = run_all_invariants(
                sheet,
                mode="minimal",
                cet_db_path=str(_CET_DB) if _CET_DB.exists() else None,
                xlsx_path=str(xlsx_path),
            )
            if failures:
                key = f"{xlsx_path.name}::{sheet.title}"
                per_file_failures[key] = failures
        wb.close()

    if per_file_failures:
        lines = [f"Invariant failures in {folder_name!r}:"]
        for key, failures in per_file_failures.items():
            lines.append(f"  {key}:")
            for name, msg in failures:
                lines.append(f"    - {name}: {msg}")
        pytest.fail("\n".join(lines))
