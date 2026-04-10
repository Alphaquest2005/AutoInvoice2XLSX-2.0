"""Unit tests for the extended XLSX invariant checklist.

Every new invariant in ``tests/pipeline/invariants.py`` gets paired tests
here:

  - ``test_<name>_passes`` — a synthetic workbook that satisfies the rule
  - ``test_<name>_fails``  — a mutated workbook that violates the rule,
    and the test asserts the expected ``AssertionError`` message fragment

The workbooks come from ``xlsx_factory.build_grouped_workbook`` and
``build_ungrouped_workbook`` — the same fixtures already used by the
variance-fixer tests.  Mutations are applied in-place on the returned
worksheet so the failing case differs from the passing one by exactly
one cell.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from tests.pipeline.invariants import (
    COL_F_TARIFF,
    COL_J_SUPPLIER_ITEM_DESC,
    COL_L_PER_UNIT,
    COL_O_COST,
    COL_P_TOTAL_COST,
    COL_Q_TOTAL,
    COL_R_TC_VS_TOTAL,
    COL_S_INVOICE_TOTAL,
    COL_T_FREIGHT,
    assert_adjustments_has_no_stacked_corrections,
    assert_column_types,
    assert_group_row_totals_match_details,
    assert_group_verification_zero,
    assert_headers_present,
    assert_invoice_totals_only_on_first_group_row,
    assert_row_q_r_formulas,
    assert_tariff_codes_valid,
    assert_totals_section_formulas,
    assert_variance_is_formula,
    assert_variance_within_tolerance,
    run_all_invariants,
)
from tests.pipeline.xlsx_factory import build_grouped_workbook, build_ungrouped_workbook

# ─── helpers ────────────────────────────────────────────────────────────


def _simple_ungrouped():
    return build_ungrouped_workbook(
        inv_total=100.0,
        items=[("Alpha", 2, 25.0), ("Beta", 1, 50.0)],
        freight=0.0,
        insurance=0.0,
        tax=0.0,
        deduction=0.0,
    )


def _simple_grouped():
    return build_grouped_workbook(
        inv_total=100.0,
        groups=[
            ("BEVERAGES (1 items)", [("Can", 2, 10.0)]),
            ("SOAPS (1 items)", [("Bar", 4, 20.0)]),
        ],
        freight=0.0,
        insurance=0.0,
        tax=0.0,
        deduction=0.0,
    )


# ─── assert_headers_present ─────────────────────────────────────────────


def test_assert_headers_present_minimal_passes():
    wb = _simple_ungrouped()
    assert_headers_present(wb.active, mode="minimal")


def test_assert_headers_present_minimal_fails_when_missing():
    wb = _simple_ungrouped()
    ws = wb.active
    ws.cell(row=1, column=COL_F_TARIFF, value="NotTariffCode")
    with pytest.raises(AssertionError, match="header mismatch at row 1, col F"):
        assert_headers_present(ws, mode="minimal")


def test_assert_headers_present_strict_fails_on_minimal_workbook():
    # The factory only writes minimal headers — strict mode should fail
    # because col A "Document Type" is missing.
    wb = _simple_ungrouped()
    with pytest.raises(AssertionError, match="Document Type"):
        assert_headers_present(wb.active, mode="strict")


# ─── assert_tariff_codes_valid ──────────────────────────────────────────


def test_assert_tariff_codes_valid_passes():
    wb = _simple_ungrouped()
    ws = wb.active
    assert_tariff_codes_valid(ws, first_data_row=2, last_data_row=3)


def test_assert_tariff_codes_valid_fails_on_short_code():
    wb = _simple_ungrouped()
    ws = wb.active
    ws.cell(row=2, column=COL_F_TARIFF, value="1234")
    with pytest.raises(AssertionError, match="is not an 8-digit string"):
        assert_tariff_codes_valid(ws, first_data_row=2, last_data_row=3)


def test_assert_tariff_codes_valid_fails_on_letters():
    wb = _simple_ungrouped()
    ws = wb.active
    ws.cell(row=2, column=COL_F_TARIFF, value="ABCDEFGH")
    with pytest.raises(AssertionError, match="is not an 8-digit string"):
        assert_tariff_codes_valid(ws, first_data_row=2, last_data_row=3)


# ─── assert_column_types ────────────────────────────────────────────────


def test_assert_column_types_passes():
    wb = _simple_ungrouped()
    assert_column_types(wb.active, first_data_row=2, last_data_row=3)


def test_assert_column_types_fails_when_L_is_numeric():  # noqa: N802
    wb = _simple_ungrouped()
    ws = wb.active
    ws.cell(row=2, column=COL_L_PER_UNIT, value=42)
    with pytest.raises(AssertionError, match=r"col L \(Per Unit\) must be string/text"):
        assert_column_types(ws, first_data_row=2, last_data_row=3)


def test_assert_column_types_fails_when_O_is_string():  # noqa: N802
    wb = _simple_ungrouped()
    ws = wb.active
    ws.cell(row=2, column=COL_O_COST, value="not a number")
    with pytest.raises(AssertionError, match=r"col O \(Cost\) must be numeric"):
        assert_column_types(ws, first_data_row=2, last_data_row=3)


# ─── assert_invoice_totals_only_on_first_group_row ──────────────────────


def test_assert_invoice_totals_only_on_first_group_row_passes():
    wb = _simple_ungrouped()
    assert_invoice_totals_only_on_first_group_row(wb.active)


def test_assert_invoice_totals_only_on_first_group_row_fails_extra_S():  # noqa: N802
    wb = _simple_ungrouped()
    ws = wb.active
    # Leak the invoice total onto the second detail row
    ws.cell(row=3, column=COL_S_INVOICE_TOTAL, value=999.0)
    with pytest.raises(AssertionError, match="is not the first row of a new invoice block"):
        assert_invoice_totals_only_on_first_group_row(ws)


def test_assert_invoice_totals_only_on_first_group_row_fails_T_without_S():  # noqa: N802
    wb = _simple_ungrouped()
    ws = wb.active
    # Put freight on row 3 without S — T without S is a violation.
    ws.cell(row=3, column=COL_T_FREIGHT, value=5.0)
    with pytest.raises(AssertionError, match="col T is populated"):
        assert_invoice_totals_only_on_first_group_row(ws)


# ─── assert_totals_section_formulas ─────────────────────────────────────


def test_assert_totals_section_formulas_passes_ungrouped():
    wb = _simple_ungrouped()
    assert_totals_section_formulas(wb.active)


def test_assert_totals_section_formulas_passes_grouped():
    wb = _simple_grouped()
    assert_totals_section_formulas(wb.active)


def test_assert_totals_section_formulas_fails_when_net_total_bad():
    wb = _simple_ungrouped()
    ws = wb.active
    from tests.pipeline.invariants import _find_label_row

    net_row = _find_label_row(ws, "NET TOTAL")
    ws.cell(row=net_row, column=COL_P_TOTAL_COST, value="=SUM(X1:X9)")
    with pytest.raises(AssertionError, match="NET TOTAL"):
        assert_totals_section_formulas(ws)


# ─── assert_row_q_r_formulas ────────────────────────────────────────────


def test_assert_row_q_r_formulas_passes_when_empty():
    # The factory leaves Q and R blank — the invariant should skip
    # cleanly rather than flag anything.
    wb = _simple_ungrouped()
    assert_row_q_r_formulas(wb.active, first_data_row=2, last_data_row=3)


def test_assert_row_q_r_formulas_fails_on_wrong_formula():
    wb = _simple_ungrouped()
    ws = wb.active
    ws.cell(row=2, column=COL_Q_TOTAL, value="=O3*K3")  # row mismatch
    with pytest.raises(AssertionError, match=r"col Q expected =O2\*K2"):
        assert_row_q_r_formulas(ws, first_data_row=2, last_data_row=3)


def test_assert_row_q_r_formulas_passes_on_correct_formula():
    wb = _simple_ungrouped()
    ws = wb.active
    ws.cell(row=2, column=COL_Q_TOTAL, value="=O2*K2")
    ws.cell(row=2, column=COL_R_TC_VS_TOTAL, value="=P2-Q2")
    assert_row_q_r_formulas(ws, first_data_row=2, last_data_row=2)


# ─── assert_group_verification_zero ─────────────────────────────────────


def test_assert_group_verification_zero_skips_when_absent():
    # Ungrouped workbook has no GROUP VERIFICATION row — should pass.
    wb = _simple_ungrouped()
    assert_group_verification_zero(wb.active)


def test_assert_group_verification_zero_passes_numeric_zero():
    wb = _simple_grouped()
    ws = wb.active
    # Manually append a GROUP VERIFICATION row with value 0.
    r = ws.max_row + 1
    ws.cell(row=r, column=COL_J_SUPPLIER_ITEM_DESC, value="GROUP VERIFICATION")
    ws.cell(row=r, column=COL_P_TOTAL_COST, value=0.0)
    assert_group_verification_zero(ws)


def test_assert_group_verification_zero_fails_on_nonzero():
    wb = _simple_grouped()
    ws = wb.active
    r = ws.max_row + 1
    ws.cell(row=r, column=COL_J_SUPPLIER_ITEM_DESC, value="GROUP VERIFICATION")
    ws.cell(row=r, column=COL_P_TOTAL_COST, value=5.0)
    with pytest.raises(AssertionError, match="GROUP VERIFICATION"):
        assert_group_verification_zero(ws)


# ─── assert_variance_within_tolerance ───────────────────────────────────


def test_assert_variance_within_tolerance_skips_when_formula_view():
    # openpyxl without data_only=True exposes the formula string — the
    # invariant skips gracefully rather than erroring.
    wb = _simple_ungrouped()
    assert_variance_within_tolerance(wb.active, tolerance=0.01)


def test_assert_variance_within_tolerance_passes_numeric_zero():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=COL_J_SUPPLIER_ITEM_DESC, value="VARIANCE CHECK")
    ws.cell(row=1, column=COL_P_TOTAL_COST, value=0.0)
    assert_variance_within_tolerance(ws, tolerance=0.01)


def test_assert_variance_within_tolerance_fails_out_of_range():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=COL_J_SUPPLIER_ITEM_DESC, value="VARIANCE CHECK")
    ws.cell(row=1, column=COL_P_TOTAL_COST, value=5.0)
    with pytest.raises(AssertionError, match=r"VARIANCE CHECK.*exceeds tolerance"):
        assert_variance_within_tolerance(ws, tolerance=0.01)


# ─── assert_group_row_totals_match_details ──────────────────────────────


def test_assert_group_row_totals_match_details_passes():
    wb = _simple_grouped()
    ws = wb.active
    assert_group_row_totals_match_details(ws, first_group_row=2, last_data_row=ws.max_row)


def test_assert_group_row_totals_match_details_fails_on_mismatch():
    wb = _simple_grouped()
    ws = wb.active
    # Tamper with the first group row P value so it no longer matches
    # the sum of its detail rows.
    ws.cell(row=2, column=COL_P_TOTAL_COST, value=999.99)
    with pytest.raises(AssertionError, match="group row 2"):
        assert_group_row_totals_match_details(ws, first_group_row=2, last_data_row=ws.max_row)


# ─── run_all_invariants aggregate ───────────────────────────────────────


def test_run_all_invariants_passes_on_clean_ungrouped():
    wb = _simple_ungrouped()
    failures = run_all_invariants(wb.active, mode="minimal")
    assert failures == [], f"unexpected failures: {failures}"


def test_run_all_invariants_passes_on_clean_grouped():
    wb = _simple_grouped()
    failures = run_all_invariants(wb.active, mode="minimal")
    assert failures == [], f"unexpected failures: {failures}"


def test_run_all_invariants_collects_multiple_failures():
    wb = _simple_ungrouped()
    ws = wb.active
    # Break two different invariants at once.
    ws.cell(row=2, column=COL_F_TARIFF, value="badcode")  # tariff
    ws.cell(row=2, column=COL_L_PER_UNIT, value=999)  # col type
    failures = run_all_invariants(ws, mode="minimal")
    names = {name for name, _ in failures}
    assert "assert_tariff_codes_valid" in names
    assert "assert_column_types" in names


# ─── sanity: the ORIGINAL three invariants still work ──────────────────


def test_original_variance_is_formula_still_passes():
    wb = _simple_ungrouped()
    assert_variance_is_formula(wb.active)


def test_original_adjustments_still_passes():
    wb = _simple_ungrouped()
    assert_adjustments_has_no_stacked_corrections(wb.active)
