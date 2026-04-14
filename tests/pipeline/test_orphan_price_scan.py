"""Phase 1 — honest OCR orphan-price recovery & uncertainty markers.

These tests guard the behaviour documented in the "Honest reconciliation"
refactor:

* ``FormatParser._scan_orphan_prices`` — recovers standalone price tokens
  from the items section that the strict regex missed, so items_sum +
  recovered ≈ subtotal.  Must never fabricate values; only promotes
  numbers that already exist in the OCR text.
* ``normalize_parse_result`` — propagates ``data_quality`` per item and
  ``data_quality_notes`` at the invoice level through to the pipeline.
* ``bl_xlsx_generator`` — stamps recovered items with a RECOVERED_FILL
  and a cell Comment, and appends a visible "INVOICE NOTES" row when
  data_quality_notes are present.
* ``variance_fixer.fix_variance(honest_mode=True)`` — absorbs residual
  variance into the ADJUSTMENTS row instead of scaling per-item cells.
"""

from __future__ import annotations

import os
import sys

import pytest
import yaml

_PIPELINE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "pipeline"))
if _PIPELINE_DIR not in sys.path:
    sys.path.insert(0, _PIPELINE_DIR)

from format_parser import FormatParser  # noqa: E402
from stages.supplier_resolver import normalize_parse_result  # noqa: E402

SHEIN_SPEC_PATH = os.path.abspath(
    os.path.join(
        os.path.dirname(__file__),
        "..",
        "..",
        "config",
        "formats",
        "shein_us_invoice.yaml",
    )
)


# Raw OCR text for ANDREA_L invoice 8 (INVUS20240728002522185).
# Kept inline so this test is self-contained and does not depend on
# /tmp scratch files.
INV8_RAW_TEXT = """8/7/24, 2:17 PM 49aquwirhtdhzq0o1bbehpyni.jpg

777 S Alameda St Fl 2, Los Angeles, CA 90021-1633

: Ss SHEIN US Services, LLC. Sales Invoice

Bill To: Invoice No.:
Andrea Lord INVUS20240728002522185
Billing Address: Invoice Date:
10813 NW 30th Street BLDG 115 GRE 4387, Miami, FLORIDA 2024-07-28
33172-2191

Delivery Address:

10813 NW 30th Street BLDG 115 GRE 4387, Miami, FLORIDA

33172-2191

Order Information

Order Number: Order Date:
GSUNJG55TOOQV7W 2024-07-28

Seller Information

Sold By: Address:
SHEIN DISTRIBUTION CORPORATION United StatesCALIFORNIALos Angeles777 S. Alamaeda St.
Suite 340, Los Angeles, CA 90021

~

Invoice Detail WY,

Description Quantity Amount(USD)

SHEIN Lady Plus Size Solid Color Round Neck Hollow Out Short Sleeve Top With Ruffle Hem And Geo

: eee tia 13.70
Printed Color Block Knit Skirt Set

SHEIN Lady Women's High Waist Pleated Detail Skirt 1 4.60

SHEIN Lady Women's Elegant Marble Printed Bodycon Skirt, Summer 1 7.54

SHEIN Lady Summer Outfitsmaxi Women Dresseswedding Guest Summerpink Women Dresses Navy Blue  1 7.06
Skirt Going Out Outfits i

Item(s) Subtotal: 33.10

Shipping/Handling: 0.00

Sales Tax: 2.33

Grand Total: 35.43
"""


@pytest.fixture(scope="module")
def shein_spec() -> dict:
    with open(SHEIN_SPEC_PATH) as f:
        return yaml.safe_load(f)


@pytest.fixture
def parsed_inv8(shein_spec: dict) -> dict:
    parser = FormatParser(shein_spec)
    return parser.parse(INV8_RAW_TEXT)


# ── Format parser orphan-scan tests ─────────────────────────────


def test_orphan_scan_recovers_13_70(parsed_inv8: dict) -> None:
    """The $13.70 line item orphaned by OCR must be recovered honestly."""
    inv = parsed_inv8["invoices"][0]
    recovered = [it for it in inv["items"] if it.get("data_quality") == "orphan_price_recovered"]
    assert len(recovered) == 1, (
        f"expected exactly 1 orphan-recovered item, got {len(recovered)}: "
        f"{[(i.get('sku'), i.get('total_cost')) for i in inv['items']]}"
    )
    rec = recovered[0]
    assert rec["total_cost"] == 13.70
    assert rec["unit_cost"] == 13.70
    assert rec["quantity"] == 1


def test_orphan_scan_items_sum_matches_subtotal(parsed_inv8: dict) -> None:
    """After recovery, items_sum should land within tolerance of subtotal."""
    inv = parsed_inv8["invoices"][0]
    items_sum = sum(it.get("total_cost", 0) or 0 for it in inv["items"])
    subtotal = inv.get("sub_total") or 0
    # 4.60 + 7.54 + 7.06 + 13.70 = 32.90 vs 33.10 subtotal → $0.20 residual
    assert abs(items_sum - subtotal) <= 0.25, (
        f"items_sum ${items_sum:.2f} vs subtotal ${subtotal:.2f}"
    )


def test_orphan_scan_writes_quality_notes(parsed_inv8: dict) -> None:
    """Orphan recovery must be visible in data_quality_notes."""
    inv = parsed_inv8["invoices"][0]
    notes = inv.get("data_quality_notes") or []
    assert any("13.70" in n for n in notes), f"expected a 13.70 recovery note, got: {notes}"


def test_orphan_scan_does_not_double_count_existing_prices(
    shein_spec: dict,
) -> None:
    """When items already sum to subtotal, no orphan must be injected."""
    clean_text = """Invoice Detail

Description Quantity Amount(USD)

Widget A 1 10.00
Widget B 1 20.00

Item(s) Subtotal: 30.00
Sales Tax: 0.00
Grand Total: 30.00

Invoice No.: INVUS1234567890
Invoice Date: 2026-01-01
Order Number: XYZ123
"""
    parser = FormatParser(shein_spec)
    result = parser.parse(clean_text)
    inv = result["invoices"][0]
    assert all(it.get("data_quality", "") != "orphan_price_recovered" for it in inv["items"]), (
        "clean invoice triggered unexpected orphan recovery"
    )
    # Structural item count mismatch notes are expected (short descriptions
    # below min_description_length cause 0 extracted items vs 2 price lines).
    # Only check that no *orphan recovery* notes leaked through.
    orphan_notes = [n for n in inv.get("data_quality_notes", []) if "orphan" in n.lower()]
    assert not orphan_notes, f"clean invoice produced orphan-scan notes: {orphan_notes}"


# ── Normalization propagation ───────────────────────────────────


def test_normalize_propagates_data_quality(parsed_inv8: dict) -> None:
    norm = normalize_parse_result(parsed_inv8)
    assert norm["invoice_total"] == 35.43
    assert any(it.get("data_quality") == "orphan_price_recovered" for it in norm["items"])
    assert norm.get("data_quality_notes"), "data_quality_notes missing after normalization"


# ── BL XLSX uncertainty markers ─────────────────────────────────


def test_bl_xlsx_marks_recovered_cells(parsed_inv8: dict, tmp_path) -> None:
    """Recovered items must get RECOVERED_FILL + comment on O/P cells,
    and the sheet must carry an INVOICE NOTES row."""
    import openpyxl
    from bl_xlsx_generator import generate_bl_xlsx

    invoice_data = normalize_parse_result(parsed_inv8)
    matched = []
    for idx, it in enumerate(invoice_data["items"], 1):
        matched.append(
            {
                "po_item_ref": "",
                "po_item_desc": "",
                "po_number": "",
                "supplier_item": it["supplier_item"] or f"ITEM-{idx}",
                "supplier_item_desc": it["description"],
                "quantity": it["quantity"],
                "unit_price": it["unit_price"],
                "total_cost": it["total"],
                "uom": "",
                "match_score": 0,
                "tariff_code": "99999999",
                "data_quality": it.get("data_quality", ""),
            }
        )

    xlsx_path = str(tmp_path / "inv8.xlsx")
    generate_bl_xlsx(
        invoice_data,
        matched,
        "SHEIN US Services, LLC",
        {"code": "SHEIN", "name": "SHEIN US Services, LLC"},
        xlsx_path,
        document_type="4000-000",
    )

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Find the recovered detail row — it's the row whose P cell is 13.70
    recovered_row = None
    for r in range(2, ws.max_row + 1):
        p_val = ws.cell(row=r, column=16).value
        if isinstance(p_val, (int, float)) and abs(float(p_val) - 13.70) < 0.01:
            # Skip group-fill rows (D9E1F2)
            fill = ws.cell(row=r, column=1).fill.start_color.rgb or ""
            if "D9E1F2" not in str(fill):
                recovered_row = r
                break
    assert recovered_row is not None, "no recovered detail row with P=13.70"

    o_cell = ws.cell(row=recovered_row, column=15)
    p_cell = ws.cell(row=recovered_row, column=16)
    o_fill = str(o_cell.fill.start_color.rgb or "")
    p_fill = str(p_cell.fill.start_color.rgb or "")
    assert "FFF2CC" in o_fill, f"recovered O cell missing fill: {o_fill!r}"
    assert "FFF2CC" in p_fill, f"recovered P cell missing fill: {p_fill!r}"
    assert o_cell.comment is not None
    assert p_cell.comment is not None
    assert "orphan_price_recovered" in p_cell.comment.text

    # INVOICE NOTES row must exist
    notes_found = False
    for r in range(2, ws.max_row + 1):
        label = ws.cell(row=r, column=10).value
        if label and "INVOICE NOTES" in str(label).upper():
            notes_found = True
            assert "13.70" in str(label)
            break
    assert notes_found, "INVOICE NOTES row not written"


# ── Honest-mode variance_fixer ──────────────────────────────────


def test_variance_fixer_honest_mode_absorbs_into_adjustments(grouped_workbook, tmp_path) -> None:
    """honest_mode must NOT call the LLM; residual goes into ADJUSTMENTS."""
    import openpyxl
    from workflow.variance_fixer import fix_variance

    # Build a grouped workbook with a deliberate $5.00 variance.
    wb = grouped_workbook(
        inv_total=100.00,
        groups=[("TEST", [("Item A", 1, 45.00), ("Item B", 1, 50.00)])],
        freight=0,
        insurance=0,
        tax=0,
        deduction=0,
    )
    xlsx_path = str(tmp_path / "variance.xlsx")
    wb.save(xlsx_path)

    result = fix_variance(
        xlsx_path=xlsx_path,
        invoice_text="dummy text (honest mode must not use LLM)",
        current_variance=5.00,
        honest_mode=True,
    )
    assert result.get("success") is True
    assert result.get("fixes_applied") == 0
    assert result.get("new_variance") == 0.00
    assert "Honest mode" in result.get("analysis", "")

    # The P/O cells on the detail rows must be untouched (no hallucinated
    # scaling).  The ADJUSTMENTS row's stored formula encodes the residual.
    wb2 = openpyxl.load_workbook(xlsx_path)
    ws2 = wb2.active
    # Detail rows are rows 3-4 in the grouped factory (row 2 is the group).
    unit_vals = [ws2.cell(row=r, column=15).value for r in (3, 4)]
    total_vals = [ws2.cell(row=r, column=16).value for r in (3, 4)]
    assert unit_vals == [45.00, 50.00], f"detail O values were modified: {unit_vals}"
    assert total_vals == [45.00, 50.00], f"detail P values were modified: {total_vals}"
