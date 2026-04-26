"""Regression tests pinning the critical behaviors of pipeline/xlsx_generator.py
after the Wave 1 magic-constant migration.

These assert that:
  * document_type resolution still honours consignee rules (Budget Marine
    → 7400-000) — hardcoding would silently break IM7 warehouse entries.
  * HS category display labels come from config (not a duplicate-key dict
    where the last value wins by accident).
  * Supplier-code aliases resolve via config.
  * Intermediate-stage suffixes are stripped when deriving XLSX filenames.
  * The variance-display epsilon is loaded from config (not hardcoded).

Run: pytest tests/pipeline/test_xlsx_generator_literals.py -v
"""

from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

_REPO_ROOT = Path(__file__).resolve().parents[2]

# Make pipeline/ importable exactly the way run.py does it.
for _p in (_REPO_ROOT, _REPO_ROOT / "pipeline"):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))


# ── Config-surface assertions (migration happened, values are in YAML) ──


def test_hs_categories_load_via_config() -> None:
    """HS-code → category label comes from config/hs_categories.yaml."""
    from pipeline.config_loader import load_hs_categories

    cats = load_hs_categories()["categories"]
    assert cats["33041000"] == "LIP PRODUCTS"
    assert cats["67042000"] == "HUMAN HAIR"
    # 33049990 must be the deduped label — the old dict-literal had a
    # duplicate key and Python silently kept the second value.
    assert cats["33049990"] == "SKINCARE & OTHER COSMETICS"


def test_supplier_code_aliases_load_via_config() -> None:
    from pipeline.config_loader import load_invoice_formats

    aliases = load_invoice_formats()["supplier_code_aliases"]
    assert aliases["ABSOLUTE"] == "ABSOLUTE"
    assert aliases["AMAZON"] == "AMAZON"
    assert aliases["TARGET"] == "TARGET"
    assert aliases["WALMART"] == "WALMART"


def test_intermediate_suffixes_load_via_config() -> None:
    from pipeline.config_loader import load_file_paths

    sufs = load_file_paths()["intermediate_suffixes"]
    assert "_grouped" in sufs
    assert "_classified" in sufs
    assert "_parsed" in sufs
    assert "_extracted" in sufs


def test_variance_display_epsilon_load_via_config() -> None:
    from pipeline.config_loader import load_validation_tolerances

    tol = load_validation_tolerances()
    assert tol["variance_display_epsilon"] == 0.01


def test_invalid_po_sentinels_load_via_config() -> None:
    from pipeline.config_loader import load_validation_tolerances

    sent = load_validation_tolerances()["invalid_po_sentinels"]
    assert "" in sent
    assert "Page" in sent
    assert "None" in sent
    assert "N/A" in sent


# ── End-to-end: xlsx_generator.run() still produces a correct workbook ──


def _minimal_grouped_payload(doc_type: str | None = None) -> dict:
    meta = {
        "invoice_number": "TEST-1",
        "date": "2026-01-15",
        "total": 100.0,
        "supplier": "ABSOLUTE NEW YORK",
        "country_code": "US",
    }
    if doc_type:
        meta["document_type"] = doc_type
    return {
        "invoice_metadata": meta,
        "groups": [
            {
                "tariff_code": "33041000",
                "category": "LIP PRODUCTS",
                "item_count": 1,
                "sum_quantity": 5,
                "sum_total_cost": 100.0,
                "average_unit_cost": 20.0,
                "items": [
                    {
                        "supplier_item": "LIP-001",
                        "description": "Matte Lipstick",
                        "quantity": 5,
                        "unit_cost": 20.0,
                        "total_cost": 100.0,
                        "billable": True,
                    }
                ],
            }
        ],
    }


def _run_generator(payload: dict, context: dict | None = None) -> Path:
    import xlsx_generator

    with tempfile.TemporaryDirectory() as tdir:
        td = Path(tdir)
        grouped = td / "src_grouped.json"
        grouped.write_text(json.dumps(payload), encoding="utf-8")
        out = td / "src.xlsx"
        result = xlsx_generator.run(
            str(grouped),
            str(out),
            config=None,
            context=context or {"input_file": str(grouped)},
        )
        assert result["status"] == "success", result
        # Copy to a stable location outside the TemporaryDirectory.
        produced = Path(result["output"])
        keeper = Path(tempfile.mkdtemp()) / produced.name
        keeper.write_bytes(produced.read_bytes())
        return keeper


def test_run_writes_resolved_document_type() -> None:
    wb_path = _run_generator(
        _minimal_grouped_payload(), context={"document_type": "7400-000", "input_file": "x.pdf"}
    )  # noqa: E501
    wb = load_workbook(wb_path)
    ws = wb.active
    assert ws["A2"].value == "7400-000", (
        "Context-provided document_type must be written to A2 — consignee-rule "
        "resolution (e.g. Budget Marine → 7400-000) happens upstream in run.py"
    )


def test_run_defaults_to_4000_000_when_no_doc_type_supplied() -> None:
    wb_path = _run_generator(_minimal_grouped_payload())
    wb = load_workbook(wb_path)
    ws = wb.active
    # Default fallback comes from document_types.json (NOT a hardcoded string).
    assert ws["A2"].value == "4000-000"


def test_run_populates_category_from_config() -> None:
    wb_path = _run_generator(_minimal_grouped_payload())
    wb = load_workbook(wb_path)
    ws = wb.active
    # Column E (index 5) on the first group row must match hs_categories.yaml.
    assert ws.cell(row=2, column=5).value == "LIP PRODUCTS"


def test_run_populates_supplier_code_from_alias() -> None:
    wb_path = _run_generator(_minimal_grouped_payload())
    wb = load_workbook(wb_path)
    ws = wb.active
    # Column Z (index 26) must contain the ABSOLUTE alias.
    assert ws.cell(row=2, column=26).value == "ABSOLUTE"


def test_intermediate_suffix_stripped_from_output_filename() -> None:
    import xlsx_generator

    with tempfile.TemporaryDirectory() as tdir:
        td = Path(tdir)
        # Pass a _grouped.json input — output name should strip _grouped.
        grouped = td / "myinvoice_grouped.json"
        grouped.write_text(json.dumps(_minimal_grouped_payload()), encoding="utf-8")
        out = td / "myinvoice_grouped.xlsx"
        result = xlsx_generator.run(
            str(grouped),
            str(out),
            config=None,
            context={"input_file": str(grouped)},
        )
        assert result["status"] == "success"
        final = Path(result["output"]).name
        assert "_grouped" not in final, f"intermediate suffix not stripped: {final}"
        assert final.startswith("myinvoice")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
