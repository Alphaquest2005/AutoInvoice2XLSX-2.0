"""Synthetic XLSX workbook factories for pipeline regression tests.

These helpers produce small, inspectable workbooks that are byte-compatible
with what ``pipeline/bl_xlsx_generator.py`` and ``pipeline/xlsx_generator.py``
emit, so the validator, variance fixer, and shipment checklist all see the
same layout they see in production.

Column layout matches ``pipeline/xlsx_validator.COL_*`` constants:

    F  (6)  TariffCode
    J  (10) Supplier Item Description
    K  (11) Quantity
    L  (12) Per Unit (text)
    O  (15) Cost (unit price, numeric)
    P  (16) Total Cost
    S  (19) InvoiceTotal            (row 2 only)
    T  (20) Freight                 (row 2 only)
    U  (21) Insurance               (row 2 only)
    V  (22) Tax                     (row 2 only)
    W  (23) Deduction               (row 2 only)
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

if TYPE_CHECKING:
    from collections.abc import Iterable

    from openpyxl.worksheet.worksheet import Worksheet

# Column indices (1-based)
COL_TARIFF = 6
COL_DESC = 10
COL_QTY = 11
COL_PER_UNIT = 12
COL_COST = 15
COL_TOTAL_COST = 16
COL_INV_TOTAL = 19
COL_FREIGHT = 20
COL_INSURANCE = 21
COL_TAX = 22
COL_DEDUCTION = 23

# Blue fill matches the "group row" fill color ``bl_xlsx_generator`` uses,
# so variance_fixer._is_group_row() detects these rows correctly.
GROUP_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _write_headers(ws: Worksheet) -> None:
    """Write the minimal set of headers the validator / fixer look for."""
    ws.cell(row=1, column=COL_TARIFF, value="TariffCode")
    ws.cell(row=1, column=COL_DESC, value="Supplier Item Description")
    ws.cell(row=1, column=COL_QTY, value="Quantity")
    ws.cell(row=1, column=COL_PER_UNIT, value="Per Unit")
    ws.cell(row=1, column=COL_COST, value="Cost")
    ws.cell(row=1, column=COL_TOTAL_COST, value="Total Cost")
    ws.cell(row=1, column=COL_INV_TOTAL, value="InvoiceTotal")
    ws.cell(row=1, column=COL_FREIGHT, value="Freight")
    ws.cell(row=1, column=COL_INSURANCE, value="Insurance")
    ws.cell(row=1, column=COL_TAX, value="Tax")
    ws.cell(row=1, column=COL_DEDUCTION, value="Deduction")


def _write_item_row(
    ws: Worksheet,
    row: int,
    description: str,
    qty: float,
    unit_price: float,
    tariff: str = "98765432",
) -> None:
    ws.cell(row=row, column=COL_TARIFF, value=tariff)
    ws.cell(row=row, column=COL_DESC, value=description)
    ws.cell(row=row, column=COL_QTY, value=qty)
    ws.cell(row=row, column=COL_PER_UNIT, value=f"    {description}")  # text
    ws.cell(row=row, column=COL_COST, value=unit_price)  # numeric unit price
    ws.cell(row=row, column=COL_TOTAL_COST, value=round(qty * unit_price, 2))


def build_ungrouped_workbook(
    *,
    inv_total: float,
    items: Iterable[tuple[str, float, float]],
    freight: float = 0.0,
    insurance: float = 0.0,
    tax: float = 0.0,
    deduction: float = 0.0,
) -> Workbook:
    """Build an ungrouped-mode workbook with plain ``SUBTOTAL`` labels.

    Args:
        inv_total: Value written to ``S2`` (InvoiceTotal on the first data row).
        items: Iterable of ``(description, qty, unit_price)`` detail rows.
        freight/insurance/tax/deduction: Written to ``T2``/``U2``/``V2``/``W2``.

    The totals rows use the labels defined under ``ungrouped_totals_section``
    in ``config/grouping.yaml``:

        SUBTOTAL         — numeric (sum of detail P cells)
        ADJUSTMENTS      — formula ``=(T2+U2+V2-W2)``
        NET TOTAL        — formula ``=(P{sub}+P{adj})``
        VARIANCE CHECK   — formula ``=(S2-P{net})``
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Invoice"
    _write_headers(ws)

    # Detail item rows start at row 2
    first_data_row = 2
    row = first_data_row
    items_list = list(items)
    for desc, qty, price in items_list:
        _write_item_row(ws, row, desc, qty, price)
        row += 1
    last_data_row = row - 1

    # Invoice-level totals live on the first data row (S2, T2, ...)
    ws.cell(row=first_data_row, column=COL_INV_TOTAL, value=inv_total)
    ws.cell(row=first_data_row, column=COL_FREIGHT, value=freight)
    ws.cell(row=first_data_row, column=COL_INSURANCE, value=insurance)
    ws.cell(row=first_data_row, column=COL_TAX, value=tax)
    ws.cell(row=first_data_row, column=COL_DEDUCTION, value=deduction)

    # Totals section (matches grouping.yaml ungrouped_totals_section)
    subtotal_row = last_data_row + 1
    adj_row = subtotal_row + 1
    net_row = subtotal_row + 2
    var_row = subtotal_row + 3

    # SUBTOTAL: numeric sum of detail items, as a plain value.  The validator
    # must NOT interpret this plain "SUBTOTAL" label as grouped-mode.
    ws.cell(row=subtotal_row, column=COL_DESC, value="SUBTOTAL")
    ws.cell(
        row=subtotal_row,
        column=COL_TOTAL_COST,
        value=round(sum(q * p for _, q, p in items_list), 2),
    )

    # ADJUSTMENTS: formula referencing row 2 of T/U/V/W
    ws.cell(row=adj_row, column=COL_DESC, value="ADJUSTMENTS")
    ws.cell(row=adj_row, column=COL_TOTAL_COST, value="=(T2+U2+V2-W2)")

    # NET TOTAL: formula
    ws.cell(row=net_row, column=COL_DESC, value="NET TOTAL")
    ws.cell(
        row=net_row,
        column=COL_TOTAL_COST,
        value=f"=(P{subtotal_row}+P{adj_row})",
    )

    # VARIANCE CHECK: formula, bold red font
    ws.cell(row=var_row, column=COL_DESC, value="VARIANCE CHECK")
    ws.cell(
        row=var_row,
        column=COL_TOTAL_COST,
        value=f"=(S2-P{net_row})",
    )
    ws.cell(row=var_row, column=COL_TOTAL_COST).font = Font(bold=True, color="FF0000")

    return wb


def build_grouped_workbook(
    *,
    inv_total: float,
    groups: Iterable[tuple[str, list[tuple[str, float, float]]]],
    freight: float = 0.0,
    insurance: float = 0.0,
    tax: float = 0.0,
    deduction: float = 0.0,
) -> Workbook:
    """Build a grouped-mode workbook with ``SUBTOTAL (GROUPED)`` labels.

    Args:
        inv_total: Written to ``S2`` on the first group row.
        groups: Iterable of ``(category_name, [(desc, qty, unit_price), ...])``.
            Each category becomes a group header row (with ``(N items)`` in J
            and the category's total in P) followed by its detail rows.
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Invoice"
    _write_headers(ws)

    row = 2
    first_group_row: int | None = None
    for category, details in groups:
        group_row = row
        if first_group_row is None:
            first_group_row = group_row
        group_total = round(sum(q * p for _, q, p in details), 2)
        ws.cell(row=group_row, column=COL_DESC, value=f"{category} ({len(details)} items)")
        ws.cell(row=group_row, column=COL_TOTAL_COST, value=group_total)
        ws.cell(row=group_row, column=1).fill = GROUP_FILL
        row += 1
        for desc, qty, price in details:
            _write_item_row(ws, row, desc, qty, price)
            row += 1

    assert first_group_row is not None
    last_data_row = row - 1

    # Invoice-level totals on the first group row (S2, T2, ...)
    ws.cell(row=first_group_row, column=COL_INV_TOTAL, value=inv_total)
    ws.cell(row=first_group_row, column=COL_FREIGHT, value=freight)
    ws.cell(row=first_group_row, column=COL_INSURANCE, value=insurance)
    ws.cell(row=first_group_row, column=COL_TAX, value=tax)
    ws.cell(row=first_group_row, column=COL_DEDUCTION, value=deduction)

    subtotal_row = last_data_row + 1
    adj_row = subtotal_row + 1
    net_row = subtotal_row + 2
    var_row = subtotal_row + 3

    # SUBTOTAL (GROUPED): sum of group-header totals
    ws.cell(row=subtotal_row, column=COL_DESC, value="SUBTOTAL (GROUPED)")
    group_sum = round(sum(round(sum(q * p for _, q, p in details), 2) for _, details in groups), 2)
    ws.cell(row=subtotal_row, column=COL_TOTAL_COST, value=group_sum)

    ws.cell(row=adj_row, column=COL_DESC, value="ADJUSTMENTS")
    ws.cell(row=adj_row, column=COL_TOTAL_COST, value="=(T2+U2+V2-W2)")

    ws.cell(row=net_row, column=COL_DESC, value="NET TOTAL")
    ws.cell(
        row=net_row,
        column=COL_TOTAL_COST,
        value=f"=(P{subtotal_row}+P{adj_row})",
    )

    ws.cell(row=var_row, column=COL_DESC, value="VARIANCE CHECK")
    ws.cell(
        row=var_row,
        column=COL_TOTAL_COST,
        value=f"=(S2-P{net_row})",
    )
    ws.cell(row=var_row, column=COL_TOTAL_COST).font = Font(bold=True, color="FF0000")

    return wb
